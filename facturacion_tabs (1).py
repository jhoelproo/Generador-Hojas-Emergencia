# facturacion_tabs.py
# Sistema de Formularios de Emergencia - Hospital General
# Version: 4.1.8 - Usuarios de turno administrables y escritura libre
# Python 3.14 compatible
import os
import re
import sys
import json
import sqlite3
from difflib import SequenceMatcher, get_close_matches
from contextlib import closing
from datetime import datetime, timedelta, date, time
import platform
import subprocess
import tempfile
import shutil
import zipfile
import zlib
import hashlib
import ctypes
import time as _time
import threading
import logging
from logging.handlers import RotatingFileHandler

SELF_TEST_MODE = "--self-test" in sys.argv
SELF_TEST_DATA_DIR = ""
if SELF_TEST_MODE:
    SELF_TEST_DATA_DIR = tempfile.mkdtemp(prefix="generador_hojas_selftest_")
    os.environ["EMERGENCIAS_DATA_DIR"] = SELF_TEST_DATA_DIR

from emergency_core.backup import BackupManager
from emergency_core.db_migrations import LATEST_SCHEMA_VERSION, migrate_database
from emergency_core.io_utils import ConfigError, atomic_write_json, load_json_file
from emergency_core.paths import data_root, harden_windows_acl, migrate_legacy_files
from emergency_core.security import AdminSecurity, SecurityError
from emergency_core.updater import APP_VERSION, UpdateError, get_latest_release, is_newer

# PDF / Excel
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from PyPDF2 import PdfReader, PdfWriter
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font as XLFont, Alignment as XLAlignment, Border as XLBorder, Side as XLSide

# UI
import tkinter as tk
from tkinter import messagebox, ttk, Toplevel, filedialog, simpledialog

# Temas
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.widgets import DateEntry as TBDateEntry


# ============================================================
# FASE 13 - Colores institucionales sobrios
# ============================================================
COLOR_PRIMARY = "#2563EB"
COLOR_SUCCESS = "#1F7A4D"
COLOR_WARNING = "#A16207"
COLOR_DANGER  = "#B42318"
COLOR_INFO    = "#0E7490"


# -------------------------------
# HELPERS DE RUTAS PARA .PY Y .EXE
# -------------------------------
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def app_data_path(*paths):
    final_path = os.path.join(str(data_root()), *paths)

    if os.path.splitext(final_path)[1]:
        os.makedirs(os.path.dirname(final_path), exist_ok=True)
    else:
        os.makedirs(final_path, exist_ok=True)

    return final_path


def output_report_path(filename):
    return app_data_path("REPORTES", filename)


# -------------------------------
# ARCHIVO DIARIO DE REPORTES Y LISTADOS
# -------------------------------
MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
}

DIAS_ES = {
    0: "lunes", 1: "martes", 2: "miércoles", 3: "jueves",
    4: "viernes", 5: "sábado", 6: "domingo"
}


def limpiar_nombre_archivo(nombre: str) -> str:
    """
    Limpia caracteres no permitidos en nombres de carpetas/archivos de Windows.
    Mantiene acentos y espacios para que el nombre se vea natural en español.
    """
    nombre = (nombre or "").strip()
    nombre = re.sub(r'[<>:"/\\|?*]+', " ", nombre)
    nombre = re.sub(r"\s+", " ", nombre).strip()
    return nombre


def fecha_larga_es(fecha_base: date) -> str:
    return f"{DIAS_ES[fecha_base.weekday()]} {fecha_base.day:02d} de {MESES_ES[fecha_base.month]} de {fecha_base.year}"


def etiqueta_turno_archivo(turno_cfg: dict) -> str:
    """
    Para días normales devuelve solo la fecha en español.
    Para turnos divididos, especialmente domingo, agrega diurno/nocturno.
    """
    if not turno_cfg:
        return datetime.now().strftime("%d-%m-%Y")

    fecha_base = turno_cfg.get("fecha_base")
    if not isinstance(fecha_base, date):
        fecha_base = fecha_base_operativa_actual()

    base = fecha_larga_es(fecha_base)
    codigo = normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM"))

    if codigo == "8AM_8PM":
        return f"{base} - diurno"
    if codigo == "8PM_8AM":
        return f"{base} - nocturno"

    return base


def carpeta_archivo_turno(turno_cfg: dict) -> str:
    etiqueta = limpiar_nombre_archivo(etiqueta_turno_archivo(turno_cfg))
    carpeta = os.path.join(ARCHIVO_DIARIO_DIR, etiqueta)
    os.makedirs(carpeta, exist_ok=True)
    return carpeta


def ruta_unica_si_existe(ruta: str) -> str:
    """
    Evita sobrescribir si por error se genera más de una vez el mismo archivo.
    """
    if not os.path.exists(ruta):
        return ruta

    base, ext = os.path.splitext(ruta)
    ts = datetime.now().strftime("%H%M%S")
    candidato = f"{base}_{ts}{ext}"
    contador = 2
    while os.path.exists(candidato):
        candidato = f"{base}_{ts}_{contador}{ext}"
        contador += 1
    return candidato


def guardar_copia_reporte_turno(ruta_pdf: str, turno_cfg: dict) -> str:
    """
    Guarda una copia organizada del reporte PDF del turno saliente.
    """
    if not ruta_pdf or not os.path.exists(ruta_pdf) or not turno_cfg:
        return ""

    carpeta = carpeta_archivo_turno(turno_cfg)
    etiqueta = limpiar_nombre_archivo(etiqueta_turno_archivo(turno_cfg))
    destino = ruta_unica_si_existe(os.path.join(carpeta, f"Reporte - {etiqueta}.pdf"))
    shutil.copy2(ruta_pdf, destino)
    return destino


def guardar_copia_excel_turno(turno_cfg: dict, ruta_excel=None) -> str:
    """
    Guarda una copia del listado Excel actual ANTES de limpiarlo/reconstruirlo.
    Se usa ruta_excel=None para evitar que EXCEL_PATH se evalúe antes de definirse.
    """
    if ruta_excel is None:
        ruta_excel = EXCEL_PATH

    if not turno_cfg or not os.path.exists(ruta_excel):
        return ""

    if not excel_tiene_registros(ruta_excel):
        return ""

    carpeta = carpeta_archivo_turno(turno_cfg)
    etiqueta = limpiar_nombre_archivo(etiqueta_turno_archivo(turno_cfg))
    destino = ruta_unica_si_existe(os.path.join(carpeta, f"Listado de pacientes - {etiqueta}.xlsx"))
    shutil.copy2(ruta_excel, destino)
    return destino


# -------------------------------
# RUTAS
# -------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

HOJAS_DIR = resource_path("HOJAS")
LOGO_PATH = resource_path("logo.jpg")

REPORTES_DIR = app_data_path("REPORTES")
ARCHIVO_DIARIO_DIR = app_data_path("ARCHIVO_DIARIO")
TURNOS_CFG = app_data_path("turnos_config.json")
EXCEL_PATH = app_data_path("LISTADO DE PACIENTES EN EMERGENCIA.xlsx")
APP_SETTINGS_PATH = app_data_path("app_settings.json")
ARS_CATALOGO_PATH = app_data_path("ars_catalogo.json")
NSS_FORMATOS_PATH = app_data_path("nss_formatos_ars.json")
REPRESENTANTES_PATH = app_data_path("representantes.json")
SECURITY_CONFIG_PATH = app_data_path("security.json")
LOGS_DIR = app_data_path("LOGS")
BACKUPS_DIR = app_data_path("BACKUPS")
DOCUMENTOS_DIR = app_data_path("DOCUMENTOS")
SUMATRA_PATH_CACHE = None
RESUMEN_TURNO_PATH = app_data_path("resumen_turno.json")

APP_LOG = logging.getLogger("emergencias")
APP_LOG.setLevel(logging.INFO)
if not APP_LOG.handlers:
    _log_path = os.path.join(LOGS_DIR, "app.log")
    try:
        _log_handler = RotatingFileHandler(
            _log_path,
            maxBytes=5 * 1024 * 1024,
            backupCount=5,
            encoding="utf-8",
        )
    except PermissionError:
        harden_windows_acl(data_root())
        try:
            _log_handler = RotatingFileHandler(
                _log_path,
                maxBytes=5 * 1024 * 1024,
                backupCount=5,
                encoding="utf-8",
            )
        except (OSError, PermissionError):
            _fallback_log = os.path.join(tempfile.gettempdir(), "GeneradorHojasEmergencia.log")
            _log_handler = RotatingFileHandler(
                _fallback_log,
                maxBytes=2 * 1024 * 1024,
                backupCount=2,
                encoding="utf-8",
            )
    _log_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    APP_LOG.addHandler(_log_handler)
APP_LOG.propagate = False

RUTA_HOJAS = {
    "GENERAL": os.path.join(HOJAS_DIR, "EMERGENCIA GENERAL.pdf"),
    "GINECOLOGIA": os.path.join(HOJAS_DIR, "EMERGENCIA GINECOLOGIA.pdf"),
    "PEDIATRIA": os.path.join(HOJAS_DIR, "EMERGENCIA PEDIATRICA.pdf"),
}

for key, ruta in RUTA_HOJAS.items():
    if not os.path.exists(ruta):
        print(f"[AVISO] Plantilla no encontrada para {key}: {ruta}")
    else:
        print(f"[OK] Plantilla {key} -> {ruta}")


# -------------------------------
# AJUSTE GLOBAL PARA REDUCCIÓN DE TEXTO
# -------------------------------
THRESHOLD_LEN = 38


# -------------------------------
# SEGUROS
# -------------------------------
SEGUROS_DISPLAY = {
    "SENASA SUBSIDIADO": "SUB",
    "SENASA CONTRIBUTIVO": "S.CONTRIBUTIVO",
    "SENASA PENSIONADOS": "S.PENSIONADOS",
    "APS": "APS",
    "ASEMAP": "ASEMAP",
    "CMD": "CMD",
    "GMA": "GMA",
    "RENACER": "RENACER",
    "RESERVAS": "RESERVAS",
    "SEMMA": "SEMMA",
    "FUTURO": "FUTURO",
    "HUMANO": "HUMANO",
    "PRIMERA": "PRIMERA",
    "ABEL GONZALEZ/SIMAG": "ABEL GONZALEZ/SIMAG",
    "METASALUD": "METASALUD",
    "MONUMENTAL": "MONUMENTAL",
    "MAPFRE/PALIC": "MAPFRE/PALIC",
    "UNIVERSAL": "UNIVERSAL",
    "BANCO CENTRAL": "BANCO CENTRAL",
    "YUNEN": "YUNEN",
    "SIN SEGURO": "SIN SEGURO",
}


# Catálogo editable de ARS y alias. Permite corregir equivalencias sin tocar el código.
DEFAULT_ARS_CATALOGO = {
    "SENASA SUBSIDIADO": [
        "SUB", "SUBS", "SUBSI", "SUBSID", "SUBSIDIADO", "SUNB",
        "SENASA SUBSIDIADO", "SENASA SUB", "SENASA REGIMEN SUBSIDIADO",
        "ARS SENASA SUBSIDIADO", "SESANA SUBSIADO", "SENASA SUBSIADO", "SUBSIADO"
    ],
    "SENASA CONTRIBUTIVO": [
        "CONTRIBUTIVO", "CONTRIB", "CONTRI", "CONT", "COTIZANTE",
        "SENASA CONTRIBUTIVO", "SENASA CONTRIB", "ARS SENASA CONTRIBUTIVO",
        "SENASA AVANZADA", "AVANZADA",
        "SENASA MAXIMO", "MAXIMO", "MÁXIMO",
        "SENASA ESPECIAL", "ESPECIAL"
    ],
    "SENASA PENSIONADOS": [
        "PENSIONADO", "PENSIONADOS", "PENS",
        "SENASA PENSIONADO", "SENASA PENSIONADOS", "ARS SENASA PENSIONADOS"
    ],
    "HUMANO": [
        "HUMANO", "ARS HUMANO", "HUM", "HUMANA", "HUMAO"
    ],
    "MAPFRE/PALIC": [
        "MAPFRE", "PALIC", "MAPFRE PALIC", "MAPFRE/PALIC",
        "ARS MAPFRE", "ARS PALIC", "MAPHRE"
    ],
    "UNIVERSAL": [
        "UNIVERSAL", "ARS UNIVERSAL", "UNI"
    ],
    "RESERVAS": [
        "RESERVAS", "ARS RESERVAS", "BANRESERVAS", "BANRESERVA", "RESERVA"
    ],
    "MONUMENTAL": [
        "MONUMENTAL", "ARS MONUMENTAL", "MONU"
    ],
    "PRIMERA": [
        "PRIMERA", "ARS PRIMERA", "PRIMERA ARS"
    ],
    "RENACER": [
        "RENACER", "ARS RENACER"
    ],
    "GMA": [
        "GMA"
    ],
    "YUNEN": [
        "YUNEN", "ARS YUNEN"
    ],
    "ABEL GONZALEZ/SIMAG": [
        "SIMAG", "ABEL GONZALEZ", "ABEL GONZÁLEZ", "ABEL",
        "ARS ABEL GONZALEZ", "CENTRO MEDICO ABEL GONZALEZ",
        "CENTRO MÉDICO ABEL GONZÁLEZ"
    ],
    "CMD": [
        "CMD", "COLEGIO MEDICO", "COLEGIO MÉDICO", "COLEGIO MEDICO DOMINICANO"
    ],
    "SEMMA": [
        "SEMMA", "ARS SEMMA", "SEGURO MAESTROS", "MAESTROS"
    ],
    "FUTURO": [
        "FUTURO", "ARS FUTURO"
    ],
    "APS": [
        "APS", "ARS APS"
    ],
    "ASEMAP": [
        "ASEMAP"
    ],
    "METASALUD": [
        "METASALUD", "META SALUD"
    ],
    "BANCO CENTRAL": [
        "BANCO CENTRAL", "BC", "BANCENTRAL"
    ],
    "MEDICA": [
        "MEDICA", "MÉDICA", "ARS MEDICA", "ARS MÉDICA"
    ],
    "SIN SEGURO": [
        "SIN SEGURO", "NO TIENE", "NO", "N/S", "NS", "N\\S",
        "INACTIVO", "INACTIVA", "PENDIENTE", "PENDIENTES",
        "NO ACTIVO", "NO VIGENTE", "VENCIDO", "CANCELADO",
        "DESAFILIADO", "PARTICULAR", "PRIVADO", "NINGUNO", "NINGUNA",
        "N/A", "NA", "S/N", "SN", "NO APLICA", "NO USA", "NO POSEE",
        "NINGUN SEGURO"
    ],
}

# FASE 1: Límites de paginación del historial.
DEFAULT_APP_SETTINGS = {
    # Visual básico
    "font_size": 11,
    "theme": "oscuro",
    "high_contrast": False,
    "window_size": "1280x740",
    "auto_print": True,

    # Preferencias de impresión
    "print_auto_hoja": True,
    "print_auto_reporte_turno": True,
    "print_auto_excel_turno": True,
    "print_copies_hoja": 1,
    "print_copies_reporte": 2,
    "print_copies_excel": 2,
    "print_pdf_orientation": "Horizontal",
    "print_excel_orientation": "Horizontal",
    "print_behavior_hoja": "Imprimir y abrir PDF",

    # Preferencias de validación
    "validation_confirm_before_generate": True,
    "validation_warn_nss_incomplete": True,
    "validation_warn_ars_sin_seguro": True,
    "validation_block_short_ars": True,
    "validation_allow_missing_cedula": True,
    "validation_allow_missing_phone": False,
    "validation_warn_duplicate_turno": True,

    # Preferencias de recién nacido
    "rn_strip_db": True,
    "rn_show_pdf": True,
    "rn_warn": True,
    "rn_format_display": "RN- NOMBRE DE LA MADRE",

    # Preferencias visuales avanzadas
    "button_size": "Normal",
    "table_row_height": 29,
    "compact_mode": False,
    "small_screen_mode": False,
    "show_side_panel": True,
    "show_turno_summary": True,
    "accent_color": "Azul hospitalario",

    # Preferencias de historial (FASE 1: 100 / 150)
    "hist_initial_limit": 100,
    "hist_next_limit": 150,
    "hist_default_filter": "Todos",
    "hist_order": "Más reciente primero",

    # Preferencias de turnos
    "turno_default": "8AM_8AM",
    "turnos_ask_representante_start": False,
    "turnos_generate_report": True,
    "turnos_save_excel_copy": True,
    "turnos_print_empty_report": False,
    "turnos_open_archive_folder": False,

    # Preferencias de PDF
    "pdf_nss_guiones": True,
    "pdf_ars_display_mode": "Abreviada",
    "pdf_nombre_font_size": 12,
    "pdf_direccion_font_size": 12,
    "pdf_open_after_generate": True,
    "pdf_keep_temp": False,
}

ACCENT_COLOR_PRESETS = {
    "Profesional sobrio": "#4f6472",
    "Azul hospitalario": "#68A9D8",
    "Azul profundo": "#4A8CC3",
    "Celeste suave": "#76BDE8",
    "Turquesa clínico": "#55B7B0",
    "Verde salud": "#5BAA70",
    "Gris profesional": "#7D8DA1",
}

def mezclar_color_hex(color_a, color_b, porcentaje=0.15):
    """
    Mezcla dos colores HEX.
    porcentaje=0.15 significa 15% del color_a sobre 85% del color_b.
    Se usa para que el color principal afecte fondos de forma suave.
    """
    try:
        color_a = str(color_a or "").strip().lstrip("#")
        color_b = str(color_b or "").strip().lstrip("#")
        if len(color_a) != 6 or len(color_b) != 6:
            return "#" + color_b

        porcentaje = max(0.0, min(1.0, float(porcentaje)))
        ra, ga, ba = int(color_a[0:2], 16), int(color_a[2:4], 16), int(color_a[4:6], 16)
        rb, gb, bb = int(color_b[0:2], 16), int(color_b[2:4], 16), int(color_b[4:6], 16)

        r = int((ra * porcentaje) + (rb * (1 - porcentaje)))
        g = int((ga * porcentaje) + (gb * (1 - porcentaje)))
        b = int((ba * porcentaje) + (bb * (1 - porcentaje)))
        return f"#{r:02X}{g:02X}{b:02X}"
    except Exception:
        return str(color_b if str(color_b).startswith("#") else "#" + str(color_b))


def resolver_color_principal(valor):
    val = str(valor or "Azul hospitalario").strip()
    if val in ACCENT_COLOR_PRESETS:
        return ACCENT_COLOR_PRESETS[val]
    if re.match(r"^#[0-9A-Fa-f]{6}$", val):
        return val
    return ACCENT_COLOR_PRESETS["Azul hospitalario"]

def nombre_color_principal(valor):
    val = str(valor or "Azul hospitalario").strip()
    if val in ACCENT_COLOR_PRESETS:
        return val
    for nombre, hexv in ACCENT_COLOR_PRESETS.items():
        if val.lower() == hexv.lower():
            return nombre
    return "Azul hospitalario"

def cargar_app_settings():
    try:
        data = load_json_file(APP_SETTINGS_PATH, default={})
        out = dict(DEFAULT_APP_SETTINGS)
        if isinstance(data, dict):
            out.update(data)
        return out
    except ConfigError as exc:
        logging.getLogger("emergencias").error("Configuracion invalida: %s", exc)
    return dict(DEFAULT_APP_SETTINGS)

def guardar_app_settings(settings: dict):
    try:
        data = dict(DEFAULT_APP_SETTINGS)
        data.update(settings or {})
        atomic_write_json(APP_SETTINGS_PATH, data)
        return True
    except (OSError, TypeError, ValueError) as exc:
        logging.getLogger("emergencias").exception("No se pudieron guardar las preferencias: %s", exc)
        return False

def app_setting(key, default=None):
    try:
        return cargar_app_settings().get(key, DEFAULT_APP_SETTINGS.get(key, default))
    except Exception:
        return DEFAULT_APP_SETTINGS.get(key, default)

def cargar_catalogo_ars():
    try:
        merged = {k: list(v) for k, v in DEFAULT_ARS_CATALOGO.items()}

        if os.path.exists(ARS_CATALOGO_PATH):
            with open(ARS_CATALOGO_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)

            if isinstance(data, dict):
                for k, v in data.items():
                    if not k:
                        continue

                    key = _limpiar_texto_seguro(k)
                    if isinstance(v, str):
                        aliases = [x.strip() for x in v.split(",") if x.strip()]
                    elif isinstance(v, list):
                        aliases = [str(x).strip() for x in v if str(x).strip()]
                    else:
                        aliases = []

                    base_aliases = merged.get(key, [])
                    vistos = set(_limpiar_texto_seguro(a) for a in base_aliases)
                    final_aliases = list(base_aliases)
                    for alias in aliases:
                        alias_clean = _limpiar_texto_seguro(alias)
                        if alias_clean and alias_clean not in vistos:
                            final_aliases.append(alias)
                            vistos.add(alias_clean)
                    merged[key] = final_aliases

        return merged
    except Exception:
        pass
    return {k: list(v) for k, v in DEFAULT_ARS_CATALOGO.items()}

def guardar_catalogo_ars(catalogo: dict):
    try:
        data = {}
        for k, v in (catalogo or {}).items():
            key = _limpiar_texto_seguro(k)
            if not key: continue
            aliases = v if isinstance(v, list) else str(v).split(",")
            data[key] = [str(a).strip() for a in aliases if str(a).strip()]
        with open(ARS_CATALOGO_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def cargar_formatos_nss_ars():
    default = {
        "RENACER": "5-5-2"
    }
    try:
        if os.path.exists(NSS_FORMATOS_PATH):
            with open(NSS_FORMATOS_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                for k, v in data.items():
                    key = _limpiar_texto_seguro(k)
                    patron = normalizar_patron_nss(v)
                    if key and patron:
                        default[key] = patron
    except Exception:
        pass
    return default


def guardar_formatos_nss_ars(formatos: dict):
    try:
        data = {}
        for k, v in (formatos or {}).items():
            key = _limpiar_texto_seguro(k)
            patron = normalizar_patron_nss(v)
            if key and patron:
                data[key] = patron
        with open(NSS_FORMATOS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def normalizar_patron_nss(patron) -> str:
    raw = str(patron or "").strip()
    if not raw:
        return ""
    nums = re.findall(r"\d+", raw)
    if not nums:
        return ""
    valores = []
    for n in nums:
        try:
            val = int(n)
            if val <= 0:
                return ""
            valores.append(str(val))
        except Exception:
            return ""
    return "-".join(valores)


def patron_desde_ejemplo_nss(ejemplo: str) -> str:
    raw = str(ejemplo or "").strip()
    if "-" not in raw:
        return ""

    partes = [p for p in raw.split("-") if p != ""]
    if not partes:
        return ""

    if not all(re.sub(r"\D", "", p) for p in partes):
        return ""

    return "-".join(str(len(re.sub(r"\D", "", p))) for p in partes)


def aplicar_patron_nss(nss: str, patron: str) -> str:
    original = str(nss or "").strip().upper()
    digitos = re.sub(r"\D", "", original)
    patron_norm = normalizar_patron_nss(patron)
    if not digitos or not patron_norm:
        return original

    partes = [int(x) for x in patron_norm.split("-")]
    if sum(partes) != len(digitos):
        return original

    out = []
    pos = 0
    for size in partes:
        out.append(digitos[pos:pos + size])
        pos += size
    return "-".join(out)


def formatear_nss_para_pdf(nss: str, ars: str) -> str:
    original = str(nss or "").strip().upper()
    if not original:
        return ""

    ars_canon = normalizar_seguro(ars or "", original)
    formatos = cargar_formatos_nss_ars()
    patron = formatos.get(ars_canon)

    if not patron:
        ars_limpia = _limpiar_texto_seguro(ars or "")
        for key, value in formatos.items():
            if key == ars_limpia or key in ars_limpia or ars_limpia in key:
                patron = value
                break

    return aplicar_patron_nss(original, patron) if patron else original


def limpiar_nombre_rn_para_db(nombre: str) -> str:
    txt = str(nombre or "").strip()
    txt = re.sub(r"^\s*RN\s*[-–—:]\s*", "", txt, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", txt).strip()


def nombre_tiene_prefijo_rn(nombre: str) -> bool:
    return bool(re.match(r"^\s*RN\s*[-–—:]\s*", str(nombre or ""), flags=re.IGNORECASE))


# -------------------------------
# HELPERS DE VALIDACIÓN
# -------------------------------
def is_all_zeros(s: str) -> bool:
    s = (s or "").strip()
    return len(s) > 0 and set(s) == {"0"}


def is_valid_nss_key(nss: str) -> bool:
    if not nss:
        return False
    nss = nss.strip().upper()
    if nss in ["N/S", r"N\\S", "NS", "NO", "SIN SEGURO", ""]:
        return False
    if not nss.isdigit():
        return False
    if is_all_zeros(nss):
        return False
    if len(nss) < 3:
        return False
    return True


def is_valid_cedula_key(cedula: str) -> bool:
    if not cedula:
        return False
    ced = cedula.strip()
    return ced.isdigit() and len(ced) == 11 and not is_all_zeros(ced)


def normalizar_nombre_clave(nombre: str) -> str:
    return re.sub(r"\s+", " ", str(nombre or "").strip()).upper()


def get_patient_key(nss: str, cedula: str):
    if is_valid_nss_key(nss):
        return nss.strip().upper()
    if is_valid_cedula_key(cedula):
        return cedula.strip()
    return None


def patient_identity_key_from_row(row: dict):
    nss = (row.get("nss") or "").strip().upper()
    cedula = (row.get("cedula") or "").strip()
    if is_valid_nss_key(nss):
        return ("NSS", nss)
    if is_valid_cedula_key(cedula):
        return ("CEDULA", cedula)
    return ("ROW", row.get("id"))


def _detectar_campos_invertidos(nss_raw: str, ars_raw: str):
    nss = (nss_raw or "").strip()
    ars = (ars_raw or "").strip()

    nss_valid_text = ["N/S", r"N\\S", "NS", "NO", "SIN SEGURO", ""]
    nss_is_text = any(c.isalpha() for c in nss) and nss.upper() not in nss_valid_text
    ars_sin_espacios = re.sub(r"\s+", "", ars)
    ars_is_numeric = ars_sin_espacios.isdigit() and len(ars_sin_espacios) >= 3

    if nss_is_text and ars_is_numeric:
        return (
            True,
            "Campos invertidos:\n\nColocaste el nombre del seguro en el campo 'NSS' "
            "y el número en el campo 'Aseguradora'.\n\nPor favor, intercámbialos."
        )
    if nss_is_text:
        return (
            True,
            f"Error en NSS:\n\nEl campo NSS debe ser estrictamente numérico. "
            f"Si no tiene NSS, déjalo vacío o escribe 'SIN SEGURO'. No coloques el nombre de la aseguradora aquí."
        )
    if ars_is_numeric:
        return (
            True,
            f"Error en Aseguradora (ARS):\n\nEl campo Aseguradora debe contener el nombre "
            f"del seguro (texto), no números largos."
        )

    return (False, "")


# -------------------------------
# FECHAS / DÍA OPERATIVO
# -------------------------------
def parse_fecha_ddmmyyyy(fecha_str: str):
    try:
        return datetime.strptime((fecha_str or "").strip(), "%d/%m/%Y").date()
    except Exception:
        return None


def parse_hora_12h(hora_str: str):
    try:
        return datetime.strptime((hora_str or "").strip(), "%I:%M %p").time()
    except Exception:
        try:
            txt = (hora_str or "").strip().upper().replace(".", "")
            match = re.search(r"(\d{1,2}):(\d{2})\s*(A\s*M|P\s*M|A|P)", txt)
            if match:
                h = int(match.group(1))
                m = int(match.group(2))
                p = match.group(3).replace(" ", "")
                if p.startswith("P") and h < 12:
                    h += 12
                elif p.startswith("A") and h == 12:
                    h = 0
                return time(h, m)
        except Exception:
            pass
        return None


def parse_datetime_local(dt_str: str):
    try:
        return datetime.strptime((dt_str or "").strip(), "%d/%m/%Y %I:%M %p")
    except Exception:
        return None


def format_datetime_local(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y %I:%M %p")


def construir_datetime_operativo(fecha_str: str, hora_str: str):
    f = parse_fecha_ddmmyyyy(fecha_str)
    h = parse_hora_12h(hora_str)
    if not f or not h:
        return None

    dt = datetime.combine(f, h)
    if h < time(8, 0):
        dt = dt - timedelta(days=1)
    return dt


def construir_datetime_real(fecha_str: str, hora_str: str):
    f = parse_fecha_ddmmyyyy(fecha_str)
    h = parse_hora_12h(hora_str)
    if not f or not h:
        return None
    return datetime.combine(f, h)


def obtener_rango_operativo_desde_fecha(base_date: date):
    inicio = datetime.combine(base_date, time(8, 0))
    fin = inicio + timedelta(days=1)
    return inicio, fin


def fecha_base_operativa_actual(momento: datetime = None) -> date:
    actual = momento or datetime.now()
    return actual.date() - timedelta(days=1) if actual.time() < time(8, 0) else actual.date()


def normalizar_turno_codigo(turno_codigo: str) -> str:
    raw = str(turno_codigo or "").strip()
    if raw in ("8AM_8AM", "8AM_8PM", "8PM_8AM"):
        return raw

    up = raw.upper()
    up = up.replace("→", " A ").replace("-", " A ").replace("–", " A ").replace("—", " A ")
    up = up.replace(".", "")
    up = up.replace(":00", "")
    up = re.sub(r"\s+", "", up)

    if "8PM" in up and "8AM" in up:
        return "8PM_8AM"
    if "8AM" in up and "8PM" in up:
        return "8AM_8PM"
    if "8AM" in up:
        return "8AM_8AM"

    return "8AM_8AM"


def limpiar_nombre_representante(valor: str) -> str:
    txt = str(valor or "").strip()
    if not txt:
        return ""

    txt = re.sub(r"\b\d{1,2}/\d{1,2}/\d{4}\b", " ", txt)
    txt = re.sub(r"\bAL\b", " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"\bDEL\b", " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"\bDESDE\b", " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"\bHASTA\b", " ", txt, flags=re.IGNORECASE)

    txt = re.sub(r"[-–—:|]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()

    if txt.upper() in {"AL", "DEL", "DESDE", "HASTA"}:
        return ""
    return txt


REPRESENTANTES_NO_VALIDOS = {
    "NO DISPONIBLE",
    "NO CONFIGURADO",
    "NOMBRE DEL REPRESENTANTE",
    "NOMBRE REPRESENTANTE",
    "SIN REPRESENTANTE",
}


def es_representante_valido(valor: str) -> bool:
    limpio = limpiar_nombre_representante(valor)
    return bool(limpio and limpio.upper() not in REPRESENTANTES_NO_VALIDOS)


def cargar_representantes(db=None):
    nombres = set()
    catalogo_existente = os.path.exists(REPRESENTANTES_PATH)
    try:
        if catalogo_existente:
            with open(REPRESENTANTES_PATH, "r", encoding="utf-8") as archivo:
                data = json.load(archivo)
            for nombre in data if isinstance(data, list) else []:
                limpio = limpiar_nombre_representante(nombre)
                if es_representante_valido(limpio):
                    nombres.add(limpio)
    except Exception:
        APP_LOG.exception("No se pudo leer el catálogo de representantes")

    try:
        cfg = cargar_turno_config(permitir_vencido=True)
        actual = limpiar_nombre_representante((cfg or {}).get("representante", ""))
        if es_representante_valido(actual):
            nombres.add(actual)
    except Exception:
        APP_LOG.exception("No se pudo recuperar el representante del turno actual")

    # La BD histórica solo se usa para inicializar el catálogo una vez. Después,
    # un usuario eliminado no debe reaparecer por existir en turnos antiguos.
    if db is not None and not catalogo_existente:
        try:
            nombres.update(
                nombre for nombre in db.listar_representantes()
                if es_representante_valido(nombre)
            )
        except Exception:
            APP_LOG.exception("No se pudieron consultar representantes históricos")

    return sorted(nombres, key=lambda valor: valor.casefold())


def guardar_catalogo_representantes(nombres) -> bool:
    limpios = {}
    for nombre in nombres or []:
        limpio = limpiar_nombre_representante(nombre)
        if es_representante_valido(limpio):
            limpios.setdefault(limpio.casefold(), limpio)
    try:
        atomic_write_json(
            REPRESENTANTES_PATH,
            sorted(limpios.values(), key=lambda valor: valor.casefold()),
        )
        return True
    except (OSError, TypeError, ValueError):
        APP_LOG.exception("No se pudo guardar el catálogo de representantes")
        return False


def guardar_representante_catalogo(nombre: str, db=None):
    limpio = limpiar_nombre_representante(nombre)
    if not es_representante_valido(limpio):
        return ""

    nombres = cargar_representantes(db)
    if limpio.casefold() not in {valor.casefold() for valor in nombres}:
        nombres.append(limpio)
    return limpio if guardar_catalogo_representantes(nombres) else ""

def descripcion_turno_config(turno_cfg: dict) -> str:
    if not turno_cfg:
        return "No configurado"
    try:
        fecha_base = turno_cfg.get("fecha_base")
        if not isinstance(fecha_base, date):
            return "No configurado"
        codigo = normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM"))
        return obtener_datos_turno_visual(fecha_base, codigo)["turno_resumen"]
    except Exception:
        return "No configurado"


def obtener_rango_turno_real(fecha_base: date, turno_codigo: str):
    turno_codigo = normalizar_turno_codigo(turno_codigo)
    if turno_codigo == "8AM_8PM":
        inicio = datetime.combine(fecha_base, time(8, 0))
        fin = datetime.combine(fecha_base, time(20, 0))
        return inicio, fin

    if turno_codigo == "8PM_8AM":
        inicio = datetime.combine(fecha_base, time(20, 0))
        fin = datetime.combine(fecha_base + timedelta(days=1), time(8, 0))
        return inicio, fin

    inicio = datetime.combine(fecha_base, time(8, 0))
    fin = datetime.combine(fecha_base + timedelta(days=1), time(8, 0))
    return inicio, fin


def obtener_datos_turno_visual(fecha_base: date, turno: str):
    turno = normalizar_turno_codigo(turno)
    siguiente = fecha_base + timedelta(days=1)

    if turno == "8AM_8PM":
        return {
            "turno_label": "DESDE 8:00 AM A 8:00 PM",
            "fecha_label": fecha_base.strftime("%d/%m/%Y"),
            "turno_resumen": "8:00 AM → 8:00 PM",
        }

    if turno == "8PM_8AM":
        return {
            "turno_label": "DESDE 8:00 PM A 8:00 AM",
            "fecha_label": f"{fecha_base.strftime('%d/%m/%Y')} AL {siguiente.strftime('%d/%m/%Y')}",
            "turno_resumen": "8:00 PM → 8:00 AM",
        }

    return {
        "turno_label": "DESDE 8:00 AM A 8:00 AM",
        "fecha_label": f"{fecha_base.strftime('%d/%m/%Y')} AL {siguiente.strftime('%d/%m/%Y')}",
        "turno_resumen": "8:00 AM → 8:00 AM",
    }


def obtener_inicio_real_turno(turno_cfg: dict):
    if not turno_cfg:
        return None

    inicio_nominal, _ = obtener_rango_turno_real(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
    inicio_real_guardado = turno_cfg.get("inicio_real_dt")

    if inicio_real_guardado and inicio_real_guardado > inicio_nominal:
        return inicio_real_guardado
    return inicio_nominal


def obtener_rango_turno_efectivo(turno_cfg: dict, fin_override: datetime = None):
    if not turno_cfg:
        return None, None
    inicio_nominal, fin_nominal = obtener_rango_turno_real(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
    inicio_real = obtener_inicio_real_turno(turno_cfg)
    fin_real = fin_override if fin_override else fin_nominal
    if fin_real < inicio_real:
        fin_real = inicio_real
    return inicio_real, fin_real


def turno_config_es_vigente(turno_cfg: dict, momento: datetime = None) -> bool:
    if not turno_cfg:
        return False
    try:
        inicio, fin = obtener_rango_turno_efectivo(turno_cfg)
        actual = momento or datetime.now()
        return bool(inicio and fin and inicio <= actual < fin)
    except (KeyError, TypeError, ValueError):
        return False


# -------------------------------
# NORMALIZACIÓN DE SEGUROS
# -------------------------------
def _limpiar_texto_seguro(txt: str) -> str:
    txt = (txt or "").strip().upper()
    txt = txt.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    txt = re.sub(r"\bARS\b", "", txt)
    txt = re.sub(r"[^A-Z0-9/ ]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _compact(txt: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", _limpiar_texto_seguro(txt))


def _score(a: str, b: str) -> float:
    return SequenceMatcher(None, _compact(a), _compact(b)).ratio()


def _es_sin_seguro_por_texto(seguro_raw: str) -> bool:
    txt = _limpiar_texto_seguro(seguro_raw)
    valores = {
        "", "N/S", "NS", "NO", "SIN SEGURO", "NINGUNO", "NINGUNA",
        "N/A", "NA", "S/N", "SN", "NO APLICA", "NO TIENE",
        "NO USA", "NO POSEE", "NINGUN SEGURO",
        "INACTIVO", "INACTIVA", "PENDIENTE", "PENDIENTES",
        "NO ACTIVO", "NO VIGENTE", "VENCIDO", "CANCELADO",
        "DESAFILIADO", "PARTICULAR", "PRIVADO"
    }
    if txt in valores:
        return True
    if txt.startswith("SIN SEG"):
        return True
    return False


def _mejor_seguro_por_similitud(txt: str):
    txt_clean = _limpiar_texto_seguro(txt)
    txt_comp = _compact(txt_clean)

    if not txt_comp:
        return None

    candidatos = {
        "SENASA SUBSIDIADO": [
            "SUB", "SUNB", "SUBSIDIADO", "SENASA SUB", "SENASA SUBSIDIADO",
            "ARS SENASA SUBSIDIADO", "SESANA SUBSIADO", "SENASA SUBSIADO",
            "SUBSIADO", "SUBSI", "SENASA SUBSIDIAD", "SUBS"
        ],
        "SENASA CONTRIBUTIVO": [
            "CONTRIBUTIVO", "SENASA CONTRIBUTIVO", "ARS SENASA CONTRIBUTIVO",
            "CONTRI", "CONTRIB", "CONTRIBUT", "CONT", "COTIZANTE",
            "SENASA AVANZADA", "AVANZADA", "SENASA MAXIMO", "MAXIMO", "MÁXIMO",
            "SENASA ESPECIAL", "ESPECIAL"
        ],
        "SENASA PENSIONADOS": [
            "PENSIONADOS", "SENASA PENSIONADOS", "ARS SENASA PENSIONADOS", "PENSIONADO"
        ],
        "APS": ["APS"],
        "ASEMAP": ["ASEMAP"],
        "CMD": ["CMD"],
        "GMA": ["GMA"],
        "RENACER": ["RENACER"],
        "RESERVAS": ["RESERVAS", "ARS RESERVAS"],
        "SEMMA": ["SEMMA"],
        "FUTURO": ["FUTURO"],
        "HUMANO": ["HUMANO", "ARS HUMANO", "HUMAO"],
        "PRIMERA": ["PRIMERA", "ARS PRIMERA"],
        "ABEL GONZALEZ/SIMAG": ["ABEL GONZALEZ", "SIMAG", "ABEL", "ARS ABEL GONZALEZ"],
        "METASALUD": ["METASALUD"],
        "MONUMENTAL": ["MONUMENTAL"],
        "MAPFRE/PALIC": ["MAPFRE", "PALIC", "ARS MAPFRE", "ARS PALIC", "MAPHRE"],
        "UNIVERSAL": ["UNIVERSAL"],
        "BANCO CENTRAL": ["BANCO CENTRAL"],
        "YUNEN": ["YUNEN"],
    }

    try:
        for canonico, aliases in cargar_catalogo_ars().items():
            canon = _limpiar_texto_seguro(canonico)
            if not canon: continue
            candidatos.setdefault(canon, [])
            for alias in aliases:
                if alias and alias not in candidatos[canon]:
                    candidatos[canon].append(alias)
    except Exception:
        pass

    if txt_comp.startswith("SU") or txt_comp in {"SUNB", "SUB", "SUBS", "SUBSI", "SUBSIDIADO", "SUBSIADO"}:
        return "SENASA SUBSIDIADO"

    if txt_comp in {"AVANZADA", "SENASAAVANZADA", "MAXIMO", "MAXIMA", "SENASAMAXIMO", "SENASAMAXIMA", "ESPECIAL", "SENASAESPECIAL"}:
        return "SENASA CONTRIBUTIVO"

    if "SENASA" in txt_clean and ("SUB" in txt_clean or "SUBSI" in txt_clean or "SUBSID" in txt_clean):
        return "SENASA SUBSIDIADO"
    if "SENASA" in txt_clean and ("CONTRI" in txt_clean or "CONTRIB" in txt_clean):
        return "SENASA CONTRIBUTIVO"
    if "SENASA" in txt_clean and "PENSION" in txt_clean:
        return "SENASA PENSIONADOS"

    if "MAPFRE" in txt_clean or "PALIC" in txt_clean:
        return "MAPFRE/PALIC"
    if "ABEL" in txt_clean or "SIMAG" in txt_clean:
        return "ABEL GONZALEZ/SIMAG"
    if "BANCO CENTRAL" in txt_clean:
        return "BANCO CENTRAL"

    mejor_nombre = None
    mejor_score = 0.0

    for canonico, aliases in candidatos.items():
        for alias in aliases:
            s = _score(txt_clean, alias)
            if s > mejor_score:
                mejor_score = s
                mejor_nombre = canonico

    if mejor_score >= 0.72:
        return mejor_nombre

    if len(txt_comp) <= 8 and mejor_score >= 0.58:
        return mejor_nombre

    flat_aliases = []
    alias_to_canon = {}
    for canonico, aliases in candidatos.items():
        for alias in aliases:
            flat_aliases.append(alias)
            alias_to_canon[alias] = canonico

    match = get_close_matches(txt_clean, flat_aliases, n=1, cutoff=0.55)
    if match:
        return alias_to_canon[match[0]]

    return None


def ars_es_corta_invalida(seguro_raw: str) -> bool:
    txt = _limpiar_texto_seguro(seguro_raw)
    comp = _compact(txt)
    if not comp or _es_sin_seguro_por_texto(txt):
        return False
    if comp.isdigit() or len(comp) == 1:
        return True
    if _mejor_seguro_por_similitud(txt):
        return False
    return len(comp) < 4

from functools import lru_cache
@lru_cache(maxsize=5000)
def normalizar_seguro(seguro_raw: str, nss_raw: str = "") -> str:
    txt = _limpiar_texto_seguro(seguro_raw)
    nss = (nss_raw or "").strip().upper()

    if _es_sin_seguro_por_texto(txt):
        return "SIN SEGURO"

    comp_txt = _compact(txt)
    if comp_txt.isdigit() or len(comp_txt) == 1:
        return "SIN SEGURO"

    parecido = _mejor_seguro_por_similitud(txt)
    if parecido:
        return parecido

    if ars_es_corta_invalida(txt):
        return "SIN SEGURO"

    if not is_valid_nss_key(nss):
        return "SIN SEGURO"

    return txt if txt else "SIN SEGURO"


def seguro_para_mostrar(seguro_canonico: str) -> str:
    return SEGUROS_DISPLAY.get(seguro_canonico, (seguro_canonico or "").strip().upper() or "SIN SEGURO")


# -------------------------------
# CONFIG TURNO
# -------------------------------
def cargar_turno_config(permitir_vencido=False):
    try:
        data = load_json_file(TURNOS_CFG, default=None)
        if not isinstance(data, dict):
            return None
        fecha_base = parse_fecha_ddmmyyyy(data.get("fecha_base", ""))
        if not fecha_base:
            return None

        inicio_real_dt = parse_datetime_local(data.get("inicio_real", ""))

        representante = limpiar_nombre_representante(data.get("representante", ""))
        if not es_representante_valido(representante):
            representante = ""
        config = {
            "representante": representante,
            "turno_codigo": normalizar_turno_codigo(data.get("turno_codigo", "8AM_8AM")),
            "fecha_base": fecha_base,
            "inicio_real": data.get("inicio_real", ""),
            "inicio_real_dt": inicio_real_dt,
        }
        if not permitir_vencido and not turno_config_es_vigente(config):
            APP_LOG.warning("Se ignoro una configuracion de turno vencida: %s", data)
            return None
        return config
    except ConfigError as exc:
        APP_LOG.error("No se pudo leer la configuracion del turno: %s", exc)
        return None


def guardar_turno_config(representante: str, turno_codigo: str, fecha_base: date, inicio_real: datetime = None):
    try:
        representante = limpiar_nombre_representante(representante)
        if not es_representante_valido(representante):
            APP_LOG.warning(
                "Se rechazó un representante inválido para el turno: %r",
                representante,
            )
            return False
        payload = {
            "representante": representante,
            "turno_codigo": normalizar_turno_codigo(turno_codigo),
            "fecha_base": fecha_base.strftime("%d/%m/%Y"),
            "inicio_real": format_datetime_local(inicio_real or datetime.now()),
        }
        atomic_write_json(TURNOS_CFG, payload)
        return True
    except (OSError, TypeError, ValueError) as exc:
        APP_LOG.exception("No se pudo guardar la configuracion del turno: %s", exc)
        return False


def crear_turno_desde_excel_existente_si_aplica() -> bool:
    try:
        if os.path.exists(TURNOS_CFG):
            return False

        if not os.path.exists(EXCEL_PATH):
            return False

        if not excel_tiene_registros(EXCEL_PATH):
            return False

        representante = ""
        turno_codigo = "8AM_8AM"
        fecha_base = datetime.now().date()

        try:
            wb = abrir_excel_workbook_seguro(EXCEL_PATH, read_only=True, data_only=True)
            ws = wb.active

            a3 = str(ws["A3"].value or "").strip()
            a4 = str(ws["A4"].value or "").strip()

            if a4:
                turno_codigo = normalizar_turno_codigo(a4)

            m = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", a3)
            if m:
                f = parse_fecha_ddmmyyyy(m.group(1))
                if f:
                    fecha_base = f

            rep_limpio = limpiar_nombre_representante(a3)
            if es_representante_valido(rep_limpio):
                representante = rep_limpio

            try:
                wb.close()
            except Exception:
                pass
        except Exception:
            pass

        hora_inicio = time(20, 0) if normalizar_turno_codigo(turno_codigo) == "8PM_8AM" else time(8, 0)
        inicio_real = datetime.combine(fecha_base, hora_inicio)

        if not es_representante_valido(representante):
            APP_LOG.warning(
                "El Excel existente no contiene un representante válido; "
                "se conservará sin crear una configuración de turno."
            )
            return False

        return guardar_turno_config(
            representante=representante,
            turno_codigo=turno_codigo,
            fecha_base=fecha_base,
            inicio_real=inicio_real
        )

    except Exception:
        return False


# -------------------------------
# DB MANAGER
# -------------------------------
class TurnoNoVigenteError(RuntimeError):
    pass


class DatabaseManager:
    SCHEMA_VERSION = LATEST_SCHEMA_VERSION

    def __init__(self, db_name='pacientes.db'):
        self.db_name = app_data_path(db_name)
        self.backup_manager = BackupManager(
            self.db_name,
            BACKUPS_DIR,
            related_paths=(TURNOS_CFG, APP_SETTINGS_PATH, EXCEL_PATH),
            retention_days=4,
        )
        self._init_db()

    def _connect(self):
        conn = sqlite3.connect(self.db_name, timeout=15)
        conn.execute("PRAGMA foreign_keys = ON;")
        conn.execute("PRAGMA busy_timeout = 15000;")
        conn.execute("PRAGMA journal_mode = WAL;")
        return conn

    def _init_db(self):
        result = migrate_database(self.db_name, self.backup_manager, APP_LOG)
        APP_LOG.info("Base de datos lista en esquema v%s: %s", self.SCHEMA_VERSION, result)

    def _crear_respaldo_migracion(self, conn, version_actual):
        if version_actual >= self.SCHEMA_VERSION or not os.path.exists(self.db_name):
            return ""
        tiene_datos = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='atenciones'"
        ).fetchone()
        if not tiene_datos:
            return ""
        marca = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = os.path.join(
            BACKUPS_DIR,
            f"pacientes_pre_migracion_v{version_actual}_a_v{self.SCHEMA_VERSION}_{marca}.db",
        )
        with closing(sqlite3.connect(destino)) as copia:
            conn.backup(copia)
        APP_LOG.info("Respaldo previo a migración creado: %s", destino)
        return destino

    @staticmethod
    def _asegurar_columna(cursor, tabla, columna, definicion):
        existentes = {fila[1] for fila in cursor.execute(f"PRAGMA table_info({tabla})")}
        if columna not in existentes:
            cursor.execute(f"ALTER TABLE {tabla} ADD COLUMN {columna} {definicion}")

    def _init_db_legacy_unused(self):
        with closing(self._connect()) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS schema_version (
                    id INTEGER PRIMARY KEY CHECK (id = 1),
                    version INTEGER NOT NULL
                );
            ''')
            cursor.execute("INSERT OR IGNORE INTO schema_version (id, version) VALUES (1, 0);")
            conn.commit()

            cursor.execute('SELECT version FROM schema_version WHERE id = 1;')
            row = cursor.fetchone()
            current_version = int(row[0]) if row else 0

            self._crear_respaldo_migracion(conn, current_version)

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS pacientes (
                    cedula TEXT CHECK(cedula IS NULL OR LENGTH(cedula) = 11),
                    nombre TEXT NOT NULL,
                    telefono TEXT CHECK(telefono IS NULL OR LENGTH(telefono) = 10),
                    direccion TEXT,
                    nacionalidad TEXT,
                    ars TEXT,
                    nss TEXT UNIQUE,
                    PRIMARY KEY (nss)
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS atenciones (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nss TEXT,
                    nombre TEXT NOT NULL,
                    sexo TEXT,
                    edad_num INTEGER,
                    unidad TEXT,
                    cedula TEXT,
                    telefono TEXT,
                    direccion TEXT,
                    nacionalidad TEXT,
                    ars TEXT,
                    hoja TEXT,
                    fecha TEXT,
                    hora TEXT,
                    tipo_atencion TEXT DEFAULT 'EMERGENCIA',
                    created_at TEXT DEFAULT (datetime('now','localtime')),
                    updated_at TEXT,
                    turno_id INTEGER,
                    nss_clean TEXT,
                    cedula_clean TEXT,
                    telefono_clean TEXT
                );
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS turnos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fecha_inicio TEXT NOT NULL,
                    fecha_fin TEXT NOT NULL,
                    representante TEXT NOT NULL,
                    tipo_turno TEXT NOT NULL,
                    estado TEXT NOT NULL DEFAULT 'ABIERTO',
                    fecha_cierre TEXT,
                    created_at TEXT DEFAULT (datetime('now','localtime')),
                    UNIQUE(fecha_inicio, tipo_turno)
                );
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS atenciones_auditoria (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    atencion_id INTEGER NOT NULL,
                    accion TEXT NOT NULL,
                    motivo TEXT,
                    usuario TEXT,
                    snapshot_json TEXT NOT NULL,
                    created_at TEXT DEFAULT (datetime('now','localtime'))
                );
            ''')

            for columna, definicion in (
                ("tipo_atencion", "TEXT DEFAULT 'EMERGENCIA'"),
                ("edad_num", "INTEGER"),
                ("unidad", "TEXT"),
                ("updated_at", "TEXT"),
                ("turno_id", "INTEGER"),
                ("nss_clean", "TEXT"),
                ("cedula_clean", "TEXT"),
                ("telefono_clean", "TEXT"),
            ):
                self._asegurar_columna(cursor, "atenciones", columna, definicion)

            cursor.execute("""
                UPDATE atenciones
                SET tipo_atencion = COALESCE(NULLIF(TRIM(tipo_atencion), ''), 'EMERGENCIA'),
                    nss_clean = REPLACE(REPLACE(REPLACE(IFNULL(nss,''),'-',''),' ',''),'.',''),
                    cedula_clean = REPLACE(REPLACE(REPLACE(IFNULL(cedula,''),'-',''),' ',''),'.',''),
                    telefono_clean = REPLACE(REPLACE(REPLACE(IFNULL(telefono,''),'-',''),' ',''),'.','')
                WHERE tipo_atencion IS NULL OR TRIM(tipo_atencion) = ''
                   OR nss_clean IS NULL OR cedula_clean IS NULL OR telefono_clean IS NULL
            """)

            indices = (
                "CREATE INDEX IF NOT EXISTS idx_atenciones_fecha ON atenciones(fecha)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_nombre ON atenciones(nombre)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_nss ON atenciones(nss)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_cedula ON atenciones(cedula)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_ars ON atenciones(ars)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_hoja ON atenciones(hoja)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_tipo ON atenciones(tipo_atencion)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_fecha_id ON atenciones(fecha, id)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_fecha_hora ON atenciones(fecha, hora)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_telefono ON atenciones(telefono)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_created_at ON atenciones(created_at)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_turno_id ON atenciones(turno_id)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_nss_clean ON atenciones(nss_clean)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_cedula_clean ON atenciones(cedula_clean)",
                "CREATE INDEX IF NOT EXISTS idx_atenciones_telefono_clean ON atenciones(telefono_clean)",
                "CREATE INDEX IF NOT EXISTS idx_turnos_estado ON turnos(estado)",
                "CREATE INDEX IF NOT EXISTS idx_auditoria_atencion ON atenciones_auditoria(atencion_id)",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_atencion_turno_cedula ON atenciones(turno_id, cedula_clean) "
                "WHERE turno_id IS NOT NULL AND LENGTH(cedula_clean) = 11",
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_atencion_turno_nss ON atenciones(turno_id, nss_clean) "
                "WHERE turno_id IS NOT NULL AND LENGTH(nss_clean) >= 3",
            )
            for sql in indices:
                cursor.execute(sql)

            cursor.execute(
                "UPDATE schema_version SET version = ? WHERE id = 1",
                (self.SCHEMA_VERSION,),
            )
            conn.commit()

            integrity = cursor.execute("PRAGMA integrity_check").fetchone()
            if not integrity or integrity[0] != "ok":
                raise sqlite3.DatabaseError(f"Falló integrity_check: {integrity}")
            APP_LOG.info("Base de datos lista en esquema v%s", self.SCHEMA_VERSION)

    def listar_representantes(self):
        with closing(self._connect()) as conn:
            filas = conn.execute(
                "SELECT DISTINCT representante FROM turnos "
                "WHERE TRIM(IFNULL(representante,'')) <> '' ORDER BY representante"
            ).fetchall()
        return [
            limpiar_nombre_representante(fila[0])
            for fila in filas
            if es_representante_valido(fila[0])
        ]

    def actualizar_representante_turno(self, turno_id: int, representante: str) -> bool:
        representante = limpiar_nombre_representante(representante)
        if not es_representante_valido(representante):
            raise ValueError(
                "Escriba un nombre de representante válido; 'No disponible' no se admite."
            )
        with closing(self._connect()) as conn:
            cur = conn.execute(
                """
                UPDATE turnos SET representante=?,updated_at=datetime('now','localtime')
                WHERE id=?
                """,
                (representante, int(turno_id)),
            )
            conn.commit()
            return cur.rowcount == 1

    def obtener_o_crear_turno(self, turno_cfg, conn=None):
        if not turno_cfg or not turno_config_es_vigente(turno_cfg):
            raise TurnoNoVigenteError("Debe abrir un turno vigente antes de registrar atenciones.")
        inicio, fin = obtener_rango_turno_efectivo(turno_cfg)
        representante = limpiar_nombre_representante(turno_cfg.get("representante", ""))
        if not es_representante_valido(representante):
            raise ValueError("El turno requiere un representante válido.")
        tipo = normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM"))
        fecha_base = turno_cfg["fecha_base"]
        dia_inicio = datetime.combine(fecha_base, time(8, 0))
        dia_fin = dia_inicio + timedelta(days=1)
        propia = conn is None
        conexion = conn or self._connect()
        try:
            conexion.execute(
                """
                INSERT INTO dias_operativos(fecha_base,fecha_inicio,fecha_fin,estado,origen,requiere_revision)
                VALUES (?,?,?,'ABIERTO','OPERATIVO',0)
                ON CONFLICT(fecha_base) DO UPDATE SET
                    fecha_inicio=excluded.fecha_inicio,
                    fecha_fin=excluded.fecha_fin,
                    estado='ABIERTO'
                """,
                (
                    fecha_base.isoformat(),
                    dia_inicio.strftime("%Y-%m-%d %H:%M:%S"),
                    dia_fin.strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            dia_row = conexion.execute(
                "SELECT id FROM dias_operativos WHERE fecha_base=?",
                (fecha_base.isoformat(),),
            ).fetchone()
            dia_id = int(dia_row[0])
            inicio_txt = inicio.strftime("%Y-%m-%d %H:%M:%S")
            fin_txt = fin.strftime("%Y-%m-%d %H:%M:%S")
            cierre_txt = (turno_cfg.get("inicio_real_dt") or datetime.now()).strftime(
                "%Y-%m-%d %H:%M:%S"
            )
            conexion.execute(
                """
                UPDATE turnos SET estado='CERRADO', fecha_cierre=?, updated_at=datetime('now','localtime')
                WHERE estado='ABIERTO'
                  AND NOT (dia_operativo_id=? AND fecha_inicio=? AND tipo_turno=?)
                """,
                (cierre_txt, dia_id, inicio_txt, tipo),
            )
            conexion.execute(
                "UPDATE dias_operativos SET estado='CERRADO' WHERE estado='ABIERTO' AND id<>?",
                (dia_id,),
            )
            conexion.execute('''
                INSERT INTO turnos (
                    dia_operativo_id,fecha_inicio,fecha_fin,fecha_inicio_real,representante,
                    tipo_turno,estado,origen,requiere_revision
                ) VALUES (?, ?, ?, ?, ?, ?, 'ABIERTO', 'OPERATIVO', 0)
                ON CONFLICT(dia_operativo_id,fecha_inicio,tipo_turno) DO UPDATE SET
                    fecha_fin=excluded.fecha_fin,
                    fecha_inicio_real=excluded.fecha_inicio_real,
                    representante=excluded.representante,
                    estado='ABIERTO',
                    fecha_cierre=NULL,
                    updated_at=datetime('now','localtime')
            ''', (
                dia_id,
                inicio_txt,
                fin_txt,
                (turno_cfg.get("inicio_real_dt") or inicio).strftime("%Y-%m-%d %H:%M:%S"),
                representante,
                tipo,
            ))
            fila = conexion.execute(
                "SELECT id FROM turnos WHERE dia_operativo_id=? AND fecha_inicio=? AND tipo_turno=?",
                (dia_id, inicio_txt, tipo),
            ).fetchone()
            if propia:
                conexion.commit()
            return int(fila[0]) if fila else None
        finally:
            if propia:
                conexion.close()

    def obtener_contexto_turno(self, turno_cfg, conn=None):
        turno_id = self.obtener_o_crear_turno(turno_cfg, conn=conn)
        propia = conn is None
        conexion = conn or self._connect()
        try:
            fila = conexion.execute(
                "SELECT id, dia_operativo_id, fecha_inicio, fecha_fin, representante, tipo_turno FROM turnos WHERE id=?",
                (turno_id,),
            ).fetchone()
            if not fila:
                raise TurnoNoVigenteError("No se pudo resolver el turno vigente.")
            return {
                "turno_id": int(fila[0]),
                "dia_operativo_id": int(fila[1]),
                "fecha_inicio": fila[2],
                "fecha_fin": fila[3],
                "representante": fila[4],
                "tipo_turno": fila[5],
            }
        finally:
            if propia:
                conexion.close()

    def buscar_contexto_turno_existente(self, turno_cfg):
        if not turno_cfg:
            return None
        inicio, _fin = obtener_rango_turno_efectivo(turno_cfg)
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                """
                SELECT t.id AS turno_id,t.dia_operativo_id,t.fecha_inicio,t.fecha_fin,
                       t.representante,t.tipo_turno,t.estado,d.fecha_base
                FROM turnos t
                JOIN dias_operativos d ON d.id=t.dia_operativo_id
                WHERE d.fecha_base=? AND t.fecha_inicio=? AND t.tipo_turno=?
                ORDER BY t.id DESC LIMIT 1
                """,
                (
                    turno_cfg["fecha_base"].isoformat(),
                    inicio.strftime("%Y-%m-%d %H:%M:%S"),
                    normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM")),
                ),
            ).fetchone()
            if not row:
                # La hora real puede variar si la configuración fue recuperada o
                # guardada nuevamente. El día y el tipo identifican el mismo turno.
                row = conn.execute(
                    """
                    SELECT t.id AS turno_id,t.dia_operativo_id,t.fecha_inicio,t.fecha_fin,
                           t.representante,t.tipo_turno,t.estado,d.fecha_base
                    FROM turnos t
                    JOIN dias_operativos d ON d.id=t.dia_operativo_id
                    WHERE d.fecha_base=? AND t.tipo_turno=?
                    ORDER BY CASE WHEN t.estado='ABIERTO' THEN 0 ELSE 1 END,t.id DESC
                    LIMIT 1
                    """,
                    (
                        turno_cfg["fecha_base"].isoformat(),
                        normalizar_turno_codigo(turno_cfg.get("turno_codigo", "8AM_8AM")),
                    ),
                ).fetchone()
        return dict(row) if row else None

    def cerrar_turno_existente(self, turno_cfg, momento_cierre=None):
        contexto = self.buscar_contexto_turno_existente(turno_cfg)
        if not contexto:
            return False
        cierre = (momento_cierre or datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
        with closing(self._connect()) as conn:
            conn.execute("BEGIN IMMEDIATE")
            conn.execute(
                """
                UPDATE turnos SET estado='CERRADO',fecha_cierre=?,updated_at=datetime('now','localtime')
                WHERE id=?
                """,
                (cierre, int(contexto["turno_id"])),
            )
            abiertos = int(
                conn.execute(
                    "SELECT COUNT(*) FROM turnos WHERE dia_operativo_id=? AND estado='ABIERTO'",
                    (int(contexto["dia_operativo_id"]),),
                ).fetchone()[0]
            )
            if not abiertos:
                conn.execute(
                    "UPDATE dias_operativos SET estado='CERRADO' WHERE id=?",
                    (int(contexto["dia_operativo_id"]),),
                )
            conn.commit()
        return True

    def _dedupe_rows_keep_latest(self, rows):
        latest = {}
        for row in sorted(rows, key=lambda r: int(r.get("id", 0))):
            latest[patient_identity_key_from_row(row)] = row
        result = list(latest.values())
        result.sort(key=lambda r: int(r.get("id", 0)), reverse=True)
        return result

    def buscar_atencion_en_turno(
        self,
        nss,
        cedula,
        inicio_turno,
        fin_turno,
        turno_id=None,
        nombre="",
        telefono="",
        dia_operativo_id=None,
    ):
        nss_limpio = re.sub(r"\D", "", (nss or "").strip().upper())
        cedula_limpia = re.sub(r"\D", "", (cedula or "").strip())
        telefono_limpio = re.sub(r"\D", "", (telefono or "").strip())
        nombre_limpio = normalizar_nombre_clave(nombre)

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            paciente_id = None
            if is_valid_cedula_key(cedula_limpia):
                row = conn.execute(
                    """
                    SELECT i.paciente_id
                    FROM paciente_identificadores i
                    JOIN pacientes p ON p.id=i.paciente_id
                    WHERE i.tipo='CEDULA' AND i.valor_normalizado=? AND i.activo=1
                      AND p.estado='ACTIVO'
                    ORDER BY i.conflicto,COALESCE(p.updated_at,p.created_at) DESC,p.id DESC
                    LIMIT 1
                    """,
                    (cedula_limpia,),
                ).fetchone()
                paciente_id = int(row["paciente_id"]) if row else None
            elif is_valid_nss_key(nss_limpio):
                rows = conn.execute(
                    """
                    SELECT p.id,p.nombre,p.telefono_clean,p.telefono
                    FROM paciente_identificadores i
                    JOIN pacientes p ON p.id=i.paciente_id
                    WHERE i.tipo='NSS' AND i.valor_normalizado=? AND i.activo=1
                      AND p.estado='ACTIVO'
                    ORDER BY i.conflicto,COALESCE(p.updated_at,p.created_at) DESC,p.id DESC
                    """,
                    (nss_limpio,),
                ).fetchall()
                for row in rows:
                    mismo_nombre = bool(
                        nombre_limpio
                        and nombre_limpio == normalizar_nombre_clave(row["nombre"])
                    )
                    telefono_row = re.sub(
                        r"\D", "", row["telefono_clean"] or row["telefono"] or ""
                    )
                    mismo_telefono = bool(
                        len(telefono_limpio) == 10 and telefono_limpio == telefono_row
                    )
                    if mismo_nombre or mismo_telefono:
                        paciente_id = int(row["id"])
                        break
            if paciente_id is None and not (nombre_limpio and len(telefono_limpio) == 10):
                return None

            if dia_operativo_id is None and turno_id is not None:
                turno = conn.execute("SELECT dia_operativo_id FROM turnos WHERE id=?", (int(turno_id),)).fetchone()
                dia_operativo_id = int(turno[0]) if turno else None

            parametros = []
            if dia_operativo_id is not None:
                rango_sql = "dia_operativo_id = ?"
                parametros.append(int(dia_operativo_id))
            else:
                rango_sql = "datetime(created_at) >= datetime(?) AND datetime(created_at) < datetime(?)"
                parametros.extend([
                    inicio_turno.strftime("%Y-%m-%d %H:%M:%S"),
                    fin_turno.strftime("%Y-%m-%d %H:%M:%S"),
                ])

            if paciente_id is not None:
                identidad_sql = "paciente_id = ?"
                parametros.insert(0, int(paciente_id))
            else:
                identidad_sql = "UPPER(TRIM(nombre)) = ? AND telefono_clean = ?"
                parametros[0:0] = [nombre_limpio, telefono_limpio]

            sql = f'''
                SELECT * FROM atenciones
                WHERE estado='ACTIVA' AND {identidad_sql} AND {rango_sql}
                ORDER BY id DESC LIMIT 1
            '''
            fila = conn.execute(sql, parametros).fetchone()
        return dict(fila) if fila else None

    def buscar_paciente(self, cedula):
        cedula_limpia = re.sub(r"\D", "", (cedula or ""))
        if not is_valid_cedula_key(cedula_limpia):
            return None
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            return conn.execute(
                """
                SELECT p.* FROM pacientes p
                JOIN paciente_identificadores i ON i.paciente_id=p.id
                WHERE i.tipo='CEDULA' AND i.valor_normalizado=? AND i.activo=1
                ORDER BY p.updated_at DESC, p.id DESC LIMIT 1
                """,
                (cedula_limpia,),
            ).fetchone()

    def buscar_por_nss(self, nss):
        nss_limpio = re.sub(r"\D", "", (nss or ""))
        if not is_valid_nss_key(nss_limpio):
            return None
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            return conn.execute(
                """
                SELECT p.* FROM pacientes p
                JOIN paciente_identificadores i ON i.paciente_id=p.id
                WHERE i.tipo='NSS' AND i.valor_normalizado=? AND i.activo=1
                ORDER BY p.updated_at DESC, p.id DESC LIMIT 1
                """,
                (nss_limpio,),
            ).fetchone()

    def _resolver_o_crear_paciente_conn(self, conn, datos, revisiones_nss=None):
        conn.row_factory = sqlite3.Row
        revisiones_nss = revisiones_nss if revisiones_nss is not None else []
        nss = (datos.get("NSS", "") or "").strip().upper()
        cedula = (datos.get("Cédula", "") or "").strip()
        nss_clean = re.sub(r"\D", "", nss)
        cedula_clean = re.sub(r"\D", "", cedula)
        telefono = (datos.get("Teléfono", "") or "").strip()
        telefono_clean = re.sub(r"\D", "", telefono)
        nombre = (datos.get("Nombre", "") or "").strip() or "SIN NOMBRE"
        nombre_clean = normalizar_nombre_clave(nombre)
        cedula_valida = is_valid_cedula_key(cedula_clean)
        nss_valido = is_valid_nss_key(nss_clean)

        def propietarios(tipo, valor):
            if not valor:
                return []
            return conn.execute(
                """
                SELECT p.*,i.conflicto
                FROM paciente_identificadores i
                JOIN pacientes p ON p.id=i.paciente_id
                WHERE i.tipo=? AND i.valor_normalizado=? AND i.activo=1
                  AND p.estado='ACTIVO'
                ORDER BY i.conflicto,COALESCE(p.updated_at,p.created_at) DESC,p.id DESC
                """,
                (tipo, valor),
            ).fetchall()

        def coincide_demografia(row):
            mismo_nombre = bool(
                nombre_clean
                and nombre_clean != "SIN NOMBRE"
                and nombre_clean == normalizar_nombre_clave(row["nombre"])
            )
            mismo_telefono = bool(
                len(telefono_clean) == 10
                and telefono_clean == re.sub(r"\D", "", row["telefono"] or "")
            )
            return mismo_nombre or mismo_telefono

        cedula_rows = propietarios("CEDULA", cedula_clean) if cedula_valida else []
        nss_rows = propietarios("NSS", nss_clean) if nss_valido else []
        patient_id = int(cedula_rows[0]["id"]) if cedula_rows else None

        # Una ficha creada inicialmente sin cédula puede completarse después.
        if patient_id is None and nss_rows:
            coincidencias = [row for row in nss_rows if coincide_demografia(row)]
            if coincidencias:
                patient_id = int(coincidencias[0]["id"])

        revision_nss = bool(not cedula_valida and nss_valido and nss_rows and patient_id is None)
        referencia_id = int(nss_rows[0]["id"]) if revision_nss else None

        direccion = (datos.get("Dirección", "") or "").strip()
        nacionalidad = (datos.get("Nacionalidad", "") or "").strip()
        ars_canonico = normalizar_seguro(datos.get("Aseguradora (ARS)", ""), nss)
        cedula_db = cedula_clean if cedula_valida else None
        nss_db = nss_clean if nss_valido else None
        telefono_db = telefono_clean if len(telefono_clean) == 10 else None

        if patient_id is not None:
            if cedula_valida:
                # La cédula manda: el NSS nuevo reemplaza todos los NSS anteriores
                # de esta ficha, incluso cuando el campo se deja vacío.
                conn.execute(
                    "DELETE FROM paciente_identificadores WHERE paciente_id=? AND tipo='NSS'",
                    (patient_id,),
                )
            conn.execute(
                """
                UPDATE pacientes SET
                    nombre=?,cedula=COALESCE(?,cedula),telefono=COALESCE(?,telefono),
                    direccion=COALESCE(NULLIF(?,''),direccion),
                    nacionalidad=COALESCE(NULLIF(?,''),nacionalidad),ars=?,
                    nss=?,nss_clean=?,cedula_clean=COALESCE(?,cedula_clean),
                    telefono_clean=COALESCE(?,telefono_clean),provisional=0,
                    requiere_revision=0,updated_at=datetime('now','localtime')
                WHERE id=?
                """,
                (
                    nombre,cedula_db,telefono_db,direccion,nacionalidad,ars_canonico,
                    nss_db,nss_db,cedula_db,telefono_db,patient_id,
                ),
            )
        else:
            provisional = int(not cedula_valida and not nss_valido)
            cursor = conn.execute(
                """
                INSERT INTO pacientes(
                    nombre,cedula,telefono,direccion,nacionalidad,ars,nss,
                    nss_clean,cedula_clean,telefono_clean,provisional,requiere_revision
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    nombre,cedula_db,telefono_db,direccion,nacionalidad,ars_canonico,
                    nss_db,nss_db,cedula_db,telefono_db,provisional,int(revision_nss),
                ),
            )
            patient_id = int(cursor.lastrowid)

        identificadores = []
        if cedula_valida:
            identificadores.append(("CEDULA", cedula_clean, 0))
        if nss_valido:
            identificadores.append(("NSS", nss_clean, int(revision_nss)))
        if revision_nss:
            conn.execute(
                "UPDATE paciente_identificadores SET conflicto=1 "
                "WHERE tipo='NSS' AND valor_normalizado=? AND activo=1",
                (nss_clean,),
            )
        for tipo, valor, conflicto in identificadores:
            conn.execute(
                """
                INSERT INTO paciente_identificadores(
                    paciente_id,tipo,valor_normalizado,activo,conflicto
                ) VALUES (?,?,?,1,?)
                ON CONFLICT(paciente_id,tipo,valor_normalizado)
                DO UPDATE SET activo=1,conflicto=excluded.conflicto
                """,
                (patient_id,tipo,valor,conflicto),
            )

        if revision_nss:
            revisiones_nss.append(
                {
                    "nss": nss_clean,
                    "paciente_nuevo_id": patient_id,
                    "paciente_referencia_id": referencia_id,
                    "detalle": (
                        "El NSS fue registrado sin cédula para datos demográficos diferentes. "
                        "La atención continuó y requiere revisión administrativa."
                    ),
                }
            )
        return patient_id

    def guardar_paciente(self, datos):
        with closing(self._connect()) as conn:
            conn.execute("BEGIN IMMEDIATE")
            revisiones_nss = []
            patient_id = self._resolver_o_crear_paciente_conn(conn, datos, revisiones_nss)
            for revision in revisiones_nss:
                conn.execute(
                    """
                    INSERT INTO nss_conflictos(
                        nss_normalizado,paciente_nuevo_id,paciente_referencia_id,detalle
                    ) VALUES (?,?,?,?)
                    """,
                    (
                        revision["nss"],revision["paciente_nuevo_id"],
                        revision["paciente_referencia_id"],revision["detalle"],
                    ),
                )
            conn.commit()
            return patient_id

    def _registrar_auditoria_conn(
        self,
        conn,
        atencion_id,
        accion,
        motivo,
        usuario,
        snapshot_before,
        snapshot_after=None,
        actor_rol="ADMISION",
    ):
        previous_row = conn.execute(
            "SELECT event_hash FROM atenciones_auditoria WHERE event_hash IS NOT NULL ORDER BY id DESC LIMIT 1"
        ).fetchone()
        previous_hash = str(previous_row[0] or "") if previous_row else ""
        before_json = json.dumps(snapshot_before or {}, ensure_ascii=False, sort_keys=True, default=str)
        after_json = json.dumps(snapshot_after, ensure_ascii=False, sort_keys=True, default=str) if snapshot_after is not None else None
        payload = {
            "atencion_id": int(atencion_id) if atencion_id is not None else None,
            "accion": accion,
            "motivo": (motivo or "").strip(),
            "usuario": limpiar_nombre_representante(usuario),
            "actor_rol": actor_rol,
            "snapshot_before": before_json,
            "snapshot_after": after_json,
            "previous_hash": previous_hash,
            "workstation": platform.node(),
        }
        event_hash = hashlib.sha256(
            json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()
        conn.execute(
            """
            INSERT INTO atenciones_auditoria(
                atencion_id,accion,motivo,usuario,actor_rol,snapshot_json,snapshot_after_json,
                previous_hash,event_hash,workstation
            ) VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            (
                payload["atencion_id"], payload["accion"], payload["motivo"], payload["usuario"],
                actor_rol, before_json, after_json, previous_hash, event_hash, platform.node(),
            ),
        )
        return event_hash

    def guardar_atencion(self, datos, hoja, turno_cfg=None):
        if not turno_cfg or not turno_config_es_vigente(turno_cfg):
            raise TurnoNoVigenteError("El turno no existe o está vencido. Abra el turno actual.")
        ars_canonico = normalizar_seguro(
            datos.get('Aseguradora (ARS)', ''),
            datos.get('NSS', '')
        )

        nss = (datos.get('NSS', '') or '').strip().upper()
        cedula = (datos.get('Cédula', '') or '').strip()
        tipo_atencion = (datos.get("TipoAtencion") or datos.get("tipo_atencion") or "EMERGENCIA").strip().upper()
        if tipo_atencion not in ("EMERGENCIA", "URGENCIA"):
            tipo_atencion = "EMERGENCIA"

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            cur = conn.cursor()
            contexto = self.obtener_contexto_turno(turno_cfg, conn=conn)
            turno_id = contexto["turno_id"]
            dia_operativo_id = contexto["dia_operativo_id"]
            nss_clean = re.sub(r"\D", "", nss)
            cedula_clean = re.sub(r"\D", "", cedula)
            telefono = (datos.get('Teléfono', '') or '').strip()
            telefono_clean = re.sub(r"\D", "", telefono)
            es_reingreso = int(bool(datos.get("EsReingreso") or datos.get("es_reingreso")))
            atencion_origen_id = datos.get("AtencionOrigenId") or datos.get("atencion_origen_id")
            motivo_reingreso = (datos.get("MotivoReingreso") or datos.get("motivo_reingreso") or "").strip()
            autorizado_por = limpiar_nombre_representante(
                datos.get("AutorizadoPor") or datos.get("autorizado_por") or ""
            )
            revisiones_nss = []
            if es_reingreso and (not atencion_origen_id or not motivo_reingreso or not autorizado_por):
                raise ValueError("El reingreso requiere atención original, motivo y autorización.")
            if es_reingreso:
                origen = cur.execute(
                    """
                    SELECT * FROM atenciones
                    WHERE id=? AND dia_operativo_id=? AND estado='ACTIVA'
                    """,
                    (int(atencion_origen_id), dia_operativo_id),
                ).fetchone()
                if not origen:
                    raise ValueError("La atención original no está activa en este día operativo.")
                paciente_id = int(origen["paciente_id"])
            else:
                paciente_id = self._resolver_o_crear_paciente_conn(
                    conn, datos, revisiones_nss
                )

            try:
                cur.execute('''
                    INSERT INTO atenciones (
                        paciente_id,dia_operativo_id,turno_id,nss,nombre,sexo,edad_num,unidad,
                        cedula,telefono,direccion,nacionalidad,ars,hoja,fecha,hora,tipo_atencion,
                        estado,es_reingreso,atencion_origen_id,motivo_reingreso,autorizado_por,
                        identidad_estado,requiere_revision,nss_clean,cedula_clean,telefono_clean
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'ACTIVA',?,?,?,?,?,?,?,?,?)
                ''', (
                    paciente_id,
                    dia_operativo_id,
                    turno_id,
                    nss,
                    datos.get('Nombre', ''),
                    datos.get('Sexo', ''),
                    int(datos.get('Edad_num', 0) or 0),
                    datos.get('Unidad', ''),
                    cedula,
                    telefono,
                    datos.get('Dirección', ''),
                    datos.get('Nacionalidad', ''),
                    ars_canonico,
                    hoja,
                    datos.get('Fecha', ''),
                    datos.get('Hora', ''),
                    tipo_atencion,
                    es_reingreso,
                    int(atencion_origen_id) if atencion_origen_id else None,
                    motivo_reingreso or None,
                    autorizado_por or None,
                    "NSS_EN_REVISION" if revisiones_nss else "VALIDADA",
                    int(bool(revisiones_nss)),
                    nss_clean or None,
                    cedula_clean or None,
                    telefono_clean or None,
                ))
            except sqlite3.IntegrityError as exc:
                conn.rollback()
                if "uq_atencion_dia_paciente" in str(exc) or "UNIQUE constraint failed" in str(exc):
                    raise sqlite3.IntegrityError(
                        "Este paciente ya tiene una atención activa en el día operativo."
                    ) from exc
                raise
            atencion_id = int(cur.lastrowid)
            for revision in revisiones_nss:
                conn.execute(
                    """
                    INSERT INTO nss_conflictos(
                        nss_normalizado,paciente_nuevo_id,paciente_referencia_id,
                        atencion_id,detalle
                    ) VALUES (?,?,?,?,?)
                    """,
                    (
                        revision["nss"],revision["paciente_nuevo_id"],
                        revision["paciente_referencia_id"],atencion_id,
                        revision["detalle"],
                    ),
                )
            snapshot = dict(conn.execute("SELECT * FROM atenciones WHERE id=?", (atencion_id,)).fetchone())
            self._registrar_auditoria_conn(
                conn,
                atencion_id,
                "CREACION",
                "Registro desde formulario principal",
                contexto["representante"],
                {},
                snapshot,
                "ADMISION",
            )
            cur.execute("INSERT INTO trabajos_salida(atencion_id) VALUES (?)", (atencion_id,))
            conn.commit()
            return atencion_id

    def obtener_revision_nss_atencion(self, atencion_id):
        with closing(self._connect()) as conn:
            row = conn.execute(
                "SELECT id FROM nss_conflictos WHERE atencion_id=? AND estado='PENDIENTE' LIMIT 1",
                (int(atencion_id),),
            ).fetchone()
        return int(row[0]) if row else None

    def listar_revisiones_nss(self, solo_pendientes=True, limite=500):
        where = "WHERE c.estado='PENDIENTE'" if solo_pendientes else ""
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                f"""
                SELECT c.*,pn.nombre AS nombre_nuevo,pr.nombre AS nombre_referencia,
                       a.fecha,a.hora,a.estado AS atencion_estado
                FROM nss_conflictos c
                LEFT JOIN pacientes pn ON pn.id=c.paciente_nuevo_id
                LEFT JOIN pacientes pr ON pr.id=c.paciente_referencia_id
                LEFT JOIN atenciones a ON a.id=c.atencion_id
                {where}
                ORDER BY CASE WHEN c.estado='PENDIENTE' THEN 0 ELSE 1 END,c.id DESC
                LIMIT ?
                """,
                (max(1, int(limite)),),
            ).fetchall()
        return [dict(row) for row in rows]

    def resolver_revision_nss(self, revision_id, resolucion, actor, motivo):
        resolucion = str(resolucion or "").strip().upper()
        if resolucion not in {
            "MANTENER_AMBOS", "DESVINCULAR_NSS", "FUSIONAR_CON_EXISTENTE"
        }:
            raise ValueError("Resolución NSS desconocida.")
        actor = limpiar_nombre_representante(actor)
        motivo = str(motivo or "").strip()
        if not actor or len(motivo) < 8:
            raise ValueError("Indique el responsable y un motivo de al menos 8 caracteres.")

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            revision = conn.execute(
                "SELECT * FROM nss_conflictos WHERE id=?",
                (int(revision_id),),
            ).fetchone()
            if not revision or revision["estado"] != "PENDIENTE":
                conn.rollback()
                raise ValueError("La revisión ya no está pendiente.")
            nuevo_id = revision["paciente_nuevo_id"]
            referencia_id = revision["paciente_referencia_id"]
            nss = revision["nss_normalizado"]

            if resolucion == "MANTENER_AMBOS":
                pass
            elif resolucion == "DESVINCULAR_NSS":
                if nuevo_id:
                    conn.execute(
                        "DELETE FROM paciente_identificadores "
                        "WHERE paciente_id=? AND tipo='NSS' AND valor_normalizado=?",
                        (int(nuevo_id), nss),
                    )
                    conn.execute(
                        "UPDATE pacientes SET nss=NULL,nss_clean=NULL,requiere_revision=0,"
                        "updated_at=datetime('now','localtime') "
                        "WHERE id=? AND nss_clean=?",
                        (int(nuevo_id), nss),
                    )
                if revision["atencion_id"]:
                    aid = int(revision["atencion_id"])
                    before = conn.execute("SELECT * FROM atenciones WHERE id=?", (aid,)).fetchone()
                    conn.execute(
                        "UPDATE atenciones SET nss=NULL,nss_clean=NULL,identidad_estado='VALIDADA',"
                        "requiere_revision=0,updated_at=datetime('now','localtime') WHERE id=?",
                        (aid,),
                    )
                    after = conn.execute("SELECT * FROM atenciones WHERE id=?", (aid,)).fetchone()
                    if before and after:
                        self._registrar_auditoria_conn(
                            conn,aid,"CORRECCION_NSS",motivo,actor,dict(before),dict(after),"ADMINISTRADOR"
                        )
            else:
                if not nuevo_id or not referencia_id:
                    conn.rollback()
                    raise ValueError("No existen ambas fichas para realizar la fusión.")
                atenciones = conn.execute(
                    "SELECT * FROM atenciones WHERE paciente_id=? ORDER BY id",
                    (int(nuevo_id),),
                ).fetchall()
                for atencion in atenciones:
                    aid = int(atencion["id"])
                    existente = conn.execute(
                        """
                        SELECT id FROM atenciones
                        WHERE paciente_id=? AND dia_operativo_id=? AND estado='ACTIVA'
                          AND es_reingreso=0 AND id<>? ORDER BY id LIMIT 1
                        """,
                        (int(referencia_id),atencion["dia_operativo_id"],aid),
                    ).fetchone()
                    conn.execute(
                        """
                        UPDATE atenciones SET paciente_id=?,es_reingreso=?,atencion_origen_id=?,
                            motivo_reingreso=?,autorizado_por=?,identidad_estado='VALIDADA',
                            requiere_revision=0,updated_at=datetime('now','localtime')
                        WHERE id=?
                        """,
                        (
                            int(referencia_id),int(bool(existente)),
                            int(existente[0]) if existente else atencion["atencion_origen_id"],
                            motivo if existente else atencion["motivo_reingreso"],
                            actor if existente else atencion["autorizado_por"],aid,
                        ),
                    )
                    after = conn.execute("SELECT * FROM atenciones WHERE id=?", (aid,)).fetchone()
                    self._registrar_auditoria_conn(
                        conn,aid,"FUSION_NSS",motivo,actor,dict(atencion),dict(after),"ADMINISTRADOR"
                    )
                conn.execute(
                    "DELETE FROM paciente_identificadores WHERE paciente_id=?",
                    (int(nuevo_id),),
                )
                conn.execute("DELETE FROM pacientes WHERE id=?", (int(nuevo_id),))

            otros_pendientes = int(
                conn.execute(
                    "SELECT COUNT(*) FROM nss_conflictos "
                    "WHERE nss_normalizado=? AND estado='PENDIENTE' AND id<>?",
                    (nss, int(revision_id)),
                ).fetchone()[0]
            )
            if not otros_pendientes:
                conn.execute(
                    "UPDATE paciente_identificadores SET conflicto=0 "
                    "WHERE tipo='NSS' AND valor_normalizado=? AND activo=1",
                    (nss,),
                )

            if nuevo_id:
                conn.execute(
                    "UPDATE pacientes SET requiere_revision=0 WHERE id=?",
                    (int(nuevo_id),),
                )
            if revision["atencion_id"]:
                conn.execute(
                    "UPDATE atenciones SET identidad_estado='VALIDADA',requiere_revision=0 WHERE id=?",
                    (int(revision["atencion_id"]),),
                )
            conn.execute(
                """
                UPDATE nss_conflictos SET estado='RESUELTO',resolucion=?,motivo_resolucion=?,
                    resuelto_por=?,resuelto_at=datetime('now','localtime') WHERE id=?
                """,
                (resolucion,motivo,actor,int(revision_id)),
            )
            conn.commit()
        return True

    def actualizar_trabajo_salida(
        self,
        atencion_id,
        etapa,
        estado,
        *,
        error="",
        pdf_path=None,
        pdf_sha256=None,
        incrementar_intento=False,
    ):
        columnas = {
            "excel": "excel_estado",
            "pdf": "pdf_estado",
            "impresion": "impresion_estado",
        }
        columna = columnas.get(str(etapa).lower())
        if not columna:
            raise ValueError(f"Etapa de salida desconocida: {etapa}")
        estado = str(estado or "").upper()
        if estado not in {"PENDIENTE", "PROCESANDO", "COMPLETADO", "FALLIDO", "ENVIADO", "DESCONOCIDO"}:
            raise ValueError(f"Estado de salida desconocido: {estado}")
        assignments = [f"{columna}=?", "updated_at=datetime('now','localtime')"]
        params = [estado]
        if error:
            assignments.append("ultimo_error=?")
            params.append(str(error)[:2000])
        if pdf_path is not None:
            assignments.append("pdf_path=?")
            params.append(str(pdf_path))
        if pdf_sha256 is not None:
            assignments.append("pdf_sha256=?")
            params.append(str(pdf_sha256))
        if incrementar_intento:
            assignments.append("intentos=intentos+1")
        params.append(int(atencion_id))
        with closing(self._connect()) as conn:
            conn.execute(
                f"UPDATE trabajos_salida SET {', '.join(assignments)} WHERE atencion_id=?",
                params,
            )
            conn.commit()

    def limpiar_error_trabajo_salida(self, atencion_id):
        with closing(self._connect()) as conn:
            conn.execute(
                """
                UPDATE trabajos_salida SET ultimo_error=NULL,updated_at=datetime('now','localtime')
                WHERE atencion_id=?
                """,
                (int(atencion_id),),
            )
            conn.commit()

    def obtener_trabajo_salida(self, atencion_id):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT * FROM trabajos_salida WHERE atencion_id=?", (int(atencion_id),)).fetchone()
            return dict(row) if row else None

    def listar_trabajos_salida_pendientes(self, limite=100):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute(
                """
                UPDATE trabajos_salida SET
                    excel_estado=CASE WHEN excel_estado='PROCESANDO' THEN 'PENDIENTE' ELSE excel_estado END,
                    pdf_estado=CASE WHEN pdf_estado='PROCESANDO' THEN 'PENDIENTE' ELSE pdf_estado END,
                    impresion_estado=CASE WHEN impresion_estado='PROCESANDO' THEN 'PENDIENTE' ELSE impresion_estado END,
                    updated_at=datetime('now','localtime')
                WHERE excel_estado='PROCESANDO' OR pdf_estado='PROCESANDO' OR impresion_estado='PROCESANDO'
                """
            )
            conn.commit()
            rows = conn.execute(
                """
                SELECT t.*, a.nombre, a.hoja, a.fecha, a.hora
                FROM trabajos_salida t
                JOIN atenciones a ON a.id=t.atencion_id
                WHERE a.estado='ACTIVA' AND (
                    t.excel_estado IN ('PENDIENTE','FALLIDO') OR
                    t.pdf_estado IN ('PENDIENTE','FALLIDO') OR
                    t.impresion_estado IN ('PENDIENTE','FALLIDO')
                )
                ORDER BY t.updated_at, t.atencion_id
                LIMIT ?
                """,
                (max(1, int(limite)),),
            ).fetchall()
            return [dict(row) for row in rows]

    def registrar_documento(self, atencion_id, tipo, ruta, plantilla=""):
        digest = hashlib.sha256()
        with open(ruta, "rb") as stream:
            for block in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(block)
        sha256 = digest.hexdigest()
        with closing(self._connect()) as conn:
            conn.execute(
                "INSERT INTO documentos(atencion_id,tipo,ruta,sha256,plantilla) VALUES (?,?,?,?,?)",
                (int(atencion_id), str(tipo).upper(), os.path.abspath(ruta), sha256, plantilla),
            )
            conn.commit()
        return sha256

    def obtener_documento_atencion(self, atencion_id, tipo="HOJA_EMERGENCIA"):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                """
                SELECT * FROM documentos
                WHERE atencion_id=? AND tipo=?
                ORDER BY id DESC LIMIT 1
                """,
                (int(atencion_id), str(tipo).upper()),
            ).fetchone()
        if not row:
            return None
        documento = dict(row)
        ruta = documento.get("ruta") or ""
        if not os.path.isfile(ruta):
            return None
        digest = hashlib.sha256()
        with open(ruta, "rb") as stream:
            for block in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(block)
        if digest.hexdigest() != (documento.get("sha256") or ""):
            APP_LOG.error("El PDF archivado de la atención #%s no supera SHA-256", atencion_id)
            return None
        return documento

    def borrar_atencion(self, atencion_id: int, motivo="Eliminada desde el historial", usuario="") -> bool:
        motivo = str(motivo or "").strip()
        usuario = limpiar_nombre_representante(usuario)
        if len(motivo) < 5:
            raise ValueError("La anulación requiere un motivo de al menos 5 caracteres.")
        if not usuario:
            raise ValueError("La anulación requiere identificar al operador.")
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            cur = conn.cursor()
            fila = cur.execute('SELECT * FROM atenciones WHERE id = ?', (atencion_id,)).fetchone()
            if not fila or str(fila["estado"] or "ACTIVA").upper() == "ANULADA":
                conn.rollback()
                return False
            reingresos_activos = int(
                cur.execute(
                    """
                    SELECT COUNT(*) FROM atenciones
                    WHERE atencion_origen_id=? AND estado='ACTIVA'
                    """,
                    (int(atencion_id),),
                ).fetchone()[0]
            )
            if reingresos_activos:
                conn.rollback()
                raise ValueError(
                    f"No se puede anular: existen {reingresos_activos} reingreso(s) activo(s) vinculados."
                )
            before = dict(fila)
            cur.execute(
                """
                UPDATE atenciones SET
                    estado='ANULADA', anulada_at=datetime('now','localtime'),
                    anulada_por=?, anulada_motivo=?, updated_at=datetime('now','localtime')
                WHERE id=? AND estado='ACTIVA'
                """,
                (usuario, motivo, int(atencion_id)),
            )
            changed = cur.rowcount > 0
            after = dict(cur.execute("SELECT * FROM atenciones WHERE id=?", (atencion_id,)).fetchone())
            self._registrar_auditoria_conn(
                conn,
                atencion_id,
                "ANULACION",
                motivo,
                usuario,
                before,
                after,
                "OPERADOR",
            )
            conn.commit()
            return changed

    def reordenar_ids_atenciones_despues_de_eliminar(self, atencion_id_eliminada: int) -> bool:
        """Los ID son referencias permanentes y nunca se renumeran."""
        APP_LOG.warning("Se ignoró una solicitud de renumeración después de eliminar #%s", atencion_id_eliminada)
        return False

    def listar_atenciones(self, filtro_texto=None, limite=200, offset=0):
        limite = max(1, int(limite or 200))
        offset = max(0, int(offset or 0))
        q = (filtro_texto or "").strip()
        q_digits = re.sub(r"\D", "", q)

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            if q:
                like = f"%{q.upper()}%"
                like_digits = f"%{q_digits or q}%"

                cur.execute('''
                    SELECT id, fecha, hora, nombre, hoja, ars, nss, cedula, edad_num, unidad, tipo_atencion
                    FROM atenciones
                    WHERE estado='ACTIVA' AND (
                        UPPER(IFNULL(nombre,'')) LIKE ?
                        OR UPPER(IFNULL(ars,'')) LIKE ?
                        OR UPPER(IFNULL(nss,'')) LIKE ?
                        OR UPPER(IFNULL(nss,'')) LIKE ?
                        OR IFNULL(cedula,'') LIKE ?
                        OR IFNULL(telefono,'') LIKE ?
                    )
                    ORDER BY id DESC
                    LIMIT ? OFFSET ?
                ''', (like, like, like, like_digits, like_digits, like_digits, limite, offset))
            else:
                cur.execute('''
                    SELECT id, fecha, hora, nombre, hoja, ars, nss, cedula, edad_num, unidad, tipo_atencion
                    FROM atenciones
                    WHERE estado='ACTIVA'
                    ORDER BY id DESC
                    LIMIT ? OFFSET ?
                ''', (limite, offset))

            return [dict(r) for r in cur.fetchall()]

    def obtener_atencion_por_id(self, atencion_id):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute('SELECT * FROM atenciones WHERE id = ?', (atencion_id,))
            r = cur.fetchone()
            return dict(r) if r else None

    def obtener_turno_config_atencion(self, atencion_id):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                """
                SELECT d.fecha_base,t.tipo_turno,t.representante,t.fecha_inicio
                FROM atenciones a
                JOIN dias_operativos d ON d.id=a.dia_operativo_id
                JOIN turnos t ON t.id=a.turno_id
                WHERE a.id=?
                """,
                (int(atencion_id),),
            ).fetchone()
        if not row:
            return None
        fecha_base = date.fromisoformat(row["fecha_base"])
        inicio_real = parse_datetime_local(row["fecha_inicio"])
        return {
            "representante": limpiar_nombre_representante(row["representante"] or ""),
            "turno_codigo": normalizar_turno_codigo(row["tipo_turno"] or "8AM_8AM"),
            "fecha_base": fecha_base,
            "inicio_real": format_datetime_local(inicio_real) if inicio_real else "",
            "inicio_real_dt": inicio_real,
        }

    def listar_atenciones_sin_seguro(self, filtro_texto=None, limite=200, offset=0):
        """
        FASE 2: Historial sin seguro dedicado, sin pasar por listar_atenciones_filtradas.
        SQLite filtra directamente. Sin dedupe (id DESC es único).
        """
        limite = max(1, int(limite or 200))
        offset = max(0, int(offset or 0))

        q = (filtro_texto or "").strip()
        q_digits = re.sub(r"\D", "", q)

        sin_seguro_aliases = [
            "SIN SEGURO", "NO TIENE", "NO", "N/S", "NS", "N\\S",
            "INACTIVO", "INACTIVA", "PENDIENTE", "PENDIENTES",
            "NO ACTIVO", "NO VIGENTE", "VENCIDO", "CANCELADO",
            "DESAFILIADO", "PARTICULAR", "PRIVADO", "NINGUNO", "NINGUNA",
            "N/A", "NA", "S/N", "SN", "NO APLICA", "NO USA", "NO POSEE",
            "NINGUN SEGURO"
        ]

        where = [
            "estado='ACTIVA'",
            """(
                TRIM(IFNULL(ars,'')) = ''
                OR UPPER(TRIM(IFNULL(ars,''))) IN ({})
                OR TRIM(IFNULL(nss,'')) = ''
                OR UPPER(TRIM(IFNULL(nss,''))) IN ('SIN SEGURO','NO','N/S','NS','N\\S')
                OR LENGTH(TRIM(IFNULL(ars,''))) = 1
                OR (
                    TRIM(IFNULL(ars,'')) <> ''
                    AND TRIM(IFNULL(ars,'')) NOT GLOB '*[^0-9]*'
                )
            )""".format(",".join(["?"] * len(sin_seguro_aliases)))
        ]
        params = [a.upper() for a in sin_seguro_aliases]

        if q:
            like = f"%{q.upper()}%"
            like_digits = f"%{q_digits or q}%"
            where.append("""(
                UPPER(IFNULL(nombre,'')) LIKE ?
        OR UPPER(IFNULL(ars,'')) LIKE ?
        OR UPPER(IFNULL(nss,'')) LIKE ?
        OR IFNULL(nss_clean,'') LIKE ?
        OR IFNULL(cedula_clean,'') LIKE ?
        OR IFNULL(telefono_clean,'') LIKE ?            
            )""")
            params.extend([like, like, like, like_digits, like_digits, like_digits])

        sql = f"""
            SELECT id, fecha, hora, nombre, hoja, ars, nss, cedula, edad_num, unidad, tipo_atencion
            FROM atenciones
            WHERE {' AND '.join(where)}
            ORDER BY id DESC
            LIMIT ? OFFSET ?
        """
        params.extend([limite, offset])

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(sql, params)
            filas = [dict(r) for r in cur.fetchall()]

        for f in filas:
            f["ars"] = "SIN SEGURO"
        return filas

    def listar_ars_distintas(self):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("""
                SELECT DISTINCT ars
                FROM atenciones
                WHERE estado='ACTIVA' AND TRIM(IFNULL(ars,'')) <> ''
                ORDER BY ars ASC
            """)
            valores_unicos = [r["ars"] for r in cur.fetchall()]
            valores_norm = set()
            for ars_raw in valores_unicos:
                canon = normalizar_seguro(ars_raw, "9999")
                if canon and canon != "SIN SEGURO":
                    valores_norm.add(canon)
            return sorted(valores_norm)

    def buscar_paciente_para_edicion(self, identidad: str):
        ident = (identidad or "").strip().replace("-", "").upper()
        if not ident:
            return None

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            patient_id = None
            attention = None
            if ident.startswith("A:") and ident[2:].isdigit():
                attention = conn.execute(
                    "SELECT * FROM atenciones WHERE id=? LIMIT 1", (int(ident[2:]),)
                ).fetchone()
                if attention:
                    patient_id = int(attention["paciente_id"])
            elif ident.startswith("P:") and ident[2:].isdigit():
                patient_id = int(ident[2:])
            else:
                ident_normalizado = re.sub(r"\D", "", ident)
                rows = conn.execute(
                    """
                    SELECT paciente_id FROM paciente_identificadores
                    WHERE activo=1 AND valor_normalizado=?
                    ORDER BY conflicto,id DESC LIMIT 1
                    """,
                    (ident_normalizado,),
                ).fetchall()
                if rows:
                    patient_id = int(rows[0][0])
            if patient_id is None:
                return None

            patient = conn.execute("SELECT * FROM pacientes WHERE id=?", (patient_id,)).fetchone()
            if not patient:
                return None
            if attention is None:
                attention = conn.execute(
                    "SELECT * FROM atenciones WHERE paciente_id=? ORDER BY id DESC LIMIT 1",
                    (patient_id,),
                ).fetchone()
            data = dict(patient)
            data["paciente_id"] = patient_id
            if attention:
                data.update(dict(attention))
                data["paciente_id"] = patient_id
            return data

    def _resolver_paciente_para_eliminacion(self, conn, paciente_id: int):
        try:
            paciente_id = int(paciente_id)
        except (TypeError, ValueError):
            return None
        conn.row_factory = sqlite3.Row
        paciente = conn.execute("SELECT * FROM pacientes WHERE id=?", (paciente_id,)).fetchone()
        if not paciente:
            return None
        atenciones = [
            dict(row) for row in conn.execute(
                "SELECT * FROM atenciones WHERE paciente_id=? ORDER BY id", (paciente_id,)
            ).fetchall()
        ]
        ids = [int(row["id"]) for row in atenciones]
        auditorias = 0
        documentos = []
        if ids:
            marks = ",".join("?" for _ in ids)
            auditorias = int(conn.execute(
                f"SELECT COUNT(*) FROM atenciones_auditoria WHERE atencion_id IN ({marks})", ids
            ).fetchone()[0])
            documentos = [row[0] for row in conn.execute(
                f"SELECT ruta FROM documentos WHERE atencion_id IN ({marks})", ids
            ).fetchall()]
        return {
            "seguro": True,
            "paciente_id": paciente_id,
            "paciente": dict(paciente),
            "atenciones": atenciones,
            "fichas": 1,
            "auditorias": auditorias,
            "documentos": documentos,
        }

    def previsualizar_eliminacion_paciente(self, paciente_id: int):
        with closing(self._connect()) as conn:
            return self._resolver_paciente_para_eliminacion(conn, paciente_id)

    def eliminar_paciente_completo(
        self,
        paciente_id: int,
        motivo: str,
        actor: str,
        *,
        confirmado_prueba: bool = False,
    ):
        motivo = (motivo or "").strip()
        actor = limpiar_nombre_representante(actor)
        if not confirmado_prueba:
            raise PermissionError("La purga física solo está permitida para datos confirmados como prueba.")
        if len(motivo) < 8:
            raise ValueError("Debe indicar un motivo de eliminación de al menos 8 caracteres.")
        if not actor:
            raise PermissionError("La purga requiere un actor administrativo identificado.")

        preview = self.previsualizar_eliminacion_paciente(paciente_id)
        if not preview:
            return None
        backup_folder = self.backup_manager.create(
            "antes_purga_paciente",
            label=f"paciente_id={int(paciente_id)}; actor={actor}",
        )
        cuarentena = os.path.join(
            BACKUPS_DIR,
            "CUARENTENA_PURGA",
            f"paciente_{int(paciente_id)}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}",
        )
        os.makedirs(cuarentena, exist_ok=False)
        archivos_movidos = []
        try:
            for indice, ruta in enumerate(preview.get("documentos", []), start=1):
                if not ruta or not os.path.isfile(ruta):
                    continue
                destino = os.path.join(cuarentena, f"{indice:04d}_{os.path.basename(ruta)}")
                os.replace(os.path.abspath(ruta), destino)
                archivos_movidos.append((os.path.abspath(ruta), destino))
        except Exception:
            for original, temporal in reversed(archivos_movidos):
                try:
                    if os.path.isfile(temporal):
                        os.makedirs(os.path.dirname(original), exist_ok=True)
                        os.replace(temporal, original)
                except OSError:
                    APP_LOG.critical("No se pudo revertir la cuarentena de %s", original)
            shutil.rmtree(cuarentena, ignore_errors=True)
            raise RuntimeError(
                "No se pudieron aislar todos los documentos; la base no fue modificada."
            )

        try:
            with closing(self._connect()) as conn:
                conn.row_factory = sqlite3.Row
                conn.execute("BEGIN IMMEDIATE")
                try:
                    resumen = self._resolver_paciente_para_eliminacion(conn, int(paciente_id))
                    if not resumen or not resumen.get("seguro"):
                        raise RuntimeError("La ficha dejó de estar disponible antes de la purga.")

                    ids = [int(row["id"]) for row in resumen["atenciones"]]
                    auditoria_ids = []
                    if ids:
                        marks = ",".join("?" for _ in ids)
                        auditoria_ids = [
                            int(row[0])
                            for row in conn.execute(
                                f"SELECT id FROM atenciones_auditoria WHERE atencion_id IN ({marks})",
                                ids,
                            ).fetchall()
                        ]
                    conn.execute("DELETE FROM atenciones WHERE paciente_id=?", (int(paciente_id),))
                    atenciones_eliminadas = int(conn.execute("SELECT changes()").fetchone()[0])
                    conn.execute("DELETE FROM pacientes WHERE id=?", (int(paciente_id),))
                    fichas_eliminadas = int(conn.execute("SELECT changes()").fetchone()[0])

                    previous_row = conn.execute(
                        "SELECT event_hash FROM purga_eventos ORDER BY id DESC LIMIT 1"
                    ).fetchone()
                    previous_hash = str(previous_row[0] or "") if previous_row else ""
                    paciente_hash = hashlib.sha256(
                        os.urandom(32) + str(int(paciente_id)).encode("ascii")
                    ).hexdigest()
                    event_payload = {
                        "paciente_hash": paciente_hash,
                        "motivo": motivo,
                        "actor": actor,
                        "actor_rol": "ADMINISTRADOR",
                        "backup_path": str(backup_folder),
                        "atenciones_eliminadas": atenciones_eliminadas,
                        "fichas_eliminadas": fichas_eliminadas,
                        "previous_hash": previous_hash,
                        "workstation": platform.node(),
                    }
                    event_hash = hashlib.sha256(
                        json.dumps(event_payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
                    ).hexdigest()
                    conn.execute(
                        """
                        INSERT INTO purga_eventos(
                            paciente_hash,motivo,actor,actor_rol,backup_path,
                            atenciones_eliminadas,fichas_eliminadas,previous_hash,event_hash,workstation
                        ) VALUES (?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            paciente_hash, motivo, actor, "ADMINISTRADOR", str(backup_folder),
                            atenciones_eliminadas, fichas_eliminadas,
                            previous_hash, event_hash, platform.node(),
                        ),
                    )
                    auditorias_redactadas = 0
                    if auditoria_ids:
                        audit_marks = ",".join("?" for _ in auditoria_ids)
                        conn.execute(
                            f"""
                            UPDATE atenciones_auditoria SET
                                accion='REDACTADO_POR_PURGA',motivo=?,usuario=?,
                                snapshot_json='{{}}',snapshot_after_json='{{}}'
                            WHERE id IN ({audit_marks})
                            """,
                            [f"Purga administrativa: {event_hash}", actor, *auditoria_ids],
                        )
                        auditorias_redactadas = int(
                            conn.execute("SELECT changes()").fetchone()[0]
                        )
                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise
        except Exception:
            for original, temporal in reversed(archivos_movidos):
                try:
                    if os.path.isfile(temporal):
                        os.makedirs(os.path.dirname(original), exist_ok=True)
                        os.replace(temporal, original)
                except OSError:
                    APP_LOG.critical("No se pudo restaurar el documento %s", original)
            shutil.rmtree(cuarentena, ignore_errors=True)
            APP_LOG.exception("Falló la eliminación total de un paciente")
            raise

        archivos_pendientes = []
        for _original, temporal in archivos_movidos:
            try:
                eliminar_archivo_sensible(temporal)
            except OSError:
                archivos_pendientes.append(temporal)
                APP_LOG.critical("Documento de purga pendiente de eliminar: %s", temporal)
        try:
            if not os.listdir(cuarentena):
                os.rmdir(cuarentena)
        except OSError:
            pass

        resumen.pop("paciente", None)
        resumen.pop("atenciones", None)
        resumen.pop("documentos", None)
        resumen.update(
            {
                "atenciones_eliminadas": atenciones_eliminadas,
                "fichas_eliminadas": fichas_eliminadas,
                "auditorias_redactadas": auditorias_redactadas,
                "documentos_eliminados": len(archivos_movidos) - len(archivos_pendientes),
                "documentos_pendientes": archivos_pendientes,
                "backup_path": str(backup_folder),
                "purga_event_hash": event_hash,
            }
        )
        APP_LOG.warning(
            "Purga administrativa completada: event_hash=%s, atenciones=%s, fichas=%s, pendientes=%s",
            event_hash,
            atenciones_eliminadas,
            fichas_eliminadas,
            len(archivos_pendientes),
        )
        return resumen

    def actualizar_datos_paciente_por_identidad(self, identidad_original: str, nuevos: dict, actualizar_ficha=True):
        ident = (identidad_original or "").strip().upper()
        if not ident:
            return 0, 0

        nombre = (nuevos.get("Nombre") or "").strip()
        cedula = (nuevos.get("Cédula") or "").strip().replace("-", "")
        telefono = (nuevos.get("Teléfono") or "").strip().replace("-", "")
        telefono_db = telefono if telefono.isdigit() and len(telefono) == 10 else None
        direccion = (nuevos.get("Dirección") or "").strip()
        nacionalidad = (nuevos.get("Nacionalidad") or "").strip()
        nss = (nuevos.get("NSS") or "").strip().upper()
        ars = normalizar_seguro(nuevos.get("Aseguradora (ARS)", ""), nss)

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            cur = conn.cursor()
            objetivo = None
            patient_id = None
            if ident.startswith("P:") and ident[2:].isdigit():
                patient_id = int(ident[2:])
            elif ident.startswith("A:") and ident[2:].isdigit():
                objetivo = cur.execute(
                    "SELECT * FROM atenciones WHERE id=? LIMIT 1", (int(ident[2:]),)
                ).fetchone()
            else:
                clean_ident = re.sub(r"\D", "", ident)
                patient_rows = cur.execute(
                    "SELECT paciente_id FROM paciente_identificadores "
                    "WHERE valor_normalizado=? AND activo=1 "
                    "ORDER BY conflicto,id DESC LIMIT 1",
                    (clean_ident,),
                ).fetchall()
                patient_id = int(patient_rows[0][0]) if patient_rows else None
            if objetivo:
                patient_id = int(objetivo["paciente_id"])
            elif patient_id is not None:
                objetivo = cur.execute('''
                    SELECT * FROM atenciones
                    WHERE paciente_id=?
                    ORDER BY id DESC LIMIT 1
                ''', (patient_id,)).fetchone()
            if patient_id is None:
                conn.rollback()
                return 0, 0

            nss_clean = re.sub(r"\D", "", nss)
            cedula_clean = re.sub(r"\D", "", cedula)
            nuevos_ids = []
            if is_valid_nss_key(nss_clean):
                nuevos_ids.append(("NSS", nss_clean))
            if is_valid_cedula_key(cedula_clean):
                nuevos_ids.append(("CEDULA", cedula_clean))
            atenciones_actualizadas = 0
            if objetivo:
                objetivo_id = int(objetivo["id"])
                cur.execute('''
                    UPDATE atenciones
                    SET nombre=?, cedula=?, telefono=?, direccion=?, nacionalidad=?, nss=?, ars=?,
                        nss_clean=?, cedula_clean=?, telefono_clean=?, identidad_estado=?,
                        requiere_revision=?, updated_at=datetime('now','localtime')
                    WHERE id=?
                ''', (
                    nombre, cedula, telefono, direccion, nacionalidad, nss, ars,
                    nss_clean or None, cedula_clean or None,
                    re.sub(r"\D", "", telefono) or None, "VALIDADA", 0, objetivo_id,
                ))
                atenciones_actualizadas = cur.rowcount
                after = dict(cur.execute("SELECT * FROM atenciones WHERE id=?", (objetivo_id,)).fetchone())
                self._registrar_auditoria_conn(
                    conn,
                    objetivo_id,
                    "MODIFICACION",
                    "Edición de una atención y/o ficha del paciente",
                    limpiar_nombre_representante((cargar_turno_config(permitir_vencido=True) or {}).get("representante", "")),
                    dict(objetivo),
                    after,
                    "OPERADOR",
                )

            pacientes_actualizados = 0
            if actualizar_ficha:
                for kind, value in nuevos_ids:
                    if kind != "CEDULA":
                        continue
                    conflict = cur.execute(
                        "SELECT paciente_id FROM paciente_identificadores "
                        "WHERE tipo=? AND valor_normalizado=? AND activo=1 AND paciente_id<>? LIMIT 1",
                        (kind, value, patient_id),
                    ).fetchone()
                    if conflict:
                        conn.rollback()
                        raise ValueError(
                            f"El {kind} indicado ya está asignado a otra ficha."
                        )
                cur.execute('''
                    UPDATE pacientes SET
                        cedula=?, nombre=?, telefono=?, direccion=?, nacionalidad=?, ars=?, nss=?,
                        nss_clean=?, cedula_clean=?, telefono_clean=?, provisional=?,
                        updated_at=datetime('now','localtime')
                    WHERE id=?
                ''', (
                    cedula_clean if is_valid_cedula_key(cedula_clean) else None,
                    nombre, telefono_db, direccion, nacionalidad, ars,
                    nss_clean if is_valid_nss_key(nss_clean) else None,
                    nss_clean if is_valid_nss_key(nss_clean) else None,
                    cedula_clean if is_valid_cedula_key(cedula_clean) else None,
                    re.sub(r"\D", "", telefono) or None,
                    int(not nuevos_ids),
                    patient_id,
                ))
                pacientes_actualizados = cur.rowcount
                for kind in ("NSS", "CEDULA"):
                    new_value = next((value for item_kind, value in nuevos_ids if item_kind == kind), None)
                    cur.execute(
                        "DELETE FROM paciente_identificadores WHERE paciente_id=? AND tipo=?",
                        (patient_id, kind),
                    )
                    if new_value:
                        cur.execute(
                            "INSERT INTO paciente_identificadores(" 
                            "paciente_id,tipo,valor_normalizado,activo,conflicto) VALUES (?,?,?,1,0) "
                            "ON CONFLICT(paciente_id,tipo,valor_normalizado) "
                            "DO UPDATE SET activo=1,conflicto=0",
                            (patient_id, kind, new_value),
                        )

            conn.commit()
            return atenciones_actualizadas, pacientes_actualizados

    def listar_ars_conteo(self):
        conteo = {}
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                "SELECT ars, nss, COUNT(*) AS cantidad "
                "FROM atenciones "
                "WHERE estado='ACTIVA' AND TRIM(IFNULL(ars,'')) <> '' "
                "GROUP BY ars, nss"
            )
            for r in cur.fetchall():
                canon = normalizar_seguro(r["ars"], r["nss"])
                if canon:
                    conteo[canon] = conteo.get(canon, 0) + int(r["cantidad"] or 0)
        return sorted(conteo.items(), key=lambda x: (-x[1], x[0]))

    def reemplazar_ars_global(self, ars_actual: str, ars_nueva: str):
        actual = normalizar_seguro(ars_actual, "9999")
        nueva = normalizar_seguro(ars_nueva, "9999")
        if not actual or not nueva:
            return 0

        total = 0
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ars, nss FROM pacientes WHERE TRIM(IFNULL(ars,'')) <> ''")
            raw_to_update = [r["ars"] for r in cur.fetchall() if normalizar_seguro(r["ars"], r["nss"]) == actual]
            for raw in set(raw_to_update):
                cur.execute("UPDATE pacientes SET ars = ? WHERE ars = ?", (nueva, raw))
                total += cur.rowcount

            conn.commit()
        return total

    def normalizar_todas_ars(self):
        total = 0
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ars, nss FROM pacientes WHERE TRIM(IFNULL(ars,'')) <> ''")
            for r in cur.fetchall():
                canon = normalizar_seguro(r["ars"], r["nss"])
                if canon and canon != r["ars"]:
                    cur.execute("UPDATE pacientes SET ars = ?,updated_at=datetime('now','localtime') WHERE ars = ?", (canon, r["ars"]))
                    total += cur.rowcount
            conn.commit()
        return total

    def limpiar_ars_cortas_invalidas(self):
        total = 0
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ars FROM pacientes WHERE TRIM(IFNULL(ars,'')) <> ''")
            for (raw,) in cur.fetchall():
                if ars_es_corta_invalida(raw):
                    cur.execute("UPDATE pacientes SET ars='SIN SEGURO',updated_at=datetime('now','localtime') WHERE ars=?", (raw,))
                    total += cur.rowcount
            conn.commit()
        return total

    def listar_atenciones_filtradas(self, filtro_texto=None, modo="Todos", ars=None, especialidad=None, fecha_txt=None, limite=200, offset=0):
        """
        FASE 5: Sin dedupe (id DESC ya es único).
        """
        limite = max(1, int(limite or 200))
        offset = max(0, int(offset or 0))
        modo = (modo or "Todos").strip()
        ars = (ars or "").strip()
        especialidad = (especialidad or "").strip().upper()
        fecha_obj = parse_fecha_ddmmyyyy(fecha_txt) if fecha_txt else None

        where = ["estado='ACTIVA'"]
        params = []

        q = (filtro_texto or "").strip()
        if q:
            q_digits = re.sub(r"\D", "", q)
            like = f"%{q.upper()}%"
            like_digits = f"%{q_digits or q}%"
            where.append('''(
                UPPER(IFNULL(nombre,'')) LIKE ?
                OR UPPER(IFNULL(ars,'')) LIKE ?
                OR UPPER(IFNULL(nss,'')) LIKE ?
                OR UPPER(IFNULL(nss,'')) LIKE ?
                OR IFNULL(cedula,'') LIKE ?
                OR IFNULL(telefono,'') LIKE ?
            )''')
            params.extend([like, like, like, like_digits, like_digits, like_digits])

        if modo == "Hoy":
            where.append(
                "dia_operativo_id=(SELECT id FROM dias_operativos WHERE fecha_base=? LIMIT 1)"
            )
            params.append(fecha_base_operativa_actual().isoformat())

        if modo == "Por fecha" and fecha_obj:
            where.append(
                "dia_operativo_id=(SELECT id FROM dias_operativos WHERE fecha_base=? LIMIT 1)"
            )
            params.append(fecha_obj.isoformat())

        if modo == "Por especialidad" and especialidad and especialidad != "(TODAS)":
            where.append("UPPER(IFNULL(hoja,'')) = ?")
            params.append(especialidad)

        if modo == "Por ARS" and ars and ars != "(Todas)":
            canon_ars = normalizar_seguro(ars, "9999")
            alias_values = [canon_ars]
            try:
                alias_values.extend(cargar_catalogo_ars().get(canon_ars, []))
                alias_values.append(seguro_para_mostrar(canon_ars))
            except Exception:
                pass
            alias_values = sorted({str(a).strip().upper() for a in alias_values if str(a).strip()})
            if alias_values:
                where.append("(" + " OR ".join(["UPPER(IFNULL(ars,'')) LIKE ?"] * len(alias_values)) + ")")
                params.extend([f"%{a}%" for a in alias_values])

        if modo == "Turno actual":
            turno_cfg = cargar_turno_config()
            if not turno_cfg:
                return []
            contexto = self.buscar_contexto_turno_existente(turno_cfg)
            if not contexto:
                return []
            where.append("turno_id=?")
            params.append(int(contexto["turno_id"]))

        sql = '''
            SELECT id, fecha, hora, nombre, hoja, ars, nss, cedula, edad_num, unidad, tipo_atencion
            FROM atenciones
        '''
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY id DESC LIMIT ? OFFSET ?"

        necesita_python = modo == "Sin seguro"
        fetch_limit = limite * 3 if necesita_python else limite
        params_sql = params + [fetch_limit, offset]

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(sql, params_sql)
            filas = [dict(r) for r in cur.fetchall()]

        if not necesita_python:
            return filas
        
        result = []
        for f in filas:
            canon = normalizar_seguro(f.get("ars", ""), f.get("nss", ""))
            if modo == "Sin seguro" and canon != "SIN SEGURO":
                continue
            result.append(f)
            if len(result) >= limite:
                break

        return result

    def buscar_pacientes_avanzado(self, texto: str, limite=80):
        q = (texto or "").strip().replace("-", "")
        if not q: return []
        like = f"%{q}%"
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("""
                SELECT id, paciente_id, fecha, hora, nombre, hoja, ars, nss, cedula, telefono, direccion, nacionalidad
                FROM atenciones
                WHERE estado='ACTIVA' AND (nombre LIKE ? OR nss LIKE ? OR cedula LIKE ? OR telefono LIKE ?)
                ORDER BY id DESC
                LIMIT ?
            """, (like, like, like, like, limite))
            rows = [dict(r) for r in cur.fetchall()]
        latest_by_patient = {}
        for row in rows:
            patient_id = int(row.get("paciente_id") or 0)
            if patient_id not in latest_by_patient:
                latest_by_patient[patient_id] = row
        return list(latest_by_patient.values())

    def actualizar_atencion_especifica(
        self,
        atencion_id: int,
        nuevos: dict,
        actualizar_ficha=False,
        usuario="",
        motivo="Corrección de atención específica",
    ):
        nombre = (nuevos.get("Nombre") or "").strip()
        fecha = (nuevos.get("Fecha") or "").strip()
        hora = (nuevos.get("Hora") or "").strip()
        hoja = (nuevos.get("Hoja") or "").strip().upper()
        ars = normalizar_seguro(nuevos.get("Aseguradora (ARS)", ""), nuevos.get("NSS", ""))
        nss = (nuevos.get("NSS") or "").strip().upper()
        cedula = (nuevos.get("Cédula") or "").strip().replace("-", "")
        telefono = (nuevos.get("Teléfono") or "").strip().replace("-", "")
        telefono_db = telefono if telefono.isdigit() and len(telefono) == 10 else None
        direccion = (nuevos.get("Dirección") or "").strip()
        nacionalidad = (nuevos.get("Nacionalidad") or "").strip()
        sexo = (nuevos.get("Sexo") or "").strip()
        if sexo not in ("Masculino", "Femenino"):
            raise ValueError("El sexo debe ser Masculino o Femenino.")
        edad_num = int(nuevos.get("Edad_num", 0) or 0)
        unidad = (nuevos.get("Unidad") or "Años").strip()
        tipo_atencion = (nuevos.get("TipoAtencion") or "EMERGENCIA").strip().upper()
        if tipo_atencion not in ("EMERGENCIA", "URGENCIA"):
            tipo_atencion = "EMERGENCIA"

        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("BEGIN IMMEDIATE")
            cur = conn.cursor()
            snapshot = cur.execute("SELECT * FROM atenciones WHERE id=?", (int(atencion_id),)).fetchone()
            if not snapshot:
                conn.rollback()
                return 0
            if str(snapshot["estado"] or "ACTIVA").upper() != "ACTIVA":
                conn.rollback()
                raise ValueError("No se puede editar una atención anulada.")

            nss_clean = re.sub(r"\D", "", nss)
            cedula_clean = re.sub(r"\D", "", cedula)
            telefono_clean = re.sub(r"\D", "", telefono)
            paciente_id = int(snapshot["paciente_id"])
            identificadores = []
            if is_valid_nss_key(nss_clean):
                identificadores.append(("NSS", nss_clean))
            if is_valid_cedula_key(cedula_clean):
                identificadores.append(("CEDULA", cedula_clean))
            revision_nss = None
            if not is_valid_cedula_key(cedula_clean) and is_valid_nss_key(nss_clean):
                revision_nss = cur.execute(
                    """
                    SELECT p.id,p.nombre
                    FROM paciente_identificadores i
                    JOIN pacientes p ON p.id=i.paciente_id
                    WHERE i.tipo='NSS' AND i.valor_normalizado=?
                      AND i.activo=1 AND p.id<>?
                    ORDER BY i.conflicto,p.id DESC LIMIT 1
                    """,
                    (nss_clean, paciente_id),
                ).fetchone()
            nss_clean_atencion = None if revision_nss else (nss_clean or None)
            identidad_estado = "NSS_EN_REVISION" if revision_nss else "VALIDADA"
            cur.execute("""
                UPDATE atenciones
                SET nombre=?, sexo=?, fecha=?, hora=?, hoja=?, ars=?, nss=?, cedula=?, telefono=?, direccion=?, nacionalidad=?,
                    edad_num=?, unidad=?, tipo_atencion=?, nss_clean=?, cedula_clean=?, telefono_clean=?,
                    identidad_estado=?, requiere_revision=?,
                    updated_at=datetime('now','localtime')
                WHERE id=?
            """, (nombre, sexo, fecha, hora, hoja, ars, nss, cedula, telefono, direccion, nacionalidad,
                  edad_num, unidad, tipo_atencion, nss_clean_atencion,
                  cedula_clean or None, telefono_clean or None, identidad_estado,
                  int(bool(revision_nss)), int(atencion_id)))
            rowcount = cur.rowcount

            # Una edición tampoco debe detener el flujo. Si queda un NSS sin
            # cédula compartido con otra ficha, se registra para revisión superior.
            cur.execute(
                """
                UPDATE nss_conflictos SET
                    estado='RESUELTO',resolucion='CORREGIDO_EN_EDICION',
                    motivo_resolucion='La atención fue editada posteriormente',
                    resuelto_por='SISTEMA',resuelto_at=datetime('now','localtime')
                WHERE atencion_id=? AND estado='PENDIENTE'
                """,
                (int(atencion_id),),
            )
            if revision_nss:
                detalle = (
                    f"NSS compartido tras edición con la ficha "
                    f"#{int(revision_nss['id'])} ({revision_nss['nombre'] or 'SIN NOMBRE'})."
                )
                cur.execute(
                    """
                    INSERT INTO nss_conflictos(
                        nss_normalizado,paciente_nuevo_id,paciente_referencia_id,
                        atencion_id,detalle
                    ) VALUES (?,?,?,?,?)
                    """,
                    (
                        nss_clean,
                        paciente_id,
                        int(revision_nss["id"]),
                        int(atencion_id),
                        detalle,
                    ),
                )
            # El PDF archivado representa el snapshot anterior. Se invalida en la
            # misma transacción para impedir que vuelva a abrirse desactualizado.
            cur.execute(
                "DELETE FROM documentos WHERE atencion_id=? AND tipo='HOJA_EMERGENCIA'",
                (int(atencion_id),),
            )
            cur.execute(
                """
                UPDATE trabajos_salida SET
                    pdf_estado='PENDIENTE',pdf_path=NULL,pdf_sha256=NULL,
                    ultimo_error=NULL,updated_at=datetime('now','localtime')
                WHERE atencion_id=?
                """,
                (int(atencion_id),),
            )

            if actualizar_ficha:
                for tipo, valor in identificadores:
                    if tipo != "CEDULA":
                        continue
                    ajenos = cur.execute(
                        """
                        SELECT DISTINCT paciente_id FROM paciente_identificadores
                        WHERE tipo=? AND valor_normalizado=? AND activo=1 AND paciente_id<>?
                        """,
                        (tipo, valor, paciente_id),
                    ).fetchall()
                    if ajenos:
                        conn.rollback()
                        raise ValueError(
                            f"El {tipo} ya está asignado a otra ficha."
                        )

                cur.execute(
                    """
                    UPDATE pacientes SET
                        nombre=?,cedula=?,telefono=?,direccion=?,nacionalidad=?,ars=?,nss=?,
                        nss_clean=?,cedula_clean=?,telefono_clean=?,provisional=?,
                        requiere_revision=0,updated_at=datetime('now','localtime')
                    WHERE id=?
                    """,
                    (
                        nombre,
                        cedula_clean if is_valid_cedula_key(cedula_clean) else None,
                        telefono_db,
                        direccion,
                        nacionalidad,
                        ars,
                        nss_clean if is_valid_nss_key(nss_clean) else None,
                        nss_clean if is_valid_nss_key(nss_clean) else None,
                        cedula_clean if is_valid_cedula_key(cedula_clean) else None,
                        telefono_clean if len(telefono_clean) == 10 else None,
                        int(not identificadores),
                        paciente_id,
                    ),
                )
                for tipo in ("NSS", "CEDULA"):
                    cur.execute(
                        "DELETE FROM paciente_identificadores WHERE paciente_id=? AND tipo=?",
                        (paciente_id, tipo),
                    )
                for tipo, valor in identificadores:
                    cur.execute(
                        """
                        INSERT INTO paciente_identificadores(
                            paciente_id,tipo,valor_normalizado,activo,conflicto
                        ) VALUES (?,?,?,1,0)
                        ON CONFLICT(paciente_id,tipo,valor_normalizado)
                        DO UPDATE SET activo=1,conflicto=0
                        """,
                        (paciente_id, tipo, valor),
                    )

            after = dict(cur.execute("SELECT * FROM atenciones WHERE id=?", (int(atencion_id),)).fetchone())
            actor = limpiar_nombre_representante(
                usuario or (cargar_turno_config(permitir_vencido=True) or {}).get("representante", "")
            )
            self._registrar_auditoria_conn(
                conn,
                int(atencion_id),
                "MODIFICACION",
                (motivo or "Corrección de atención específica").strip(),
                actor,
                dict(snapshot),
                after,
                "OPERADOR",
            )

            conn.commit()
            return rowcount

    def eliminar_ars_global(self, ars_actual: str):
        actual = normalizar_seguro(ars_actual, "9999")
        if not actual: return 0
        total = 0
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT ars, nss FROM pacientes WHERE TRIM(IFNULL(ars,'')) <> ''")
            raw_to_update = [r["ars"] for r in cur.fetchall() if normalizar_seguro(r["ars"], r["nss"]) == actual]
            for raw in set(raw_to_update):
                cur.execute("UPDATE pacientes SET ars='SIN SEGURO',updated_at=datetime('now','localtime') WHERE ars=?", (raw,))
                total += cur.rowcount
            conn.commit()
        return total

    def restaurar_atencion_snapshot(self, atencion: dict):
        raise RuntimeError(
            "La reinserción heredada está deshabilitada: las anulaciones conservan el registro y su ID."
        )


    def resumen_turno_actual(self):
        turno_cfg = cargar_turno_config() or cargar_turno_config(permitir_vencido=True)
        base = {"total": 0, "sin_seguro": 0, "GENERAL": 0, "PEDIATRIA": 0, "GINECOLOGIA": 0, "URGENCIAS": 0}
        if not turno_cfg:
            return base
        contexto = self.buscar_contexto_turno_existente(turno_cfg)
        if not contexto:
            excel = resumen_excel_actual_simple(turno_cfg=turno_cfg)
            if int(excel.get("total", 0) or 0):
                excel["URGENCIAS"] = 0
                excel["_fuente"] = "EXCEL_RECUPERADO"
                return excel
            return base
        filas = self.obtener_atenciones_para_rango_real(
            turno_id=int(contexto["turno_id"])
        )
        resumen = dict(base)
        for f in filas:
            tipo = (f.get("tipo_atencion") or "EMERGENCIA").strip().upper()
            if tipo == "URGENCIA":
                resumen["URGENCIAS"] += 1
                continue

            resumen["total"] += 1
            if normalizar_seguro(f.get("ars", ""), f.get("nss", "")) == "SIN SEGURO":
                resumen["sin_seguro"] += 1
            hoja = (f.get("hoja", "") or "").upper()
            if hoja in resumen:
                resumen[hoja] += 1
        resumen["_fuente"] = "BD"

        # Si se recuperó un Excel del mismo turno y contiene más filas que la BD,
        # se conserva visible su conteo en vez de presentar un cero engañoso.
        excel = resumen_excel_actual_simple(turno_cfg=turno_cfg)
        if int(excel.get("total", 0) or 0) > int(resumen.get("total", 0) or 0):
            excel["URGENCIAS"] = resumen.get("URGENCIAS", 0)
            excel["_fuente"] = "EXCEL_RECUPERADO"
            return excel
        if int(excel.get("total", 0) or 0) == int(resumen.get("total", 0) or 0):
            resumen["_fuente"] = "BD_EXCEL"
        return resumen

    def obtener_atenciones_para_reporte(self, fecha_inicio=None, fecha_fin=None):
        data = self.obtener_atenciones_para_rango_real(fecha_inicio, fecha_fin)
        for item in data:
            dt_real = item.get("dt_real")
            item["dt_operativo"] = (
                dt_real - timedelta(days=1) if dt_real and dt_real.time() < time(8, 0) else dt_real
            )
        return sorted(data, key=lambda item: int(item.get("id", 0)), reverse=True)

    def obtener_atenciones_para_rango_real(
        self,
        fecha_inicio=None,
        fecha_fin=None,
        *,
        turno_id=None,
        dia_operativo_id=None,
    ):
        with closing(self._connect()) as conn:
            conn.row_factory = sqlite3.Row
            where = ["a.estado='ACTIVA'"]
            params = []
            if fecha_inicio:
                where.append("datetime(a.created_at) >= datetime(?)")
                params.append(fecha_inicio.strftime("%Y-%m-%d %H:%M:%S"))
            if fecha_fin:
                where.append("datetime(a.created_at) < datetime(?)")
                params.append(fecha_fin.strftime("%Y-%m-%d %H:%M:%S"))
            if turno_id is not None:
                where.append("a.turno_id=?")
                params.append(int(turno_id))
            if dia_operativo_id is not None:
                where.append("a.dia_operativo_id=?")
                params.append(int(dia_operativo_id))
            rows = conn.execute(
                """
                SELECT a.id,a.paciente_id,a.dia_operativo_id,a.turno_id,a.fecha,a.hora,a.created_at,
                       a.nombre,a.hoja,a.ars,a.nss,a.cedula,a.edad_num,a.unidad,a.tipo_atencion,
                       t.representante,t.tipo_turno,d.fecha_base
                FROM atenciones a
                JOIN turnos t ON t.id=a.turno_id
                JOIN dias_operativos d ON d.id=a.dia_operativo_id
                WHERE """ + " AND ".join(where) + " ORDER BY a.id DESC",
                params,
            ).fetchall()

            data = []
            for r in rows:
                item = dict(r)
                try:
                    dt_real = datetime.strptime(str(item.get("created_at") or "")[:19], "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    dt_real = construir_datetime_real(item.get("fecha", ""), item.get("hora", ""))

                canon = normalizar_seguro(item.get("ars", ""), item.get("nss", ""))
                item["dt_real"] = dt_real
                item["ars_normalizado"] = canon
                item["ars_display"] = seguro_para_mostrar(canon)
                item["hoja_normalizada"] = (item.get("hoja", "") or "").strip().upper() or "SIN ESPECIALIDAD"
                item["tipo_atencion"] = (item.get("tipo_atencion") or "EMERGENCIA").strip().upper()
                data.append(item)

            return sorted(data, key=lambda item: int(item.get("id", 0)), reverse=True)

    def obtener_metadatos_reporte(self, registros):
        representantes = sorted(
            {
                limpiar_nombre_representante(row.get("representante") or "")
                for row in registros
                if limpiar_nombre_representante(row.get("representante") or "")
            }
        )
        turnos = sorted(
            {
                (
                    row.get("fecha_base") or "",
                    normalizar_turno_codigo(row.get("tipo_turno") or "8AM_8AM"),
                )
                for row in registros
            }
        )
        if len(representantes) == 1:
            representante = representantes[0]
        elif representantes:
            representante = "Varios: " + ", ".join(representantes)
        else:
            representante = ""
        if len(turnos) == 1:
            turno_resumen = obtener_datos_turno_visual(
                date.fromisoformat(turnos[0][0]), turnos[0][1]
            )["turno_resumen"]
        elif turnos:
            turno_resumen = f"{len(turnos)} turnos en el período"
        else:
            turno_resumen = None
        return turno_resumen, representante


# -------------------------------
# EXCEL
# -------------------------------
def guardar_excel_seguro(wb, ruta_excel=EXCEL_PATH, accion="guardar el Excel", interactivo=True):
    while True:
        try:
            wb.save(ruta_excel)
            return True
        except PermissionError:
            if not interactivo:
                try:
                    wb.close()
                except Exception:
                    pass
                raise
            try:
                retry = messagebox.askretrycancel(
                    "Excel abierto",
                    "El listado de Excel está abierto.\n\n"
                    "Cierre el archivo y presione 'Reintentar'.\n\n"
                    f"Acción pendiente: {accion}."
                )
            except Exception:
                retry = False
            if retry:
                continue
            try: wb.close()
            except Exception: pass
            return False
        except Exception as e:
            try: wb.close()
            except Exception: pass
            if not interactivo:
                raise
            try: messagebox.showwarning("Excel", f"No se pudo {accion}:\n{e}")
            except Exception: print(f"[AVISO] No se pudo {accion}: {e}")
            return False


def aplicar_formato_excel(ws):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    for row in range(1, 6):
        ws.row_dimensions[row].height = 18

    thin = XLSide(style="thin", color="D9D9D9")
    border = XLBorder(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 5):
        cell = ws[f"A{row}"]
        cell.font = XLFont(name="Calibri", size=11, bold=False)
        cell.alignment = XLAlignment(horizontal="left", vertical="center")

    for col in range(1, 5):
        cell = ws.cell(row=5, column=col)
        cell.font = XLFont(name="Calibri", size=11, bold=False)
        cell.alignment = XLAlignment(horizontal="left", vertical="center")
        cell.border = border

    for row in range(6, ws.max_row + 1):
        ws.row_dimensions[row].height = 18
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = XLFont(name="Calibri", size=11, bold=False)
            cell.border = border
            if col == 1:
                cell.alignment = XLAlignment(horizontal="right", vertical="center")
            else:
                cell.alignment = XLAlignment(horizontal="left", vertical="center", wrap_text=False)

    try:
        orientacion_excel = str(app_setting("print_excel_orientation", "Horizontal")).lower()
        ws.page_setup.orientation = "landscape" if orientacion_excel.startswith("h") else "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_area = f"A1:D{max(ws.max_row, 6)}"
        ws.freeze_panes = "A6"
    except Exception:
        pass


def es_error_excel_corrupto(exc) -> bool:
    msg = str(exc or "").lower()
    return (
        isinstance(exc, (zipfile.BadZipFile, zlib.error))
        or "decompress" in msg
        or "decompressing data" in msg
        or "invalid distance" in msg
        or "bad crc" in msg
        or "file is not a zip file" in msg
    )


def recrear_excel_basico_por_corrupcion():
    try:
        if os.path.exists(EXCEL_PATH):
            corrupt_name = os.path.join(
                os.path.dirname(EXCEL_PATH),
                f"LISTADO_CORRUPTO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            try:
                shutil.move(EXCEL_PATH, corrupt_name)
            except Exception:
                try:
                    os.remove(EXCEL_PATH)
                except Exception:
                    pass

        wb = Workbook()
        ws = wb.active
        ws.title = "Pacientes"

        ws.merge_cells('A1:D1')
        ws.merge_cells('A2:D2')
        ws.merge_cells('A3:D3')
        ws.merge_cells('A4:D4')

        ws['A1'] = "ASISTENCIA DE PACIENTES A EMERGENCIA"
        ws['A2'] = "ASEGURADOS Y NO ASEGURADOS"
        ws['A3'] = ""
        ws['A4'] = ""

        ws['A5'] = "NO."
        ws['B5'] = "NOMBRE"
        ws['C5'] = "ESPECIALIDAD"
        ws['D5'] = "ARS"

        aplicar_formato_excel(ws)
        guardar_excel_seguro(wb, EXCEL_PATH, "recrear el listado de Excel")
        return True
    except Exception:
        return False


def abrir_excel_workbook_seguro(ruta_excel=None, mostrar_error=True, **kwargs):
    if ruta_excel is None:
        ruta_excel = EXCEL_PATH

    try:
        return openpyxl.load_workbook(ruta_excel, **kwargs)
    except Exception as e:
        if es_error_excel_corrupto(e):
            if not mostrar_error:
                raise
            messagebox.showwarning(
                "Excel dañado",
                "El listado de Excel presentó un error de compresión o corrupción.\n\n"
                "Se creará un Excel nuevo y se intentará reconstruir con los datos del turno actual."
            )
            recrear_excel_basico_por_corrupcion()
            return openpyxl.load_workbook(ruta_excel, **kwargs)
        raise

def verificar_o_crear_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pacientes"

        ws.merge_cells('A1:D1')
        ws.merge_cells('A2:D2')
        ws.merge_cells('A3:D3')
        ws.merge_cells('A4:D4')

        ws['A1'] = "ASISTENCIA DE PACIENTES A EMERGENCIA"
        ws['A2'] = "ASEGURADOS Y NO ASEGURADOS"
        ws['A3'] = ""
        ws['A4'] = ""

        ws['A5'] = "NO."
        ws['B5'] = "NOMBRE"
        ws['C5'] = "ESPECIALIDAD"
        ws['D5'] = "ARS"

        aplicar_formato_excel(ws)
        guardar_excel_seguro(wb, EXCEL_PATH, "crear el listado de Excel")


def actualizar_encabezado_excel(representante: str, turno_label: str, fecha_label: str):
    if not os.path.exists(EXCEL_PATH):
        verificar_o_crear_excel()

    wb = abrir_excel_workbook_seguro(EXCEL_PATH)
    ws = wb.active

    for rng in ('A1:D1', 'A2:D2', 'A3:D3', 'A4:D4'):
        try:
            ws.unmerge_cells(rng)
        except Exception:
            pass
        ws.merge_cells(rng)

    ws['A1'] = "ASISTENCIA DE PACIENTES A EMERGENCIA"
    ws['A2'] = "ASEGURADOS Y NO ASEGURADOS"
    ws['A3'] = f"{limpiar_nombre_representante(representante)} {fecha_label}".strip()
    ws['A4'] = turno_label

    ws['A5'] = "NO."
    ws['B5'] = "NOMBRE"
    ws['C5'] = "ESPECIALIDAD"
    ws['D5'] = "ARS"

    aplicar_formato_excel(ws)
    guardar_excel_seguro(wb, EXCEL_PATH, "actualizar el encabezado del Excel")


def limpiar_filas_excel():
    if not os.path.exists(EXCEL_PATH):
        verificar_o_crear_excel()

    wb = abrir_excel_workbook_seguro(EXCEL_PATH)
    ws = wb.active
    max_row = ws.max_row

    if max_row >= 6:
        ws.delete_rows(6, max_row - 5)

    ws['A5'] = "NO."
    ws['B5'] = "NOMBRE"
    ws['C5'] = "ESPECIALIDAD"
    ws['D5'] = "ARS"

    aplicar_formato_excel(ws)
    guardar_excel_seguro(wb, EXCEL_PATH, "limpiar las filas del Excel")


def _obtener_numeros_usados_excel(ws):
    usados = set()
    for fila in range(6, ws.max_row + 1):
        valor = ws.cell(row=fila, column=1).value
        try:
            if valor is not None and str(valor).strip() != "":
                usados.add(int(valor))
        except Exception:
            pass
    return usados


def _primer_numero_libre(usados):
    n = 1
    while n in usados:
        n += 1
    return n


def reconstruir_excel_turno(db: DatabaseManager, turno_cfg: dict):
    if not turno_cfg:
        raise TurnoNoVigenteError("No hay turno vigente para reconstruir el listado.")

    datos_turno = obtener_datos_turno_visual(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
    inicio, fin = obtener_rango_turno_efectivo(turno_cfg)
    contexto = db.buscar_contexto_turno_existente(turno_cfg)
    filas = db.obtener_atenciones_para_rango_real(
        fecha_inicio=None if contexto else inicio,
        fecha_fin=None if contexto else fin,
        turno_id=int(contexto["turno_id"]) if contexto else None,
    )

    if os.path.exists(EXCEL_PATH):
        wb = abrir_excel_workbook_seguro(EXCEL_PATH, mostrar_error=False)
    else:
        wb = Workbook()
        wb.active.title = "Pacientes"
    ws = wb.active

    for rng in ('A1:D1', 'A2:D2', 'A3:D3', 'A4:D4'):
        try:
            ws.unmerge_cells(rng)
        except (KeyError, ValueError):
            pass
        ws.merge_cells(rng)
    ws['A1'] = "ASISTENCIA DE PACIENTES A EMERGENCIA"
    ws['A2'] = "ASEGURADOS Y NO ASEGURADOS"
    ws['A3'] = (
        f"{limpiar_nombre_representante(turno_cfg.get('representante', ''))} "
        f"{datos_turno['fecha_label']}"
    ).strip()
    ws['A4'] = datos_turno["turno_label"]
    ws['A5'] = "NO."
    ws['B5'] = "NOMBRE"
    ws['C5'] = "ESPECIALIDAD"
    ws['D5'] = "ARS"
    if ws.max_row >= 6:
        ws.delete_rows(6, ws.max_row - 5)

    numero = 1
    for fila in sorted(filas, key=lambda x: int(x.get("id", 0))):
        if (fila.get("tipo_atencion") or "EMERGENCIA").strip().upper() == "URGENCIA":
            continue
        ws.append([
            numero,
            (fila.get("nombre", "") or "").upper(),
            fila.get("hoja_normalizada", fila.get("hoja", "")),
            fila.get("ars_display", "SIN SEGURO")
        ])
        numero += 1

    aplicar_formato_excel(ws)
    temp_excel = EXCEL_PATH + ".tmp.xlsx"
    try:
        wb.save(temp_excel)
        wb.close()
        os.replace(temp_excel, EXCEL_PATH)
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        try:
            if os.path.exists(temp_excel):
                os.remove(temp_excel)
        except OSError:
            pass
        raise
    return numero - 1


def actualizar_representante_turno_actual(
    db: DatabaseManager,
    representante: str,
) -> dict:
    """Cambia solo el responsable del turno, su encabezado y reportes futuros."""
    representante = limpiar_nombre_representante(representante)
    if not es_representante_valido(representante):
        raise ValueError(
            "Escriba un nombre válido. 'No disponible' no puede guardarse como usuario."
        )
    turno_cfg = cargar_turno_config(permitir_vencido=True)
    if not turno_cfg:
        raise TurnoNoVigenteError("No existe un turno configurado para actualizar.")
    contexto = db.buscar_contexto_turno_existente(turno_cfg)
    if not contexto:
        raise TurnoNoVigenteError("No se encontró el turno asociado a la configuración.")

    datos_turno = obtener_datos_turno_visual(
        turno_cfg["fecha_base"], turno_cfg["turno_codigo"]
    )
    # Se comprueba primero que el Excel pueda actualizarse. Si está abierto, no
    # se modifica la configuración ni la base y el operador puede reintentar.
    actualizar_encabezado_excel(
        representante,
        datos_turno["turno_label"],
        datos_turno["fecha_label"],
    )
    inicio_real = turno_cfg.get("inicio_real_dt")
    if not guardar_turno_config(
        representante,
        turno_cfg["turno_codigo"],
        turno_cfg["fecha_base"],
        inicio_real=inicio_real,
    ):
        raise OSError("No se pudo actualizar la configuración del turno.")
    if not db.actualizar_representante_turno(
        int(contexto["turno_id"]), representante
    ):
        raise RuntimeError("No se pudo actualizar el representante en el turno.")

    guardar_representante_catalogo(representante, db)
    actualizado = cargar_turno_config(permitir_vencido=True) or dict(turno_cfg)
    actualizado["representante"] = representante
    actualizado["turno_id"] = int(contexto["turno_id"])
    return actualizado


def agregar_excel_temporal(nombre, especialidad, ars_canonico):
    try:
        if not os.path.exists(EXCEL_PATH):
            verificar_o_crear_excel()

        wb = abrir_excel_workbook_seguro(EXCEL_PATH)
        ws = wb.active

        usados = _obtener_numeros_usados_excel(ws)
        siguiente_numero = _primer_numero_libre(usados)

        ws.append([
            siguiente_numero,
            (nombre or '').upper(),
            especialidad,
            seguro_para_mostrar(ars_canonico)
        ])
        aplicar_formato_excel(ws)
        if not guardar_excel_seguro(wb, EXCEL_PATH, "agregar el paciente al Excel"):
            return False
        return True
    except PermissionError:
        messagebox.showwarning("Archivo abierto", "Cierre el Excel para actualizar.")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar Excel: {str(e)}")
        return False


# -------------------------------
# UTILIDADES GENERALES
# -------------------------------
def formatear_cedula(cedula):
    cedula = (cedula or "").strip()
    if not cedula or is_all_zeros(cedula):
        return "N/A"
    if len(cedula) != 11 or not cedula.isdigit():
        return "N/A"
    return f"{cedula[:3]}-{cedula[3:10]}-{cedula[10:]}"

def formatear_telefono(telefono):
    telefono = (telefono or "").strip()
    if len(telefono) != 10 or not telefono.isdigit():
        return "N/A"
    return f"{telefono[:3]}-{telefono[3:6]}-{telefono[6:]}"

def escribir_log_impresion(mensaje):
    try:
        os.makedirs(LOGS_DIR, exist_ok=True)
        log_path = os.path.join(LOGS_DIR, "impresion.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%d/%m/%Y %I:%M:%S %p')}] {mensaje}\n")
    except Exception:
        pass

def imprimir_pdf(ruta_pdf, copias=1, mostrar_error=False):
    try:
        escribir_log_impresion("Solicitud de impresión recibida.")

        if not ruta_pdf or not os.path.exists(ruta_pdf):
            escribir_log_impresion(f"ERROR: PDF no encontrado: {ruta_pdf}")
            if mostrar_error:
                messagebox.showwarning("Impresión", "No se encontró el PDF para imprimir.")
            return False

        ruta_abs = os.path.abspath(ruta_pdf)
        copias = max(1, int(copias or 1))
        sis = platform.system()

        escribir_log_impresion(f"PDF: {ruta_abs}")
        escribir_log_impresion(f"Copias solicitadas: {copias}")
        escribir_log_impresion(f"Sistema: {sis}")

        if sis == "Windows":
            exe_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else BASE_DIR
            meipass_dir = getattr(sys, "_MEIPASS", exe_dir)
            cwd_dir = os.getcwd()

            nombres_sumatra = [
                "SumatraPDF.exe",
                "SumatraPDF-3.6.1-64.exe",
                "SumatraPDF.exe.exe",
            ]

            script_dir = os.path.dirname(os.path.abspath(sys.argv[0])) if sys.argv and sys.argv[0] else BASE_DIR
            appdata_dir = os.path.dirname(EXCEL_PATH)

            carpetas_busqueda = []
            for carpeta in [
                exe_dir,
                meipass_dir,
                BASE_DIR,
                cwd_dir,
                script_dir,
                appdata_dir,
                os.path.dirname(os.path.abspath(ruta_abs)),
            ]:
                if carpeta and carpeta not in carpetas_busqueda:
                    carpetas_busqueda.append(carpeta)

            posibles_sumatra = []
            for carpeta in carpetas_busqueda:
                for nombre in nombres_sumatra:
                    posibles_sumatra.append(os.path.join(carpeta, nombre))

            escribir_log_impresion("Buscando SumatraPDF en:")
            for p in posibles_sumatra:
                escribir_log_impresion(f"  - {p}")

            global SUMATRA_PATH_CACHE
            sumatra_path = SUMATRA_PATH_CACHE if SUMATRA_PATH_CACHE and os.path.exists(SUMATRA_PATH_CACHE) else None
            if not sumatra_path:
                for p in posibles_sumatra:
                    if p and os.path.exists(p):
                        sumatra_path = p
                        SUMATRA_PATH_CACHE = p
                        break

            if not sumatra_path:
                escribir_log_impresion("ERROR: No se encontró SumatraPDF en ninguna ruta.")
                if mostrar_error:
                    messagebox.showwarning(
                        "SumatraPDF no encontrado",
                        "No se encontró SumatraPDF.exe.\n\n"
                        "Coloque SumatraPDF.exe en la misma carpeta de la app o inclúyalo al compilar.\n\n"
                        "Se creó/actualizó el archivo debug_impresion.txt para diagnóstico."
                    )
                return False

            escribir_log_impresion(f"SumatraPDF encontrado: {sumatra_path}")

            ok_general = False
            for i in range(copias):
                comando = [
                    sumatra_path,
                    "-print-to-default",
                    "-silent",
                    ruta_abs
                ]
                escribir_log_impresion(f"Ejecutando copia {i + 1}/{copias}: {' '.join(comando)}")

                try:
                    proc = subprocess.run(
                        comando,
                        check=False,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL
                    )
                    escribir_log_impresion(f"Resultado SumatraPDF returncode: {proc.returncode}")
                    ok_general = ok_general or (proc.returncode == 0)
                except Exception as e:
                    escribir_log_impresion(f"ERROR ejecutando SumatraPDF: {str(e)}")

                if copias > 1 and i < copias - 1:
                    try:
                        _time.sleep(0.7)
                    except Exception:
                        pass

            if ok_general:
                escribir_log_impresion("Impresión enviada correctamente según SumatraPDF.")
            else:
                escribir_log_impresion("ADVERTENCIA: SumatraPDF no confirmó impresión correcta.")

            if mostrar_error and not ok_general:
                messagebox.showwarning(
                    "Impresión",
                    "SumatraPDF fue encontrado, pero no confirmó la impresión.\n\n"
                    "Revise que exista una impresora predeterminada y que esté disponible.\n\n"
                    "También revise debug_impresion.txt junto a la app."
                )

            return ok_general

        ok_general = False
        for i in range(copias):
            if sis == "Darwin":
                ok = subprocess.run(["lp", ruta_abs], check=False).returncode == 0
            elif sis == "Linux":
                ok = subprocess.run(["lpr", ruta_abs], check=False).returncode == 0
            else:
                ok = False

            ok_general = ok_general or ok

            if copias > 1 and i < copias - 1:
                try:
                    _time.sleep(0.7)
                except Exception:
                    pass

        return ok_general

    except Exception as e:
        escribir_log_impresion(f"ERROR general en imprimir_pdf: {str(e)}")
        if mostrar_error:
            messagebox.showwarning(
                "Impresión",
                f"No se pudo imprimir automáticamente el PDF:\n{str(e)}"
            )
        return False

def resumen_excel_actual_simple(turno_cfg=None):
    resumen = {
        "total": 0,
        "sin_seguro": 0,
        "GENERAL": 0,
        "PEDIATRIA": 0,
        "GINECOLOGIA": 0,
    }
    try:
        if not os.path.exists(EXCEL_PATH):
            return resumen

        wb = abrir_excel_workbook_seguro(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active

        if turno_cfg:
            encabezado_fecha = str(ws["A3"].value or "")
            encabezado_turno = str(ws["A4"].value or "")
            fecha_esperada = turno_cfg["fecha_base"].strftime("%d/%m/%Y")
            turno_esperado = normalizar_turno_codigo(
                turno_cfg.get("turno_codigo", "8AM_8AM")
            )
            if (
                fecha_esperada not in encabezado_fecha
                or normalizar_turno_codigo(encabezado_turno) != turno_esperado
            ):
                try:
                    wb.close()
                except Exception:
                    pass
                return resumen

        for fila in range(6, ws.max_row + 1):
            nombre = str(ws.cell(row=fila, column=2).value or "").strip()
            if not nombre:
                continue
            esp = str(ws.cell(row=fila, column=3).value or "").strip().upper()
            ars = str(ws.cell(row=fila, column=4).value or "").strip()

            resumen["total"] += 1
            if esp in resumen:
                resumen[esp] += 1

            if normalizar_seguro(ars, "") == "SIN SEGURO":
                resumen["sin_seguro"] += 1

        try:
            wb.close()
        except Exception:
            pass
    except Exception:
        pass
    return resumen

def excel_tiene_registros(ruta_excel=None):
    if ruta_excel is None:
        ruta_excel = EXCEL_PATH

    try:
        if not os.path.exists(ruta_excel):
            return False

        wb = abrir_excel_workbook_seguro(ruta_excel, read_only=True, data_only=True)
        ws = wb.active

        for fila in range(6, ws.max_row + 1):
            nombre = ws.cell(row=fila, column=2).value
            if nombre is not None and str(nombre).strip():
                try:
                    wb.close()
                except Exception:
                    pass
                return True

        try:
            wb.close()
        except Exception:
            pass
        return False
    except Exception:
        return False

def reintentar_si_excel_abierto(accion):
    while True:
        resp = messagebox.askretrycancel(
            "Excel abierto",
            "El listado de Excel está abierto.\n\n"
            "Cierre el archivo y presione “Reintentar”."
        )
        if not resp:
            return False
        try:
            return bool(accion())
        except PermissionError:
            continue
        except Exception as e:
            messagebox.showwarning("Aviso", f"No se pudo completar la acción:\n{str(e)}")
            return False

def imprimir_excel(ruta_excel=None, copias=1):
    if ruta_excel is None:
        ruta_excel = EXCEL_PATH

    try:
        if not os.path.exists(ruta_excel):
            return False

        if not excel_tiene_registros(ruta_excel):
            return False

        ruta_abs = os.path.abspath(ruta_excel)
        sis = platform.system()
        copias = max(1, int(copias or 1))

        for _ in range(copias):
            if sis == "Windows":
                os.startfile(ruta_abs, "print")
            elif sis == "Darwin":
                subprocess.run(["lp", ruta_abs], check=False)
            elif sis == "Linux":
                subprocess.run(["lpr", ruta_abs], check=False)

        return True

    except PermissionError:
        return reintentar_si_excel_abierto(lambda: imprimir_excel(ruta_excel, copias))
    except Exception as e:
        messagebox.showwarning(
            "Aviso",
            f"No se pudo imprimir automáticamente el listado de Excel:\n{str(e)}"
        )
        return False

def abrir_pdf(ruta_pdf):
    try:
        sis = platform.system()
        if sis == "Windows":
            os.startfile(ruta_pdf)
        elif sis == "Darwin":
            subprocess.run(["open", ruta_pdf], check=False)
        elif sis == "Linux":
            subprocess.run(["xdg-open", ruta_pdf], check=False)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el PDF: {str(e)}")


def sanitize_filename(name: str) -> str:
    keep = "-_.() "
    return "".join(c for c in (name or "") if c.isalnum() or c in keep).strip().replace("  ", " ")


def poner_hora_entre_parentesis(texto: str) -> str:
    if not texto:
        return ""
    return re.sub(r'(?<!\()(\b\d{1,2}:\d{2}\s?[AP]M\b)(?!\))', r'(\1)', texto)


def draw_text_auto(c, x, y, text, base_font="Helvetica", base_size=12, max_len=THRESHOLD_LEN, min_size=8):
    if text is None:
        return base_size
    txt = str(text).strip()
    size = float(base_size)

    if len(txt) > max_len:
        factor = max(min_size / base_size, max_len / len(txt))
        size = max(round(base_size * factor, 1), min_size)

    c.setFont(base_font, size)
    c.drawString(x, y, txt)
    c.setFont("Helvetica", 12)
    return size


def draw_direccion_auto(c, x, y, text, max_width=200, base_font="Helvetica", base_size=12, min_size=6, normal_limit=18):
    txt = str(text or "").strip()
    if not txt:
        return base_size

    size = float(base_size)
    
    if len(txt) > normal_limit:
        exceso = len(txt) - normal_limit
        size = max(base_size - (exceso * 0.3), min_size)

    c.setFont(base_font, size)

    while size > min_size and c.stringWidth(txt, base_font, size) > max_width:
        size -= 0.2
        c.setFont(base_font, size)

    c.drawString(x, y, txt)
    c.setFont("Helvetica", 12)
    return size


def preparar_datos_pdf(datos):
    ars_canon = normalizar_seguro(datos.get('Aseguradora (ARS)', ''), datos.get('NSS', ''))
    usar_guiones = bool(app_setting("pdf_nss_guiones", True))
    modo_ars = str(app_setting("pdf_ars_display_mode", "Abreviada"))
    ars_display = ars_canon if modo_ars.lower().startswith("completa") else seguro_para_mostrar(ars_canon)

    return {
        **datos,
        "Cédula": formatear_cedula(datos.get('Cédula', '')),
        "Teléfono": formatear_telefono(datos.get('Teléfono', '')),
        "NSS": formatear_nss_para_pdf(datos.get('NSS', ''), ars_canon) if usar_guiones else str(datos.get('NSS', '') or '').strip().upper(),
        "Edad": f"{datos.get('Edad_num', 0)}{datos.get('Unidad', 'Años')[0].upper()}",
        "ARS_CANONICO": ars_canon,
        "ARS_DISPLAY": ars_display,
    }


def crear_pdf_temporal(hoja, datos, mostrar_error=True):
    ruta_hoja = RUTA_HOJAS.get(hoja)
    writer = PdfWriter()
    if not ruta_hoja or not os.path.exists(ruta_hoja):
        APP_LOG.error("Plantilla no encontrada para %s: %s", hoja, ruta_hoja)
        if mostrar_error:
            messagebox.showerror("Error", f"Plantilla no encontrada: {ruta_hoja}")
        return None
    temp_overlay = None
    try:
        reader = PdfReader(ruta_hoja)
        fd_overlay, temp_overlay = tempfile.mkstemp(suffix=".pdf")
        os.close(fd_overlay)
        c = canvas.Canvas(temp_overlay, pagesize=letter)
        first_page = reader.pages[0]
        page_width = float(first_page.mediabox[2])
        page_height = float(first_page.mediabox[3])
        c.setPageSize((page_width, page_height))
        c.setFont("Helvetica", 12)
        c.setFillColorRGB(0, 0, 0)

        d = preparar_datos_pdf(datos)

        if hoja == "GENERAL":
            c.drawString(495, 680, f"{d['Fecha']}")
            c.drawString(275, 665, f"{d['Hora']}")
            draw_text_auto(c, 63, 579, f"{d['Nombre']}", base_size=12, max_len=32, min_size=8)
            c.drawString(53, 560, f"{d['Sexo']}")
            c.drawString(135, 560, f"{d['Edad']}")
            
            draw_text_auto(c, 83, 542, f"{d['ARS_DISPLAY']}", base_size=12, max_len=20, min_size=8)
            
            c.drawString(228, 542, f"{d['NSS']}")
            c.drawString(495, 560, f"{d['Cédula']}")
            c.drawString(340, 493, f"{d['Teléfono']}")
            draw_direccion_auto(c, 70, 510, f"{d['Dirección']}", max_width=450, base_size=12, min_size=7)
            draw_text_auto(c, 100, 495, f"{d['Nacionalidad']}", base_size=12, max_len=18, min_size=8)

        elif hoja == "GINECOLOGIA":
            c.drawString(182, 720, f"{d['Fecha']}")
            c.drawString(273, 720, f"{d['Hora']}")
            draw_text_auto(c, 128, 705, f"{d['Nombre']}", base_size=12, max_len=30, min_size=8)
            c.drawString(406, 707, f"{d['Edad']}")
            
            ars_visual = d['ARS_DISPLAY']
            if ars_visual == "S.CONTRIBUTIVO":
                ars_visual = "S.CONT"
            elif ars_visual == "S.PENSIONADOS":
                ars_visual = "S.PENSIONAD"
                
            draw_text_auto(c, 110, 692, f"{ars_visual}", base_size=9.5, max_len=20, min_size=6)
            
            c.drawString(330, 693, f"{d['Teléfono']}")
            
            c.drawString(177, 693, f"{d['NSS']}")
            
            c.drawString(120, 667, f"{d['Cédula']}")
            
            draw_direccion_auto(c, 133, 679, f"{d['Dirección']}", max_width=185, base_size=12, min_size=6, normal_limit=17)
            
            draw_text_auto(c, 330, 680, f"{d['Nacionalidad']}", base_size=12, max_len=16, min_size=8)

        elif hoja == "PEDIATRIA":
            draw_text_auto(c, 110, 639, f"{d['Nombre']}", base_size=12, max_len=30, min_size=8)
            c.drawString(305, 625, f"{d['Teléfono']}")
            c.drawString(418, 639, f"{d['Edad']}")
            
            c.drawString(165, 624, f"{d['NSS']}")
            
            ars_visual = d['ARS_DISPLAY']
            if ars_visual == "S.CONTRIBUTIVO":
                ars_visual = "S.CONT"
            elif ars_visual == "S.PENSIONADOS":
                ars_visual = "S.PENSIONAD"
                
            draw_text_auto(c, 95, 624, f"{ars_visual}", base_size=9.5, max_len=13, min_size=5.5)
            
            draw_direccion_auto(c, 121, 610, f"{d['Dirección']}", max_width=185, base_size=12, min_size=6, normal_limit=17)
            
            draw_text_auto(c, 320, 610, f"{d['Nacionalidad']}", base_size=12, max_len=14, min_size=8)
            c.drawString(187, 657, f"{d['Fecha']}")
            c.drawString(297, 657, f"{d['Hora']}")
            c.drawString(469, 637, f"{d['Sexo']}")

        c.save()

        overlay_pages = [PdfReader(temp_overlay).pages[0]]
        temp_overlay2 = None
        
        if hoja == "GENERAL" and len(reader.pages) >= 2:
            try:
                fd_overlay2, temp_overlay2 = tempfile.mkstemp(suffix=".pdf")
                os.close(fd_overlay2)
                c2 = canvas.Canvas(temp_overlay2, pagesize=letter)
                c2.setPageSize((page_width, page_height))
                c2.setFont("Helvetica", 10)
                c2.setFillColorRGB(0, 0, 0)
                
                campos_examen = [
                    {"nombre": "Cabeza", "x_check": 499, "y_check": 609, "x_texto": 510, "y_texto": 610},
                    {"nombre": "Cuello", "x_check": 499, "y_check": 584, "x_texto": 510, "y_texto": 585},
                    {"nombre": "Corazón", "x_check": 499, "y_check": 557, "x_texto": 510, "y_texto": 558},
                    {"nombre": "Tórax", "x_check": 499, "y_check": 529, "x_texto": 510, "y_texto": 530},
                    {"nombre": "Abdomen", "x_check": 499, "y_check": 500, "x_texto": 510, "y_texto": 501},
                    {"nombre": "Genitales", "x_check": 499, "y_check": 469, "x_texto": 510, "y_texto": 470},
                    {"nombre": "Pulmones", "x_check": 499, "y_check": 441, "x_texto": 510, "y_texto": 442},
                    {"nombre": "Extremidades", "x_check": 499, "y_check": 415, "x_texto": 510, "y_texto": 416},
                    {"nombre": "Ex. Neurológico", "x_check": 499, "y_check": 389, "x_texto": 510, "y_texto": 390},
                    {"nombre": "Tacto Rectal", "x_check": 499, "y_check": 363, "x_texto": 510, "y_texto": 364},
                    {"nombre": "Tacto Vaginal", "x_check": 499, "y_check": 337, "x_texto": 510, "y_texto": 338},
                ]
                
                checkbox_size = 9
                
                for campo in campos_examen:
                    c2.rect(campo["x_check"], campo["y_check"] - 2, checkbox_size, checkbox_size, stroke=1, fill=0)
                    c2.drawString(campo["x_texto"], campo["y_texto"] - 2, "Sin patología aparente")
                
                c2.save()
                overlay_pages.append(PdfReader(temp_overlay2).pages[0])
            except Exception:
                pass
            finally:
                if temp_overlay2 and os.path.exists(temp_overlay2):
                    try:
                        os.remove(temp_overlay2)
                    except Exception:
                        pass

        for idx, page in enumerate(reader.pages):
            if idx < len(overlay_pages):
                page.merge_page(overlay_pages[idx])
            writer.add_page(page)

        fd_final, temp_final = tempfile.mkstemp(
            suffix=f"_{sanitize_filename(d.get('Nombre', 'PACIENTE'))}_{hoja}.pdf"
        )
        os.close(fd_final)
        with open(temp_final, "wb") as out:
            writer.write(out)

        try:
            os.remove(temp_overlay)
        except Exception:
            pass

        return temp_final

    except Exception as e:
        APP_LOG.exception("Error generando PDF temporal")
        if mostrar_error:
            messagebox.showerror("Error", f"Error generando PDF temporal: {str(e)}")
        if temp_overlay and os.path.exists(temp_overlay):
            try:
                os.remove(temp_overlay)
            except Exception:
                pass
        return None


def archivar_pdf_atencion(ruta_temporal, atencion_id, fecha=None):
    if not ruta_temporal or not os.path.isfile(ruta_temporal):
        raise FileNotFoundError("No existe el PDF generado para archivar.")
    momento = fecha if isinstance(fecha, datetime) else datetime.now()
    folder = os.path.join(DOCUMENTOS_DIR, momento.strftime("%Y"), momento.strftime("%m"))
    os.makedirs(folder, exist_ok=True)
    destino = os.path.join(folder, f"atencion_{int(atencion_id)}.pdf")
    temporal_destino = destino + ".tmp"
    shutil.copy2(ruta_temporal, temporal_destino)
    os.replace(temporal_destino, destino)
    return destino


def regenerar_pdf_archivado(db: DatabaseManager, atencion_id: int, mostrar_error=False):
    """Regenera el PDF desde el snapshot vigente y actualiza su referencia íntegra."""
    atencion = db.obtener_atencion_por_id(int(atencion_id))
    if not atencion or str(atencion.get("estado") or "").upper() != "ACTIVA":
        return None
    datos = {
        "Fecha": atencion.get("fecha", ""),
        "Hora": atencion.get("hora", ""),
        "Nombre": atencion.get("nombre", ""),
        "Sexo": atencion.get("sexo", "") or "Femenino",
        "Edad_num": int(atencion.get("edad_num") or 0),
        "Unidad": atencion.get("unidad", "Años"),
        "Cédula": atencion.get("cedula", ""),
        "Teléfono": atencion.get("telefono", "") or "",
        "Dirección": atencion.get("direccion", ""),
        "Nacionalidad": atencion.get("nacionalidad", ""),
        "Aseguradora (ARS)": atencion.get("ars", ""),
        "NSS": atencion.get("nss", ""),
        "TipoAtencion": atencion.get("tipo_atencion", "EMERGENCIA"),
    }
    hoja = atencion.get("hoja") or "GENERAL"
    ruta_temporal = crear_pdf_temporal(hoja, datos, mostrar_error=mostrar_error)
    if not ruta_temporal:
        return None
    try:
        ruta = archivar_pdf_atencion(ruta_temporal, int(atencion_id))
        sha256 = db.registrar_documento(
            int(atencion_id), "HOJA_EMERGENCIA", ruta, hoja
        )
        db.actualizar_trabajo_salida(
            int(atencion_id),
            "pdf",
            "COMPLETADO",
            pdf_path=ruta,
            pdf_sha256=sha256,
        )
        return ruta
    finally:
        try:
            os.remove(ruta_temporal)
        except OSError:
            pass


def eliminar_archivo_sensible(ruta):
    """Sobrescribe y retira un archivo regular; falla de forma explícita si no puede hacerlo."""
    ruta = os.path.abspath(ruta)
    if not os.path.isfile(ruta):
        return
    size = os.path.getsize(ruta)
    with open(ruta, "r+b", buffering=0) as stream:
        bloque = b"\0" * (1024 * 1024)
        restante = size
        while restante:
            chunk = bloque if restante >= len(bloque) else bloque[:restante]
            stream.write(chunk)
            restante -= len(chunk)
        stream.flush()
        os.fsync(stream.fileno())
    os.remove(ruta)


def crear_selector_fecha(parent, width=16):
    widget = TBDateEntry(
        parent,
        bootstyle="primary",
        dateformat="%d/%m/%Y",
        firstweekday=0,
        width=width,
        startdate=datetime.now()
    )
    return widget


def obtener_fecha_selector(widget):
    try:
        return widget.entry.get().strip()
    except Exception:
        try:
            return widget.get().strip()
        except Exception:
            return ""


def establecer_fecha_selector(widget, fecha_obj: date):
    texto = fecha_obj.strftime("%d/%m/%Y")
    try:
        widget.entry.delete(0, tk.END)
        widget.entry.insert(0, texto)
        return
    except Exception:
        pass

    try:
        widget.delete(0, tk.END)
        widget.insert(0, texto)
    except Exception:
        pass


def construir_resumen_desde_registros(registros, periodo_texto, turno_resumen=None, representante=""):
    conteo_seguro = {}
    conteo_esp = {}
    urgencias = 0
    total_emergencia = 0

    for r in registros:
        tipo = (r.get("tipo_atencion") or "EMERGENCIA").strip().upper()
        if tipo == "URGENCIA":
            urgencias += 1
            continue

        total_emergencia += 1
        seguro = r.get("ars_display", "SIN SEGURO")
        especialidad = r.get("hoja_normalizada", "SIN ESPECIALIDAD")
        conteo_seguro[seguro] = conteo_seguro.get(seguro, 0) + 1
        conteo_esp[especialidad] = conteo_esp.get(especialidad, 0) + 1

    if urgencias:
        conteo_esp["URGENCIAS"] = urgencias

    cantidad_sin_seguro = conteo_seguro.get("SIN SEGURO", 0)

    por_seguro = sorted(conteo_seguro.items(), key=lambda x: (-x[1], x[0]))
    por_especialidad = sorted(conteo_esp.items(), key=lambda x: (-x[1], x[0]))

    return {
        "periodo_texto": poner_hora_entre_parentesis(periodo_texto),
        "total_general": total_emergencia,
        "cantidad_sin_seguro": cantidad_sin_seguro,
        "cantidad_urgencias": urgencias,
        "por_seguro": por_seguro,
        "por_especialidad": por_especialidad,
        "turno_resumen": turno_resumen,
        "representante": (representante or "").strip(),
    }

def construir_resumen_desde_excel_actual(turno_cfg: dict, periodo_texto: str):
    registros = []
    try:
        if not os.path.exists(EXCEL_PATH) or not excel_tiene_registros(EXCEL_PATH):
            return None

        wb = abrir_excel_workbook_seguro(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active

        for fila in range(6, ws.max_row + 1):
            nombre = str(ws.cell(row=fila, column=2).value or "").strip()
            hoja = str(ws.cell(row=fila, column=3).value or "").strip().upper()
            ars = str(ws.cell(row=fila, column=4).value or "").strip()

            if not nombre:
                continue

            ars_canon = normalizar_seguro(ars, "999999999")
            registros.append({
                "nombre": nombre,
                "hoja": hoja,
                "hoja_normalizada": hoja or "SIN ESPECIALIDAD",
                "ars": ars_canon,
                "ars_display": seguro_para_mostrar(ars_canon),
                "nss": "",
                "cedula": "",
                "fecha": "",
                "hora": "",
            })

        try:
            wb.close()
        except Exception:
            pass

        if not registros:
            return None

        datos_turno = obtener_datos_turno_visual(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
        return construir_resumen_desde_registros(
            registros,
            periodo_texto,
            turno_resumen=datos_turno["turno_resumen"],
            representante=turno_cfg.get("representante", "")
        )
    except Exception:
        return None


def construir_resumen_turno(db: DatabaseManager, turno_cfg: dict, fin_override: datetime = None):
    if not turno_cfg:
        return None

    inicio, fin = obtener_rango_turno_efectivo(turno_cfg, fin_override=fin_override)
    contexto = db.buscar_contexto_turno_existente(turno_cfg)
    registros = db.obtener_atenciones_para_rango_real(
        inicio,
        fin,
        turno_id=int(contexto["turno_id"]) if contexto else None,
    )
    datos_turno = obtener_datos_turno_visual(turno_cfg["fecha_base"], turno_cfg["turno_codigo"])
    periodo_texto = f"{inicio.strftime('%d/%m/%Y %I:%M %p')} a {fin.strftime('%d/%m/%Y %I:%M %p')}"

    if not registros:
        resumen_excel = construir_resumen_desde_excel_actual(turno_cfg, periodo_texto)
        if resumen_excel:
            return resumen_excel

    return construir_resumen_desde_registros(
        registros,
        periodo_texto,
        turno_resumen=datos_turno["turno_resumen"],
        representante=turno_cfg.get("representante", "")
    )


def crear_pdf_reporte(resumen, destino=None):
    os.makedirs(REPORTES_DIR, exist_ok=True)

    if destino:
        pdf_path = destino
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = output_report_path(f"Reporte_Pacientes_{ts}.pdf")

    orientacion_pdf = str(app_setting("print_pdf_orientation", "Horizontal")).lower()
    pagesize_reporte = (letter[1], letter[0]) if orientacion_pdf.startswith("h") else letter
    c = canvas.Canvas(pdf_path, pagesize=pagesize_reporte)
    width, height = pagesize_reporte
    margen_x = 40
    margen_y = 42
    usable_w = width - (margen_x * 2)
    titulo_institucional = "REPORTE ESTADÍSTICO DE EMERGENCIA"
    subtitulo = "HOSPITAL PROVINCIAL DR. ÁNGEL CONTRERAS"
    representante_mayus = (resumen.get("representante") or "").strip().upper()

    def draw_center_text(txt, y, font="Helvetica-Bold", size=13):
        c.setFont(font, size)
        c.drawCentredString(width / 2, y, txt)

    def draw_table_header(y, headers, widths):
        h = 22
        x = margen_x
        c.setFillColorRGB(0.89, 0.93, 0.98)
        c.rect(x, y - h, sum(widths), h, stroke=1, fill=1)
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 10)
        cx = x
        for idx, header in enumerate(headers):
            c.rect(cx, y - h, widths[idx], h, stroke=1, fill=0)
            c.drawCentredString(cx + widths[idx] / 2, y - 15, str(header))
            cx += widths[idx]
        return y - h

    def draw_table_row(y, values, widths, bold=False):
        h = 20
        x = margen_x
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 9.5)
        cx = x
        for idx, value in enumerate(values):
            c.rect(cx, y - h, widths[idx], h, stroke=1, fill=0)
            text = str(value)
            if idx == len(values) - 1:
                c.drawCentredString(cx + widths[idx] / 2, y - 14, text)
            else:
                size = 9.5
                while size > 7 and c.stringWidth(text, "Helvetica-Bold" if bold else "Helvetica", size) > widths[idx] - 8:
                    size -= 0.5
                c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
                c.drawString(cx + 5, y - 14, text)
                c.setFont("Helvetica-Bold" if bold else "Helvetica", 9.5)
            cx += widths[idx]
        return y - h

    def draw_section_title(txt, y):
        c.setFillColorRGB(0.08, 0.18, 0.30)
        c.rect(margen_x, y - 21, usable_w, 21, stroke=0, fill=1)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margen_x + 8, y - 15, txt)
        c.setFillColorRGB(0, 0, 0)
        return y - 28

    def draw_summary_box(y):
        box_h = 48
        gap = 8
        box_w = (usable_w - gap * 2) / 3
        items = [
            ("TOTAL GENERAL", resumen.get("total_general", 0)),
            ("SIN SEGURO", resumen.get("cantidad_sin_seguro", 0)),
            ("GENERADO", datetime.now().strftime("%d/%m/%Y %I:%M %p")),
        ]
        x = margen_x
        for title, value in items:
            c.rect(x, y - box_h, box_w, box_h, stroke=1, fill=0)
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(x + box_w / 2, y - 15, str(title))
            c.setFont("Helvetica-Bold", 13 if isinstance(value, int) else 9.5)
            c.drawCentredString(x + box_w / 2, y - 35, str(value))
            x += box_w + gap
        return y - box_h - 18

    def dibujar_encabezado(page_no=1):
        y = height - 38
        if os.path.exists(LOGO_PATH):
            try:
                logo_w = usable_w + 20
                logo_h = 82
                x_logo = margen_x - 10
                c.drawImage(LOGO_PATH, x_logo, height - logo_h - 6,
                            width=logo_w, height=logo_h, preserveAspectRatio=True, anchor='c', mask='auto')
                y = height - logo_h - 22
            except Exception:
                pass

        draw_center_text(titulo_institucional, y, size=14)
        y -= 16
        draw_center_text(subtitulo, y, font="Helvetica-Bold", size=11)
        y -= 18

        c.setFont("Helvetica", 9.5)
        c.drawString(margen_x, y, f"Período: {poner_hora_entre_parentesis(resumen.get('periodo_texto', ''))}")
        y -= 13
        if resumen.get("turno_resumen"):
            c.drawString(margen_x, y, f"Turno: {resumen.get('turno_resumen')}")
            y -= 13
        if representante_mayus:
            c.drawString(margen_x, y, f"Auxiliar de facturación: {representante_mayus}")
            y -= 13
        c.drawRightString(width - margen_x, height - 30, f"Página {page_no}")
        c.line(margen_x, y - 4, width - margen_x, y - 4)
        return y - 20

    page_no = 1
    y = dibujar_encabezado(page_no)
    y = draw_summary_box(y)

    por_seguro = resumen.get("por_seguro", []) or []
    sin_seguro_rows = [(s, n) for s, n in por_seguro if str(s).upper() == "SIN SEGURO"]
    asegurados_rows = [(s, n) for s, n in por_seguro if str(s).upper() != "SIN SEGURO"]

    y = draw_section_title("PACIENTES ASEGURADOS POR ARS", y)
    widths = [usable_w * 0.76, usable_w * 0.24]
    y = draw_table_header(y, ["ARS / Seguro", "Cantidad"], widths)
    if asegurados_rows:
        for seguro, cantidad in asegurados_rows:
            if y < margen_y + 80:
                c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)
                y = draw_section_title("PACIENTES ASEGURADOS POR ARS", y)
                y = draw_table_header(y, ["ARS / Seguro", "Cantidad"], widths)
            y = draw_table_row(y, [seguro, cantidad], widths)
    else:
        y = draw_table_row(y, ["Sin registros de asegurados", 0], widths)

    y -= 14
    if y < margen_y + 100:
        c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)

    y = draw_section_title("PACIENTES SIN SEGURO", y)
    y = draw_table_header(y, ["Categoría", "Cantidad"], widths)
    cantidad_ss = sin_seguro_rows[0][1] if sin_seguro_rows else resumen.get("cantidad_sin_seguro", 0)
    y = draw_table_row(y, ["SIN SEGURO", cantidad_ss], widths, bold=True)

    y -= 14
    if y < margen_y + 120:
        c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)

    y = draw_section_title("PACIENTES POR ESPECIALIDAD", y)
    y = draw_table_header(y, ["Especialidad", "Cantidad"], widths)
    por_especialidad = resumen.get("por_especialidad", []) or []
    if por_especialidad:
        for esp, cantidad in por_especialidad:
            if y < margen_y + 60:
                c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)
                y = draw_section_title("PACIENTES POR ESPECIALIDAD", y)
                y = draw_table_header(y, ["Especialidad", "Cantidad"], widths)
            y = draw_table_row(y, [esp, cantidad], widths)
    else:
        y = draw_table_row(y, ["Sin registros de especialidad", 0], widths)

    if y < margen_y + 60:
        c.showPage(); page_no += 1; y = dibujar_encabezado(page_no)
    y -= 20
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margen_x, y, f"Total general del período: {resumen.get('total_general', 0)}")
    if representante_mayus:
        c.drawRightString(width - margen_x, y, f"AUXILIAR: {representante_mayus}")

    c.save()
    return pdf_path


def crear_excel_reporte_estadistico(resumen, destino=None):
    os.makedirs(REPORTES_DIR, exist_ok=True)
    if destino: xlsx_path = destino
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = output_report_path(f"Reporte_Estadistico_{ts}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen por ARS"
    def style_sheet(ws):
        thin = XLSide(style="thin", color="D9D9D9")
        border = XLBorder(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = XLAlignment(vertical="center", wrap_text=True)
                cell.font = XLFont(name="Calibri", size=11)
        for cell in ws[1]: cell.font = XLFont(name="Calibri", size=12, bold=True)
        for col in ws.columns:
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = max(14, min(45, max(len(str(c.value or "")) for c in col) + 3))
    ws.append(["ARS / Seguro", "Cantidad"])
    for seguro, cantidad in resumen.get("por_seguro", []): ws.append([seguro, cantidad])
    style_sheet(ws)
    ws2 = wb.create_sheet("Resumen por especialidad")
    ws2.append(["Especialidad", "Cantidad"])
    for esp, cantidad in resumen.get("por_especialidad", []): ws2.append([esp, cantidad])
    style_sheet(ws2)
    ws3 = wb.create_sheet("Lista de pacientes")
    ws3.append(["Fecha", "Hora", "Nombre", "Especialidad", "ARS", "NSS", "Cédula"])
    sin_seguro_rows = []
    for r in resumen.get("registros", []) or []:
        row = [r.get("fecha",""), r.get("hora",""), r.get("nombre",""), r.get("hoja_normalizada", r.get("hoja","")), r.get("ars_display", ""), r.get("nss",""), r.get("cedula","")]
        ws3.append(row)
        if str(r.get("ars_display", "")).upper() == "SIN SEGURO": sin_seguro_rows.append(row)
    style_sheet(ws3)
    ws4 = wb.create_sheet("Sin seguro")
    ws4.append(["Fecha", "Hora", "Nombre", "Especialidad", "ARS", "NSS", "Cédula"])
    for row in sin_seguro_rows: ws4.append(row)
    style_sheet(ws4)
    ws5 = wb.create_sheet("Resumen general", 0)
    ws5.append(["Campo", "Valor"])
    ws5.append(["Período", resumen.get("periodo_texto", "")])
    ws5.append(["Turno", resumen.get("turno_resumen", "")])
    ws5.append(["Representante", resumen.get("representante", "")])
    ws5.append(["Total general", resumen.get("total_general", 0)])
    ws5.append(["Pacientes sin seguro", resumen.get("cantidad_sin_seguro", 0)])
    style_sheet(ws5)
    guardar_excel_seguro(wb, xlsx_path, "exportar el reporte estadístico a Excel")
    return xlsx_path


# -------------------------------
# APP
# -------------------------------
class App:
    def __init__(self):
        try:
            migrate_legacy_files((
                "pacientes.db",
                "LISTADO DE PACIENTES EN EMERGENCIA.xlsx",
                "turnos_config.json",
                "app_settings.json",
                "ars_catalogo.json",
                "nss_formatos_ars.json",
                "representantes.json",
                "security.json",
            ))
        except Exception:
            APP_LOG.exception("No se pudo completar la preparacion del directorio de datos")
        self.db = DatabaseManager()
        try:
            self.db.backup_manager.ensure_daily()
        except Exception:
            APP_LOG.exception("No se pudo crear el respaldo diario verificado")
        self.security = AdminSecurity(
            SECURITY_CONFIG_PATH,
            os.path.join(LOGS_DIR, "security_audit.jsonl"),
        )
        self._admin_authorized_until = None
        self._admin_authorized_actor = ""
        self.app_settings = cargar_app_settings()
        self._asegurar_preferencias_impresion_hoja()
        verificar_o_crear_excel()
        self._turno_creado_desde_excel_existente = crear_turno_desde_excel_existente_si_aplica()

        self._temp_files = set()
        self._updating_period_dates = False
        self._undo_stack = []
        self._undo_limit = 20

        # FASE 3: Cache de rendimiento
        self._cache_ars = []
        self._cache_ars_time = 0
        self._cache_resumen_turno = None
        self._cache_resumen_time = 0
        self._cache_especialidades = ["GENERAL", "PEDIATRIA", "GINECOLOGIA"]
        self._sumatra_path_cache = None

        self.root = tb.Window(themename="superhero")
        self.root.title("Generador de Formularios de Emergencia - Hospital General")
        self.root.geometry(self.app_settings.get("window_size", "1280x740"))
        self.root.minsize(760, 600)
        self.root.resizable(True, True)
        self.root.configure(bg="#07111f")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        style = tb.Style()
        style.configure("TLabel", font=("Arial", 11), foreground="#EAF2FF")
        style.configure("Muted.TLabel", font=("Arial", 10), foreground="#A9B8CC")
        style.configure("Title.TLabel", font=("Arial", 18, "bold"), foreground="#F5F9FF")
        style.configure("Subtitle.TLabel", font=("Arial", 12), foreground="#B7C6DA")
        style.configure("Section.TLabel", font=("Arial", 13, "bold"), foreground="#5CB6FF")
        style.configure("TButton", font=("Arial", 11, "bold"))
        style.configure("TEntry", font=("Arial", 11), fieldbackground="#111E2E", foreground="#F5F9FF")
        style.configure("TCombobox", font=("Arial", 11), fieldbackground="#111E2E", foreground="#F5F9FF")
        style.configure("Card.TFrame", background="#0E1B2B")
        style.configure("Root.TFrame", background="#07111f")
        try:
            style.configure("Treeview", font=("Arial", 10), rowheight=28)
            style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
            style.configure(
                "Modern.Treeview",
                font=("Arial", 10),
                rowheight=29,
                background="#0B1624",
                foreground="#EAF2FF",
                fieldbackground="#0B1624",
                bordercolor="#203348",
                lightcolor="#203348",
                darkcolor="#203348"
            )
            style.configure(
                "Modern.Treeview.Heading",
                font=("Arial", 10, "bold"),
                background="#12243A",
                foreground="#FFFFFF",
                bordercolor="#203348"
            )
            style.map(
                "Modern.Treeview",
                background=[("selected", "#1D6EFF")],
                foreground=[("selected", "#FFFFFF")]
            )
        except Exception:
            pass

        self.style = style
        self._configurar_estilos_desde_preferencias()

        self.val_cedula = (self.root.register(lambda P, *_: self.validar_numerico(P, 'cedula')), '%P')
        self.val_telefono = (self.root.register(lambda P, *_: self.validar_numerico(P, 'telefono')), '%P')

        self.main = tb.Frame(self.root, padding=14, style="Root.TFrame")
        self.main.pack(fill="both", expand=True)

        header = tb.Frame(self.main, style="Root.TFrame")
        header.pack(fill="x", pady=(0, 14))
        header.columnconfigure(1, weight=1, minsize=320)
        self.header = header

        logo_box = tb.Frame(header, padding=10, style="Card.TFrame")
        logo_box.grid(row=0, column=0, rowspan=2, sticky="nw", padx=(0, 14))
        try:
            logo_original = tk.PhotoImage(file=resource_path("istipo_hospitales.png"))
            factor = max(1, int(max(logo_original.width(), logo_original.height()) / 48))
            self.header_logo_image = logo_original.subsample(factor, factor)
            tb.Label(logo_box, image=self.header_logo_image, background="#0E1B2B").pack()
        except Exception:
            tb.Label(
                logo_box,
                text="HG",
                font=("Arial", 18, "bold"),
                foreground="#58A6FF",
                background="#0E1B2B",
            ).pack()

        title_lbl = tb.Label(
            header,
            text="GENERADOR DE FORMULARIOS DE EMERGENCIA",
            style="Title.TLabel",
            background="#07111f",
            wraplength=650,
            justify="left"
        )
        title_lbl.grid(row=0, column=1, sticky="w")
        self.title_lbl = title_lbl
        tb.Label(header, text="Sistema de Admisión en Emergencia", style="Subtitle.TLabel", background="#07111f")\
            .grid(row=1, column=1, sticky="w")

        info_header = tb.Frame(header, style="Root.TFrame")
        info_header.grid(row=0, column=2, rowspan=2, sticky="e")
        self.info_header = info_header

        fecha_card = tb.Frame(info_header, padding=(14, 10), style="Card.TFrame")
        fecha_card.pack(side="left", padx=6)
        tb.Label(fecha_card, text="▣  Fecha actual", style="Muted.TLabel", background="#0E1B2B")\
            .pack(anchor="w")
        self.fecha_actual_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        tb.Label(fecha_card, textvariable=self.fecha_actual_var, font=("Arial", 11, "bold"), foreground="#FFFFFF", background="#0E1B2B")\
            .pack(anchor="w")

        turno_cfg = cargar_turno_config()
        turno_txt = descripcion_turno_config(turno_cfg)
        self.turno_header_var = tk.StringVar(value=turno_txt)
        turno_card = tb.Frame(info_header, padding=(14, 10), style="Card.TFrame")
        turno_card.pack(side="left", padx=6)
        tb.Label(turno_card, text="◷  Turno actual", style="Muted.TLabel", background="#0E1B2B")\
            .pack(anchor="w")
        tb.Label(turno_card, textvariable=self.turno_header_var, font=("Arial", 11, "bold"), foreground="#FFFFFF", background="#0E1B2B")\
            .pack(anchor="w")

        tb.Button(info_header, text="↻ Cambiar Turno", bootstyle=SECONDARY, command=self.reiniciar_datos_excel, width=16)\
            .pack(side="left", padx=(8, 0), ipady=5)

        self.actions_menu_button = tb.Menubutton(
            header, text="Menú", bootstyle=SECONDARY, width=10
        )
        self.actions_menu_button.grid(row=0, column=3, rowspan=2, sticky="e", padx=(10, 0))
        self.actions_menu = tk.Menu(self.actions_menu_button, tearoff=0)
        self.actions_menu.add_command(label="Cambiar turno", command=self.reiniciar_datos_excel)
        self.actions_menu.add_command(label="Historial", command=self.abrir_historial)
        self.actions_menu.add_command(label="Reporte estadístico", command=self.abrir_ventana_reporte)
        self.actions_menu.add_command(label="Listado de Excel", command=self._abrir_excel_actual)
        self.actions_menu.add_separator()
        self.actions_menu.add_command(label="Editar paciente", command=self._abrir_edicion_paciente)
        self.actions_menu.add_command(label="Impresiones y documentos pendientes", command=self.abrir_trabajos_salida_pendientes)
        self.actions_menu.add_command(label="Buscar actualizaciones", command=self.buscar_actualizaciones)
        self.actions_menu.add_command(label="Configuración", command=self._abrir_configuracion_interna)
        self.actions_menu_button.configure(menu=self.actions_menu)

        content_area = tb.Frame(self.main, style="Root.TFrame")
        content_area.pack(fill="both", expand=True)
        content_area.columnconfigure(0, weight=1)
        content_area.columnconfigure(1, weight=0)
        content_area.rowconfigure(0, weight=1)
        self.content_area = content_area

        self.form_host = tb.Frame(content_area, style="Card.TFrame")
        self.form_host.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        self.form_canvas = tk.Canvas(
            self.form_host,
            borderwidth=0,
            highlightthickness=0,
            background="#0E1B2B",
        )
        self.form_scrollbar = ttk.Scrollbar(
            self.form_host, orient="vertical", command=self.form_canvas.yview
        )
        self.form_canvas.configure(yscrollcommand=self.form_scrollbar.set)
        self.form_canvas.pack(side="left", fill="both", expand=True)
        self.form_scrollbar.pack(side="right", fill="y")

        self.frame = tb.Frame(self.form_canvas, padding=18, style="Card.TFrame")
        self._form_canvas_window = self.form_canvas.create_window(
            (0, 0), window=self.frame, anchor="nw"
        )
        self.frame.bind(
            "<Configure>",
            lambda _event: self.form_canvas.configure(
                scrollregion=self.form_canvas.bbox("all")
            ),
        )
        self.form_canvas.bind(
            "<Configure>",
            lambda event: self.form_canvas.itemconfigure(
                self._form_canvas_window, width=event.width
            ),
        )
        self.form_canvas.bind(
            "<Enter>",
            lambda _event: self.root.bind_all(
                "<MouseWheel>",
                lambda event: self.form_canvas.yview_scroll(
                    int(-1 * (event.delta / 120)), "units"
                ),
            ),
        )
        self.form_canvas.bind(
            "<Leave>", lambda _event: self.root.unbind_all("<MouseWheel>")
        )
        for col in range(6):
            self.frame.columnconfigure(col, weight=1)

        self.quick_panel = tb.Frame(content_area, padding=14, style="Card.TFrame")
        self.quick_panel.grid(row=0, column=1, sticky="ns")
        self.quick_panel.configure(width=280)
        try:
            self.quick_panel.grid_propagate(False)
        except Exception:
            pass
        self._crear_panel_acciones_rapidas(self.quick_panel)

        title_row = tb.Frame(self.frame, style="Card.TFrame")
        title_row.grid(row=0, column=0, columnspan=6, sticky="ew", pady=(0, 12))
        title_row.columnconfigure(0, weight=1)
        tb.Label(title_row, text="DATOS DEL PACIENTE", font=("Arial", 13, "bold"), foreground="#EAF2FF", background="#0E1B2B")\
            .grid(row=0, column=0, sticky="w")
        self.boton_historial = tb.Button(title_row, text="Historial", command=self.abrir_historial, width=15, bootstyle=INFO)
        self.boton_historial.grid(row=0, column=1, sticky="e")
        sep = ttk.Separator(self.frame, orient="horizontal")
        sep.grid(row=1, column=0, columnspan=6, sticky="ew", pady=(0, 14))

        def lbl(text, row, col, colspan=1):
            widget = tb.Label(
                self.frame,
                text=text,
                font=("Arial", 10, "bold"),
                foreground="#EAF2FF",
                background="#0E1B2B",
            )
            widget.grid(row=row, column=col, columnspan=colspan, sticky="w", padx=(4, 10), pady=(2, 3))
            return widget

        self.lbl_nombre = lbl("Nombre  ·  Nombre completo del paciente", 2, 0, 3)
        self.entry_nombre = tb.Entry(self.frame, width=54)
        self.entry_nombre.grid(row=3, column=0, columnspan=3, sticky="ew", padx=(4, 24), pady=(0, 10), ipady=6)

        self.lbl_sexo = lbl("Sexo", 2, 3, 2)
        self.var_sexo = tk.StringVar(value="Femenino")
        
        sexo_frame = tb.Frame(self.frame)
        self.sexo_frame = sexo_frame
        sexo_frame.grid(row=3, column=3, columnspan=3, sticky="w", padx=(4, 4), pady=(0, 10))

        self.var_embarazada = tk.BooleanVar(value=False)

        self.lbl_sexo_m = tb.Radiobutton(sexo_frame, text="Masculino", variable=self.var_sexo, value="Masculino")
        self.lbl_sexo_m.pack(side="left", padx=(0, 10))

        self.lbl_sexo_f = tb.Radiobutton(sexo_frame, text="Femenino", variable=self.var_sexo, value="Femenino")
        self.lbl_sexo_f.pack(side="left", padx=(0, 10))

        self.check_embarazada = tb.Checkbutton(sexo_frame, text="Embarazada", variable=self.var_embarazada, bootstyle=INFO)
        self.check_embarazada.pack(side="left", padx=(8, 0))

        self.var_sexo.trace_add("write", lambda *args: self.actualizar_embarazada())
        self.var_embarazada.trace_add("write", lambda *args: self.actualizar_embarazada())
        self.actualizar_embarazada()

        self.lbl_edad = lbl("Edad  ·  Ej: 25", 4, 0)
        self.entry_edad = tb.Entry(self.frame, width=14)
        self.entry_edad.grid(row=5, column=0, sticky="ew", padx=(4, 8), pady=(0, 10), ipady=6)
        self.unidad_edad = tk.StringVar(value="Años")
        self.combo_unidad = tb.Combobox(self.frame, textvariable=self.unidad_edad,
                                        values=["Días", "Meses", "Años"], state="readonly", width=12)
        self.combo_unidad.grid(row=5, column=1, columnspan=2, sticky="ew", padx=(0, 24), pady=(0, 10), ipady=6)

        self.var_urgencia = tk.BooleanVar(value=False)
        self.check_urgencia = tb.Checkbutton(
            self.frame,
            text="Atención de urgencia (conteo aparte)",
            variable=self.var_urgencia,
            bootstyle=INFO
        )
        self.check_urgencia.grid(row=5, column=3, columnspan=3, sticky="w", padx=(4, 4), pady=(0, 10))

        self.lbl_cedula = lbl("Cédula  ·  11 dígitos", 6, 0, 3)
        self.entry_cedula = tb.Entry(self.frame, validate="key", validatecommand=self.val_cedula)
        self.entry_cedula.grid(row=7, column=0, columnspan=3, sticky="ew", padx=(4, 24), pady=(0, 10), ipady=6)
        self.entry_cedula.bind("<KeyRelease>", lambda e: (self.limitar_caracteres(self.entry_cedula, 11), self._try_autocomplete_cedula()))
        self.entry_cedula.bind("<FocusOut>", self.auto_completar)
        self.entry_cedula.bind("<Return>", self.auto_completar)

        self.lbl_telefono = lbl("Teléfono  ·  10 dígitos", 6, 3, 3)
        self.entry_telefono = tb.Entry(self.frame, validate="key", validatecommand=self.val_telefono)
        self.entry_telefono.grid(row=7, column=3, columnspan=3, sticky="ew", padx=(4, 4), pady=(0, 10), ipady=6)
        self.entry_telefono.bind("<KeyRelease>", lambda e: self.limitar_caracteres(self.entry_telefono, 10))

        self.lbl_direccion = lbl("Dirección", 8, 0, 6)
        self.entry_direccion = tb.Entry(self.frame)
        self.entry_direccion.grid(row=9, column=0, columnspan=6, sticky="ew", padx=(4, 4), pady=(0, 10), ipady=6)

        self.lbl_nacionalidad = lbl("Nacionalidad", 10, 0, 3)
        self.entry_nacionalidad = tb.Entry(self.frame)
        self.entry_nacionalidad.grid(row=11, column=0, columnspan=3, sticky="ew", padx=(4, 24), pady=(0, 10), ipady=6)

        self.lbl_ars = lbl("Aseguradora (ARS)  ·  Escriba: SUB, HUMANO, MAPFRE…", 10, 3, 3)
        self.entry_ars = tb.Entry(self.frame)
        self.entry_ars.grid(row=11, column=3, columnspan=3, sticky="ew", padx=(4, 4), pady=(0, 0), ipady=6)
        self.entry_ars.bind("<KeyRelease>", self._on_ars_keyrelease)
        self.entry_ars.bind("<FocusOut>", lambda e: (self._actualizar_deteccion_seguro(), self.root.after(180, self._ocultar_sugerencias_ars)))
        self.entry_ars.bind("<Down>", self._focus_sugerencias_ars)

        self.ars_detectado_var = tk.StringVar(value="Detectado como: pendiente")
        self.ars_detectado_label = tb.Label(
            self.frame,
            textvariable=self.ars_detectado_var,
            font=("Arial", 9, "bold"),
            foreground="#8FA6BF",
            background="#0E1B2B"
        )
        self.ars_detectado_label.grid(row=12, column=3, columnspan=3, sticky="w", padx=(4, 4), pady=(2, 4))

        self.ars_suggestions = tk.Listbox(
            self.frame,
            height=5,
            bg="#0B1624",
            fg="#EAF2FF",
            selectbackground="#1D6EFF",
            selectforeground="#FFFFFF",
            highlightthickness=1,
            highlightbackground="#254260",
            relief="flat",
            font=("Arial", 10)
        )
        self.ars_suggestions.grid(row=13, column=3, columnspan=3, sticky="ew", padx=(4, 4), pady=(2, 8))
        self.ars_suggestions.grid_remove()
        self.ars_suggestions.bind("<ButtonRelease-1>", self._seleccionar_sugerencia_ars)
        self.ars_suggestions.bind("<Return>", self._seleccionar_sugerencia_ars)
        self.ars_suggestions.bind("<Escape>", lambda e: self._ocultar_sugerencias_ars())
        self._ars_catalogo = self._obtener_catalogo_ars()

        self.lbl_nss = lbl("NSS  ·  Número de seguro o SIN SEGURO", 13, 0, 1)

        self.nss_detectado_var = tk.StringVar(value="NSS: pendiente")
        self.nss_detectado_label = tb.Label(
            self.frame,
            textvariable=self.nss_detectado_var,
            font=("Arial", 9, "bold"),
            foreground="#8FA6BF",
            background="#0E1B2B"
        )
        self.nss_detectado_label.grid(row=13, column=1, columnspan=2, sticky="w", padx=(4, 24), pady=(0, 2))

        self.entry_nss = tb.Entry(self.frame)
        self.entry_nss.grid(row=14, column=0, columnspan=3, sticky="ew", padx=(4, 24), pady=(0, 6), ipady=6)
        self.entry_nss.bind("<KeyRelease>", lambda e: (self._actualizar_deteccion_seguro(), self._try_autocomplete_nss()))
        self.entry_nss.bind("<FocusOut>", self.auto_completar_por_nss)
        self.entry_nss.bind("<Return>", self.auto_completar_por_nss)

        self.form_actions_separator = ttk.Separator(self.frame, orient="horizontal")
        self.form_actions_separator.grid(row=15, column=0, columnspan=6, sticky="ew", pady=(6, 6))

        btns = tb.Frame(self.frame, style="Card.TFrame")
        self.form_buttons = btns
        btns.grid(row=16, column=0, columnspan=6, sticky="e", pady=(0, 10))
        self.boton_limpiar = tb.Button(btns, text="Limpiar", command=self.limpiar_campos, width=14, bootstyle=SECONDARY)
        self.boton_limpiar.pack(side="left", padx=(0, 12), ipady=4)
        self.boton_generar_pdf = tb.Button(btns, text="Registrar e imprimir", command=self.generar_pdf, width=22, bootstyle=PRIMARY)
        self.boton_generar_pdf.pack(side="left", ipady=4)

        status = tb.Frame(self.root, padding=(18, 8), style="Root.TFrame")
        status.pack(fill="x", side="bottom", before=self.main)

        self.shortcuts_var = tk.StringVar(value="Generador de Hojas 4.1")
        self.shortcuts_label = tb.Label(
            status,
            textvariable=self.shortcuts_var,
            style="Muted.TLabel",
            background="#07111f"
        )
        self.shortcuts_label.pack(side="left")

        self.connection_var = tk.StringVar(value="Conectado a: verificando…")
        self.connection_label = tb.Label(
            status,
            textvariable=self.connection_var,
            font=("Arial", 10, "bold"),
            foreground="#72E39B",
            background="#07111f"
        )
        self.connection_label.pack(side="right")

        self.status_var = tk.StringVar(value="Listo para registrar")
        self.status_label = tb.Label(
            status,
            textvariable=self.status_var,
            font=("Arial", 10, "bold"),
            foreground="#5CB6FF",
            background="#07111f"
        )
        self.status_label.pack(side="right", padx=(0, 24))

        self.notif_frame = tb.Frame(self.root, padding=(10, 6), style="Root.TFrame")
        self.notif_frame.pack(fill="x", side="bottom")
        self.notif_frame.pack_forget()
        self.notif_label = tb.Label(self.notif_frame, text="", anchor="w", background="#07111f")
        self.notif_label.pack(side="left", fill="x", expand=True)
        self.btn_deshacer = tb.Button(self.notif_frame, text="Deshacer", bootstyle=WARNING)
        self.btn_deshacer.pack(side="right", padx=6)

        self._notif_after_id = None
        self._ultimo_atencion_id = None
        self._last_report_summary = None

        self.historial_win = None
        self.historial_sin_seguro_win = None
        self.reporte_win = None
        self.turno_win = None
        self.dialogo_unico_win = None
        self.edicion_paciente_win = None
        self.configuracion_interna_win = None
        self.salida_pendiente_win = None
        self.trabajos_salida_win = None
        self._output_payloads = {}

        self.menu_contextual = tk.Menu(self.root, tearoff=0)
        self.menu_contextual.add_command(label="Copiar", command=self._copiar)
        self.menu_contextual.add_command(label="Pegar", command=self._pegar)
        self.menu_contextual.add_command(label="Cortar", command=self._cortar)
        for w in [self.entry_nombre, self.entry_edad, self.entry_cedula, self.entry_telefono,
                  self.entry_direccion, self.entry_nacionalidad, self.entry_ars, self.entry_nss]:
            w.bind("<Button-3>", self.mostrar_menu_contextual)

        self.root.bind('<F5>', lambda e: self.reiniciar_datos_excel())
        self._suspend_autocomplete = False
        self.entry_nombre.focus_set()

        self.all_entries = [
            self.entry_nombre,
            self.entry_edad,
            self.entry_cedula,
            self.entry_telefono,
            self.entry_direccion,
            self.entry_nacionalidad,
            self.entry_ars,
            self.entry_nss
        ]
        self._initial_styles = {}
        self._pending_restores = {}
        self._capture_initial_styles()
        self._configurar_accesibilidad_teclado()

        # FASE 9B: Vincular FocusIn/Out para Ctrl+Z por campo
        for _w in (self.entry_nombre, self.entry_edad, self.entry_cedula,
                   self.entry_telefono, self.entry_direccion,
                   self.entry_nacionalidad, self.entry_ars, self.entry_nss):
            try:
                _w.bind("<FocusIn>",  self._on_field_focus_in)
                _w.bind("<FocusOut>", self._on_field_focus_out)
            except Exception:
                pass

        self._aplicar_preferencias_en_vivo()
        self._responsive_after_id = None
        self.root.bind("<Configure>", self._programar_modo_responsivo, add="+")
        try:
            self.root.after(400, self._actualizar_resumen_turno_panel)
            self.root.after(1800, self._programar_refresco_resumen_en_vivo)
            self.root.after(1000, self._actualizar_fecha_actual)
            self.root.after(2200, self._avisar_trabajos_salida_pendientes)
            self.root.after(5000, lambda: self.buscar_actualizaciones(manual=False))
        except Exception:
            pass

        if turno_cfg and not getattr(self, "_turno_creado_desde_excel_existente", False):
            try:
                self.root.after(1200, lambda cfg=turno_cfg: self._reconstruir_excel_inicio_diferido(cfg))
            except Exception:
                pass
        elif getattr(self, "_turno_creado_desde_excel_existente", False):
            try:
                self.set_status("Turno recuperado desde Excel existente; no se reconstruyó para conservar datos.", "ok")
            except Exception:
                pass

    def _asegurar_preferencias_impresion_hoja(self):
        try:
            changed = False
            if self.app_settings.get("auto_print") is not True:
                self.app_settings["auto_print"] = True
                changed = True
            if self.app_settings.get("print_auto_hoja") is not True:
                self.app_settings["print_auto_hoja"] = True
                changed = True
            if str(self.app_settings.get("print_behavior_hoja", "")).strip() not in ["Solo imprimir", "Imprimir y abrir PDF"]:
                self.app_settings["print_behavior_hoja"] = "Imprimir y abrir PDF"
                changed = True
            try:
                copias = int(self.app_settings.get("print_copies_hoja", 1) or 1)
                if copias < 1:
                    self.app_settings["print_copies_hoja"] = 1
                    changed = True
            except Exception:
                self.app_settings["print_copies_hoja"] = 1
                changed = True
            if changed:
                guardar_app_settings(self.app_settings)
        except Exception:
            pass

    def _reconstruir_excel_inicio_diferido(self, turno_cfg):
        try:
            verificar_o_crear_excel()
            contexto = self.db.buscar_contexto_turno_existente(turno_cfg)
            filas_bd = (
                self.db.obtener_atenciones_para_rango_real(
                    turno_id=int(contexto["turno_id"])
                )
                if contexto
                else []
            )
            filas_excel = int(
                resumen_excel_actual_simple(turno_cfg).get("total", 0) or 0
            )
            if filas_bd:
                reconstruir_excel_turno(self.db, turno_cfg)
                self.set_status(
                    f"Listado verificado: {len(filas_bd)} paciente(s) sincronizados.",
                    "ok",
                )
            elif filas_excel:
                self.set_status(
                    f"Listado recuperado: {filas_excel} paciente(s) visibles; "
                    "no se sobrescribió el Excel.",
                    "warning",
                )
            else:
                self.set_status("Inicio rápido: Excel verificado", "ok")
            self._actualizar_turno_visual_en_vivo()
            self._refrescar_resumen_en_vivo()
        except PermissionError:
            self.set_status("Excel abierto. Cierre el listado para actualizarlo.", "warning")
        except Exception as e:
            self.set_status(f"Aviso al verificar Excel: {e}", "warning")

    def _paleta_visual_actual(self):
        """
        FASE 13: Paleta sobria con colores institucionales.
        """
        theme = str(self.app_settings.get("theme", "oscuro") or "oscuro").lower()
        high = bool(self.app_settings.get("high_contrast", False))
        accent_pref = self.app_settings.get("accent_color", "Azul hospitalario")
        accent_hex = resolver_color_principal(accent_pref)

        accent_dark_soft = mezclar_color_hex(accent_hex, "#111E2E", 0.18)
        accent_dark_heading = mezclar_color_hex(accent_hex, "#172A3E", 0.25)
        accent_dark_selected = mezclar_color_hex(accent_hex, "#2D5F93", 0.55)

        accent_light_soft = mezclar_color_hex(accent_hex, "#FFFFFF", 0.10)
        accent_light_heading = mezclar_color_hex(accent_hex, "#DDE6F0", 0.18)
        accent_light_selected = mezclar_color_hex(accent_hex, "#4B83C2", 0.50)

        accent_high_dark_soft = mezclar_color_hex(accent_hex, "#10151C", 0.22)
        accent_high_light_soft = mezclar_color_hex(accent_hex, "#FFFFFF", 0.12)

        if high and theme == "claro":
            return {
                "mode": "claro_alto",
                "root": "#FAFAFA",
                "card": "#FFFFFF",
                "card2": accent_high_light_soft,
                "entry": "#FFFFFF",
                "tree": "#FFFFFF",
                "heading": mezclar_color_hex(accent_hex, "#E1E6EE", 0.18),
                "text": "#111827",
                "muted": "#374151",
                "accent": accent_hex,
                "accent2": COLOR_INFO,
                "border": "#111827",
                "separator": "#6B7280",
                "selected_bg": mezclar_color_hex(accent_hex, "#E6B800", 0.30),
                "selected_fg": "#111827",
                "button_fg": "#FFFFFF",
                "danger": "#8B4C4C",
                "warning": "#7A6B4A",
                "success": "#4F7B55",
                "info": COLOR_INFO,
            }

        if high and theme != "claro":
            return {
                "mode": "oscuro_alto",
                "root": "#080A0D",
                "card": "#10151C",
                "card2": accent_high_dark_soft,
                "entry": "#080A0D",
                "tree": "#080A0D",
                "heading": mezclar_color_hex(accent_hex, "#1B2532", 0.22),
                "text": "#F8FAFC",
                "muted": "#D7DEE8",
                "accent": accent_hex,
                "accent2": COLOR_INFO,
                "border": "#E5E7EB",
                "separator": "#B7C0CA",
                "selected_bg": mezclar_color_hex(accent_hex, "#D8B536", 0.30),
                "selected_fg": "#111827",
                "button_fg": "#FFFFFF",
                "danger": "#C85858",
                "warning": "#D7A24A",
                "success": "#5BAA70",
                "info": COLOR_INFO,
            }

        if theme == "claro":
            return {
                "mode": "claro",
                "root": "#EEF2F6",
                "card": "#FFFFFF",
                "card2": accent_light_soft,
                "entry": "#FFFFFF",
                "tree": "#FFFFFF",
                "heading": accent_light_heading,
                "text": "#1F2A37",
                "muted": "#566678",
                "accent": accent_hex,
                "accent2": COLOR_INFO,
                "border": "#A9BACB",
                "separator": "#CCD6E2",
                "selected_bg": accent_light_selected,
                "selected_fg": "#FFFFFF",
                "button_fg": "#FFFFFF",
                "danger": "#9B6464",
                "warning": "#8B7B5A",
                "success": "#5F7B69",
                "info": COLOR_INFO,
            }

        return {
            "mode": "oscuro",
            "root": "#0A1420",
            "card": "#111E2E",
            "card2": accent_dark_soft,
            "entry": "#0F1B2A",
            "tree": "#0B1724",
            "heading": accent_dark_heading,
            "text": "#E5EEF8",
            "muted": "#AAB8C8",
            "accent": accent_hex,
            "accent2": COLOR_INFO,
            "border": "#36516A",
            "separator": "#40566D",
            "selected_bg": accent_dark_selected,
            "selected_fg": "#FFFFFF",
            "button_fg": "#FFFFFF",
            "danger": COLOR_DANGER,
            "warning": COLOR_WARNING,
            "success": COLOR_SUCCESS,
            "info": COLOR_INFO,
        }

    def _font_size_pref(self):
        try:
            return max(10, min(18, int(self.app_settings.get("font_size", 11))))
        except Exception:
            return 11

    def _configurar_estilos_desde_preferencias(self):
        pal = self._paleta_visual_actual()
        fs = self._font_size_pref()
        style = getattr(self, "style", None) or tb.Style()

        try:
            self.root.configure(bg=pal["root"])
            self.root.option_add("*Font", f"Arial {fs}")
            self.root.option_add("*TCombobox*Listbox.background", pal["entry"])
            self.root.option_add("*TCombobox*Listbox.foreground", pal["text"])
            self.root.option_add("*TCombobox*Listbox.selectBackground", pal["selected_bg"])
            self.root.option_add("*TCombobox*Listbox.selectForeground", pal["selected_fg"])
            self.root.option_add("*TCombobox*Listbox.font", f"Arial {fs}")
            self.root.option_add("*Listbox.background", pal["entry"])
            self.root.option_add("*Listbox.foreground", pal["text"])
            self.root.option_add("*Listbox.selectBackground", pal["selected_bg"])
            self.root.option_add("*Listbox.selectForeground", pal["selected_fg"])
        except Exception:
            pass

        try:
            style.configure("Root.TFrame", background=pal["root"])
            style.configure("Card.TFrame", background=pal["card"])
            style.configure("TFrame", background=pal["card"])

            style.configure(
                "TLabelframe",
                background=pal["card"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.configure(
                "TLabelframe.Label",
                background=pal["card"],
                foreground=pal["accent"],
                font=("Arial", fs, "bold")
            )

            style.configure("TLabel", font=("Arial", fs), foreground=pal["text"], background=pal["card"])
            style.configure("Muted.TLabel", font=("Arial", max(fs - 1, 10)), foreground=pal["muted"], background=pal["card"])
            style.configure("Title.TLabel", font=("Arial", fs + 7, "bold"), foreground=pal["text"], background=pal["root"])
            style.configure("Subtitle.TLabel", font=("Arial", fs + 1), foreground=pal["muted"], background=pal["root"])
            style.configure("Section.TLabel", font=("Arial", fs + 2, "bold"), foreground=pal["accent"], background=pal["card"])

            style.configure("TButton", font=("Arial", fs, "bold"), foreground=pal["button_fg"])
            style.configure("primary.TButton", font=("Arial", fs, "bold"), background=COLOR_PRIMARY, foreground=pal["button_fg"])
            style.configure("success.TButton", font=("Arial", fs, "bold"), background=COLOR_SUCCESS, foreground=pal["button_fg"])
            style.configure("warning.TButton", font=("Arial", fs, "bold"), background=COLOR_WARNING, foreground=pal["button_fg"])
            style.configure("danger.TButton", font=("Arial", fs, "bold"), background=COLOR_DANGER, foreground=pal["button_fg"])
            style.configure("info.TButton", font=("Arial", fs, "bold"), background=COLOR_INFO, foreground=pal["button_fg"])
            style.configure("secondary.TButton", font=("Arial", fs, "bold"), background=mezclar_color_hex("#000000", pal["border"], 0.30), foreground=pal["button_fg"])

            style.configure(
                "TEntry",
                font=("Arial", fs),
                fieldbackground=pal["entry"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"],
                insertcolor=pal["text"]
            )
            style.map(
                "TEntry",
                fieldbackground=[("focus", pal["entry"]), ("!disabled", pal["entry"])],
                foreground=[("!disabled", pal["text"])]
            )

            style.configure(
                "TCombobox",
                font=("Arial", fs),
                fieldbackground=pal["entry"],
                background=pal["entry"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"],
                arrowcolor=pal["accent"]
            )
            style.map(
                "TCombobox",
                fieldbackground=[("readonly", pal["entry"]), ("!disabled", pal["entry"])],
                background=[("readonly", pal["entry"]), ("!disabled", pal["entry"])],
                foreground=[("readonly", pal["text"]), ("!disabled", pal["text"])],
                selectbackground=[("readonly", pal["selected_bg"]), ("!disabled", pal["selected_bg"])],
                selectforeground=[("readonly", pal["selected_fg"]), ("!disabled", pal["selected_fg"])],
                arrowcolor=[("readonly", pal["accent"]), ("!disabled", pal["accent"])]
            )

            style.configure("TCheckbutton", font=("Arial", fs), foreground=pal["text"], background=pal["card"])
            style.configure("TRadiobutton", font=("Arial", fs), foreground=pal["text"], background=pal["card"])

            style.configure("TNotebook", background=pal["card"], bordercolor=pal["border"])
            style.configure(
                "TNotebook.Tab",
                font=("Arial", max(fs - 1, 10), "bold"),
                padding=(10, 5),
                background=pal["card2"],
                foreground=pal["text"]
            )
            style.map(
                "TNotebook.Tab",
                background=[("selected", pal["heading"]), ("active", pal["card2"])],
                foreground=[("selected", pal["text"]), ("active", pal["text"])]
            )

            style.configure(
                "Treeview",
                font=("Arial", fs),
                rowheight=max(int(self.app_settings.get("table_row_height", 29) or 29), fs + 19),
                background=pal["tree"],
                foreground=pal["text"],
                fieldbackground=pal["tree"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.configure(
                "Treeview.Heading",
                font=("Arial", fs, "bold"),
                background=pal["heading"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.map(
                "Treeview",
                background=[("selected", pal["selected_bg"])],
                foreground=[("selected", pal["selected_fg"])]
            )

            style.configure(
                "Modern.Treeview",
                font=("Arial", fs),
                rowheight=max(int(self.app_settings.get("table_row_height", 29) or 29), fs + 20),
                background=pal["tree"],
                foreground=pal["text"],
                fieldbackground=pal["tree"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.configure(
                "Modern.Treeview.Heading",
                font=("Arial", fs, "bold"),
                background=pal["heading"],
                foreground=pal["text"],
                bordercolor=pal["border"],
                lightcolor=pal["border"],
                darkcolor=pal["border"]
            )
            style.map(
                "Modern.Treeview",
                background=[("selected", pal["selected_bg"])],
                foreground=[("selected", pal["selected_fg"])]
            )

            style.configure("TSeparator", background=pal["separator"])
        except Exception:
            pass

    def _aplicar_preferencias_a_widgets(self, base=None):
        pal = self._paleta_visual_actual()
        fs = self._font_size_pref()
        base = base or self.root

        old_root_colors = {
            "#07111f", "#07111F", "#000000", "#050505", "#080A0D", "#F3F6FA",
            "#f3f6fa", "#EEF3F8", "#eef3f8", "#FDFDFD", "#fdfdfd", "#FAFAFA", "#fafafa"
        }
        old_card_colors = {
            "#0E1B2B", "#0e1b2b", "#111E2E", "#111e2e", "#10151C", "#10151c",
            "#0A0A0A", "#0a0a0a", "#FFFFFF", "#ffffff", "#F6F9FC", "#f6f9fc",
            "#121212", "#132337", "#17283B", "#17283b"
        }
        old_entry_colors = {
            "#111E2E", "#111e2e", "#0F1D2E", "#0f1d2e", "#0B1624", "#0b1624",
            "#FFFFFF", "#ffffff", "#000000", "#080A0D", "#080a0d", "#0F1B2A", "#0f1b2a"
        }

        def safe_conf(w, **kwargs):
            for k, v in kwargs.items():
                try:
                    w.configure(**{k: v})
                except Exception:
                    pass

        def get_bg(w):
            for opt in ("background", "bg"):
                try:
                    return str(w.cget(opt))
                except Exception:
                    continue
            return ""

        def decide_bg(w, text=""):
            current_bg = get_bg(w)
            upper = (text or "").upper()

            if isinstance(w, (tk.Tk, tk.Toplevel)):
                return pal["root"]

            if any(x in upper for x in ["GENERADOR DE FORMULARIOS", "SISTEMA DE ADMISIÓN", "SISTEMA DE ADMISION"]):
                return pal["root"]

            if current_bg in old_root_colors:
                return pal["root"]

            if isinstance(w, (tk.Entry, tk.Text, tk.Listbox)):
                return pal["entry"]

            if current_bg in old_entry_colors and not isinstance(w, (tk.Label, ttk.Label)):
                return pal["entry"]

            if current_bg in old_card_colors:
                return pal["card"]

            return pal["card"]

        def walk(w):
            try:
                text = str(w.cget("text"))
            except Exception:
                text = ""

            bg = decide_bg(w, text)

            if isinstance(w, (tk.Tk, tk.Toplevel)):
                safe_conf(w, bg=pal["root"])

            elif isinstance(w, tk.Frame):
                safe_conf(w, bg=bg, highlightbackground=pal["border"], highlightcolor=pal["accent"])

            elif isinstance(w, tk.LabelFrame):
                safe_conf(w, bg=pal["card"], fg=pal["text"], font=("Arial", fs, "bold"),
                          highlightbackground=pal["border"], highlightcolor=pal["accent"])

            elif isinstance(w, tk.Label):
                upper = text.upper()
                fsize, weight = fs, "normal"
                fg = pal["text"]

                if any(x in upper for x in ["GENERADOR", "CONFIGURACIÓN", "CONFIGURACION", "HISTORIAL", "REPORTE"]):
                    fsize, weight = fs + 4, "bold"
                    fg = pal["text"]
                elif any(x in upper for x in ["DATOS", "ACCIONES", "INFORMACIÓN", "INFORMACION", "RESUMEN"]):
                    fsize, weight = fs + 1, "bold"
                    fg = pal["accent"]

                safe_conf(w, bg=bg, fg=fg, font=("Arial", fsize, weight),
                          activebackground=bg, activeforeground=pal["accent"])

            elif isinstance(w, ttk.Label):
                upper = text.upper()
                style_name = "TLabel"
                fg = pal["text"]
                if any(x in upper for x in ["GENERADOR", "CONFIGURACIÓN", "CONFIGURACION", "HISTORIAL", "REPORTE"]):
                    style_name = "Title.TLabel"
                    bg = pal["root"]
                elif any(x in upper for x in ["DATOS", "ACCIONES", "INFORMACIÓN", "INFORMACION", "RESUMEN"]):
                    style_name = "Section.TLabel"
                    fg = pal["accent"]

                safe_conf(w, style=style_name, background=bg, foreground=fg)

            elif isinstance(w, (tk.Entry, tk.Text, tk.Listbox)):
                safe_conf(
                    w,
                    bg=pal["entry"],
                    fg=pal["text"],
                    insertbackground=pal["text"],
                    font=("Arial", fs),
                    highlightbackground=pal["border"],
                    highlightcolor=pal["accent"],
                    selectbackground=pal["selected_bg"],
                    selectforeground=pal["selected_fg"]
                )

            elif isinstance(w, tk.Button):
                safe_conf(
                    w,
                    bg=pal["heading"],
                    fg=pal["text"],
                    activebackground=pal["accent"],
                    activeforeground=pal["selected_fg"],
                    font=("Arial", fs, "bold"),
                    highlightbackground=pal["border"],
                    highlightcolor=pal["accent"]
                )

            try:
                if isinstance(w, ttk.Treeview):
                    w.configure(style="Modern.Treeview")
            except Exception:
                pass

            try:
                for child in w.winfo_children():
                    walk(child)
            except Exception:
                pass

        try:
            walk(base)
        except Exception:
            pass

    def _aplicar_preferencias_en_vivo(self, ventana_actual=None, forzar_todo=False):
        """
        FASE 13: Aplica preferencias sin recorrer todos los widgets cada vez.
        - Al iniciar: root.
        - Al guardar preferencias: usar forzar_todo=True.
        """
        self._configurar_estilos_desde_preferencias()

        if forzar_todo:
            self._aplicar_preferencias_a_widgets(self.root)
            for attr in (
                "historial_win",
                "historial_sin_seguro_win",
                "reporte_win",
                "turno_win",
                "dialogo_unico_win",
                "configuracion_interna_win",
                "edicion_paciente_win",
                "salida_pendiente_win",
                "trabajos_salida_win",
            ):
                try:
                    w = getattr(self, attr, None)
                    if w is not None and w.winfo_exists():
                        self._aplicar_preferencias_a_widgets(w)
                except Exception:
                    pass
        elif ventana_actual is not None:
            self._aplicar_preferencias_a_widgets(ventana_actual)

        try:
            self.root.geometry(self.app_settings.get("window_size", "1280x740"))
        except Exception:
            pass

        try:
            self.root.update_idletasks()
        except Exception:
            pass
        self._aplicar_modo_responsivo()

    def _programar_modo_responsivo(self, event=None):
        if event is not None and event.widget is not self.root:
            return
        previo = getattr(self, "_responsive_after_id", None)
        if previo:
            try:
                self.root.after_cancel(previo)
            except Exception:
                pass
        self._responsive_after_id = self.root.after(80, self._aplicar_modo_responsivo)

    def _aplicar_modo_responsivo(self):
        try:
            self._responsive_after_id = None
            ancho = max(1, int(self.root.winfo_width()))
            forzar_pequeno = bool(self.app_settings.get("small_screen_mode", False))
            compacto = bool(self.app_settings.get("compact_mode", False)) or ancho < 1050
            mostrar_panel = (
                bool(self.app_settings.get("show_side_panel", True))
                and ancho >= 1180
                and not forzar_pequeno
            )

            if mostrar_panel:
                self.quick_panel.grid()
            else:
                self.quick_panel.grid_remove()

            if ancho >= 1040 and not forzar_pequeno:
                self.info_header.grid()
            else:
                self.info_header.grid_remove()

            margen = 8 if compacto else 14
            padding_form = 10 if compacto else 18
            self.main.configure(padding=margen)
            self.frame.configure(padding=padding_form)
            self.form_host.grid_configure(padx=(0, 8 if mostrar_panel else 0))
            self.title_lbl.configure(wraplength=430 if compacto else 650)
            self._configurar_columnas_formulario(una_columna=ancho < 850)

            pal = self._paleta_visual_actual()
            self.form_canvas.configure(background=pal["card"])
            tamano = str(self.app_settings.get("button_size", "Normal"))
            padding_boton = {
                "Compacto": (8, 3),
                "Normal": (10, 5),
                "Grande": (12, 7),
            }.get(tamano, (10, 5))
            self.style.configure("TButton", padding=padding_boton)
        except (AttributeError, tk.TclError):
            pass

    def _configurar_columnas_formulario(self, una_columna=False):
        if getattr(self, "_formulario_una_columna", None) is bool(una_columna):
            return
        self._formulario_una_columna = bool(una_columna)
        if una_columna:
            layout = (
                (self.lbl_nombre, 2, 0, 6, "w", (4, 4)),
                (self.entry_nombre, 3, 0, 6, "ew", (4, 4)),
                (self.lbl_sexo, 4, 0, 6, "w", (4, 4)),
                (self.sexo_frame, 5, 0, 6, "w", (4, 4)),
                (self.lbl_edad, 6, 0, 6, "w", (4, 4)),
                (self.entry_edad, 7, 0, 4, "ew", (4, 8)),
                (self.combo_unidad, 7, 4, 2, "ew", (0, 4)),
                (self.check_urgencia, 8, 0, 6, "w", (4, 4)),
                (self.lbl_cedula, 9, 0, 6, "w", (4, 4)),
                (self.entry_cedula, 10, 0, 6, "ew", (4, 4)),
                (self.lbl_telefono, 11, 0, 6, "w", (4, 4)),
                (self.entry_telefono, 12, 0, 6, "ew", (4, 4)),
                (self.lbl_direccion, 13, 0, 6, "w", (4, 4)),
                (self.entry_direccion, 14, 0, 6, "ew", (4, 4)),
                (self.lbl_nacionalidad, 15, 0, 6, "w", (4, 4)),
                (self.entry_nacionalidad, 16, 0, 6, "ew", (4, 4)),
                (self.lbl_ars, 17, 0, 6, "w", (4, 4)),
                (self.entry_ars, 18, 0, 6, "ew", (4, 4)),
                (self.ars_detectado_label, 19, 0, 6, "w", (4, 4)),
                (self.ars_suggestions, 20, 0, 6, "ew", (4, 4)),
                (self.lbl_nss, 21, 0, 4, "w", (4, 4)),
                (self.nss_detectado_label, 21, 4, 2, "w", (4, 4)),
                (self.entry_nss, 22, 0, 6, "ew", (4, 4)),
                (self.form_actions_separator, 23, 0, 6, "ew", (4, 4)),
                (self.form_buttons, 24, 0, 6, "e", (4, 4)),
            )
        else:
            layout = (
                (self.lbl_nombre, 2, 0, 3, "w", (4, 10)),
                (self.entry_nombre, 3, 0, 3, "ew", (4, 24)),
                (self.lbl_sexo, 2, 3, 2, "w", (4, 10)),
                (self.sexo_frame, 3, 3, 3, "w", (4, 4)),
                (self.lbl_edad, 4, 0, 1, "w", (4, 10)),
                (self.entry_edad, 5, 0, 1, "ew", (4, 8)),
                (self.combo_unidad, 5, 1, 2, "ew", (0, 24)),
                (self.check_urgencia, 5, 3, 3, "w", (4, 4)),
                (self.lbl_cedula, 6, 0, 3, "w", (4, 10)),
                (self.entry_cedula, 7, 0, 3, "ew", (4, 24)),
                (self.lbl_telefono, 6, 3, 3, "w", (4, 10)),
                (self.entry_telefono, 7, 3, 3, "ew", (4, 4)),
                (self.lbl_direccion, 8, 0, 6, "w", (4, 10)),
                (self.entry_direccion, 9, 0, 6, "ew", (4, 4)),
                (self.lbl_nacionalidad, 10, 0, 3, "w", (4, 10)),
                (self.entry_nacionalidad, 11, 0, 3, "ew", (4, 24)),
                (self.lbl_ars, 10, 3, 3, "w", (4, 10)),
                (self.entry_ars, 11, 3, 3, "ew", (4, 4)),
                (self.ars_detectado_label, 12, 3, 3, "w", (4, 4)),
                (self.ars_suggestions, 13, 3, 3, "ew", (4, 4)),
                (self.lbl_nss, 13, 0, 1, "w", (4, 10)),
                (self.nss_detectado_label, 13, 1, 2, "w", (4, 24)),
                (self.entry_nss, 14, 0, 3, "ew", (4, 24)),
                (self.form_actions_separator, 15, 0, 6, "ew", (4, 4)),
                (self.form_buttons, 16, 0, 6, "e", (4, 4)),
            )
        suggestions_visible = self.ars_suggestions.winfo_ismapped()
        for widget, row, column, span, sticky, padx in layout:
            widget.grid_configure(
                row=row,
                column=column,
                columnspan=span,
                sticky=sticky,
                padx=padx,
            )
        if not suggestions_visible:
            self.ars_suggestions.grid_remove()


    def _ejecutar_en_segundo_plano(self, mensaje, funcion, al_terminar=None, al_error=None):
        def worker():
            try:
                resultado = funcion()
                if al_terminar:
                    self.root.after(0, lambda: al_terminar(resultado))
            except Exception as e:
                if al_error:
                    self.root.after(0, lambda err=e: al_error(err))
                else:
                    self.root.after(0, lambda err=e: messagebox.showerror("Error", str(err)))

        try:
            self.set_status(mensaje, "process")
        except Exception:
            pass

        hilo = threading.Thread(target=worker, daemon=True)
        hilo.start()
        return hilo

    def _actor_actual(self):
        turno = cargar_turno_config(permitir_vencido=True) or {}
        return limpiar_nombre_representante(turno.get("representante", "")) or os.environ.get("USERNAME", "OPERADOR")

    def _solicitar_autorizacion_admin(self, accion, parent=None, force=False):
        parent = parent or self.root
        actor = self._actor_actual()
        if (
            not force
            and self._admin_authorized_until
            and datetime.now() < self._admin_authorized_until
            and self._admin_authorized_actor
        ):
            return self._admin_authorized_actor
        try:
            configured = self.security.is_configured()
        except (ConfigError, SecurityError) as exc:
            messagebox.showerror("Seguridad", str(exc), parent=parent)
            return None
        if not configured:
            if not messagebox.askyesno(
                "Configurar administración",
                "Esta acción requiere un PIN administrativo y todavía no existe uno.\n\n"
                "¿Desea configurarlo ahora?",
                parent=parent,
            ):
                return None
            first = simpledialog.askstring(
                "Nuevo PIN administrativo",
                "Defina un PIN numérico de al menos 6 dígitos:",
                show="*",
                parent=parent,
            )
            if first is None:
                return None
            second = simpledialog.askstring(
                "Confirmar PIN",
                "Repita el PIN administrativo:",
                show="*",
                parent=parent,
            )
            if first != second:
                messagebox.showerror("Seguridad", "Los PIN no coinciden.", parent=parent)
                return None
            try:
                self.security.setup(first, actor=actor)
            except SecurityError as exc:
                messagebox.showerror("Seguridad", str(exc), parent=parent)
                return None

        pin = simpledialog.askstring(
            "Autorización administrativa",
            f"Ingrese el PIN para autorizar:\n{accion}",
            show="*",
            parent=parent,
        )
        if pin is None:
            return None
        try:
            if not self.security.verify(pin, actor=actor, action=accion):
                messagebox.showerror("Seguridad", "PIN administrativo incorrecto.", parent=parent)
                return None
        except SecurityError as exc:
            messagebox.showerror("Seguridad", str(exc), parent=parent)
            return None
        self._admin_authorized_until = datetime.now() + timedelta(minutes=5)
        self._admin_authorized_actor = actor
        return actor

    def buscar_actualizaciones(self, manual=True):
        """Consulta GitHub sin bloquear la interfaz y ofrece una instalacion verificada."""
        def consultar():
            return get_latest_release()

        def completado(release):
            self.set_status("Actualizaciones comprobadas", "ok")
            if not is_newer(release.version, APP_VERSION):
                if manual:
                    messagebox.showinfo(
                        "Actualizaciones",
                        f"Ya tiene la version mas reciente ({APP_VERSION}).",
                        parent=self.root,
                    )
                return
            notas = release.notes[:900].strip()
            detalle = f"\n\n{notas}" if notas else ""
            aceptar = messagebox.askyesno(
                "Actualizacion disponible",
                f"Esta disponible la version {release.version}.\n\n"
                "La aplicacion se cerrara, verificara el paquete y volvera a abrirse. "
                "La base de datos y los documentos no se reemplazaran."
                f"{detalle}\n\n¿Desea actualizar ahora?",
                parent=self.root,
            )
            if aceptar:
                self._iniciar_actualizador_externo()

        def fallo(exc):
            APP_LOG.warning("No se pudo consultar la actualizacion: %s", exc)
            self.set_status("No se pudo comprobar actualizaciones", "warning")
            if manual:
                messagebox.showwarning(
                    "Actualizaciones",
                    f"No se pudo comprobar si hay una version nueva:\n\n{exc}",
                    parent=self.root,
                )

        self._ejecutar_en_segundo_plano(
            "Comprobando actualizaciones...",
            consultar,
            al_terminar=completado,
            al_error=fallo,
        )

    def _iniciar_actualizador_externo(self):
        if not getattr(sys, "frozen", False):
            messagebox.showinfo(
                "Actualizaciones",
                "La actualizacion automatica se prueba desde la version empaquetada.",
                parent=self.root,
            )
            return
        install_dir = os.path.dirname(os.path.abspath(sys.executable))
        updater = os.path.join(install_dir, "ACTUALIZADOR.exe")
        if not os.path.isfile(updater):
            messagebox.showerror(
                "Actualizaciones",
                "No se encontro ACTUALIZADOR.exe junto a la aplicacion.",
                parent=self.root,
            )
            return
        try:
            temp_updater = os.path.join(
                tempfile.gettempdir(),
                f"GeneradorHojas_Actualizador_{os.getpid()}.exe",
            )
            shutil.copy2(updater, temp_updater)
            subprocess.Popen(
                [
                    temp_updater,
                    "--install-dir",
                    install_dir,
                    "--current-version",
                    APP_VERSION,
                    "--wait-pid",
                    str(os.getpid()),
                ],
                cwd=tempfile.gettempdir(),
                close_fds=True,
            )
            self.on_close()
        except OSError as exc:
            APP_LOG.exception("No se pudo iniciar el actualizador")
            messagebox.showerror(
                "Actualizaciones",
                f"No se pudo iniciar el actualizador:\n\n{exc}",
                parent=self.root,
            )

    def set_status(self, mensaje, tipo="info"):
        colores = {
            "info": "#72E39B",
            "ok": "#72E39B",
            "warning": "#FFD166",
            "error": "#FF6B6B",
            "process": "#5CB6FF",
        }
        try:
            self.status_var.set(mensaje)
            self.status_label.configure(foreground=colores.get(tipo, "#72E39B"))
            self.root.update_idletasks()
        except Exception:
            pass

    def _configurar_accesibilidad_teclado(self):
        self._tab_order = [
            self.entry_nombre,
            self.lbl_sexo_m,
            self.lbl_sexo_f,
            self.entry_edad,
            self.combo_unidad,
            self.check_urgencia,
            self.entry_cedula,
            self.entry_telefono,
            self.entry_direccion,
            self.entry_nacionalidad,
            self.entry_ars,
            self.entry_nss,
            self.boton_limpiar,
            self.boton_generar_pdf,
            self.boton_historial,
            self.actions_menu_button,
        ]
        for w in self._tab_order:
            try:
                w.configure(takefocus=True)
            except Exception:
                pass
            try:
                w.bind("<Tab>", lambda e, widget=w: self._focus_siguiente(widget), add="+")
                w.bind("<Shift-Tab>", lambda e, widget=w: self._focus_anterior(widget), add="+")
                w.bind("<ISO_Left_Tab>", lambda e, widget=w: self._focus_anterior(widget), add="+")
            except Exception:
                pass

        self.root.bind("<Control-l>", lambda e: (self.limpiar_campos(), self.set_status("✓  Campos limpiados", "ok"), "break"))
        self.root.bind("<Control-L>", lambda e: (self.limpiar_campos(), self.set_status("✓  Campos limpiados", "ok"), "break"))
        self.root.bind("<Control-h>", lambda e: (self.abrir_historial(), "break"))
        self.root.bind("<Control-H>", lambda e: (self.abrir_historial(), "break"))
        self.root.bind("<Control-z>", self._undo_last_action)
        self.root.bind("<Control-Z>", self._undo_last_action)
        def registrar_desde_teclado(_event=None):
            self.generar_pdf()
            return "break"

        self.root.bind("<Control-Return>", registrar_desde_teclado)
        self.root.bind("<F9>", registrar_desde_teclado)

    def _focus_siguiente(self, widget):
        try:
            idx = self._tab_order.index(widget)
            self._tab_order[(idx + 1) % len(self._tab_order)].focus_set()
            return "break"
        except Exception:
            return None

    def _focus_anterior(self, widget):
        try:
            idx = self._tab_order.index(widget)
            self._tab_order[(idx - 1) % len(self._tab_order)].focus_set()
            return "break"
        except Exception:
            return None

    def _campos_minimos_completos(self):
        nombre = (self.entry_nombre.get() or "").strip()
        telefono = (self.entry_telefono.get() or "").strip()
        edad = (self.entry_edad.get() or "").strip()
        unidad = self.unidad_edad.get()
        return bool(nombre and telefono.isdigit() and len(telefono) == 10 and edad.isdigit() and unidad in ("Días", "Meses", "Años"))

    def _enter_generar_si_completo(self, event=None):
        try:
            if isinstance(event.widget, tk.Listbox):
                return None
        except Exception:
            pass
        if self._campos_minimos_completos():
            self.generar_pdf()
            return "break"
        self.set_status("Complete nombre, edad y teléfono válido para generar con Enter.", "warning")
        return None

    def _obtener_ars_cache(self, forzar=False):
        try:
            ahora = datetime.now().timestamp()
            if (not forzar) and self._cache_ars and (ahora - float(self._cache_ars_time or 0) < 30):
                return list(self._cache_ars)
            self._cache_ars = self.db.listar_ars_distintas()
            self._cache_ars_time = ahora
            return list(self._cache_ars)
        except Exception:
            return []

    def _invalidar_cache_ars(self):
        self._cache_ars = []
        self._cache_ars_time = 0
        self._invalidar_cache_resumen_turno()

    def _invalidar_cache_resumen_turno(self):
        self._cache_resumen_turno = None
        self._cache_resumen_time = 0

    def _obtener_resumen_turno_cache(self, forzar=False):
        ahora = datetime.now().timestamp()
        if (not forzar) and self._cache_resumen_turno and (ahora - float(self._cache_resumen_time or 0) < 10):
            return dict(self._cache_resumen_turno)

        resumen = self.db.resumen_turno_actual()
        self._cache_resumen_turno = dict(resumen)
        self._cache_resumen_time = ahora
        self._guardar_resumen_turno_json(resumen)
        return resumen

    def _guardar_resumen_turno_json(self, resumen: dict):
        """
        FASE 7: resumen_turno.json para estadísticas rápidas.
        """
        try:
            payload = {
                "actualizado": datetime.now().strftime("%d/%m/%Y %I:%M:%S %p"),
                "total": int(resumen.get("total", 0) or 0),
                "sin_seguro": int(resumen.get("sin_seguro", 0) or 0),
                "general": int(resumen.get("GENERAL", 0) or 0),
                "pediatria": int(resumen.get("PEDIATRIA", 0) or 0),
                "ginecologia": int(resumen.get("GINECOLOGIA", 0) or 0),
                "urgencias": int(resumen.get("URGENCIAS", 0) or 0),
            }
            with open(RESUMEN_TURNO_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _invalidar_caches_datos(self):
        self._invalidar_cache_ars()
        self._invalidar_cache_resumen_turno()


    def _actualizar_turno_visual_en_vivo(self):
        try:
            cfg = cargar_turno_config()
            turno_txt = descripcion_turno_config(cfg)
            representante = limpiar_nombre_representante(cfg.get("representante") if cfg else "") or "No configurado"

            if hasattr(self, "turno_header_var"):
                self.turno_header_var.set(turno_txt)
            if hasattr(self, "turno_panel_var"):
                self.turno_panel_var.set(f"Turno: {turno_txt}")
            if hasattr(self, "representante_panel_var"):
                self.representante_panel_var.set(f"Representante: {limpiar_nombre_representante(representante) or 'No configurado'}")

            self._refrescar_resumen_en_vivo()
            try:
                self.root.update_idletasks()
            except Exception:
                pass
        except Exception:
            pass

    def _actualizar_fecha_actual(self):
        try:
            if hasattr(self, "fecha_actual_var"):
                self.fecha_actual_var.set(datetime.now().strftime("%d/%m/%Y"))
        except Exception:
            pass
        try:
            self.root.after(60000, self._actualizar_fecha_actual)
        except Exception:
            pass

    def _crear_panel_acciones_rapidas(self, parent):
        for w in parent.winfo_children():
            try:
                w.destroy()
            except Exception:
                pass

        pal = self._paleta_visual_actual()

        bottom_info = tb.Frame(parent, padding=(0, 6), style="Card.TFrame")
        bottom_info.pack(side="bottom", fill="x", pady=(8, 0))

        tb.Separator(bottom_info, orient="horizontal").pack(fill="x", pady=(0, 8))

        tb.Label(
            bottom_info,
            text="INFORMACIÓN",
            font=("Arial", 9, "bold"),
            foreground=pal["muted"],
            background=pal["card"]
        ).pack(anchor="w", pady=(0, 4))

        try:
            cfg = cargar_turno_config()
            representante = limpiar_nombre_representante(cfg.get("representante") if cfg else "") or "No configurado"
            turno = descripcion_turno_config(cfg)
        except Exception:
            representante = "No configurado"
            turno = "No configurado"

        self.representante_panel_var = tk.StringVar(value=f"Representante: {representante}")
        self.turno_panel_var = tk.StringVar(value=f"Turno: {turno}")

        tb.Label(
            bottom_info,
            textvariable=self.representante_panel_var,
            style="Muted.TLabel",
            background=pal["card"],
            wraplength=250,
            justify="left"
        ).pack(anchor="w", pady=1)

        tb.Label(
            bottom_info,
            textvariable=self.turno_panel_var,
            style="Muted.TLabel",
            background=pal["card"],
            wraplength=250,
            justify="left"
        ).pack(anchor="w", pady=(1, 6))

        tb.Label(
            bottom_info,
            text="RESUMEN DEL TURNO",
            font=("Arial", 9, "bold"),
            foreground=pal["muted"],
            background=pal["card"]
        ).pack(anchor="w", pady=(2, 3))

        self.turno_resumen_var = tk.StringVar(value="Cargando resumen…")
        tb.Label(
            bottom_info,
            textvariable=self.turno_resumen_var,
            style="Muted.TLabel",
            background=pal["card"],
            justify="left",
            wraplength=250
        ).pack(anchor="w", pady=(0, 0))

        actions_area = tb.Frame(parent, padding=(0, 0), style="Card.TFrame")
        actions_area.pack(side="top", fill="x")

        tb.Label(
            actions_area,
            text="ACCIONES RÁPIDAS",
            font=("Arial", 11, "bold"),
            foreground=pal["accent"],
            background=pal["card"]
        ).pack(anchor="w", pady=(0, 8))

        def add_action(texto, icono, comando, bootstyle=SECONDARY):
            btn = tb.Button(
                actions_area,
                text=texto,
                bootstyle=bootstyle,
                command=comando,
                width=28
            )
            btn.pack(fill="x", pady=3, ipady=4)
            return btn

        add_action("Reporte estadístico", "📊", self.abrir_ventana_reporte, INFO)
        add_action("Abrir Listado en Excel", "▤", self._abrir_excel_actual, SUCCESS)
        # FASE 14: Icono uniforme 🛡
        add_action("Ver Historial sin Seguro", "🛡", self.abrir_historial_sin_seguros, WARNING)
        add_action("Editar paciente", "🖉", self._abrir_edicion_paciente, PRIMARY)
        add_action("Impresiones pendientes", "!", self.abrir_trabajos_salida_pendientes, WARNING)
        add_action("Configuración interna", "⚙", self._abrir_configuracion_interna, SECONDARY)

        self._actualizar_resumen_turno_panel()

    def _actualizar_resumen_turno_panel(self, forzar=False):
        try:
            r = self._obtener_resumen_turno_cache(forzar=forzar)
            fuente = "Base de datos"
            if r.get("_fuente") == "BD_EXCEL":
                fuente = "BD y Excel sincronizados"
            elif os.path.exists(EXCEL_PATH):
                fuente = "BD · Excel pendiente de actualización"
            if r.get("_fuente") == "EXCEL_RECUPERADO":
                fuente = "Excel recuperado · revisión de sincronización"

            if not bool(self.app_settings.get("show_turno_summary", True)):
                texto = "Resumen oculto por preferencias."
            else:
                texto = (
                    f"Total pacientes: {r.get('total', 0)}\n"
                    f"Sin seguro: {r.get('sin_seguro', 0)}\n"
                    f"General: {r.get('GENERAL', 0)}\n"
                    f"Pediatría: {r.get('PEDIATRIA', 0)}\n"
                    f"Ginecología: {r.get('GINECOLOGIA', 0)}\n"
                    f"Urgencias: {r.get('URGENCIAS', 0)}"
                )

            if hasattr(self, "turno_resumen_var"):
                self.turno_resumen_var.set(texto)

            if hasattr(self, "connection_var"):
                self.connection_var.set(f"Conectado a: {fuente}")

            try:
                self.root.update_idletasks()
            except Exception:
                pass

        except Exception:
            if hasattr(self, "turno_resumen_var"):
                self.turno_resumen_var.set("No disponible")
            if hasattr(self, "connection_var"):
                self.connection_var.set("Conectado a: no disponible")

    def _programar_refresco_resumen_en_vivo(self):
        """
        FASE 6: Refresco cada 15 segundos y solo si la ventana principal está activa.
        """
        try:
            if getattr(self, "root", None) and self.root.winfo_exists():
                if self.root.state() != "withdrawn":
                    try:
                        activo = bool(self.root.focus_displayof())
                    except Exception:
                        activo = True
                    if activo:
                        self._actualizar_resumen_turno_panel(forzar=False)
                self.root.after(15000, self._programar_refresco_resumen_en_vivo)
        except Exception:
            pass

    def _refrescar_resumen_en_vivo(self, delay_ms=0):
        try:
            self._invalidar_cache_resumen_turno()
            if delay_ms and delay_ms > 0:
                self.root.after(delay_ms, lambda: self._actualizar_resumen_turno_panel(forzar=True))
            else:
                self._actualizar_resumen_turno_panel(forzar=True)
        except Exception:
            pass

    def _abrir_excel_actual(self):
        try:
            verificar_o_crear_excel()
            ruta = os.path.abspath(EXCEL_PATH)
            if platform.system() == "Windows":
                os.startfile(ruta)
            elif platform.system() == "Darwin":
                subprocess.run(["open", ruta], check=False)
            else:
                subprocess.run(["xdg-open", ruta], check=False)
            self.set_status("Listado de Excel abierto", "ok")
        except Exception as e:
            self.set_status("Error: Excel abierto o no disponible", "error")
            messagebox.showerror("Error", f"No se pudo abrir el listado de Excel:\n{str(e)}")

    def abrir_ventana_reporte(self):
        win = self._crear_toplevel_estable("Reporte estadístico", "960x720", "reporte_win")
        if win is None:
            return

        self._bind_esc_cerrar(win)

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Reporte estadístico",
            "Genera el reporte de pacientes por período, ARS, especialidad y conteo general.",
            "📊"
        )

        try:
            win.minsize(940, 690)
            win.resizable(True, True)
        except Exception:
            pass

        barra = tb.Frame(cont, padding=(8, 8), style="Card.TFrame")
        barra.pack(side="bottom", fill="x", pady=(8, 0))

        panel = tb.Frame(cont, padding=12, style="Card.TFrame")
        panel.pack(fill="x", pady=(0, 10))
        panel.columnconfigure(1, weight=1)
        panel.columnconfigure(3, weight=1)

        periodo_var = tk.StringVar(value="Diario")
        fecha_inicio = crear_selector_fecha(panel, width=14)
        fecha_fin = crear_selector_fecha(panel, width=14)

        hoy = datetime.now().date()
        establecer_fecha_selector(fecha_inicio, hoy)
        establecer_fecha_selector(fecha_fin, hoy)

        tb.Label(panel, text="Período:", background="#0E1B2B", font=("Arial", 10, "bold")).grid(
            row=0, column=0, sticky="w", padx=6, pady=6
        )
        combo_periodo = tb.Combobox(
            panel,
            textvariable=periodo_var,
            state="readonly",
            values=["Diario", "Semanal", "Mensual", "Anual", "Personalizado"],
            width=18
        )
        combo_periodo.grid(row=0, column=1, sticky="w", padx=6, pady=6)

        tb.Label(panel, text="Desde:", background="#0E1B2B", font=("Arial", 10, "bold")).grid(
            row=1, column=0, sticky="w", padx=6, pady=6
        )
        fecha_inicio.grid(row=1, column=1, sticky="w", padx=6, pady=6)

        tb.Label(panel, text="Hasta:", background="#0E1B2B", font=("Arial", 10, "bold")).grid(
            row=1, column=2, sticky="w", padx=6, pady=6
        )
        fecha_fin.grid(row=1, column=3, sticky="w", padx=6, pady=6)

        estado_var = tk.StringVar(value="Seleccione el período y presione Generar reporte.")
        tb.Label(panel, textvariable=estado_var, style="Muted.TLabel", background="#0E1B2B").grid(
            row=2, column=0, columnspan=4, sticky="w", padx=6, pady=(4, 0)
        )

        vista = ttk.Treeview(
            cont,
            columns=("seccion", "concepto", "cantidad"),
            show="headings",
            height=10,
            style="Modern.Treeview"
        )
        vista.pack(fill="both", expand=True, pady=(0, 6))
        vista.heading("seccion", text="Sección")
        vista.heading("concepto", text="Concepto")
        vista.heading("cantidad", text="Cantidad")
        vista.column("seccion", width=170, anchor="w")
        vista.column("concepto", width=420, anchor="w")
        vista.column("cantidad", width=110, anchor="center")

        ultimo_resumen = {"data": None, "ruta": ""}

        def _rango_periodo():
            base_txt = obtener_fecha_selector(fecha_inicio)
            base = parse_fecha_ddmmyyyy(base_txt) or datetime.now().date()
            modo = periodo_var.get()

            if modo == "Diario":
                inicio, fin = obtener_rango_operativo_desde_fecha(base)
                etiqueta = f"Día operativo {base.strftime('%d/%m/%Y')}"
                return inicio, fin, etiqueta

            if modo == "Semanal":
                inicio_date = base - timedelta(days=base.weekday())
                fin_date = inicio_date + timedelta(days=7)
                inicio = datetime.combine(inicio_date, time(8, 0))
                fin = datetime.combine(fin_date, time(8, 0))
                etiqueta = f"Semana {inicio_date.strftime('%d/%m/%Y')} a {(fin_date - timedelta(days=1)).strftime('%d/%m/%Y')}"
                return inicio, fin, etiqueta

            if modo == "Mensual":
                inicio_date = base.replace(day=1)
                if inicio_date.month == 12:
                    siguiente = date(inicio_date.year + 1, 1, 1)
                else:
                    siguiente = date(inicio_date.year, inicio_date.month + 1, 1)
                inicio = datetime.combine(inicio_date, time(8, 0))
                fin = datetime.combine(siguiente, time(8, 0))
                etiqueta = f"Mes {inicio_date.strftime('%m/%Y')}"
                return inicio, fin, etiqueta

            if modo == "Anual":
                inicio_date = date(base.year, 1, 1)
                siguiente = date(base.year + 1, 1, 1)
                inicio = datetime.combine(inicio_date, time(8, 0))
                fin = datetime.combine(siguiente, time(8, 0))
                etiqueta = f"Año {base.year}"
                return inicio, fin, etiqueta

            ini = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or base
            fin_d = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_fin)) or ini
            if fin_d < ini:
                ini, fin_d = fin_d, ini
            inicio = datetime.combine(ini, time(0, 0))
            fin = datetime.combine(fin_d + timedelta(days=1), time(0, 0))
            etiqueta = f"Personalizado {ini.strftime('%d/%m/%Y')} a {fin_d.strftime('%d/%m/%Y')}"
            return inicio, fin, etiqueta

        def _cargar_vista_resumen(resumen):
            vista.delete(*vista.get_children())
            vista.insert("", "end", values=("Resumen general", "Total general", resumen.get("total_general", 0)))
            vista.insert("", "end", values=("Resumen general", "Sin seguro", resumen.get("cantidad_sin_seguro", 0)))
            if resumen.get("cantidad_urgencias", 0):
                vista.insert("", "end", values=("Resumen general", "Urgencias", resumen.get("cantidad_urgencias", 0)))

            for seguro, cantidad in resumen.get("por_seguro", []):
                vista.insert("", "end", values=("Por ARS", seguro, cantidad))

            for especialidad, cantidad in resumen.get("por_especialidad", []):
                vista.insert("", "end", values=("Por especialidad", especialidad, cantidad))

        def generar():
            try:
                inicio, fin, etiqueta = _rango_periodo()
                estado_var.set("Cargando datos del reporte…")
                self.set_status("Generando reporte en segundo plano…", "process")
                win.update_idletasks()

                def _trabajo():
                    registros = self.db.obtener_atenciones_para_reporte(inicio, fin)
                    turno_resumen, representante = self.db.obtener_metadatos_reporte(
                        registros
                    )

                    resumen = construir_resumen_desde_registros(
                        registros,
                        f"{inicio.strftime('%d/%m/%Y %I:%M %p')} a {fin.strftime('%d/%m/%Y %I:%M %p')}",
                        turno_resumen=turno_resumen,
                        representante=representante
                    )
                    return etiqueta, resumen

                def _ok(resultado):
                    etiqueta_ok, resumen = resultado
                    ultimo_resumen["data"] = resumen
                    ultimo_resumen["ruta"] = ""
                    _cargar_vista_resumen(resumen)
                    estado_var.set(f"{etiqueta_ok} · Reporte listo. Total emergencia: {resumen.get('total_general', 0)}")
                    self.set_status("Reporte cargado", "ok")

                def _error(e):
                    estado_var.set("Error al generar reporte.")
                    messagebox.showerror("Reporte", f"No se pudo generar el reporte:\n{str(e)}")

                self._ejecutar_en_segundo_plano("Generando reporte…", _trabajo, _ok, _error)

            except Exception as e:
                estado_var.set("Error al generar reporte.")
                messagebox.showerror("Reporte", f"No se pudo generar el reporte:\n{str(e)}")

        def guardar_pdf():
            resumen = ultimo_resumen.get("data")
            if not resumen:
                messagebox.showinfo("Reporte", "Primero presione \"Generar reporte\" y espere a que termine la carga.")
                return

            try:
                ruta = crear_pdf_reporte(resumen)
                ultimo_resumen["ruta"] = ruta
                abrir_pdf(ruta)

                if self.app_settings.get("auto_print", True) and bool(self.app_settings.get("print_auto_reporte_turno", True)):
                    copias = max(1, int(self.app_settings.get("print_copies_reporte", 2) or 2))
                    imprimir_pdf(ruta, copias=copias, mostrar_error=True)
                    self.set_status(f"Reporte generado e impreso ({copias} copia/s)", "ok")
                else:
                    self.set_status("Reporte generado", "ok")

            except Exception as e:
                messagebox.showerror("Reporte", f"No se pudo crear el PDF del reporte:\n{str(e)}")

        def _actualizar_fechas_por_periodo(*_):
            modo = periodo_var.get()
            if modo == "Diario":
                base = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or datetime.now().date()
                establecer_fecha_selector(fecha_inicio, base)
                establecer_fecha_selector(fecha_fin, base)
            elif modo == "Semanal":
                base = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or datetime.now().date()
                inicio_date = base - timedelta(days=base.weekday())
                fin_date = inicio_date + timedelta(days=6)
                establecer_fecha_selector(fecha_inicio, inicio_date)
                establecer_fecha_selector(fecha_fin, fin_date)
            elif modo == "Mensual":
                base = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or datetime.now().date()
                inicio_date = base.replace(day=1)
                if inicio_date.month == 12:
                    siguiente = date(inicio_date.year + 1, 1, 1)
                else:
                    siguiente = date(inicio_date.year, inicio_date.month + 1, 1)
                establecer_fecha_selector(fecha_inicio, inicio_date)
                establecer_fecha_selector(fecha_fin, date(siguiente.year, siguiente.month, 1) - timedelta(days=1))
            elif modo == "Anual":
                base = parse_fecha_ddmmyyyy(obtener_fecha_selector(fecha_inicio)) or datetime.now().date()
                inicio_date = date(base.year, 1, 1)
                fin_date = date(base.year, 12, 31)
                establecer_fecha_selector(fecha_inicio, inicio_date)
                establecer_fecha_selector(fecha_fin, fin_date)
            else:
                _marcar_personalizado()

        def _marcar_personalizado(*_):
            periodo_var.set("Personalizado")

        combo_periodo.bind("<<ComboboxSelected>>", _actualizar_fechas_por_periodo)
        fecha_inicio.bind("<<DateEntrySelected>>", _marcar_personalizado)
        fecha_inicio.bind("<KeyRelease>", _marcar_personalizado)
        fecha_fin.bind("<<DateEntrySelected>>", _marcar_personalizado)
        fecha_fin.bind("<KeyRelease>", _marcar_personalizado)

        tb.Button(barra, text="📊  Generar reporte", bootstyle=PRIMARY, command=generar, width=20).pack(side="left", padx=5, ipady=5)
        tb.Button(barra, text="📄  Crear / abrir PDF", bootstyle=SUCCESS, command=guardar_pdf, width=20).pack(side="left", padx=5, ipady=5)
        tb.Button(barra, text="Cerrar", bootstyle=SECONDARY, command=win.destroy, width=12).pack(side="right", padx=5, ipady=5)

        try:
            estado_var.set("Seleccione el período y presione Generar reporte.")
        except Exception:
            pass


    def _generar_reporte_del_dia_rapido(self):
        try:
            fecha_base = fecha_base_operativa_actual()
            inicio, fin = obtener_rango_operativo_desde_fecha(fecha_base)
            registros = self.db.obtener_atenciones_para_reporte(inicio, fin)
            turno_resumen, representante = self.db.obtener_metadatos_reporte(registros)

            resumen = construir_resumen_desde_registros(
                registros,
                f"{inicio.strftime('%d/%m/%Y %I:%M %p')} a {fin.strftime('%d/%m/%Y %I:%M %p')}",
                turno_resumen=turno_resumen,
                representante=representante
            )

            if resumen["total_general"] == 0:
                messagebox.showinfo("Sin registros", "No hay registros para el día operativo actual.")
                return

            ruta = crear_pdf_reporte(resumen)
            abrir_pdf(ruta)
            self.set_status("Reporte generado correctamente", "ok")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el reporte del día:\n{str(e)}")

    def _obtener_catalogo_ars(self):
        catalogo = set()
        try:
            for k, v in SEGUROS_DISPLAY.items():
                if k and k != "SIN SEGURO":
                    catalogo.add(seguro_para_mostrar(k))
                    catalogo.add(k)
                if v and v != "SIN SEGURO":
                    catalogo.add(v)
        except Exception:
            pass
        try:
            for a in self.db.listar_ars_distintas():
                if a and a != "(Todas)":
                    catalogo.add(seguro_para_mostrar(a))
                    catalogo.add(a)
        except Exception:
            pass
        canonicos = [
            "SENASA SUBSIDIADO", "SENASA CONTRIBUTIVO", "SENASA PENSIONADOS",
            "MAPFRE/PALIC", "HUMANO", "PRIMERA", "SEMMA", "RENACER",
            "RESERVAS", "ASEMAP", "UNIVERSAL", "MONUMENTAL", "ABEL GONZALEZ/SIMAG",
            "METASALUD"
        ]
        catalogo.update(canonicos)
        return sorted({c.strip() for c in catalogo if c and c.strip()})

    def _configurar_label_deteccion(self, label, estado="neutral"):
        try:
            colores = {
                "ok": "#72E39B",
                "warning": "#F6B860",
                "error": "#FF7A7A",
                "neutral": "#8FA6BF",
            }
            label.configure(foreground=colores.get(estado, colores["neutral"]))
        except Exception:
            pass

    def _on_ars_keyrelease(self, event=None):
        self._actualizar_deteccion_seguro()
        return self._actualizar_sugerencias_ars(event)

    def _actualizar_deteccion_seguro(self, event=None):
        try:
            ars_txt = (self.entry_ars.get() or "").strip()
            nss_txt = (self.entry_nss.get() or "").strip().upper()
        except Exception:
            return

        try:
            if ars_txt:
                canon_alias = _mejor_seguro_por_similitud(ars_txt)
                if not canon_alias:
                    canon_alias = normalizar_seguro(ars_txt, nss_txt)
            else:
                canon_alias = "SIN SEGURO" if nss_txt in ["SIN SEGURO", "NO", "N/S", "NS", "N\\S"] else ""

            if not ars_txt and not nss_txt:
                self.ars_detectado_var.set("Detectado como: pendiente")
                self._configurar_label_deteccion(self.ars_detectado_label, "neutral")
            else:
                display = seguro_para_mostrar(canon_alias or "SIN SEGURO")
                self.ars_detectado_var.set(f"Detectado como: {display}")

                if display == "SIN SEGURO":
                    self._configurar_label_deteccion(self.ars_detectado_label, "warning")
                else:
                    self._configurar_label_deteccion(self.ars_detectado_label, "ok")

            if not nss_txt:
                if ars_txt and canon_alias and canon_alias != "SIN SEGURO":
                    self.nss_detectado_var.set("NSS: pendiente; sin NSS se guardará como SIN SEGURO")
                    self._configurar_label_deteccion(self.nss_detectado_label, "warning")
                else:
                    self.nss_detectado_var.set("NSS: pendiente")
                    self._configurar_label_deteccion(self.nss_detectado_label, "neutral")
            elif nss_txt in ["SIN SEGURO", "NO", "N/S", "NS", "N\\S"]:
                self.nss_detectado_var.set("NSS: marcado como SIN SEGURO")
                self._configurar_label_deteccion(self.nss_detectado_label, "warning")
            elif not nss_txt.isdigit():
                self.nss_detectado_var.set("NSS: inválido; debe ser numérico o SIN SEGURO")
                self._configurar_label_deteccion(self.nss_detectado_label, "error")
            elif is_all_zeros(nss_txt):
                self.nss_detectado_var.set("NSS: inválido; no puede ser todo ceros")
                self._configurar_label_deteccion(self.nss_detectado_label, "error")
            else:
                self.nss_detectado_var.set("NSS: válido")
                self._configurar_label_deteccion(self.nss_detectado_label, "ok")

        except Exception:
            pass

    def _actualizar_sugerencias_ars(self, event=None):
        texto = (self.entry_ars.get() or "").strip()
        if not texto:
            self._ocultar_sugerencias_ars()
            return
        q = _limpiar_texto_seguro(texto)
        resultados = []
        for opcion in self._ars_catalogo:
            limpio = _limpiar_texto_seguro(opcion)
            if not limpio:
                continue
            score = SequenceMatcher(None, q, limpio).ratio()
            if q in limpio or limpio.startswith(q) or score >= 0.45:
                resultados.append((score + (0.4 if q in limpio else 0), opcion))
        resultados = [op for _, op in sorted(resultados, reverse=True)[:8]]
        self.ars_suggestions.delete(0, tk.END)
        if not resultados:
            self._ocultar_sugerencias_ars()
            return
        for item in resultados:
            self.ars_suggestions.insert(tk.END, item)
        self.ars_suggestions.grid()

    def _seleccionar_sugerencia_ars(self, event=None):
        try:
            idx = self.ars_suggestions.curselection()
            if not idx:
                return
            val = self.ars_suggestions.get(idx[0])
            self.entry_ars.delete(0, tk.END)
            self.entry_ars.insert(0, val)
            self._actualizar_deteccion_seguro()
            self._ocultar_sugerencias_ars()
            self.entry_nss.focus_set()
        except Exception:
            pass

    def _focus_sugerencias_ars(self, event=None):
        try:
            if self.ars_suggestions.winfo_ismapped() and self.ars_suggestions.size() > 0:
                self.ars_suggestions.focus_set()
                self.ars_suggestions.selection_clear(0, tk.END)
                self.ars_suggestions.selection_set(0)
                self.ars_suggestions.activate(0)
                return "break"
        except Exception:
            pass

    def _ocultar_sugerencias_ars(self):
        try:
            self.ars_suggestions.grid_remove()
        except Exception:
            pass

    def _capture_initial_styles(self):
        for e in self.all_entries:
            try:
                self._initial_styles[e] = {
                    "bootstyle": e.cget("bootstyle") if "bootstyle" in e.keys() else "",
                    "style": e.cget("style") if "style" in e.keys() else "",
                }
            except Exception:
                self._initial_styles[e] = {"bootstyle": "", "style": ""}

    def _restore_widget_style(self, widget):
        try:
            if widget in self._pending_restores and self._pending_restores[widget]:
                try:
                    self.root.after_cancel(self._pending_restores[widget])
                except Exception:
                    pass
                self._pending_restores[widget] = None
            init = self._initial_styles.get(widget, {"bootstyle": "", "style": ""})
            try:
                widget.configure(bootstyle=init.get("bootstyle", ""))
            except Exception:
                pass
            try:
                widget.configure(style=init.get("style", ""))
            except Exception:
                pass
        except Exception:
            pass

    def _restore_all_styles(self):
        for w, after_id in list(self._pending_restores.items()):
            if after_id:
                try:
                    self.root.after_cancel(after_id)
                except Exception:
                    pass
                self._pending_restores[w] = None
        for e in self.all_entries:
            self._restore_widget_style(e)
        self.root.update_idletasks()

    def _ventana_activa(self, ventana):
        try:
            return ventana is not None and ventana.winfo_exists()
        except Exception:
            return False

    def _enfocar_ventana(self, ventana):
        try:
            ventana.deiconify()
            ventana.lift()
            ventana.focus_force()
        except Exception:
            pass

    def _crear_toplevel_estable(self, titulo, geometry, attr_name):
        ventana_existente = getattr(self, attr_name, None)
        if self._ventana_activa(ventana_existente):
            self._enfocar_ventana(ventana_existente)
            return None

        win = Toplevel(self.root)
        win.title(titulo)
        win.geometry(geometry)
        win.configure(bg="#07111f")
        try:
            base_geo = geometry.split("+")[0]
            mw, mh = [int(x) for x in base_geo.lower().split("x")[:2]]
            win.minsize(min(max(mw, 760), 1180), min(max(mh, 460), 800))
        except Exception:
            try:
                win.minsize(900, 480)
            except Exception:
                pass
        win.transient(self.root)
        win.bind("<Escape>", lambda e: _cerrar())

        def _cerrar():
            try:
                setattr(self, attr_name, None)
                win.destroy()
            except Exception:
                setattr(self, attr_name, None)

        win.protocol("WM_DELETE_WINDOW", _cerrar)
        setattr(self, attr_name, win)
        try:
            win.lift()
            win.focus_set()
            win.after(10, lambda w=win: (w.lift(), w.focus_set()))
        except Exception:
            pass
        return win

    def _crear_header_ventana(self, parent, titulo, subtitulo="", icono="▣"):
        header = tb.Frame(parent, padding=(12, 10), style="Card.TFrame")
        header.pack(fill="x", pady=(0, 12))
        header.columnconfigure(1, weight=1)

        tb.Label(
            header,
            text=icono,
            font=("Arial", 22, "bold"),
            foreground="#5CB6FF",
            background="#0E1B2B"
        ).grid(row=0, column=0, rowspan=2, sticky="nw", padx=(0, 12))

        tb.Label(
            header,
            text=titulo,
            font=("Arial", 16, "bold"),
            foreground="#F5F9FF",
            background="#0E1B2B",
            wraplength=760,
            justify="left"
        ).grid(row=0, column=1, sticky="w")

        if subtitulo:
            tb.Label(
                header,
                text=subtitulo,
                style="Muted.TLabel",
                background="#0E1B2B",
                wraplength=760,
                justify="left"
            ).grid(row=1, column=1, sticky="w", pady=(3, 0))
        return header

    def _crear_card(self, parent, padding=12):
        return tb.Frame(parent, padding=padding, style="Card.TFrame")

    def _mostrar_dialogo_modal_unico(self, titulo, mensaje):
        if self._ventana_activa(self.dialogo_unico_win):
            self._enfocar_ventana(self.dialogo_unico_win)
            return

        win = Toplevel(self.root)
        self.dialogo_unico_win = win
        win.title(titulo)
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        win.bind("<Escape>", lambda e: self._cerrar_dialogo_modal_unico())

        frame = tb.Frame(win, padding=16)
        frame.pack(fill="both", expand=True)

        tb.Label(frame, text=mensaje, wraplength=360, justify="left").pack(anchor="w", pady=(0, 14))
        tb.Button(frame, text="Cerrar", bootstyle=PRIMARY, command=lambda: self._cerrar_dialogo_modal_unico()).pack()

        def on_close():
            self._cerrar_dialogo_modal_unico()

        win.protocol("WM_DELETE_WINDOW", on_close)
        win.update_idletasks()

        w = win.winfo_width()
        h = win.winfo_height()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2) - (w // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2) - (h // 2)
        try:
            win.geometry(f"+{max(x, 0)}+{max(y, 0)}")
        except Exception:
            pass

        try:
            win.focus_force()
        except Exception:
            pass

    def _cerrar_dialogo_modal_unico(self):
        try:
            if self.dialogo_unico_win and self.dialogo_unico_win.winfo_exists():
                self.dialogo_unico_win.grab_release()
                self.dialogo_unico_win.destroy()
        except Exception:
            pass
        finally:
            self.dialogo_unico_win = None

    def _register_temp(self, path: str):
        if path and os.path.exists(path):
            self._temp_files.add(path)

    def _cleanup_temp_files(self):
        for p in list(self._temp_files):
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass
            finally:
                self._temp_files.discard(p)

    def on_close(self):
        self._cleanup_temp_files()
        try:
            self.root.destroy()
        except Exception:
            pass

    def _highlight_error(self, widget, mensaje: str):
        if widget not in self._initial_styles:
            try:
                self._initial_styles[widget] = {
                    "bootstyle": widget.cget("bootstyle") if "bootstyle" in widget.keys() else "",
                    "style": widget.cget("style") if "style" in widget.keys() else "",
                }
            except Exception:
                self._initial_styles[widget] = {"bootstyle": "", "style": ""}

        try:
            widget.configure(bootstyle=DANGER)
        except Exception:
            try:
                widget.configure(style="danger.TEntry")
            except Exception:
                pass

        widget.focus_set()
        self.set_status(f"Corrija el campo marcado: {mensaje}", "error")
        self._mostrar_notificacion(mensaje, autohide_ms=9000, tipo="error")

        try:
            after_id = self.root.after(2000, lambda w=widget: self._restore_widget_style(w))
            self._pending_restores[widget] = after_id
        except Exception:
            pass

    def _ocultar_notificacion(self):
        if self._notif_after_id:
            try:
                self.root.after_cancel(self._notif_after_id)
            except Exception:
                pass
            self._notif_after_id = None
        self.notif_frame.pack_forget()
        self.notif_label.config(text="")
        self.btn_deshacer.config(command=lambda: None)

    def _push_undo_action(self, descripcion: str, callback):
        if not callable(callback):
            return
        try:
            self._undo_stack.append({"descripcion": descripcion or "última acción", "callback": callback})
            if len(self._undo_stack) > getattr(self, "_undo_limit", 20):
                self._undo_stack = self._undo_stack[-self._undo_limit:]
        except Exception:
            self._undo_stack = [{"descripcion": descripcion or "última acción", "callback": callback}]

    def _on_field_focus_in(self, event):
        """
        FASE 9B: Captura el valor del campo al recibir foco.
        """
        try:
            w = event.widget
            if isinstance(w, (tk.Entry, tb.Entry)):
                w._undo_prev_value = w.get()
                w._undo_has_prev = True
        except Exception:
            pass

    def _on_field_focus_out(self, event):
        """
        FASE 9B: Si el valor del campo cambió, registra undo de campo.
        """
        try:
            w = event.widget
            if not isinstance(w, (tk.Entry, tb.Entry)):
                return
            if not getattr(w, "_undo_has_prev", False):
                return
            prev = getattr(w, "_undo_prev_value", None)
            curr = w.get()
            if prev is not None and prev != curr:
                def _undo_campo(ww=w, valor_anterior=prev):
                    try:
                        ww.delete(0, tk.END)
                        ww.insert(0, valor_anterior)
                        ww.focus_set()
                    except Exception:
                        pass
                    return "break"

                nombre_campo = "campo"
                try:
                    mapping = {
                        "entry_nombre": "Nombre",
                        "entry_edad": "Edad",
                        "entry_cedula": "Cédula",
                        "entry_telefono": "Teléfono",
                        "entry_direccion": "Dirección",
                        "entry_nacionalidad": "Nacionalidad",
                        "entry_ars": "ARS",
                        "entry_nss": "NSS",
                    }
                    nombre_campo = mapping.get(str(w), "campo")
                except Exception:
                    pass

                self._push_undo_action(f"cambio en {nombre_campo}", _undo_campo)
            w._undo_has_prev = False
        except Exception:
            pass

    def _undo_last_action(self, event=None):
        """
        FASE 9: Deshacer global con Ctrl+Z.
        - Restaura último campo modificado / texto eliminado / formulario limpiado.
        - Devuelve el foco al widget original cuando aplica.
        """
        try:
            if not getattr(self, "_undo_stack", None):
                self.set_status("No hay acciones para deshacer", "warning")
                try:
                    self.root.bell()
                except Exception:
                    pass
                return "break"

            accion = self._undo_stack.pop()
            callback = accion.get("callback")
            descripcion = accion.get("descripcion", "última acción")

            try:
                foco_antes = self.root.focus_get()
            except Exception:
                foco_antes = None

            if callable(callback):
                try:
                    resultado = callback()
                except Exception as e:
                    self.set_status(f"Error al deshacer: {e}", "error")
                    return "break"

                try:
                    foco_actual = self.root.focus_get()
                except Exception:
                    foco_actual = None

                if foco_actual in (None, "") and foco_antes is not None:
                    try:
                        if foco_antes.winfo_exists():
                            foco_antes.focus_set()
                    except Exception:
                        pass

                try:
                    self._invalidar_caches_datos()
                except Exception:
                    pass
                self.set_status(f"Deshecho: {descripcion}", "ok")
                self._mostrar_notificacion(
                    f"Deshecho: {descripcion}",
                    on_undo=None,
                    autohide_ms=3000
                )
            return "break"
        except Exception as e:
            self.set_status(f"No se pudo deshacer: {str(e)}", "error")
            messagebox.showerror("Deshacer", f"No se pudo deshacer la última acción:\n{str(e)}")
            return "break"


    def _mostrar_notificacion(self, texto: str, on_undo=None, autohide_ms=7000, tipo="info"):
        colores = {
            "info": self._paleta_visual_actual()["text"],
            "ok": "#72E39B",
            "warning": "#FFD166",
            "error": "#FF8A80",
        }
        self.notif_label.config(text=texto, foreground=colores.get(tipo, colores["info"]))
        if on_undo:
            self.btn_deshacer.config(command=lambda: (on_undo(), self._ocultar_notificacion()))
            self.btn_deshacer.pack(side="right", padx=6)
        else:
            self.btn_deshacer.pack_forget()
        self.notif_frame.pack(fill="x", side="bottom")
        if self._notif_after_id:
            try:
                self.root.after_cancel(self._notif_after_id)
            except Exception:
                pass
        self._notif_after_id = self.root.after(autohide_ms, self._ocultar_notificacion)

    def validar_numerico(self, value, tipo):
        if value == "":
            return True
        if tipo == 'cedula':
            return value.isdigit() and len(value) <= 11
        elif tipo == 'telefono':
            return value.isdigit() and len(value) <= 10
        return True

    def limitar_caracteres(self, entry, max_len):
        current = entry.get()
        if len(current) > max_len:
            entry.delete(max_len, tk.END)

    def actualizar_embarazada(self):
        try:
            if self.var_sexo.get() == "Femenino":
                if not self.check_embarazada.winfo_ismapped():
                    self.check_embarazada.pack(side="left", padx=(8, 0))
            else:
                if self.check_embarazada.winfo_ismapped():
                    self.check_embarazada.pack_forget()
                self.var_embarazada.set(False)
        except Exception:
            pass

    def determinar_hoja(self):
        if self.var_sexo.get() == "Femenino" and self.var_embarazada.get():
            return "GINECOLOGIA"

        try:
            edad = int(self.entry_edad.get())
            unidad = self.unidad_edad.get()[0].upper()
        except Exception:
            return None

        if unidad == "D":
            meses = edad / 30
        elif unidad == "M":
            meses = edad
        else:
            meses = edad * 12

        if meses <= 180:
            return "PEDIATRIA"
        return "GENERAL"

    def _registro_afecta_excel_turno(self, atencion):
        if not atencion:
            return False
        if (atencion.get("tipo_atencion") or "EMERGENCIA").strip().upper() == "URGENCIA":
            return False
        return self._registro_esta_en_turno_actual(atencion)

    def _cambio_requiere_reconstruir_excel(self, antes: dict, despues: dict) -> bool:
        """
        FASE 8: Reconstruir Excel SOLO si cambia un campo crítico.
        Campos críticos: nombre, hoja, ARS, tipo_atencion, fecha.
        Campos NO críticos: telefono, direccion, sexo, nacionalidad, cedula, edad, nss, hora.
        URGENCIA nunca reconstruye.
        """
        if not antes and not despues:
            return False

        antes = antes or {}
        despues = despues or {}

        tipo_despues = str(
            despues.get("TipoAtencion", despues.get("tipo_atencion",
            antes.get("tipo_atencion", "EMERGENCIA"))) or "EMERGENCIA"
        ).strip().upper()

        if tipo_despues == "URGENCIA":
            return False

        campos_clave = [
            ("nombre", "Nombre"),
            ("hoja",   "Hoja"),
            ("ars",    "Aseguradora (ARS)"),
            ("tipo_atencion", "TipoAtencion"),
            ("fecha",  "Fecha"),
        ]

        for k_antes, k_despues in campos_clave:
            a = str(antes.get(k_antes, "") or "").strip().upper()
            d = str(despues.get(k_despues, despues.get(k_antes, "")) or "").strip().upper()
            if k_antes == "ars":
                a = normalizar_seguro(a, antes.get("nss", ""))
                d = normalizar_seguro(d, despues.get("NSS", despues.get("nss", "")))
            if a != d:
                return True

        if self._registro_afecta_excel_turno(antes):
            return True

        try:
            tmp = dict(antes)
            tmp.update({
                "fecha": despues.get("Fecha", despues.get("fecha", antes.get("fecha", ""))),
                "hora":  despues.get("Hora",  despues.get("hora",  antes.get("hora",  ""))),
                "tipo_atencion": tipo_despues,
            })
            if self._registro_afecta_excel_turno(tmp):
                return True
        except Exception:
            pass

        return False

    def _reconstruir_excel_si_necesario(self, razon="", antes=None, despues=None, forzar=False):
        try:
            turno_cfg = cargar_turno_config()
            if not turno_cfg:
                return False

            if not forzar:
                if antes is not None or despues is not None:
                    if not self._cambio_requiere_reconstruir_excel(antes or {}, despues or {}):
                        return False

            reconstruir_excel_turno(self.db, turno_cfg)
            self._refrescar_resumen_en_vivo()
            return True
        except PermissionError:
            self.set_status("Excel abierto. Cierre el listado para actualizarlo.", "warning")
            raise
        except Exception as e:
            self.set_status(f"Aviso al actualizar Excel: {e}", "warning")
            return False

    def _registro_esta_en_turno_actual(self, atencion):
        turno_cfg = cargar_turno_config()
        if not turno_cfg or not atencion:
            return False
        contexto = self.db.buscar_contexto_turno_existente(turno_cfg)
        if not contexto:
            return False
        turno_id = atencion.get("turno_id")
        if turno_id is None and atencion.get("id"):
            actual = self.db.obtener_atencion_por_id(int(atencion["id"]))
            turno_id = actual.get("turno_id") if actual else None
        return turno_id is not None and int(turno_id) == int(contexto["turno_id"])

    def _generar_y_abrir_reporte_turno(self, turno_cfg, fin_corte=None):
        if not turno_cfg:
            return ""

        resumen = construir_resumen_turno(self.db, turno_cfg, fin_override=fin_corte or datetime.now())
        if not resumen:
            self.set_status("No se pudo generar el reporte: no se encontraron datos del turno.", "warning")
            return ""

        ruta = crear_pdf_reporte(resumen)

        try:
            guardar_copia_reporte_turno(ruta, turno_cfg)
        except Exception as e:
            messagebox.showwarning(
                "Archivo diario",
                f"El reporte se generó, pero no se pudo guardar la copia clasificada:\n{str(e)}"
            )

        abrir_pdf(ruta)
        if self.app_settings.get("auto_print", True) and bool(self.app_settings.get("print_auto_reporte_turno", True)):
            copias_reporte = max(1, int(self.app_settings.get("print_copies_reporte", 2) or 2))
            imprimir_pdf(ruta, copias=copias_reporte, mostrar_error=True)
            self.set_status(f"Reporte impreso {copias_reporte} vez/veces", "ok")
        else:
            self.set_status("Reporte generado (impresión automática desactivada)", "ok")
        return ruta

    def auto_completar(self, event=None):
        if self._suspend_autocomplete:
            return
        cedula = (self.entry_cedula.get() or "").strip().replace("-", "")
        if not cedula or is_all_zeros(cedula):
            return
        if len(cedula) == 11 and cedula.isdigit():
            try:
                self._suspend_autocomplete = True
                paciente = self.db.buscar_paciente(cedula)
                if paciente:
                    self.entry_nombre.delete(0, tk.END)
                    self.entry_nombre.insert(0, paciente['nombre'] or "")
                    self.entry_telefono.delete(0, tk.END)
                    self.entry_telefono.insert(0, paciente['telefono'] or "")
                    self.entry_direccion.delete(0, tk.END)
                    self.entry_direccion.insert(0, paciente['direccion'] or "")
                    self.entry_nacionalidad.delete(0, tk.END)
                    self.entry_nacionalidad.insert(0, paciente['nacionalidad'] or "")
                    self.entry_ars.delete(0, tk.END)
                    self.entry_ars.insert(0, seguro_para_mostrar(paciente['ars'] or ""))
                    self.entry_nss.delete(0, tk.END)
                    self.entry_nss.insert(0, (paciente['nss'] or "").upper())
            except Exception as e:
                messagebox.showerror("Error", f"Error al autocompletar: {str(e)}")
            finally:
                self._suspend_autocomplete = False

    def auto_completar_por_nss(self, event=None):
        if self._suspend_autocomplete:
            return
        nss = (self.entry_nss.get() or "").strip().upper()

        if not is_valid_nss_key(nss):
            return
        try:
            self._suspend_autocomplete = True
            paciente = self.db.buscar_por_nss(nss)
            if paciente:
                self.entry_cedula.delete(0, tk.END)
                self.entry_cedula.insert(0, paciente['cedula'] or "")
                self.entry_nombre.delete(0, tk.END)
                self.entry_nombre.insert(0, paciente['nombre'] or "")
                self.entry_telefono.delete(0, tk.END)
                self.entry_telefono.insert(0, paciente['telefono'] or "")
                self.entry_direccion.delete(0, tk.END)
                self.entry_direccion.insert(0, paciente['direccion'] or "")
                self.entry_nacionalidad.delete(0, tk.END)
                self.entry_nacionalidad.insert(0, paciente['nacionalidad'] or "")
                self.entry_ars.delete(0, tk.END)
                self.entry_ars.insert(0, seguro_para_mostrar(paciente['ars'] or ""))
        except Exception as e:
            messagebox.showerror("Error", f"Error al autocompletar por NSS: {str(e)}")
        finally:
            self._suspend_autocomplete = False
            try:
                self._actualizar_deteccion_seguro()
            except Exception:
                pass

    def _try_autocomplete_cedula(self):
        cedula = (self.entry_cedula.get() or "").strip().replace("-", "")
        if len(cedula) == 11 and cedula.isdigit() and not is_all_zeros(cedula):
            self.auto_completar()

    def _try_autocomplete_nss(self):
        nss = (self.entry_nss.get() or "").strip().upper()
        if is_valid_nss_key(nss):
            self.auto_completar_por_nss()

    def _guardar_formulario_actual(self):
        """
        FASE 9/10: Guarda estado completo del formulario + widget con foco.
        """
        try:
            try:
                widget_con_foco = self.root.focus_get()
            except Exception:
                widget_con_foco = None

            self._ultimo_formulario = {
                "nombre":      self.entry_nombre.get(),
                "edad":        self.entry_edad.get(),
                "edad_unit":   self.unidad_edad.get(),
                "cedula":      self.entry_cedula.get(),
                "telefono":    self.entry_telefono.get(),
                "direccion":   self.entry_direccion.get(),
                "nacionalidad":self.entry_nacionalidad.get(),
                "ars":         self.entry_ars.get(),
                "nss":         self.entry_nss.get(),
                "sexo":        self.var_sexo.get(),
                "embarazada":  self.var_embarazada.get(),
                "urgencia":    self.var_urgencia.get(),
                "hoja":        self.determinar_hoja() or "GENERAL",
                "_widget_foco": widget_con_foco,
            }
        except Exception as e:
            self.set_status(f"Aviso: no se pudo guardar formulario para Ctrl+Z: {e}", "warning")

    def _restaurar_formulario(self):
        """
        FASE 9/10: Restaura el formulario + devuelve el foco al widget original.
        """
        try:
            form = getattr(self, "_ultimo_formulario", None)
            if not form:
                self.set_status("No hay formulario anterior para restaurar", "warning")
                return

            self.entry_nombre.delete(0, tk.END);       self.entry_nombre.insert(0, form.get("nombre", ""))
            self.entry_edad.delete(0, tk.END);         self.entry_edad.insert(0, form.get("edad", ""))
            self.unidad_edad.set(form.get("edad_unit", "Años"))
            self.entry_cedula.delete(0, tk.END);       self.entry_cedula.insert(0, form.get("cedula", ""))
            self.entry_telefono.delete(0, tk.END);     self.entry_telefono.insert(0, form.get("telefono", ""))
            self.entry_direccion.delete(0, tk.END);    self.entry_direccion.insert(0, form.get("direccion", ""))
            self.entry_nacionalidad.delete(0, tk.END); self.entry_nacionalidad.insert(0, form.get("nacionalidad", ""))
            self.entry_ars.delete(0, tk.END);          self.entry_ars.insert(0, form.get("ars", ""))
            self.entry_nss.delete(0, tk.END);          self.entry_nss.insert(0, form.get("nss", ""))
            self.var_sexo.set(form.get("sexo") or "Femenino")
            self.var_embarazada.set(form.get("embarazada", False))
            self.var_urgencia.set(form.get("urgencia", False))

            self._restore_all_styles()
            try:
                self._actualizar_deteccion_seguro()
            except Exception:
                pass
            try:
                self.actualizar_embarazada()
            except Exception:
                pass

            try:
                widget_foco = form.get("_widget_foco")
                if widget_foco is not None and widget_foco.winfo_exists():
                    widget_foco.focus_set()
                else:
                    self.entry_nombre.focus_set()
            except Exception:
                self.entry_nombre.focus_set()

            self.set_status("✓  Formulario restaurado", "ok")
        except Exception as e:
            self.set_status(f"No se pudo restaurar formulario: {e}", "error")

    def limpiar_campos(self):
        """
        FASE 9/10: Antes de limpiar, guarda el estado y registra undo.
        """
        estado_previo = (
            (self.entry_nombre.get() or "").strip()
            + (self.entry_edad.get() or "").strip()
            + (self.entry_cedula.get() or "").strip()
            + (self.entry_telefono.get() or "").strip()
            + (self.entry_nss.get() or "").strip()
        )

        if estado_previo:
            self._guardar_formulario_actual()

            def _undo():
                self._restaurar_formulario()
            self._push_undo_action("restaurar formulario limpiado", _undo)

        self.entry_nombre.delete(0, tk.END)
        self.entry_edad.delete(0, tk.END)
        self.entry_cedula.delete(0, tk.END)
        self.entry_telefono.delete(0, tk.END)
        self.entry_direccion.delete(0, tk.END)
        self.entry_nacionalidad.delete(0, tk.END)
        self.entry_ars.delete(0, tk.END)
        self.entry_nss.delete(0, tk.END)
        self.var_sexo.set("Femenino")
        self.var_embarazada.set(False)
        try:
            self.var_urgencia.set(False)
        except Exception:
            pass
        self.unidad_edad.set("Años")
        try:
            self.ars_detectado_var.set("Detectado como: pendiente")
            self.nss_detectado_var.set("NSS: pendiente")
            self._configurar_label_deteccion(self.ars_detectado_label, "neutral")
            self._configurar_label_deteccion(self.nss_detectado_label, "neutral")
        except Exception:
            pass
        self.entry_nombre.focus_set()

    def _validar_campos_o_alertar(self):
        nombre = (self.entry_nombre.get() or "").strip()
        telefono = (self.entry_telefono.get() or "").strip()
        cedula = (self.entry_cedula.get() or "").strip()
        nss = (self.entry_nss.get() or "").strip().upper()
        direccion = (self.entry_direccion.get() or "").strip()
        nacionalidad = (self.entry_nacionalidad.get() or "").strip()
        ars = (self.entry_ars.get() or "").strip()
        sexo = self.var_sexo.get()

        if not nombre:
            self._highlight_error(self.entry_nombre, "El nombre es obligatorio.")
            return None

        if sexo not in ("Masculino", "Femenino"):
            self._highlight_error(
                self.lbl_sexo_m,
                "Seleccione explícitamente el sexo del paciente.",
            )
            return None

        if telefono:
            if not (telefono.isdigit() and len(telefono) == 10):
                self._highlight_error(self.entry_telefono, "El teléfono debe tener exactamente 10 dígitos numéricos.")
                return None
        elif not bool(self.app_settings.get("validation_allow_missing_phone", False)):
            self._highlight_error(self.entry_telefono, "El teléfono es obligatorio según las preferencias actuales.")
            return None

        edad_txt = (self.entry_edad.get() or "").strip()
        if edad_txt == "" or not edad_txt.isdigit():
            self._highlight_error(self.entry_edad, "La edad es obligatoria y debe ser un número entero.")
            return None
        edad = int(edad_txt)
        if edad < 0 or edad > 130:
            self._highlight_error(self.entry_edad, "La edad debe estar entre 0 y 130.")
            return None

        unidad = self.unidad_edad.get()
        if unidad not in ("Días", "Meses", "Años"):
            self._highlight_error(self.combo_unidad, "Selecciona una unidad de edad válida (Días/Meses/Años).")
            return None

        if cedula:
            if not (cedula.isdigit() and len(cedula) == 11):
                self._highlight_error(self.entry_cedula, "La cédula debe tener 11 dígitos.")
                return None
        elif not bool(self.app_settings.get("validation_allow_missing_cedula", True)):
            self._highlight_error(self.entry_cedula, "La cédula es obligatoria según las preferencias actuales.")
            return None

        if nss:
            if not nss.isdigit() and nss not in ["N/S", "N\\S", "NS", "NO", "SIN SEGURO"]:
                self._highlight_error(
                    self.entry_nss,
                    "El NSS debe ser numérico. Si el paciente no tiene seguro, deje el campo vacío o escriba SIN SEGURO."
                )
                return None
            if is_all_zeros(nss):
                self._highlight_error(self.entry_nss, "El NSS no puede ser todo ceros. Si no tiene seguro, deje el campo vacío o escriba SIN SEGURO.")
                return None

        if bool(self.app_settings.get("validation_block_short_ars", True)) and ars_es_corta_invalida(ars):
            self._highlight_error(
                self.entry_ars,
                "La ARS escrita es demasiado corta o no reconocida. Escriba al menos 4 caracteres o use una referencia válida como SUB, HUMANO, MAPFRE o SIN SEGURO."
            )
            return None

        invertido, msg_invertido = _detectar_campos_invertidos(
            self.entry_nss.get(), self.entry_ars.get()
        )
        if invertido:
            self._highlight_error(self.entry_nss, msg_invertido)
            return None

        hoja = self.determinar_hoja()
        if hoja is None:
            self._highlight_error(self.entry_edad, "Edad inválida para determinar la hoja.")
            return None

        avisos = []
        if bool(self.app_settings.get("validation_warn_nss_incomplete", True)):
            nss_digits = re.sub(r"\D", "", nss)
            if nss_digits and len(nss_digits) < 8:
                avisos.append("El NSS parece incompleto o demasiado corto.")

        ars_canon_tmp = normalizar_seguro(ars, nss)
        if bool(self.app_settings.get("validation_warn_ars_sin_seguro", True)) and ars_canon_tmp == "SIN SEGURO":
            avisos.append("La ARS fue detectada como SIN SEGURO.")

        if bool(self.app_settings.get("rn_warn", True)) and nombre_tiene_prefijo_rn(nombre):
            avisos.append("El nombre tiene prefijo RN-. Según preferencias, puede guardarse sin RN- en la base de datos.")

        if avisos:
            if not messagebox.askyesno("Revisión de datos", "Revise estos datos antes de continuar:\n\n- " + "\n- ".join(avisos) + "\n\n¿Desea continuar?"):
                return None

        ahora = datetime.now()
        hora_12 = ahora.strftime("%I:%M")
        am_pm = "AM" if ahora.hour < 12 else "PM"

        datos = {
            "Fecha": ahora.strftime("%d/%m/%Y"),
            "Hora": f"{hora_12} {am_pm}",
            "Nombre": nombre,
            "Sexo": sexo,
            "Edad_num": edad,
            "Unidad": unidad,
            "Cédula": cedula,
            "Teléfono": telefono,
            "Dirección": direccion,
            "Nacionalidad": nacionalidad,
            "Aseguradora (ARS)": ars,
            "NSS": nss,
            "TipoAtencion": "URGENCIA" if getattr(self, "var_urgencia", tk.BooleanVar(value=False)).get() else "EMERGENCIA",
        }
        return datos, hoja

    def _buscar_duplicado_turno_actual(self, datos):
        nss = (datos.get("NSS") or "").strip().upper()
        ced = (datos.get("Cédula") or "").strip()
        nombre = (datos.get("Nombre") or "").strip()
        telefono = (datos.get("Teléfono") or "").strip()
        if (
            not is_valid_nss_key(nss)
            and not is_valid_cedula_key(ced)
            and not (nombre and len(re.sub(r"\D", "", telefono)) == 10)
        ):
            return None

        turno_cfg = cargar_turno_config()
        if not turno_cfg:
            return None

        inicio_turno, fin_turno = obtener_rango_turno_efectivo(turno_cfg)
        contexto = self.db.obtener_contexto_turno(turno_cfg)
        return self.db.buscar_atencion_en_turno(
            nss,
            ced,
            inicio_turno,
            fin_turno,
            turno_id=contexto["turno_id"],
            dia_operativo_id=contexto["dia_operativo_id"],
            nombre=nombre,
            telefono=telefono,
        )

    def _dialogo_atencion_existente(self, atencion):
        resultado = {"accion": "cancelar"}
        win = Toplevel(self.root)
        win.title("Atención ya registrada")
        win.geometry("720x360")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        cont = tb.Frame(win, padding=20, style="Root.TFrame")
        cont.pack(fill="both", expand=True)
        tb.Label(
            cont,
            text="Este paciente ya tiene una hoja en el turno actual",
            font=("Arial", 15, "bold"),
            foreground="#FFFFFF",
        ).pack(anchor="w", pady=(0, 12))
        detalle = (
            f"Atención #{atencion.get('id')}\n"
            f"Paciente: {(atencion.get('nombre') or '').upper()}\n"
            f"Registrada: {atencion.get('fecha', '')} {atencion.get('hora', '')}\n"
            f"Especialidad: {atencion.get('hoja', '')}"
        )
        tb.Label(cont, text=detalle, justify="left", wraplength=670).pack(anchor="w", pady=(0, 18))
        tb.Label(
            cont,
            text="No se creará otra atención ni se reemplazará el historial.",
            bootstyle=WARNING,
        ).pack(anchor="w", pady=(0, 18))

        botones = tb.Frame(cont, style="Root.TFrame")
        botones.pack(fill="x", side="bottom")

        def elegir(accion):
            resultado["accion"] = accion
            win.destroy()

        tb.Button(botones, text="Reimprimir", bootstyle=SUCCESS, command=lambda: elegir("reimprimir")).pack(side="left", padx=(0, 8))
        tb.Button(botones, text="Abrir PDF", bootstyle=INFO, command=lambda: elegir("abrir")).pack(side="left", padx=(0, 8))
        tb.Button(botones, text="Editar", bootstyle=SECONDARY, command=lambda: elegir("editar")).pack(side="left", padx=(0, 8))
        tb.Button(
            botones,
            text="Registrar reingreso",
            bootstyle=WARNING,
            command=lambda: elegir("reingreso"),
        ).pack(side="left", padx=(0, 8))
        tb.Button(botones, text="Cancelar", command=lambda: elegir("cancelar")).pack(side="right")
        win.protocol("WM_DELETE_WINDOW", lambda: elegir("cancelar"))
        win.wait_window()
        return resultado["accion"]

    def _atender_duplicado_turno(self, atencion):
        accion = self._dialogo_atencion_existente(atencion)
        if accion == "editar":
            self._abrir_editor_atencion(int(atencion["id"]))
            return None
        if accion == "reingreso":
            actor = self._solicitar_autorizacion_admin(
                "AUTORIZAR_REINGRESO", parent=self.root, force=True
            )
            if not actor:
                return None
            motivo = simpledialog.askstring(
                "Motivo del reingreso",
                "Explique por qué corresponde crear una segunda hoja en este día operativo:",
                parent=self.root,
            )
            motivo = (motivo or "").strip()
            if len(motivo) < 8:
                messagebox.showwarning(
                    "Reingreso",
                    "El motivo debe contener al menos 8 caracteres.",
                    parent=self.root,
                )
                return None
            return {
                "EsReingreso": True,
                "AtencionOrigenId": int(atencion["id"]),
                "MotivoReingreso": motivo,
                "AutorizadoPor": actor,
            }
        if accion not in {"abrir", "reimprimir"}:
            return None

        documento = self.db.obtener_documento_atencion(int(atencion["id"]))
        ruta = documento.get("ruta") if documento else None
        if not ruta:
            ruta_temporal = crear_pdf_temporal(
                atencion.get("hoja", "GENERAL"),
                self._snapshot_a_datos(atencion),
                mostrar_error=False,
            )
            if not ruta_temporal:
                messagebox.showerror(
                    "Hoja existente",
                    "No se pudo recuperar ni reconstruir el PDF de la atención existente.",
                )
                return None
            try:
                ruta = archivar_pdf_atencion(ruta_temporal, int(atencion["id"]))
                self.db.registrar_documento(
                    int(atencion["id"]), "HOJA_EMERGENCIA", ruta, atencion.get("hoja", "")
                )
            finally:
                try:
                    os.remove(ruta_temporal)
                except OSError:
                    pass
        if accion == "abrir":
            abrir_pdf(ruta)
            return None

        copias = max(1, int(self.app_settings.get("print_copies_hoja", 1) or 1))
        self.set_status(f"Reimprimiendo atención #{atencion['id']}...", "process")
        if imprimir_pdf(ruta, copias=copias, mostrar_error=True):
            self.set_status(f"Atención #{atencion['id']} enviada a impresión", "ok")
        else:
            self.set_status("La atención existe, pero falló la reimpresión", "error")
        return None

    def _iniciar_salida_atencion(self, atencion_id, hoja, datos_pdf, turno_cfg, abrir_pdf_final=False):
        payload = {
            "atencion_id": int(atencion_id),
            "hoja": hoja,
            "datos_pdf": dict(datos_pdf),
            "turno_cfg": dict(turno_cfg) if turno_cfg else None,
            "abrir_pdf_final": bool(abrir_pdf_final),
        }
        if not hasattr(self, "_output_payloads"):
            self._output_payloads = {}
        self._output_payloads[int(atencion_id)] = payload
        try:
            self.boton_generar_pdf.configure(state=tk.DISABLED)
        except Exception:
            pass
        hilo = threading.Thread(
            target=self._procesar_salida_atencion,
            kwargs=payload,
            name=f"salida-atencion-{int(atencion_id)}",
            daemon=True,
        )
        hilo.start()

    def _procesar_salida_atencion(
        self,
        atencion_id,
        hoja,
        datos_pdf,
        turno_cfg,
        abrir_pdf_final=False,
    ):
        errores = {}
        ruta_pdf = None
        trabajo = self.db.obtener_trabajo_salida(atencion_id) or {}

        if trabajo.get("excel_estado") != "COMPLETADO":
            try:
                self.db.actualizar_trabajo_salida(atencion_id, "excel", "PROCESANDO")
                if not turno_cfg:
                    turno_cfg = self.db.obtener_turno_config_atencion(atencion_id)
                if not turno_cfg:
                    raise TurnoNoVigenteError("No se pudo reconstruir el contexto del turno.")
                reconstruir_excel_turno(self.db, turno_cfg)
                self.db.actualizar_trabajo_salida(atencion_id, "excel", "COMPLETADO")
            except Exception as exc:
                errores["Excel"] = str(exc)
                APP_LOG.exception("Falló Excel para la atención #%s", atencion_id)
                self.db.actualizar_trabajo_salida(
                    atencion_id, "excel", "FALLIDO", error=str(exc)
                )

        trabajo = self.db.obtener_trabajo_salida(atencion_id) or {}
        documento = self.db.obtener_documento_atencion(atencion_id)
        if documento:
            ruta_pdf = documento.get("ruta")
        if trabajo.get("pdf_estado") != "COMPLETADO" or not ruta_pdf:
            ruta_temporal = None
            try:
                self.db.actualizar_trabajo_salida(atencion_id, "pdf", "PROCESANDO")
                ruta_temporal = crear_pdf_temporal(hoja, datos_pdf, mostrar_error=False)
                if not ruta_temporal:
                    raise RuntimeError("No se pudo generar el documento PDF.")
                ruta_pdf = archivar_pdf_atencion(ruta_temporal, atencion_id)
                sha256 = self.db.registrar_documento(
                    atencion_id, "HOJA_EMERGENCIA", ruta_pdf, hoja
                )
                self.db.actualizar_trabajo_salida(
                    atencion_id,
                    "pdf",
                    "COMPLETADO",
                    pdf_path=ruta_pdf,
                    pdf_sha256=sha256,
                )
            except Exception as exc:
                errores["PDF"] = str(exc)
                APP_LOG.exception("Falló el PDF para la atención #%s", atencion_id)
                self.db.actualizar_trabajo_salida(
                    atencion_id, "pdf", "FALLIDO", error=str(exc)
                )
                ruta_pdf = None
            finally:
                if ruta_temporal:
                    try:
                        os.remove(ruta_temporal)
                    except OSError:
                        pass

        trabajo = self.db.obtener_trabajo_salida(atencion_id) or {}
        if trabajo.get("impresion_estado") != "ENVIADO":
            if not ruta_pdf:
                error = "No existe un PDF íntegro para enviar a impresión."
                errores["Impresión"] = error
                self.db.actualizar_trabajo_salida(
                    atencion_id, "impresion", "FALLIDO", error=error
                )
            else:
                try:
                    self.db.actualizar_trabajo_salida(
                        atencion_id,
                        "impresion",
                        "PROCESANDO",
                        incrementar_intento=True,
                    )
                    copias = max(1, int(self.app_settings.get("print_copies_hoja", 1) or 1))
                    escribir_log_impresion(
                        f"Envío de la atención #{atencion_id} desde la cola durable."
                    )
                    if not imprimir_pdf(ruta_pdf, copias=copias, mostrar_error=False):
                        raise RuntimeError(
                            "La aplicación de impresión no confirmó el envío a la cola."
                        )
                    self.db.actualizar_trabajo_salida(
                        atencion_id, "impresion", "ENVIADO"
                    )
                except Exception as exc:
                    errores["Impresión"] = str(exc)
                    APP_LOG.exception("Falló la impresión para la atención #%s", atencion_id)
                    self.db.actualizar_trabajo_salida(
                        atencion_id, "impresion", "FALLIDO", error=str(exc)
                    )

        trabajo = self.db.obtener_trabajo_salida(atencion_id) or {}
        completo = (
            trabajo.get("excel_estado") == "COMPLETADO"
            and trabajo.get("pdf_estado") == "COMPLETADO"
            and trabajo.get("impresion_estado") == "ENVIADO"
        )
        if completo:
            self.db.limpiar_error_trabajo_salida(atencion_id)
        resultado = {
            "atencion_id": int(atencion_id),
            "ruta_pdf": ruta_pdf,
            "errores": errores,
            "trabajo": self.db.obtener_trabajo_salida(atencion_id) or trabajo,
            "abrir_pdf_final": bool(abrir_pdf_final),
        }
        try:
            self.root.after(0, lambda r=resultado: self._finalizar_salida_atencion(r))
        except Exception:
            APP_LOG.warning("La interfaz se cerró antes de finalizar la salida #%s", atencion_id)

    def _finalizar_salida_atencion(self, resultado):
        atencion_id = int(resultado["atencion_id"])
        trabajo = resultado.get("trabajo") or {}
        ruta_pdf = resultado.get("ruta_pdf")
        errores = resultado.get("errores") or {}
        try:
            self.boton_generar_pdf.configure(state=tk.NORMAL)
        except Exception:
            pass
        if resultado.get("abrir_pdf_final") and ruta_pdf:
            try:
                abrir_pdf(ruta_pdf)
            except Exception:
                APP_LOG.exception("No se pudo abrir el PDF de la atención #%s", atencion_id)

        if (
            trabajo.get("excel_estado") == "COMPLETADO"
            and trabajo.get("pdf_estado") == "COMPLETADO"
            and trabajo.get("impresion_estado") == "ENVIADO"
        ):
            self.set_status(
                f"Atención #{atencion_id} guardada y enviada a impresión", "ok"
            )
            self._output_payloads.pop(atencion_id, None)
        else:
            etapas = ", ".join(errores) or "una etapa de salida"
            self.set_status(
                f"Atención #{atencion_id} guardada; pendiente: {etapas}", "warning"
            )
            self._dialogo_salida_pendiente(atencion_id, ruta_pdf, errores)
        self._invalidar_caches_datos()
        self._refrescar_resumen_en_vivo()

    def _dialogo_salida_pendiente(self, atencion_id, ruta_pdf, errores):
        win = self._crear_toplevel_estable(
            f"Impresion o documento pendiente #{atencion_id}", "640x360", "salida_pendiente_win"
        )
        if win is None:
            return
        cont = tb.Frame(win, padding=20, style="Root.TFrame")
        cont.pack(fill="both", expand=True)
        tb.Label(
            cont,
            text=f"La atención #{atencion_id} está guardada",
            font=("Arial", 15, "bold"),
        ).pack(anchor="w", pady=(0, 10))
        detalle = "\n".join(f"{etapa}: {mensaje}" for etapa, mensaje in errores.items())
        tb.Label(
            cont,
            text=(detalle or "Queda una etapa pendiente.")
            + "\n\nReintentar no crea otra atención ni otro número.",
            justify="left",
            wraplength=590,
        ).pack(anchor="w", fill="x")
        botones = tb.Frame(cont, style="Root.TFrame")
        botones.pack(side="bottom", fill="x")

        def reintentar():
            try:
                win.destroy()
            except Exception:
                pass
            self._reintentar_trabajo_salida(atencion_id)

        tb.Button(
            botones, text="Reintentar pendientes", bootstyle=SUCCESS, command=reintentar
        ).pack(side="left", padx=(0, 8))
        if ruta_pdf:
            tb.Button(
                botones,
                text="Abrir PDF",
                bootstyle=INFO,
                command=lambda: abrir_pdf(ruta_pdf),
            ).pack(side="left", padx=(0, 8))
        tb.Button(botones, text="Continuar", command=win.destroy).pack(side="right")

    def _reintentar_trabajo_salida(self, atencion_id):
        payload = getattr(self, "_output_payloads", {}).get(int(atencion_id))
        if payload:
            self._iniciar_salida_atencion(**payload)
            return
        atencion = self.db.obtener_atencion_por_id(int(atencion_id))
        if not atencion or str(atencion.get("estado") or "").upper() != "ACTIVA":
            messagebox.showwarning("Documento pendiente", "La atención ya no está activa.")
            return
        self._iniciar_salida_atencion(
            int(atencion_id),
            atencion.get("hoja") or "GENERAL",
            self._snapshot_a_datos(atencion),
            self.db.obtener_turno_config_atencion(int(atencion_id)),
            False,
        )

    def _avisar_trabajos_salida_pendientes(self):
        try:
            pendientes = self.db.listar_trabajos_salida_pendientes(limite=100)
        except Exception:
            APP_LOG.exception("No se pudo consultar la cola de salidas")
            return
        if pendientes:
            self.set_status(
                f"Hay {len(pendientes)} atención(es) con impresión o documento pendiente", "warning"
            )

    def abrir_trabajos_salida_pendientes(self):
        win = self._crear_toplevel_estable(
            "Impresiones y documentos pendientes", "980x560", "trabajos_salida_win"
        )
        if win is None:
            return
        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)
        tb.Label(
            cont,
            text="Atenciones guardadas con tareas pendientes",
            font=("Arial", 15, "bold"),
        ).pack(anchor="w", pady=(0, 8))
        tb.Label(
            cont,
            text="Reintentar continúa desde la etapa fallida y nunca crea otra atención.",
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(0, 12))

        columnas = ("id", "paciente", "excel", "pdf", "impresion", "error")
        tree = ttk.Treeview(cont, columns=columnas, show="headings", style="Modern.Treeview")
        titulos = {
            "id": "Atención",
            "paciente": "Paciente",
            "excel": "Excel",
            "pdf": "PDF",
            "impresion": "Impresión",
            "error": "Último error",
        }
        anchos = {"id": 80, "paciente": 220, "excel": 95, "pdf": 95, "impresion": 105, "error": 300}
        for columna in columnas:
            tree.heading(columna, text=titulos[columna])
            tree.column(columna, width=anchos[columna], anchor="w")
        tree.pack(fill="both", expand=True)

        def cargar():
            for item in tree.get_children():
                tree.delete(item)
            pendientes = self.db.listar_trabajos_salida_pendientes(limite=200)
            for trabajo in pendientes:
                tree.insert(
                    "",
                    "end",
                    iid=str(trabajo["atencion_id"]),
                    values=(
                        trabajo["atencion_id"],
                        trabajo.get("nombre") or "",
                        trabajo.get("excel_estado") or "",
                        trabajo.get("pdf_estado") or "",
                        trabajo.get("impresion_estado") or "",
                        trabajo.get("ultimo_error") or "",
                    ),
                )
            if not pendientes:
                self.set_status("No hay impresiones ni documentos pendientes", "ok")

        def reintentar_seleccion():
            seleccion = tree.selection()
            if not seleccion:
                messagebox.showwarning(
                    "Impresiones y documentos pendientes", "Seleccione una atención para reintentar.", parent=win
                )
                return
            atencion_id = int(seleccion[0])
            win.destroy()
            self._reintentar_trabajo_salida(atencion_id)

        botones = tb.Frame(cont, style="Root.TFrame")
        botones.pack(fill="x", pady=(10, 0))
        tb.Button(
            botones,
            text="Reintentar selección",
            bootstyle=SUCCESS,
            command=reintentar_seleccion,
        ).pack(side="left")
        tb.Button(botones, text="Actualizar", command=cargar).pack(side="left", padx=8)
        tb.Button(botones, text="Cerrar", command=win.destroy).pack(side="right")
        tree.bind("<Double-1>", lambda _event: reintentar_seleccion())
        cargar()

    def generar_pdf(self):
        salida_iniciada = False
        try:
            self.boton_generar_pdf.config(state=tk.DISABLED)
            self.set_status("Validando atención...", "process")

            validacion = self._validar_campos_o_alertar()
            if not validacion:
                return
            datos, hoja = validacion

            ars_canon = normalizar_seguro(datos.get('Aseguradora (ARS)', ''), datos.get('NSS', ''))
            es_sin_seguro = (ars_canon == "SIN SEGURO")

            datos['Aseguradora (ARS)'] = ars_canon

            turno_cfg = cargar_turno_config()
            if not turno_cfg:
                self.set_status("Debe abrir el turno operativo actual", "warning")
                messagebox.showwarning(
                    "Turno requerido",
                    "No existe un turno vigente. Abra el turno actual antes de registrar al paciente.",
                    parent=self.root,
                )
                self._dialogo_turno()
                return

            duplicado = self._buscar_duplicado_turno_actual(datos)
            if duplicado is not None:
                try:
                    self.root.lift()
                    self.root.focus_force()
                except Exception:
                    pass
                reingreso = self._atender_duplicado_turno(duplicado)
                if not reingreso:
                    return
                datos.update(reingreso)

            if datos.get("EsReingreso"):
                msg = (
                    "Se registrará un reingreso autorizado y se conservará la referencia "
                    f"a la atención #{datos.get('AtencionOrigenId')}.\n\n¿Desea continuar?"
                )
            elif es_sin_seguro:
                msg = (
                    "Este paciente será registrado como SIN SEGURO.\n\n"
                    "Se guardará en el historial sin seguro, se agregará al Excel "
                    "y se generará/imprimirá la hoja."
                )
            else:
                msg = "¿Agregar a la lista (guardar) y generar/imprimir la hoja?"

            if bool(self.app_settings.get("validation_confirm_before_generate", True)):
                if not messagebox.askyesno("Confirmación", msg):
                    return

            datos_pdf = dict(datos)
            datos_db = dict(datos)
            if nombre_tiene_prefijo_rn(datos_db.get("Nombre", "")):
                if bool(self.app_settings.get("rn_strip_db", True)):
                    datos_db["Nombre"] = limpiar_nombre_rn_para_db(datos_db.get("Nombre", ""))
                if not bool(self.app_settings.get("rn_show_pdf", True)):
                    datos_pdf["Nombre"] = limpiar_nombre_rn_para_db(datos_pdf.get("Nombre", ""))

            self.set_status("Guardando atención en la base de datos...", "process")
            atencion_id = self.db.guardar_atencion(
                datos_db, hoja, turno_cfg=turno_cfg
            )
            self._ultimo_atencion_id = atencion_id
            revision_nss_id = self.db.obtener_revision_nss_atencion(atencion_id)
            comportamiento = self.app_settings.get(
                "print_behavior_hoja", "Imprimir y abrir PDF"
            )
            debe_abrir = (
                bool(self.app_settings.get("pdf_open_after_generate", True))
                and comportamiento == "Imprimir y abrir PDF"
            )
            self.set_status(
                f"Atención #{atencion_id} guardada; preparando documentos...", "process"
            )
            self._iniciar_salida_atencion(
                atencion_id,
                hoja,
                datos_pdf,
                turno_cfg,
                abrir_pdf_final=debe_abrir,
            )
            salida_iniciada = True

            self.limpiar_campos()
            self._restore_all_styles()
            self._refrescar_resumen_en_vivo()
            if revision_nss_id:
                aviso = (
                    f"Atención #{atencion_id} guardada; la hoja continúa normalmente. "
                    "El NSS fue enviado a revisión administrativa."
                )
                self.set_status(aviso, "warning")
                self._mostrar_notificacion(aviso, autohide_ms=12000, tipo="warning")

        except sqlite3.IntegrityError:
            APP_LOG.exception("Se bloqueó una atención duplicada por la restricción del turno")
            duplicado = self._buscar_duplicado_turno_actual(locals().get("datos", {}))
            if duplicado:
                self._atender_duplicado_turno(duplicado)
            else:
                messagebox.showwarning("Atención duplicada", "No se creó otra hoja para este paciente en el turno actual.")
        except (TurnoNoVigenteError, ValueError) as e:
            APP_LOG.warning("Registro rechazado: %s", e)
            self.set_status(str(e), "warning")
            messagebox.showwarning("No se guardó la atención", str(e), parent=self.root)
        except Exception as e:
            APP_LOG.exception("Error al generar la hoja de emergencia")
            self.set_status(f"Error: {str(e)}", "error")
            messagebox.showerror("Error", str(e))
        finally:
            if not salida_iniciada:
                self.boton_generar_pdf.config(state=tk.NORMAL)

    # ─── HISTORIALES ───────────────────────────────────────────────────────
    def abrir_historial(self):
        win = self._crear_toplevel_estable("Historial de Atenciones", "1160x720", "historial_win")
        if win is None:
            return

        # FASE 3: Constante para el caché del menú
        MENU_CACHE_SECONDS = 60

        self._bind_esc_cerrar(win)

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Historial de Atenciones",
            "Consulta registros, abre hojas PDF y anula atenciones cuando sea necesario.",
            "📁"
        )

        frm_bus = tb.Frame(cont, padding=12, style="Card.TFrame")
        frm_bus.pack(fill="x", pady=(0, 10))

        tb.Label(frm_bus, text="Buscar", font=("Arial", 10, "bold"), foreground="#EAF2FF", background="#0E1B2B").pack(side="left", padx=(0, 8))
        self.var_bus = tk.StringVar()
        ent_bus = tb.Entry(frm_bus, textvariable=self.var_bus, width=44)
        ent_bus.pack(side="left", ipady=4)
        win.after(80, lambda: (ent_bus.focus_set(), ent_bus.icursor("end")))

        filtro_rapido_var = tk.StringVar(value=self.app_settings.get("hist_default_filter", "Todos"))
        ars_filtro_var = tk.StringVar(value="(Todas)")
        esp_filtro_var = tk.StringVar(value="(Todas)")
        fecha_filtro = crear_selector_fecha(frm_bus, width=12)

        tb.Label(frm_bus, text="Filtro", font=("Arial", 10, "bold"), foreground="#EAF2FF", background="#0E1B2B").pack(side="left", padx=(10, 4))

        filtro_label_var = tk.StringVar(value="Todos  ▾")
        filtro_btn = tk.Menubutton(
            frm_bus,
            textvariable=filtro_label_var,
            relief="solid",
            bd=1,
            width=24,
            anchor="w",
            bg="#12243A",
            fg="#FFFFFF",
            activebackground="#1D6EFF",
            activeforeground="#FFFFFF",
            font=("Arial", 10)
        )
        filtro_btn.pack(side="left", padx=4, ipady=3)

        filtro_menu = tk.Menu(
            filtro_btn,
            tearoff=0,
            bg="#30475C",
            fg="#FFFFFF",
            activebackground="#506A83",
            activeforeground="#FFFFFF",
            font=("Arial", 10)
        )
        ars_menu = tk.Menu(
            filtro_menu,
            tearoff=0,
            bg="#30475C",
            fg="#FFFFFF",
            activebackground="#506A83",
            activeforeground="#FFFFFF",
            font=("Arial", 10)
        )
        esp_menu = tk.Menu(
            filtro_menu,
            tearoff=0,
            bg="#30475C",
            fg="#FFFFFF",
            activebackground="#506A83",
            activeforeground="#FFFFFF",
            font=("Arial", 10)
        )
        filtro_btn.configure(menu=filtro_menu)

        try:
            fecha_filtro.pack_forget()
        except Exception:
            pass

        def _actualizar_texto_boton_filtro():
            modo = filtro_rapido_var.get()
            if modo == "Por ARS":
                seleccionado = ars_filtro_var.get() or "(Todas)"
                filtro_label_var.set(f"Por ARS: {seleccionado}  ▾")
            elif modo == "Por especialidad":
                seleccionado = esp_filtro_var.get() or "(Todas)"
                filtro_label_var.set(f"Por especialidad: {seleccionado}  ▾")
            elif modo == "Por fecha":
                filtro_label_var.set("Por fecha  ▾")
            else:
                filtro_label_var.set(f"{modo}  ▾")

        def _actualizar_visibilidad_filtros(*_):
            try:
                fecha_filtro.pack_forget()
            except Exception:
                pass
            if filtro_rapido_var.get() == "Por fecha":
                fecha_filtro.pack(side="left", padx=4)
            try:
                _actualizar_texto_boton_filtro()
            except Exception:
                pass

        menu_state = {"last_build": 0, "busy": False}

        def _programar_busqueda_filtro():
            try:
                win.after(180, buscar)
            except Exception:
                try:
                    buscar()
                except Exception:
                    pass

        def _seleccionar_filtro_simple(modo):
            filtro_rapido_var.set(modo)
            if modo != "Por ARS":
                ars_filtro_var.set("(Todas)")
            if modo != "Por especialidad":
                esp_filtro_var.set("(Todas)")
            _actualizar_visibilidad_filtros()
            _programar_busqueda_filtro()

        def _seleccionar_ars_filtro(valor):
            filtro_rapido_var.set("Por ARS")
            ars_filtro_var.set(valor)
            esp_filtro_var.set("(Todas)")
            _actualizar_visibilidad_filtros()
            _programar_busqueda_filtro()

        def _seleccionar_especialidad_filtro(valor):
            filtro_rapido_var.set("Por especialidad")
            esp_filtro_var.set(valor)
            ars_filtro_var.set("(Todas)")
            _actualizar_visibilidad_filtros()
            _programar_busqueda_filtro()

        def _reconstruir_menu_filtros(_=None, forzar=False):
            """
            FASE 3: Reconstruye el menú usando caché de 60 segundos.
            """
            if menu_state.get("busy"):
                return
            ahora = _time.time()
            if (not forzar) and (ahora - float(menu_state.get("last_build") or 0) < MENU_CACHE_SECONDS):
                return

            menu_state["busy"] = True
            try:
                filtro_menu.delete(0, "end")
                ars_menu.delete(0, "end")
                esp_menu.delete(0, "end")

                filtro_menu.add_command(label="Todos", command=lambda: _seleccionar_filtro_simple("Todos"))
                filtro_menu.add_command(label="Hoy", command=lambda: _seleccionar_filtro_simple("Hoy"))
                filtro_menu.add_command(label="Turno actual", command=lambda: _seleccionar_filtro_simple("Turno actual"))
                filtro_menu.add_command(label="Sin seguro", command=lambda: _seleccionar_filtro_simple("Sin seguro"))
                filtro_menu.add_separator()

                ars_menu.add_command(label="(Todas)", command=lambda: _seleccionar_ars_filtro("(Todas)"))
                ars_menu.add_separator()

                ars_items = self._obtener_ars_cache(forzar=False)
                if not ars_items:
                    ars_items = sorted([a for a in DEFAULT_ARS_CATALOGO.keys() if a != "SIN SEGURO"])

                ars_items = [a for a in ars_items if a and str(a).strip()]
                total_ars = len(ars_items)

                for ars_val in ars_items[:80]:
                    ars_menu.add_command(label=ars_val, command=lambda v=ars_val: _seleccionar_ars_filtro(v))

                if total_ars > 80:
                    ars_menu.add_separator()
                    ars_menu.add_command(
                        label=f"Hay {total_ars} ARS. Use Buscar para filtrar más.",
                        state="disabled"
                    )

                filtro_menu.add_cascade(label="Por ARS", menu=ars_menu)

                for esp_val in ["(Todas)", "GENERAL", "PEDIATRIA", "GINECOLOGIA"]:
                    esp_menu.add_command(label=esp_val, command=lambda v=esp_val: _seleccionar_especialidad_filtro(v))
                filtro_menu.add_cascade(label="Por especialidad", menu=esp_menu)

                filtro_menu.add_command(label="Por fecha", command=lambda: _seleccionar_filtro_simple("Por fecha"))
                menu_state["last_build"] = ahora
            finally:
                menu_state["busy"] = False

        try:
            filtro_menu.configure(postcommand=lambda: _reconstruir_menu_filtros(forzar=False))
        except Exception:
            pass

        try:
            win.after(50, lambda: _reconstruir_menu_filtros(forzar=False))
        except Exception:
            pass
        _actualizar_texto_boton_filtro()

        cols = ("id", "fecha", "hora", "nombre", "hoja", "ars", "nss", "cedula", "tipo")
        tree = ttk.Treeview(cont, columns=cols, show="headings", height=10, style="Modern.Treeview")
        tree.pack(fill="both", expand=True, pady=(0, 10))

        cols_def = [
            ("id", "ID", 60, "center"),
            ("fecha", "Fecha", 90, "center"),
            ("hora", "Hora", 90, "center"),
            ("nombre", "Nombre", 250, "w"),
            ("hoja", "Especialidad", 120, "center"),
            ("ars", "Seguro", 160, "center"),
            ("nss", "NSS", 125, "center"),
            ("cedula", "Cédula", 125, "center"),
            ("tipo", "Tipo", 100, "center"),
        ]
        for c, title, w, anchor in cols_def:
            tree.heading(c, text=title)
            tree.column(c, width=w, anchor=anchor)

        # FASE 1: 100 / 150
        page_state = {
            "offset": 0,
            "first_limit": min(100, max(50, int(self.app_settings.get("hist_initial_limit", 100) or 100))),
            "next_limit":  min(150, max(80, int(self.app_settings.get("hist_next_limit",  150) or 150))),
            "loading": False,
            "done": False,
            "load_id": 0,
        }

        def _insertar_mensaje_tabla(mensaje):
            for i in tree.get_children():
                tree.delete(i)
            tree.insert("", "end", values=("", "", "", mensaje, "", "", "", "", ""))

        def cargar_pagina(reset=False):
            if page_state["loading"]:
                try:
                    win.after(250, lambda: cargar_pagina(reset=reset))
                except Exception:
                    pass
                return
            page_state["loading"] = True
            page_state["load_id"] += 1
            current_load = page_state["load_id"]

            if reset:
                page_state["offset"] = 0
                page_state["done"] = False
                _insertar_mensaje_tabla("Cargando datos…")

            limit = page_state["first_limit"] if reset else page_state["next_limit"]
            fecha_txt = obtener_fecha_selector(fecha_filtro) if filtro_rapido_var.get() == "Por fecha" else None

            def _finalizar_carga(filas):
                if current_load != page_state["load_id"]:
                    return
                if reset:
                    for i in tree.get_children():
                        tree.delete(i)
                for f in filas:
                    seguro_canon = normalizar_seguro(f.get("ars", ""), f.get("nss", ""))
                    tree.insert(
                        "",
                        "end",
                        values=(
                            f["id"],
                            f["fecha"],
                            f["hora"],
                            f["nombre"],
                            f["hoja"],
                            seguro_para_mostrar(seguro_canon),
                            f["nss"],
                            f.get("cedula", ""),
                            (f.get("tipo_atencion") or "EMERGENCIA")
                        )
                    )
                page_state["offset"] += len(filas)
                if len(filas) < limit:
                    page_state["done"] = True
                if reset and not filas:
                    tree.insert("", "end", values=("", "", "", "No se encontraron registros.", "", "", "", "", ""))
                page_state["loading"] = False

            def _error_carga(err):
                if current_load != page_state["load_id"]:
                    return
                page_state["loading"] = False
                self.set_status("Error cargando historial.", "error")

            def _hacer_carga():
                if page_state["done"] and not reset:
                    return []
                modo_actual = filtro_rapido_var.get()
                # FASE 2: Rama dedicada para Sin seguro
                if modo_actual == "Sin seguro":
                    return self.db.listar_atenciones_sin_seguro(
                        filtro_texto=self.var_bus.get().strip(),
                        limite=limit,
                        offset=page_state["offset"],
                    )
                
                return self.db.listar_atenciones_filtradas(
                    filtro_texto=self.var_bus.get().strip(),
                    modo=modo_actual,
                    ars=ars_filtro_var.get(),
                    especialidad=esp_filtro_var.get(),
                    fecha_txt=fecha_txt,
                    limite=limit,
                    offset=page_state["offset"],
                )

            self._ejecutar_en_segundo_plano(
                "Cargando historial…",
                _hacer_carga,
                al_terminar=_finalizar_carga,
                al_error=_error_carga,
            )

        def buscar():
            cargar_pagina(reset=True)

        def _al_mover_scroll(event=None):
            try:
                win.after(80, lambda: (not page_state["loading"] and not page_state["done"] and tree.yview()[1] >= 0.98 and cargar_pagina(reset=False)))
            except Exception:
                pass

        try:
            fecha_filtro.bind("<<DateEntrySelected>>", lambda e: buscar())
            fecha_filtro.bind("<Return>", lambda e: buscar())
        except Exception:
            pass
        tree.bind("<MouseWheel>", _al_mover_scroll)
        tree.bind("<Button-5>", _al_mover_scroll)

        tb.Button(frm_bus, text="🔎  Buscar", bootstyle=PRIMARY, command=buscar).pack(side="left", padx=6, ipady=3)
        tb.Button(
            frm_bus,
            text="Mostrar todo",
            bootstyle=SECONDARY,
            command=lambda: [
                self.var_bus.set(""),
                filtro_rapido_var.set("Todos"),
                ars_filtro_var.set("(Todas)"),
                esp_filtro_var.set("(Todas)"),
                _actualizar_visibilidad_filtros(),
                _reconstruir_menu_filtros(),
                buscar()
            ]
        ).pack(side="left", padx=6, ipady=3)
        tb.Button(frm_bus, text="📊  Reporte estadístico", bootstyle=INFO, command=self.abrir_ventana_reporte).pack(side="right", padx=(8, 0), ipady=3)

        ent_bus.bind("<Return>", lambda e: buscar())
        _actualizar_visibilidad_filtros()

        frm_btn = tb.Frame(cont, padding=(12, 10), style="Card.TFrame")
        frm_btn.pack(side="bottom", fill="x", pady=(8, 0))

        tb.Button(frm_btn, text="📄  Ver PDF", bootstyle=SUCCESS, command=lambda: self.ver_pdf_seleccionado(tree)).pack(side="left", padx=4, ipady=4)
        tb.Button(frm_btn, text="🖉  Editar atención", bootstyle=SECONDARY, command=lambda: self._abrir_editor_atencion_desde_tree(tree, buscar)).pack(side="left", padx=4, ipady=4)
        tb.Button(frm_btn, text="Anular seleccionado", bootstyle=DANGER, command=lambda: self.eliminar_atencion_seleccionada(tree, reordenar_ids=True, refrescar_callback=buscar)).pack(side="left", padx=4, ipady=4)
        tb.Button(frm_btn, text="🛡  Historial sin seguros", bootstyle=WARNING, command=self.abrir_historial_sin_seguros).pack(side="left", padx=4, ipady=4)

        try:
            win.after(10, lambda: cargar_pagina(reset=True))
        except Exception:
            cargar_pagina(reset=True)

        menu_historial = tk.Menu(win, tearoff=0)

        def _seleccionar_fila_click_derecho(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                tree.focus(item)
            return item

        def _identidad_seleccionada_para_config():
            sel = tree.selection()
            if not sel:
                self._mostrar_dialogo_modal_unico("Historial", "Seleccione un registro para editar.")
                return ""
            vals = tree.item(sel[0], "values")
            nss = (vals[6] if len(vals) > 6 else "") or ""
            ced = (vals[7] if len(vals) > 7 else "") or ""
            if is_valid_nss_key(str(nss)):
                return str(nss).strip()
            if is_valid_cedula_key(str(ced)):
                return str(ced).strip()
            self._mostrar_dialogo_modal_unico(
                "Editar paciente",
                "Este registro no tiene NSS ni cédula válida para buscarlo en configuración."
            )
            return ""

        def _editar_paciente_desde_historial():
            identidad = _identidad_seleccionada_para_config()
            if identidad:
                self._abrir_edicion_paciente(prefill_identidad=identidad)

        menu_historial.add_command(label="🖉 Editar atención", command=lambda: self._abrir_editor_atencion_desde_tree(tree, buscar))
        menu_historial.add_command(label="⚙ Editar datos del paciente", command=_editar_paciente_desde_historial)
        menu_historial.add_separator()
        menu_historial.add_command(label="📄 Ver PDF", command=lambda: self.ver_pdf_seleccionado(tree))
        menu_historial.add_command(label="Anular atención", command=lambda: self.eliminar_atencion_seleccionada(tree, reordenar_ids=True, refrescar_callback=buscar))

        def _mostrar_menu_historial(event):
            if _seleccionar_fila_click_derecho(event):
                menu_historial.tk_popup(event.x_root, event.y_root)

        tree.bind("<Button-3>", _mostrar_menu_historial)
        tree.bind("<Double-1>", lambda e: self.ver_pdf_seleccionado(tree))

    def cargar_tabla(self, tree, filtro):
        for i in tree.get_children():
            tree.delete(i)
        filas = self.db.listar_atenciones(filtro_texto=filtro, limite=200, offset=0)
        for f in filas:
            seguro_canon = normalizar_seguro(f.get("ars", ""), f.get("nss", ""))
            tree.insert(
                "",
                "end",
                values=(
                    f["id"],
                    f["fecha"],
                    f["hora"],
                    f["nombre"],
                    f["hoja"],
                    seguro_para_mostrar(seguro_canon),
                    f["nss"],
                    f.get("cedula", "")
                )
            )

    def cargar_tabla_filtrada(self, tree, filtro=None, modo="Todos", ars=None, especialidad=None, fecha_txt=None):
        for i in tree.get_children(): tree.delete(i)
        filas = self.db.listar_atenciones_filtradas(filtro_texto=filtro, modo=modo or "Todos", ars=ars, especialidad=especialidad, fecha_txt=fecha_txt, limite=200, offset=0)
        for f in filas:
            seguro_canon = normalizar_seguro(f.get("ars", ""), f.get("nss", ""))
            tree.insert("", "end", values=(f["id"], f["fecha"], f["hora"], f["nombre"], f["hoja"], seguro_para_mostrar(seguro_canon), f["nss"], f.get("cedula", "")))

    def _snapshot_a_datos(self, atencion):
        return {
            "Fecha": atencion["fecha"],
            "Hora": atencion["hora"],
            "Nombre": atencion["nombre"],
            "Sexo": atencion.get("sexo", "") or "Femenino",
            "Edad_num": int(atencion.get("edad_num") or 0),
            "Unidad": atencion.get("unidad", "Años"),
            "Cédula": atencion.get("cedula", ""),
            "Teléfono": atencion.get("telefono", "") or "",
            "Dirección": atencion.get("direccion", ""),
            "Nacionalidad": atencion.get("nacionalidad", ""),
            "Aseguradora (ARS)": atencion.get("ars", ""),
            "NSS": atencion.get("nss", ""),
            "TipoAtencion": atencion.get("tipo_atencion", "EMERGENCIA")
        }

    def ver_pdf_seleccionado(self, tree):
        sel = tree.selection()
        if not sel:
            self._mostrar_dialogo_modal_unico("Historial", "Seleccione un registro.")
            return

        vals = tree.item(sel[0], "values")
        atencion_id = int(vals[0])
        atencion = self.db.obtener_atencion_por_id(atencion_id)
        if not atencion:
            messagebox.showerror("Error", "No se encontró la atención seleccionada.")
            return

        documento = self.db.obtener_documento_atencion(atencion_id)
        if documento:
            abrir_pdf(documento["ruta"])
            return

        ruta = regenerar_pdf_archivado(
            self.db, atencion_id, mostrar_error=False
        )
        if not ruta:
            messagebox.showerror("PDF", "No se pudo recuperar ni reconstruir el documento.")
            return
        abrir_pdf(ruta)

    def eliminar_atencion_seleccionada(self, tree, reordenar_ids=False, refrescar_callback=None):
        sel = tree.selection()
        if not sel:
            self._mostrar_dialogo_modal_unico("Historial", "Seleccione un registro para anular.")
            return

        actor_autorizado = self._actor_actual()

        vals = tree.item(sel[0], "values")
        atencion_id = int(vals[0])

        atencion = self.db.obtener_atencion_por_id(atencion_id)
        if not atencion:
            messagebox.showwarning("Aviso", "No se encontró el registro.")
            return

        afecta_excel = self._registro_esta_en_turno_actual(atencion)

        mensaje_eliminar = (
            "¿Anular esta atención?\n\n"
            f"Paciente: {(atencion.get('nombre') or '').upper()}\n"
            f"Fecha: {atencion.get('fecha', '')} {atencion.get('hora', '')}\n"
            f"Especialidad: {atencion.get('hoja', '')}\n\n"
            "También se actualizará el Excel si pertenece al turno actual."
        )
        if not messagebox.askyesno("Confirmación", mensaje_eliminar):
            return

        motivo = simpledialog.askstring(
            "Motivo de anulación",
            "Indique brevemente por qué se anula esta atención:",
            parent=self.root,
        )
        if motivo is None:
            return
        motivo = motivo.strip()
        if len(motivo) < 5:
            messagebox.showwarning(
                "Motivo requerido",
                "Escriba un motivo de al menos 5 caracteres para conservar la auditoría.",
                parent=tree.winfo_toplevel(),
            )
            return
        turno_cfg = cargar_turno_config() or {}
        usuario = actor_autorizado or turno_cfg.get("representante", "")

        try:
            anulada = self.db.borrar_atencion(
                atencion_id, motivo=motivo, usuario=usuario
            )
        except ValueError as exc:
            messagebox.showwarning("No se puede anular", str(exc), parent=self.root)
            return

        if anulada:

            if afecta_excel and (atencion.get("tipo_atencion") or "EMERGENCIA").strip().upper() != "URGENCIA":
                self._reconstruir_excel_si_necesario("anular atención", antes=atencion, despues={})

            if refrescar_callback:
                try:
                    refrescar_callback()
                except Exception:
                    self.cargar_tabla_filtrada(tree)
            else:
                self.cargar_tabla_filtrada(tree)

            self._actualizar_resumen_turno_panel()
            self._mostrar_notificacion(f"Atención #{atencion_id} anulada.")
        else:
            messagebox.showwarning("Aviso", "No se pudo anular. Intente nuevamente.")

    def abrir_historial_sin_seguros(self):
        win = self._crear_toplevel_estable("Historial sin seguros", "1120x700", "historial_sin_seguro_win")
        if win is None:
            return

        self._bind_esc_cerrar(win)

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Historial sin seguros",
            "Pacientes registrados sin cobertura o sin NSS válido. Busca por nombre, consulta, abre PDF o anula.",
            "🛡"
        )

        frm_bus = tb.Frame(cont, padding=12, style="Card.TFrame")
        frm_bus.pack(fill="x", pady=(0, 10))

        tb.Label(frm_bus, text="Buscar", font=("Arial", 10, "bold"), foreground="#EAF2FF", background="#0E1B2B").pack(side="left", padx=(0, 8))
        var_bus = tk.StringVar()
        ent_bus = tb.Entry(frm_bus, textvariable=var_bus, width=44)
        ent_bus.pack(side="left", ipady=4)
        win.after(80, lambda: (ent_bus.focus_set(), ent_bus.icursor("end")))

        btn_buscar_nombre = tb.Button(frm_bus, text="🔎  Buscar nombre", bootstyle=PRIMARY, command=lambda: None)
        btn_buscar_nombre.pack(side="left", padx=8, ipady=3)

        btn_mostrar_todo = tb.Button(frm_bus, text="Mostrar todo", bootstyle=SECONDARY, command=lambda: None)
        btn_mostrar_todo.pack(side="left", padx=4, ipady=3)

        cols = ("id", "fecha", "hora", "nombre", "hoja", "ars", "nss", "cedula", "tipo")
        tree = ttk.Treeview(cont, columns=cols, show="headings", height=10, style="Modern.Treeview")
        tree.pack(fill="both", expand=True, pady=(0, 10))

        cols_def = [
            ("id", "ID", 60, "center"),
            ("fecha", "Fecha", 90, "center"),
            ("hora", "Hora", 90, "center"),
            ("nombre", "Nombre", 260, "w"),
            ("hoja", "Especialidad", 125, "center"),
            ("ars", "Seguro", 130, "center"),
            ("nss", "NSS", 125, "center"),
            ("cedula", "Cédula", 125, "center"),
            ("tipo", "Tipo", 100, "center"),
        ]
        for c, title, w, anchor in cols_def:
            tree.heading(c, text=title)
            tree.column(c, width=w, anchor=anchor)

        # FASE 1: 100 / 150
        page_state = {
            "offset": 0,
            "first_limit": min(100, max(50, int(self.app_settings.get("hist_initial_limit", 100) or 100))),
            "next_limit":  min(150, max(80, int(self.app_settings.get("hist_next_limit",  150) or 150))),
            "loading": False,
            "done": False,
            "load_id": 0,
        }

        def _insertar_mensaje_tabla(mensaje):
            for i in tree.get_children():
                tree.delete(i)
            tree.insert("", "end", values=("", "", "", mensaje, "", "", "", "", ""))

        def cargar(reset=True):
            if page_state["loading"]:
                return
            page_state["loading"] = True
            page_state["load_id"] += 1
            current_load = page_state["load_id"]

            if reset:
                page_state["offset"] = 0
                page_state["done"] = False
                _insertar_mensaje_tabla("Cargando datos…")

            limit = page_state["first_limit"] if reset else page_state["next_limit"]

            def _finalizar_carga(filas):
                if current_load != page_state["load_id"]:
                    return
                if reset:
                    for i in tree.get_children():
                        tree.delete(i)
                for f in filas:
                    tree.insert(
                        "",
                        "end",
                        values=(
                            f["id"],
                            f["fecha"],
                            f["hora"],
                            f["nombre"],
                            f["hoja"],
                            "SIN SEGURO",
                            f["nss"],
                            f.get("cedula", ""),
                            (f.get("tipo_atencion") or "EMERGENCIA")
                        )
                    )
                page_state["offset"] += len(filas)
                if len(filas) < limit:
                    page_state["done"] = True
                if reset and not filas:
                    tree.insert("", "end", values=("", "", "", "No se encontraron registros sin seguro.", "", "", "", "", ""))
                page_state["loading"] = False

            def _error_carga(err):
                if current_load != page_state["load_id"]:
                    return
                page_state["loading"] = False
                self.set_status("Error cargando historial sin seguro.", "error")

            def _hacer_carga():
                if page_state["done"] and not reset:
                    return []
                return self.db.listar_atenciones_sin_seguro(
                    var_bus.get().strip() or None,
                    limite=limit,
                    offset=page_state["offset"],
                )

            self._ejecutar_en_segundo_plano(
                "Cargando historial sin seguro…",
                _hacer_carga,
                al_terminar=_finalizar_carga,
                al_error=_error_carga,
            )

        def _al_mover_scroll_ss(event=None):
            try:
                win.after(80, lambda: (not page_state["loading"] and not page_state["done"] and tree.yview()[1] >= 0.98 and cargar(reset=False)))
            except Exception:
                pass

        tree.bind("<MouseWheel>", _al_mover_scroll_ss)
        tree.bind("<Button-5>", _al_mover_scroll_ss)

        def buscar():
            cargar(reset=True)

        def mostrar_todo():
            var_bus.set("")
            cargar(reset=True)

        try:
            btn_buscar_nombre.configure(command=buscar)
            btn_mostrar_todo.configure(command=mostrar_todo)
            ent_bus.bind("<Return>", lambda e: buscar())
        except Exception:
            pass

        frm_btn = tb.Frame(cont, padding=(12, 10), style="Card.TFrame")
        frm_btn.pack(side="bottom", fill="x", pady=(8, 0))
        tb.Button(frm_btn, text="📄  Ver PDF", bootstyle=SUCCESS, command=lambda: self.ver_pdf_seleccionado(tree)).pack(side="left", padx=4, ipady=4)
        tb.Button(frm_btn, text="🖉  Editar atención", bootstyle=SECONDARY, command=lambda: self._abrir_editor_atencion_desde_tree(tree, cargar)).pack(side="left", padx=4, ipady=4)
        tb.Button(frm_btn, text="Anular seleccionado", bootstyle=DANGER, command=lambda: self.eliminar_atencion_seleccionada(tree, reordenar_ids=False, refrescar_callback=cargar)).pack(side="left", padx=4, ipady=4)
        tb.Button(frm_btn, text="⚙ Editar paciente", bootstyle=INFO, command=lambda: self._abrir_edicion_paciente(prefill_identidad=(tree.item(tree.selection()[0], "values")[6] or tree.item(tree.selection()[0], "values")[7]) if tree.selection() else "")).pack(side="left", padx=4, ipady=4)

        menu_sin_seguro = tk.Menu(win, tearoff=0)

        def _seleccionar_fila_ss(event):
            item = tree.identify_row(event.y)
            if item:
                tree.selection_set(item)
                tree.focus(item)
            return item

        def _editar_paciente_ss():
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            identidad = ""
            if len(vals) > 6 and is_valid_nss_key(str(vals[6])):
                identidad = str(vals[6]).strip()
            elif len(vals) > 7 and is_valid_cedula_key(str(vals[7])):
                identidad = str(vals[7]).strip()
            if identidad:
                self._abrir_edicion_paciente(prefill_identidad=identidad)
            else:
                self._mostrar_dialogo_modal_unico("Editar paciente", "Este registro no tiene NSS ni cédula válida para buscarlo en configuración.")

        menu_sin_seguro.add_command(label="🖉 Editar atención", command=lambda: self._abrir_editor_atencion_desde_tree(tree, cargar))
        menu_sin_seguro.add_command(label="⚙ Editar datos del paciente", command=_editar_paciente_ss)
        menu_sin_seguro.add_separator()
        menu_sin_seguro.add_command(label="📄 Ver PDF", command=lambda: self.ver_pdf_seleccionado(tree))
        menu_sin_seguro.add_command(label="Anular atención", command=lambda: self.eliminar_atencion_seleccionada(tree, reordenar_ids=False, refrescar_callback=cargar))

        def _mostrar_menu_ss(event):
            if _seleccionar_fila_ss(event):
                menu_sin_seguro.tk_popup(event.x_root, event.y_root)

        tree.bind("<Button-3>", _mostrar_menu_ss)
        tree.bind("<Double-1>", lambda e: self.ver_pdf_seleccionado(tree))

        try:
            win.after(80, lambda: cargar(reset=True))
        except Exception:
            cargar(reset=True)

    @staticmethod
    def _bind_esc_cerrar(win: Toplevel):
        win.bind("<Escape>", lambda e: win.destroy())
    # ────────────────────────────────────────────────────────────────────────

    def _abrir_editor_atencion_desde_tree(self, tree, on_saved=None):
        sel = tree.selection()
        if not sel:
            self._mostrar_dialogo_modal_unico("Historial", "Seleccione una atención para editar.")
            return
        vals = tree.item(sel[0], "values")
        self._abrir_editor_atencion(int(vals[0]), on_saved=on_saved)

    def _abrir_editor_atencion(self, atencion_id: int, on_saved=None):
        at = self.db.obtener_atencion_por_id(atencion_id)
        if not at:
            messagebox.showerror("Error", "No se encontró la atención seleccionada.")
            return

        win = Toplevel(self.root)
        win.title(f"Editar atención #{atencion_id}")
        win.geometry("820x760")
        win.minsize(800, 720)
        win.transient(self.root)
        self._bind_esc_cerrar(win)
        try:
            win.lift()
            win.focus_set()
            win.after(20, win.focus_set)
        except Exception:
            pass

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Editar atención específica",
            "Corrige este registro del historial, incluyendo edad, unidad, seguro y tipo de atención.",
            "🖉"
        )

        footer = tb.Frame(cont, padding=(8, 10), style="Card.TFrame")
        footer.pack(side="bottom", fill="x", pady=(8, 0))

        form = tb.Frame(cont, padding=12, style="Card.TFrame")
        form.pack(fill="both", expand=True)
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        campos = {}
        campos_def = [
            ("Nombre", at.get("nombre", ""), "entry"),
            ("Sexo", at.get("sexo", "Femenino") or "Femenino", "sexo"),
            ("Fecha", at.get("fecha", ""), "entry"),
            ("Hora", at.get("hora", ""), "entry"),
            ("Hoja", at.get("hoja", ""), "hoja"),
            ("Edad", at.get("edad_num", ""), "entry"),
            ("Unidad", at.get("unidad", "Años"), "unidad"),
            ("TipoAtencion", at.get("tipo_atencion", "EMERGENCIA"), "tipo"),
            ("Aseguradora (ARS)", seguro_para_mostrar(normalizar_seguro(at.get("ars", ""), at.get("nss", ""))), "entry"),
            ("NSS", at.get("nss", ""), "entry"),
            ("Cédula", at.get("cedula", ""), "entry"),
            ("Teléfono", at.get("telefono", ""), "entry"),
            ("Dirección", at.get("direccion", ""), "entry"),
            ("Nacionalidad", at.get("nacionalidad", ""), "entry"),
        ]

        for idx, (label, value, kind) in enumerate(campos_def):
            row = idx // 2
            col = (idx % 2) * 2
            visible_label = "Tipo de atención" if label == "TipoAtencion" else label

            tb.Label(form, text=visible_label, background="#0E1B2B", font=("Arial", 10, "bold")).grid(
                row=row * 2, column=col, sticky="w", padx=6, pady=(4, 2)
            )

            if kind == "hoja":
                var = tk.StringVar(value=(value or "GENERAL").upper())
                ent = tb.Combobox(form, textvariable=var, state="readonly", values=["GENERAL", "PEDIATRIA", "GINECOLOGIA"])
                ent.after_idle(lambda ent=ent, v=(value or "GENERAL"): ent.current(["GENERAL", "PEDIATRIA", "GINECOLOGIA"].index(v.upper())))
            elif kind == "unidad":
                val_unidad = value if value in ["Días", "Meses", "Años"] else "Años"
                var = tk.StringVar(value=val_unidad)
                ent = tb.Combobox(form, textvariable=var, state="readonly", values=["Días", "Meses", "Años"])
                ent.after_idle(lambda ent=ent, v=val_unidad: ent.current(["Días", "Meses", "Años"].index(v)))
            elif kind == "tipo":
                var = tk.StringVar(value=(value or "EMERGENCIA").upper())
                ent = tb.Combobox(form, textvariable=var, state="readonly", values=["EMERGENCIA", "URGENCIA"])
                ent.after_idle(lambda ent=ent, v=(value or "EMERGENCIA"): ent.current(["EMERGENCIA", "URGENCIA"].index(v.upper())))
            elif kind == "sexo":
                val_sexo = value if value in ["Masculino", "Femenino"] else "Femenino"
                var = tk.StringVar(value=val_sexo)
                ent = tb.Combobox(
                    form,
                    textvariable=var,
                    state="readonly",
                    values=["Femenino", "Masculino"],
                )
                ent.after_idle(
                    lambda ent=ent, v=val_sexo: ent.current(
                        ["Femenino", "Masculino"].index(v)
                    )
                )
            else:
                ent = tb.Entry(form)
                ent.insert(0, value or "")

            ent.grid(row=row * 2 + 1, column=col, sticky="ew", padx=6, pady=(0, 6), ipady=3)
            campos[label] = ent

        try:
            campos["Cédula"].bind("<KeyRelease>", lambda e: self.limitar_caracteres(campos["Cédula"], 11))
            campos["Teléfono"].bind("<KeyRelease>", lambda e: self.limitar_caracteres(campos["Teléfono"], 10))
        except Exception:
            pass

        def _get(label):
            return campos[label].get().strip()

        def guardar():
            if not _get("Nombre"):
                messagebox.showerror("Validación", "El nombre no puede quedar vacío.")
                campos["Nombre"].focus_set()
                return

            edad_txt = _get("Edad")
            if not edad_txt.isdigit():
                messagebox.showerror("Validación", "La edad debe ser numérica.")
                campos["Edad"].focus_set()
                return

            edad_num = int(edad_txt)
            if edad_num < 0 or edad_num > 130:
                messagebox.showerror("Validación", "La edad debe estar entre 0 y 130.")
                campos["Edad"].focus_set()
                return

            hoja = _get("Hoja").upper().strip()
            # FASE 11: La hoja no puede quedar vacía
            if not hoja or hoja not in ["GENERAL", "PEDIATRIA", "GINECOLOGIA"]:
                messagebox.showerror("Validación", "Debe seleccionar una especialidad válida (GENERAL, PEDIATRIA o GINECOLOGIA).")
                campos["Hoja"].focus_set()
                return

            ced = _get("Cédula").replace("-", "")
            tel = _get("Teléfono").replace("-", "")

            if ced and not is_valid_cedula_key(ced):
                messagebox.showerror("Validación", "La cédula debe tener 11 dígitos o dejarse vacía.")
                campos["Cédula"].focus_set()
                return

            if tel and not (tel.isdigit() and len(tel) == 10):
                messagebox.showerror("Validación", "El teléfono debe tener 10 dígitos o dejarse vacío.")
                campos["Teléfono"].focus_set()
                return

            nss = _get("NSS").upper()
            if nss and not is_valid_nss_key(nss) and nss not in ["SIN SEGURO", "NO", "N/S", "NS", "N\\S"]:
                messagebox.showerror("Validación", "El NSS debe ser numérico o SIN SEGURO.")
                campos["NSS"].focus_set()
                return

            nuevos = {
                "Nombre": _get("Nombre"),
                "Sexo": _get("Sexo") or "Femenino",
                "Fecha": _get("Fecha"),
                "Hora": _get("Hora"),
                "Hoja": hoja,
                "Edad_num": edad_num,
                "Unidad": _get("Unidad") or "Años",
                "TipoAtencion": _get("TipoAtencion") or "EMERGENCIA",
                "Aseguradora (ARS)": _get("Aseguradora (ARS)"),
                "NSS": nss,
                "Cédula": ced,
                "Teléfono": tel,
                "Dirección": _get("Dirección"),
                "Nacionalidad": _get("Nacionalidad"),
            }

            try:
                snapshot_antes = dict(at)
                self.db.actualizar_atencion_especifica(atencion_id, nuevos)
                ruta_pdf_actualizada = regenerar_pdf_archivado(
                    self.db, atencion_id, mostrar_error=False
                )
                revision_nss_id = self.db.obtener_revision_nss_atencion(atencion_id)
                self._invalidar_cache_ars()

                def _undo_edit_atencion():
                    datos_anteriores = {
                        "Nombre": snapshot_antes.get("nombre", ""),
                        "Sexo": snapshot_antes.get("sexo", "Femenino") or "Femenino",
                        "Fecha": snapshot_antes.get("fecha", ""),
                        "Hora": snapshot_antes.get("hora", ""),
                        "Hoja": snapshot_antes.get("hoja", ""),
                        "Edad_num": int(snapshot_antes.get("edad_num") or 0),
                        "Unidad": snapshot_antes.get("unidad", "Años"),
                        "TipoAtencion": snapshot_antes.get("tipo_atencion", "EMERGENCIA"),
                        "Aseguradora (ARS)": snapshot_antes.get("ars", ""),
                        "NSS": snapshot_antes.get("nss", ""),
                        "Cédula": snapshot_antes.get("cedula", ""),
                        "Teléfono": snapshot_antes.get("telefono", ""),
                        "Dirección": snapshot_antes.get("direccion", ""),
                        "Nacionalidad": snapshot_antes.get("nacionalidad", ""),
                    }
                    self.db.actualizar_atencion_especifica(atencion_id, datos_anteriores)
                    regenerar_pdf_archivado(
                        self.db, atencion_id, mostrar_error=False
                    )
                    turno_cfg_undo = cargar_turno_config()
                    if turno_cfg_undo:
                        reconstruir_excel_turno(self.db, turno_cfg_undo)
                    self._actualizar_resumen_turno_panel()
                    if on_saved:
                        try:
                            on_saved()
                        except Exception:
                            pass
                    messagebox.showinfo("Deshacer", f"Edición de la atención #{atencion_id} revertida.")

                self._push_undo_action(f"edición de atención #{atencion_id}", _undo_edit_atencion)

                self._reconstruir_excel_si_necesario("editar atención", antes=snapshot_antes, despues=nuevos)
                self._invalidar_caches_datos()
                self._refrescar_resumen_en_vivo()
                if on_saved:
                    on_saved()

                self._mostrar_notificacion(f"Atención #{atencion_id} actualizada. Ctrl+Z para deshacer.", on_undo=_undo_edit_atencion, autohide_ms=5000)
                if revision_nss_id:
                    self._mostrar_notificacion(
                        f"Atención #{atencion_id} actualizada. El conflicto NSS "
                        "fue enviado a revisión administrativa sin detener el flujo.",
                        autohide_ms=12000,
                        tipo="warning",
                    )
                mensaje_guardado = "Atención y PDF actualizados correctamente."
                if not ruta_pdf_actualizada:
                    mensaje_guardado = (
                        "Atención actualizada. El PDF quedó pendiente y se "
                        "reconstruirá automáticamente al abrirlo."
                    )
                messagebox.showinfo("Guardado", mensaje_guardado)
                try:
                    win.destroy()
                except Exception:
                    pass
            except PermissionError:
                messagebox.showwarning("Excel abierto", "Cierre el Excel para reconstruir el listado del turno.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la atención:\n{str(e)}")

        tb.Button(footer, text="💾  Guardar cambios", bootstyle=SUCCESS, command=guardar, width=22).pack(side="left", padx=5, ipady=5)
        tb.Button(footer, text="Cerrar", bootstyle=SECONDARY, command=win.destroy, width=12).pack(side="right", padx=5, ipady=5)

    def _abrir_edicion_paciente(self, prefill_identidad=None):
        """
        Ventana independiente y ligera para editar pacientes.
        """
        win = self._crear_toplevel_estable("Editar paciente", "1160x760", "edicion_paciente_win")
        if win is None:
            return

        try:
            win.minsize(1100, 720)
        except Exception:
            pass

        cont = tb.Frame(win, padding=12, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Editar paciente",
            "Busca por NSS, cédula, nombre o teléfono y corrige datos sin generar una hoja nueva.",
            "🖉"
        )

        footer_edicion = tb.Frame(cont, padding=(8, 8), style="Card.TFrame")
        footer_edicion.pack(side="bottom", fill="x", pady=(8, 0))

        patient_card = tb.Frame(cont, padding=10, style="Card.TFrame")
        patient_card.pack(fill="both", expand=True)
        patient_card.columnconfigure(1, weight=1)
        patient_card.columnconfigure(3, weight=1)

        tb.Label(
            patient_card,
            text="Buscar por NSS, cédula, nombre o teléfono:",
            background="#0E1B2B",
            font=("Arial", 10, "bold")
        ).grid(row=0, column=0, sticky="w", padx=5, pady=(0, 6))

        buscar_var = tk.StringVar()
        ent_buscar = tb.Entry(patient_card, textvariable=buscar_var, width=34)
        ent_buscar.grid(row=0, column=1, sticky="ew", padx=5, pady=(0, 6), ipady=3)

        btn_buscar = tb.Button(patient_card, text="🔎  Buscar", bootstyle=PRIMARY)
        btn_buscar.grid(row=0, column=2, sticky="w", padx=5, pady=(0, 6), ipady=3)

        estado_var = tk.StringVar(value="Ingrese un dato del paciente y presione Buscar.")
        tb.Label(
            patient_card,
            textvariable=estado_var,
            style="Muted.TLabel",
            background="#0E1B2B"
        ).grid(row=1, column=0, columnspan=4, sticky="w", padx=5, pady=(0, 6))

        resultados_tree = ttk.Treeview(
            patient_card,
            columns=("id", "fecha", "nombre", "nss", "cedula", "telefono"),
            show="headings",
            height=4,
            style="Modern.Treeview"
        )
        resultados_tree.grid(row=2, column=0, columnspan=4, sticky="ew", padx=5, pady=(0, 8))
        for col, title, width in [
            ("id", "ID", 55),
            ("fecha", "Fecha", 90),
            ("nombre", "Nombre", 310),
            ("nss", "NSS", 140),
            ("cedula", "Cédula", 130),
            ("telefono", "Teléfono", 120),
        ]:
            resultados_tree.heading(col, text=title)
            resultados_tree.column(col, width=width, anchor="center" if col != "nombre" else "w")

        ttk.Separator(patient_card, orient="horizontal").grid(row=3, column=0, columnspan=4, sticky="ew", pady=(2, 8))

        campos = {}
        ayudas = {
            "Nombre": "Nombre completo del paciente",
            "Cédula": "11 dígitos",
            "Teléfono": "10 dígitos",
            "NSS": "Número de seguro o SIN SEGURO",
            "Dirección": "Dirección del paciente",
            "Nacionalidad": "Nacionalidad",
            "Aseguradora (ARS)": "Ej.: SUB, HUMANO, MAPFRE, AVANZADA…",
        }

        filas = [
            ("Nombre", "Cédula"),
            ("Teléfono", "NSS"),
            ("Aseguradora (ARS)", "Nacionalidad"),
            ("Dirección", None),
        ]

        row_base = 4
        for i, (campo_izq, campo_der) in enumerate(filas):
            r = row_base + i

            tb.Label(
                patient_card,
                text=f"{campo_izq}:",
                font=("Arial", 10, "bold"),
                background="#0E1B2B"
            ).grid(row=r, column=0, sticky="w", padx=(5, 4), pady=5)

            ent_izq = tb.Entry(patient_card)
            ent_izq.grid(row=r, column=1, sticky="ew", padx=(0, 12), pady=5, ipady=4)
            try:
                ent_izq.insert(0, ayudas.get(campo_izq, ""))
                ent_izq.delete(0, tk.END)
            except Exception:
                pass
            campos[campo_izq] = ent_izq

            if campo_der:
                tb.Label(
                    patient_card,
                    text=f"{campo_der}:",
                    font=("Arial", 10, "bold"),
                    background="#0E1B2B"
                ).grid(row=r, column=2, sticky="w", padx=(5, 4), pady=5)

                ent_der = tb.Entry(patient_card)
                ent_der.grid(row=r, column=3, sticky="ew", padx=(0, 5), pady=5, ipady=4)
                campos[campo_der] = ent_der
            else:
                ent_izq.grid(row=r, column=1, columnspan=3, sticky="ew", padx=(0, 5), pady=5, ipady=4)

        ayuda_var = tk.StringVar(
            value="Consejo: si un paciente estaba SIN SEGURO y luego aparece vigente, coloque NSS y ARS aquí; quedará actualizado como asegurado."
        )
        tb.Label(
            patient_card,
            textvariable=ayuda_var,
            style="Muted.TLabel",
            background="#0E1B2B",
            wraplength=980,
            justify="left"
        ).grid(row=row_base + len(filas), column=0, columnspan=4, sticky="w", padx=5, pady=(8, 0))

        original_identidad = {"valor": "", "paciente_id": None}

        def _llenar_formulario_paciente(data, ident):
            original_identidad["valor"] = ident
            original_identidad["paciente_id"] = data.get("paciente_id")
            valores = {
                "Nombre": data.get("nombre", ""),
                "Cédula": data.get("cedula", ""),
                "Teléfono": data.get("telefono", ""),
                "NSS": data.get("nss", ""),
                "Dirección": data.get("direccion", ""),
                "Nacionalidad": data.get("nacionalidad", ""),
                "Aseguradora (ARS)": seguro_para_mostrar(normalizar_seguro(data.get("ars", ""), data.get("nss", ""))),
            }
            for k, ent in campos.items():
                ent.delete(0, tk.END)
                ent.insert(0, valores.get(k, ""))

            estado_var.set(f"Registro cargado. Última atención ID: {data.get('id', 'N/A')}")
            try:
                btn_eliminar.configure(state=tk.NORMAL if original_identidad["paciente_id"] else tk.DISABLED)
            except (NameError, tk.TclError):
                pass
            self.set_status("Paciente cargado para edición", "ok")

        def cargar_paciente():
            ident = buscar_var.get().strip()
            if not ident:
                messagebox.showwarning("Buscar", "Ingrese NSS, cédula, nombre o teléfono para buscar.")
                return

            for i in resultados_tree.get_children():
                resultados_tree.delete(i)
            resultados_tree.insert("", "end", values=("", "", "Cargando datos…", "", "", ""))
            estado_var.set("Cargando datos del paciente…")
            win.update_idletasks()

            def _buscar():
                try:
                    for i in resultados_tree.get_children():
                        resultados_tree.delete(i)

                    resultados = self.db.buscar_pacientes_avanzado(ident)
                    if not resultados:
                        data = self.db.buscar_paciente_para_edicion(ident)
                        if not data:
                            estado_var.set("No se encontraron coincidencias.")
                            self.set_status("No se encontró paciente para editar", "warning")
                            return
                        resultados = [data]

                    for r in resultados:
                        resultados_tree.insert(
                            "",
                            "end",
                            values=(
                                r.get("id", ""),
                                r.get("fecha", ""),
                                r.get("nombre", ""),
                                r.get("nss", ""),
                                r.get("cedula", ""),
                                r.get("telefono", "")
                            )
                        )

                    primero = resultados[0]
                    ident_carga = (
                        f"A:{primero.get('id')}" if primero.get("id")
                        else str(primero.get("nss") or primero.get("cedula") or ident)
                    )
                    _llenar_formulario_paciente(primero, ident_carga)
                    estado_var.set(f"{len(resultados)} coincidencia(s). Seleccione una fila para cargar otra.")
                except Exception as e:
                    estado_var.set("Error al buscar paciente.")
                    messagebox.showerror("Error", f"No se pudo buscar el paciente:\n{str(e)}")

            try:
                win.after(40, _buscar)
            except Exception:
                _buscar()

        def seleccionar_resultado_paciente(_=None):
            sel = resultados_tree.selection()
            if not sel:
                return
            vals = resultados_tree.item(sel[0], "values")
            ident = f"A:{vals[0]}" if vals[0] else (vals[3] or vals[4])
            data = self.db.buscar_paciente_para_edicion(str(ident))
            if data:
                _llenar_formulario_paciente(data, str(ident))

        resultados_tree.bind("<<TreeviewSelect>>", seleccionar_resultado_paciente)

        def guardar_edicion():
            if not original_identidad["valor"]:
                messagebox.showwarning("Guardar", "Primero busque y cargue un paciente.")
                return

            nuevos = {k: ent.get().strip() for k, ent in campos.items()}

            snapshot_paciente = self.db.buscar_paciente_para_edicion(original_identidad["valor"])
            if not snapshot_paciente:
                messagebox.showwarning("Guardar", "No se pudo cargar los datos actuales del paciente.")
                return

            def _get_nombre(d):
                return (d.get("nombre", d.get("Nombre", "")) or "").strip()

            def _get_cedula(d):
                return (d.get("cedula", d.get("Cédula", "")) or "").strip().replace("-", "")

            def _get_telefono(d):
                return (d.get("telefono", d.get("Teléfono", "")) or "").strip().replace("-", "")

            def _get_nss(d):
                return re.sub(r"\D", "", (d.get("nss", d.get("NSS", "")) or "").strip().upper())

            def _get_direccion(d):
                return (d.get("direccion", d.get("Dirección", "")) or "").strip()

            def _get_nacionalidad(d):
                return (d.get("nacionalidad", d.get("Nacionalidad", "")) or "").strip()

            def _get_ars(d, nss_ref=""):
                return normalizar_seguro(
                    d.get("ars", d.get("ARS", d.get("Aseguradora (ARS)", "")) or ""),
                    nss_ref or ""
                )

            nuevo_nombre = _get_nombre(nuevos)
            viejo_nombre = _get_nombre(snapshot_paciente)

            nueva_cedula = _get_cedula(nuevos)
            vieja_cedula = _get_cedula(snapshot_paciente)

            nuevo_telefono = _get_telefono(nuevos)
            viejo_telefono = _get_telefono(snapshot_paciente)

            nuevo_nss = _get_nss(nuevos)
            viejo_nss = _get_nss(snapshot_paciente)

            nueva_direccion = _get_direccion(nuevos)
            vieja_direccion = _get_direccion(snapshot_paciente)

            nueva_nacionalidad = _get_nacionalidad(nuevos)
            vieja_nacionalidad = _get_nacionalidad(snapshot_paciente)

            nueva_ars = _get_ars(nuevos, nuevo_nss)
            vieja_ars = _get_ars(snapshot_paciente, viejo_nss)

            hay_cambios = (
                nuevo_nombre != viejo_nombre or
                nueva_cedula != vieja_cedula or
                nuevo_telefono != viejo_telefono or
                nuevo_nss != viejo_nss or
                nueva_ars != vieja_ars or
                nueva_direccion != vieja_direccion or
                nueva_nacionalidad != vieja_nacionalidad
            )

            if not hay_cambios:
                messagebox.showinfo("Guardar", "No se detectaron cambios. No se guardó nada.")
                return

            if not nuevos.get("Nombre"):
                messagebox.showerror("Validación", "El nombre no puede quedar vacío.")
                campos["Nombre"].focus_set()
                return

            ced = nuevos.get("Cédula", "").replace("-", "")
            tel = nuevos.get("Teléfono", "").replace("-", "")
            nss = nuevos.get("NSS", "").upper()
            ars = nuevos.get("Aseguradora (ARS)", "")

            if ced and not is_valid_cedula_key(ced):
                messagebox.showerror("Validación", "La cédula debe tener 11 dígitos o dejarse vacía.")
                campos["Cédula"].focus_set()
                return

            if tel and not (tel.isdigit() and len(tel) == 10):
                messagebox.showerror("Validación", "El teléfono debe tener 10 dígitos o dejarse vacío.")
                campos["Teléfono"].focus_set()
                return

            if nss and not is_valid_nss_key(nss) and nss not in ["SIN SEGURO", "NO", "N/S", "NS", "N\\S"]:
                messagebox.showerror(
                    "Validación",
                    "El NSS debe ser numérico. Si no tiene seguro, escriba SIN SEGURO o déjelo vacío."
                )
                campos["NSS"].focus_set()
                return

            if ars_es_corta_invalida(ars):
                messagebox.showerror(
                    "Validación",
                    "La ARS es demasiado corta o no reconocida. No se guardan ARS accidentales menores de 4 letras."
                )
                campos["Aseguradora (ARS)"].focus_set()
                return

            try:
                alcance = messagebox.askyesnocancel(
                    "Alcance de la modificación",
                    "¿Desea actualizar también la ficha actual del paciente?\n\n"
                    "Sí: esta atención y la ficha actual\n"
                    "No: solamente esta atención\n"
                    "Cancelar: no guardar",
                    parent=win,
                )
                if alcance is None:
                    return
                actualizar_ficha = alcance is True
                snapshot_paciente = self.db.buscar_paciente_para_edicion(original_identidad["valor"])
                at_count, pac_count = self.db.actualizar_datos_paciente_por_identidad(
                    original_identidad["valor"],
                    nuevos,
                    actualizar_ficha=actualizar_ficha,
                )

                def _undo_edit_paciente():
                    if not snapshot_paciente:
                        messagebox.showwarning("Deshacer", "No se encontró una copia anterior para revertir.")
                        return
                    anteriores = {
                        "Nombre": snapshot_paciente.get("nombre", ""),
                        "Cédula": snapshot_paciente.get("cedula", ""),
                        "Teléfono": snapshot_paciente.get("telefono", ""),
                        "NSS": snapshot_paciente.get("nss", ""),
                        "Dirección": snapshot_paciente.get("direccion", ""),
                        "Nacionalidad": snapshot_paciente.get("nacionalidad", ""),
                        "Aseguradora (ARS)": snapshot_paciente.get("ars", ""),
                    }
                    ident_undo = (
                        snapshot_paciente.get("nss")
                        or snapshot_paciente.get("cedula")
                        or f"A:{snapshot_paciente.get('id')}"
                    )
                    self.db.actualizar_datos_paciente_por_identidad(
                        str(ident_undo),
                        anteriores,
                        actualizar_ficha=actualizar_ficha,
                    )
                    self._reconstruir_excel_si_necesario("deshacer edición paciente", antes=nuevos, despues=anteriores, forzar=False)
                    self._invalidar_caches_datos()
                    self._refrescar_resumen_en_vivo()
                    messagebox.showinfo("Deshacer", "Edición del paciente revertida.")

                self._push_undo_action("edición de paciente", _undo_edit_paciente)

                self._reconstruir_excel_si_necesario("editar paciente", antes=snapshot_paciente or {}, despues=nuevos, forzar=False)
                self._invalidar_caches_datos()
                self._ars_catalogo = self._obtener_catalogo_ars()
                self._refrescar_resumen_en_vivo()

                estado_var.set(f"Guardado: {at_count} atención(es) y {pac_count} ficha(s) actualizada(s).")
                self.set_status("Datos del paciente actualizados", "ok")
                self._mostrar_notificacion("Datos del paciente actualizados. Ctrl+Z para deshacer.", on_undo=_undo_edit_paciente, autohide_ms=5000)
                messagebox.showinfo("Guardado", "Datos actualizados correctamente. No se generó una hoja nueva.")
            except PermissionError:
                messagebox.showwarning(
                    "Excel abierto",
                    "El listado de Excel está abierto.\n\nCierre el archivo y vuelva a intentar guardar."
                )
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la edición:\n{str(e)}")

        btn_buscar.configure(command=cargar_paciente)
        ent_buscar.bind("<Return>", lambda e: cargar_paciente())

        try:
            campos["Cédula"].bind("<KeyRelease>", lambda e: self.limitar_caracteres(campos["Cédula"], 11))
            campos["Teléfono"].bind("<KeyRelease>", lambda e: self.limitar_caracteres(campos["Teléfono"], 10))
        except Exception:
            pass

        def limpiar_campos_edicion():
            campos["Nombre"].delete(0, tk.END)
            campos["Cédula"].delete(0, tk.END)
            campos["Teléfono"].delete(0, tk.END)
            campos["NSS"].delete(0, tk.END)
            campos["Dirección"].delete(0, tk.END)
            campos["Nacionalidad"].delete(0, tk.END)
            campos["Aseguradora (ARS)"].delete(0, tk.END)
            buscar_var.set("")
            for i in resultados_tree.get_children():
                resultados_tree.delete(i)
            original_identidad.update({"valor": "", "paciente_id": None})
            try:
                btn_eliminar.configure(state=tk.DISABLED)
            except (NameError, tk.TclError):
                pass
            estado_var.set("Campos limpiados. Busque un paciente para cargar datos.")
            self.set_status("Campos limpiados en edición paciente", "ok")

        def eliminar_paciente_total():
            paciente_id = original_identidad.get("paciente_id")
            if not paciente_id:
                messagebox.showwarning("Eliminar paciente", "Primero busque y cargue el paciente que desea eliminar.", parent=win)
                return

            try:
                resumen = self.db.previsualizar_eliminacion_paciente(paciente_id)
            except Exception as e:
                APP_LOG.exception("No se pudo preparar la eliminación total del paciente")
                messagebox.showerror("Eliminar paciente", f"No se pudo consultar el paciente:\n{str(e)}", parent=win)
                return

            if not resumen:
                messagebox.showwarning("Eliminar paciente", "El paciente ya no existe en la base de datos.", parent=win)
                return
            if not resumen.get("seguro"):
                messagebox.showwarning(
                    "Identidad insuficiente",
                    resumen.get("motivo", "No se puede identificar al paciente con seguridad."),
                    parent=win,
                )
                return

            paciente = resumen["paciente"]
            nombre = (paciente.get("nombre") or "SIN NOMBRE").upper()
            total_atenciones = len(resumen["atenciones"])
            total_fichas = int(resumen.get("fichas", 0) or 0)
            total_auditorias = int(resumen.get("auditorias", 0) or 0)
            advertencia = (
                "Esta eliminación es permanente y está destinada a retirar pacientes de prueba.\n\n"
                f"Paciente: {nombre}\n"
                f"NSS: {paciente.get('nss') or 'N/A'}\n"
                f"Cédula: {paciente.get('cedula') or 'N/A'}\n"
                f"Atenciones que se eliminarán: {total_atenciones}\n"
                f"Fichas que se eliminarán: {total_fichas}\n"
                f"Registros clínicos de auditoría que se retirarán: {total_auditorias}\n"
                "Se conservará un evento administrativo anonimizado y un respaldo verificable.\n\n"
                "Los ID restantes no serán renumerados. ¿Desea continuar?"
            )
            if not messagebox.askyesno("Eliminar paciente por completo", advertencia, icon="warning", parent=win):
                return

            if not messagebox.askyesno(
                "Confirmar dato de prueba",
                "Confirme que esta ficha pertenece exclusivamente a un paciente de prueba.\n\n"
                "Una atención clínica real debe anularse, no purgarse.",
                icon="warning",
                parent=win,
            ):
                return

            motivo = simpledialog.askstring(
                "Motivo obligatorio",
                "Indique por qué se confirma que esta ficha es de prueba:",
                parent=win,
            )
            if motivo is None or len(motivo.strip()) < 8:
                messagebox.showwarning(
                    "Motivo requerido",
                    "La purga requiere un motivo de al menos 8 caracteres.",
                    parent=win,
                )
                return

            actor = self._solicitar_autorizacion_admin(
                "PURGAR_PACIENTE_DE_PRUEBA", parent=win, force=True
            )
            if not actor:
                return

            confirmacion = simpledialog.askstring(
                "Confirmación final",
                f"Para confirmar, escriba exactamente ELIMINAR {paciente_id}:",
                parent=win,
            )
            if (confirmacion or "").strip().upper() != f"ELIMINAR {paciente_id}":
                messagebox.showinfo("Cancelado", "No se eliminó ningún dato.", parent=win)
                return

            afecta_turno = any(
                self._registro_esta_en_turno_actual(atencion)
                for atencion in resumen["atenciones"]
            )
            try:
                resultado = self.db.eliminar_paciente_completo(
                    paciente_id,
                    motivo.strip(),
                    actor,
                    confirmado_prueba=True,
                )
                if not resultado or not resultado.get("seguro"):
                    messagebox.showwarning("Eliminar paciente", "No se pudo completar la eliminación.", parent=win)
                    return
                self.security.audit(
                    "PATIENT_PURGED",
                    actor=actor,
                    success=not bool(resultado.get("documentos_pendientes")),
                    detail=str(resultado.get("purga_event_hash") or ""),
                )

                aviso_excel = ""
                if afecta_turno:
                    turno_cfg = cargar_turno_config()
                    if turno_cfg:
                        try:
                            reconstruir_excel_turno(self.db, turno_cfg)
                        except Exception:
                            APP_LOG.exception("El paciente se eliminó, pero no se pudo reconstruir el Excel")
                            aviso_excel = (
                                "\n\nEl paciente sí fue eliminado, pero el Excel no pudo actualizarse. "
                                "Cierre el archivo y use la reconstrucción del turno."
                            )

                self._invalidar_caches_datos()
                self._refrescar_resumen_en_vivo()
                limpiar_campos_edicion()
                pendientes = resultado.get("documentos_pendientes") or []
                mensaje = (
                    f"Se eliminó la ficha de prueba seleccionada.\n\n"
                    f"Atenciones: {resultado.get('atenciones_eliminadas', 0)}\n"
                    f"Fichas: {resultado.get('fichas_eliminadas', 0)}\n"
                    f"Auditorías redactadas: {resultado.get('auditorias_redactadas', 0)}\n"
                    f"Documentos eliminados: {resultado.get('documentos_eliminados', 0)}\n\n"
                    f"Los demás ID permanecen sin cambios.{aviso_excel}"
                )
                if pendientes:
                    messagebox.showwarning(
                        "Purga incompleta",
                        mensaje
                        + "\n\nUno o más documentos quedaron aislados en cuarentena. "
                        "Revise el registro técnico antes de considerar completada la purga.",
                        parent=win,
                    )
                else:
                    messagebox.showinfo("Paciente eliminado", mensaje, parent=win)
            except PermissionError:
                messagebox.showwarning("Eliminar paciente", "No se pudo acceder a la base de datos.", parent=win)
            except Exception as e:
                messagebox.showerror("Eliminar paciente", f"No se pudo eliminar el paciente:\n{str(e)}", parent=win)

        tb.Button(footer_edicion, text="🧹  Limpiar campos", bootstyle=SECONDARY, command=limpiar_campos_edicion, width=18).pack(side="left", padx=5, ipady=5)
        tb.Button(footer_edicion, text="💾  Guardar cambios", bootstyle=SUCCESS, command=guardar_edicion, width=22).pack(side="left", padx=5, ipady=5)
        btn_eliminar = tb.Button(
            footer_edicion,
            text="Eliminar paciente de prueba",
            bootstyle=DANGER,
            command=eliminar_paciente_total,
            width=25,
            state=tk.DISABLED,
        )
        btn_eliminar.pack(side="left", padx=5, ipady=5)
        tb.Button(footer_edicion, text="Cerrar", bootstyle=SECONDARY, command=win.destroy, width=12).pack(side="right", padx=5, ipady=5)

        if prefill_identidad:
            try:
                buscar_var.set(str(prefill_identidad).strip())
                estado_var.set("Cargando datos…")
                win.after(160, cargar_paciente)
            except Exception:
                pass

        try:
            ent_buscar.focus_set()
        except Exception:
            pass

    def _abrir_configuracion_interna(self, prefill_identidad=None):
        if prefill_identidad:
            self._abrir_edicion_paciente(prefill_identidad=prefill_identidad)
            return

        if not self._solicitar_autorizacion_admin("ABRIR_CONFIGURACION", parent=self.root):
            return

        win = self._crear_toplevel_estable("Configuración interna", "1160x760", "configuracion_interna_win")
        if win is None:
            return

        try:
            self.configuracion_interna_win = win
        except Exception:
            pass

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Configuración interna",
            "Administra ARS, usuarios de turno, respaldos, revisión NSS y preferencias.",
            "⚙"
        )

        notebook = ttk.Notebook(cont)
        notebook.pack(fill="both", expand=True)

        # ---------------- TAB ARS ----------------
        tab_ars = tb.Frame(notebook, padding=12, style="Card.TFrame")
        notebook.add(tab_ars, text="Administrar ARS")
        tab_ars.columnconfigure(0, weight=1)
        tab_ars.columnconfigure(1, weight=1)

        info_ars = tk.StringVar(value="Cargando datos de ARS…")
        tb.Label(tab_ars, textvariable=info_ars, style="Muted.TLabel", background="#0E1B2B").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        tree_ars = ttk.Treeview(tab_ars, columns=("ars", "cantidad"), show="headings", height=14, style="Modern.Treeview")
        tree_ars.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 8))
        tree_ars.heading("ars", text="ARS / Seguro")
        tree_ars.heading("cantidad", text="Cantidad")
        tree_ars.column("ars", width=420, anchor="w")
        tree_ars.column("cantidad", width=120, anchor="center")
        tab_ars.rowconfigure(1, weight=1)

        tb.Label(tab_ars, text="ARS actual:", background="#0E1B2B").grid(row=2, column=0, sticky="w", padx=5, pady=(8, 2))
        ars_actual_var = tk.StringVar()
        ent_actual = tb.Entry(tab_ars, textvariable=ars_actual_var)
        ent_actual.grid(row=3, column=0, sticky="ew", padx=5, pady=(0, 6), ipady=4)

        tb.Label(tab_ars, text="Reemplazar por:", background="#0E1B2B").grid(row=2, column=1, sticky="w", padx=5, pady=(8, 2))
        ars_nueva_var = tk.StringVar()
        ent_nueva = tb.Entry(tab_ars, textvariable=ars_nueva_var)
        ent_nueva.grid(row=3, column=1, sticky="ew", padx=5, pady=(0, 6), ipady=4)

        def cargar_ars():
            for i in tree_ars.get_children():
                tree_ars.delete(i)
            tree_ars.insert("", "end", values=("Cargando datos…", ""))
            info_ars.set("Cargando datos de ARS…")
            win.update_idletasks()

            def _cargar():
                rows = list(self.db.listar_ars_conteo())
                catalogo = self._obtener_catalogo_ars()
                return rows, catalogo

            def _finalizar(resultado):
                if not win.winfo_exists():
                    return
                rows, catalogo = resultado
                tree_ars.delete(*tree_ars.get_children())
                for ars, cantidad in rows:
                    tree_ars.insert("", "end", values=(seguro_para_mostrar(ars), cantidad))
                self._ars_catalogo = catalogo
                self._invalidar_cache_ars()
                info_ars.set("Seleccione una ARS para corregirla o use las acciones de limpieza.")

            def _error(err):
                info_ars.set("No se pudieron cargar las ARS.")
                messagebox.showerror("ARS", f"No se pudo cargar la lista de ARS:\n{str(err)}")

            self._ejecutar_en_segundo_plano(
                "Cargando ARS…",
                _cargar,
                al_terminar=_finalizar,
                al_error=_error,
            )

        def seleccionar_ars(_evt=None):
            sel = tree_ars.selection()
            if not sel:
                return
            vals = tree_ars.item(sel[0], "values")
            ars_actual_var.set(vals[0])

        def reemplazar_ars():
            actual = ars_actual_var.get().strip()
            nueva = ars_nueva_var.get().strip()
            if not actual or not nueva:
                messagebox.showwarning("ARS", "Indique la ARS actual y la nueva ARS.")
                return
            if ars_es_corta_invalida(nueva):
                messagebox.showerror("ARS", "La nueva ARS es demasiado corta o no reconocida. Escriba al menos 4 letras o una referencia válida.")
                ent_nueva.focus_set()
                return
            if not messagebox.askyesno(
                "Confirmación",
                f"¿Actualizar {actual} a {nueva} en las fichas actuales?\n\n"
                "Las atenciones históricas conservarán el valor registrado en su momento.",
            ):
                return
            total = self.db.reemplazar_ars_global(actual, nueva)
            self.security.audit(
                "ARS_CURRENT_RECORDS_REPLACED",
                actor=self._admin_authorized_actor or self._actor_actual(),
                success=True,
                detail=f"{actual}->{nueva}; fichas={total}",
            )
            cargar_ars()
            self.set_status(f"ARS actualizadas: {total}", "ok")
            messagebox.showinfo("ARS", f"Registros actualizados: {total}")

        def normalizar_todo():
            total = self.db.normalizar_todas_ars()
            self.security.audit(
                "ARS_CURRENT_RECORDS_NORMALIZED",
                actor=self._admin_authorized_actor or self._actor_actual(),
                success=True,
                detail=f"fichas={total}",
            )
            cargar_ars()
            self.set_status(f"ARS normalizadas: {total}", "ok")
            messagebox.showinfo("ARS", f"ARS normalizadas: {total}")

        def limpiar_cortas():
            total = self.db.limpiar_ars_cortas_invalidas()
            self.security.audit(
                "ARS_INVALID_CURRENT_RECORDS_CLEANED",
                actor=self._admin_authorized_actor or self._actor_actual(),
                success=True,
                detail=f"fichas={total}",
            )
            cargar_ars()
            self.set_status(f"ARS inválidas limpiadas: {total}", "ok")
            messagebox.showinfo("ARS", f"ARS inválidas convertidas a SIN SEGURO: {total}")

        tree_ars.bind("<<TreeviewSelect>>", seleccionar_ars)

        menu_ars = tk.Menu(win, tearoff=0)
        menu_ars.add_command(label="✏ Cargar para reemplazar", command=seleccionar_ars)
        def eliminar_ars_sel():
            sel = tree_ars.selection()
            if not sel:
                messagebox.showwarning("ARS", "Seleccione una ARS para eliminar."); return
            actual = tree_ars.item(sel[0], "values")[0]
            if not messagebox.askyesno(
                "Eliminar ARS",
                f"¿Convertir las fichas actuales de '{actual}' a SIN SEGURO?\n\n"
                "Las atenciones históricas no se modificarán.",
            ):
                return
            total = self.db.eliminar_ars_global(actual)
            self.security.audit(
                "ARS_CURRENT_RECORDS_REMOVED",
                actor=self._admin_authorized_actor or self._actor_actual(),
                success=True,
                detail=f"ars={actual}; fichas={total}",
            )
            cargar_ars(); self._actualizar_resumen_turno_panel()
            messagebox.showinfo("ARS", f"Registros convertidos a SIN SEGURO: {total}")
        menu_ars.add_command(label="🗑 Eliminar ARS (pasar a SIN SEGURO)", command=eliminar_ars_sel)
        def mostrar_menu_ars(event):
            item = tree_ars.identify_row(event.y)
            if item:
                tree_ars.selection_set(item); tree_ars.focus(item); menu_ars.tk_popup(event.x_root, event.y_root)
        tree_ars.bind("<Button-3>", mostrar_menu_ars)

        acciones_ars = tb.Frame(tab_ars, padding=(0, 10), style="Card.TFrame")
        acciones_ars.grid(row=4, column=0, columnspan=2, sticky="ew")
        tb.Button(acciones_ars, text="↻  Actualizar lista", bootstyle=SECONDARY, command=cargar_ars, width=18).pack(side="left", padx=5, ipady=5)
        tb.Button(acciones_ars, text="✓  Normalizar ARS", bootstyle=INFO, command=normalizar_todo, width=18).pack(side="left", padx=5, ipady=5)
        tb.Button(acciones_ars, text="⌫  Limpiar cortas", bootstyle=WARNING, command=limpiar_cortas, width=18).pack(side="left", padx=5, ipady=5)
        tb.Button(acciones_ars, text="✓  Reemplazar", bootstyle=SUCCESS, command=reemplazar_ars, width=16).pack(side="left", padx=5, ipady=5)
        tb.Button(acciones_ars, text="🗑️  Eliminar ARS", bootstyle=DANGER, command=eliminar_ars_sel, width=16).pack(side="left", padx=5, ipady=5)
        tb.Button(acciones_ars, text="Cerrar", bootstyle=SECONDARY, command=win.destroy, width=12).pack(side="right", padx=5, ipady=5)

        try:
            win.after(20, cargar_ars)
        except Exception:
            cargar_ars()

        # ---------------- TAB CATÁLOGO ARS ----------------
        tab_catalogo = tb.Frame(notebook, padding=12, style="Card.TFrame")
        notebook.add(tab_catalogo, text="Catálogo ARS")
        tab_catalogo.columnconfigure(0, weight=1); tab_catalogo.columnconfigure(1, weight=1); tab_catalogo.rowconfigure(1, weight=1)
        tb.Label(tab_catalogo, text="ARS oficial", background="#0E1B2B").grid(row=0, column=0, sticky="w", padx=5, pady=4)
        tb.Label(tab_catalogo, text="Alias aceptados (separados por coma)", background="#0E1B2B").grid(row=0, column=1, sticky="w", padx=5, pady=4)
        tree_cat = ttk.Treeview(tab_catalogo, columns=("oficial", "alias"), show="headings", height=12, style="Modern.Treeview")
        tree_cat.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 8))
        tree_cat.heading("oficial", text="ARS oficial"); tree_cat.heading("alias", text="Alias aceptados")
        tree_cat.column("oficial", width=300, anchor="w"); tree_cat.column("alias", width=520, anchor="w")
        oficial_var = tk.StringVar(); alias_var = tk.StringVar()
        ent_oficial = tb.Entry(tab_catalogo, textvariable=oficial_var); ent_alias = tb.Entry(tab_catalogo, textvariable=alias_var)
        ent_oficial.grid(row=2, column=0, sticky="ew", padx=5, pady=4, ipady=4); ent_alias.grid(row=2, column=1, sticky="ew", padx=5, pady=4, ipady=4)
        catalogo_data = {k: list(v) for k, v in cargar_catalogo_ars().items()}
        def cargar_catalogo_tree():
            tree_cat.delete(*tree_cat.get_children())
            tree_cat.insert("", "end", values=("Cargando catálogo…", ""))

            def _cargar():
                return [(k, catalogo_data[k]) for k in sorted(catalogo_data)]

            def _finalizar(resultado):
                if not win.winfo_exists():
                    return
                tree_cat.delete(*tree_cat.get_children())
                for k, aliases in resultado:
                    tree_cat.insert("", "end", values=(k, ", ".join(aliases)))

            self._ejecutar_en_segundo_plano(
                "Cargando catálogo ARS…",
                _cargar,
                al_terminar=_finalizar,
            )
        def seleccionar_catalogo(_=None):
            sel = tree_cat.selection()
            if not sel: return
            vals = tree_cat.item(sel[0], "values"); oficial_var.set(vals[0]); alias_var.set(vals[1])
        def guardar_alias():
            k = _limpiar_texto_seguro(oficial_var.get())
            if not k: messagebox.showwarning("Catálogo", "Escriba una ARS oficial."); return
            aliases = [a.strip() for a in alias_var.get().split(",") if a.strip()]
            catalogo_data[k] = aliases; guardar_catalogo_ars(catalogo_data); cargar_catalogo_tree(); self._ars_catalogo = self._obtener_catalogo_ars()
            messagebox.showinfo("Catálogo", "Alias guardados correctamente.")
        def eliminar_alias():
            k = _limpiar_texto_seguro(oficial_var.get())
            if k in catalogo_data and messagebox.askyesno("Catálogo", f"¿Eliminar {k} del catálogo editable?"):
                catalogo_data.pop(k, None); guardar_catalogo_ars(catalogo_data); oficial_var.set(""); alias_var.set(""); cargar_catalogo_tree()
        tree_cat.bind("<<TreeviewSelect>>", seleccionar_catalogo)
        barra_cat = tb.Frame(tab_catalogo, style="Card.TFrame"); barra_cat.grid(row=3, column=0, columnspan=2, sticky="ew", pady=8)
        tb.Button(barra_cat, text="💾  Guardar alias", bootstyle=SUCCESS, command=guardar_alias).pack(side="left", padx=5, ipady=5)
        tb.Button(barra_cat, text="🗑️  Eliminar alias", bootstyle=DANGER, command=eliminar_alias).pack(side="left", padx=5, ipady=5)
        tb.Button(barra_cat, text="↻  Recargar", bootstyle=SECONDARY, command=cargar_catalogo_tree).pack(side="left", padx=5, ipady=5)
        try:
            win.after(20, cargar_catalogo_tree)
        except Exception:
            cargar_catalogo_tree()


        # ---------------- TAB FORMATO NSS PDF ----------------
        tab_nss = tb.Frame(notebook, padding=12, style="Card.TFrame")
        notebook.add(tab_nss, text="Formato NSS PDF")
        tab_nss.columnconfigure(0, weight=1)
        tab_nss.columnconfigure(1, weight=1)
        tab_nss.rowconfigure(1, weight=1)

        info_nss_var = tk.StringVar(
            value="Elija una ARS y escriba un ejemplo de cómo debe verse el NSS en el PDF. Ej.: 00896-00258-03."
        )
        tb.Label(tab_nss, textvariable=info_nss_var, style="Muted.TLabel", background="#0E1B2B").grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 8)
        )

        tree_nss_fmt = ttk.Treeview(
            tab_nss,
            columns=("ars", "ejemplo", "patron"),
            show="headings",
            height=12,
            style="Modern.Treeview"
        )
        tree_nss_fmt.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 8))
        tree_nss_fmt.heading("ars", text="ARS")
        tree_nss_fmt.heading("ejemplo", text="Ejemplo en PDF")
        tree_nss_fmt.heading("patron", text="Patrón automático")
        tree_nss_fmt.column("ars", width=260, anchor="w")
        tree_nss_fmt.column("ejemplo", width=300, anchor="w")
        tree_nss_fmt.column("patron", width=150, anchor="center")

        nss_fmt_data = cargar_formatos_nss_ars()
        ars_disponibles = sorted(set(self._obtener_ars_cache(forzar=True)) | set(nss_fmt_data.keys()) | {"RENACER"})

        nss_ars_var = tk.StringVar()
        nss_ejemplo_var = tk.StringVar(value="00896-00258-03")
        nss_patron_detectado_var = tk.StringVar(value="")
        nss_prueba_sin_guiones_var = tk.StringVar(value="008960025803")

        tb.Label(tab_nss, text="ARS:", background="#0E1B2B").grid(row=2, column=0, sticky="w", padx=5, pady=(8, 2))
        combo_nss_ars = tb.Combobox(tab_nss, textvariable=nss_ars_var, values=ars_disponibles, width=34)
        combo_nss_ars.grid(row=3, column=0, sticky="ew", padx=5, pady=(0, 6), ipady=4)

        tb.Label(tab_nss, text="Ejemplo de NSS con guiones para esa ARS:", background="#0E1B2B").grid(row=2, column=1, sticky="w", padx=5, pady=(8, 2))
        ent_nss_ejemplo = tb.Entry(tab_nss, textvariable=nss_ejemplo_var)
        ent_nss_ejemplo.grid(row=3, column=1, sticky="ew", padx=5, pady=(0, 6), ipady=4)

        tb.Label(tab_nss, text="NSS de prueba sin guiones:", background="#0E1B2B").grid(row=4, column=0, sticky="w", padx=5, pady=(8, 2))
        ent_nss_prueba = tb.Entry(tab_nss, textvariable=nss_prueba_sin_guiones_var)
        ent_nss_prueba.grid(row=5, column=0, sticky="ew", padx=5, pady=(0, 6), ipady=4)

        tb.Label(tab_nss, textvariable=nss_patron_detectado_var, style="Muted.TLabel", background="#0E1B2B").grid(
            row=5, column=1, sticky="w", padx=5, pady=(0, 6)
        )

        def _ejemplo_desde_patron(patron):
            base = re.sub(r"\D", "", nss_prueba_sin_guiones_var.get().strip()) or "008960025803"
            return aplicar_patron_nss(base, patron)

        def refrescar_formatos_nss():
            tree_nss_fmt.delete(*tree_nss_fmt.get_children())
            ars_todas = sorted(set(ars_disponibles) | set(nss_fmt_data.keys()))
            combo_nss_ars.configure(values=ars_todas)
            for ars_key in ars_todas:
                patron = nss_fmt_data.get(ars_key, "")
                ejemplo = _ejemplo_desde_patron(patron) if patron else "Sin formato"
                tree_nss_fmt.insert("", "end", values=(ars_key, ejemplo, patron or ""))
            probar_ejemplo_nss()

        def seleccionar_formato_nss(_=None):
            sel = tree_nss_fmt.selection()
            if not sel:
                return
            vals = tree_nss_fmt.item(sel[0], "values")
            nss_ars_var.set(vals[0])
            if vals[1] and vals[1] != "Sin formato":
                nss_ejemplo_var.set(vals[1])
            probar_ejemplo_nss()

        def probar_ejemplo_nss(*_):
            ejemplo = nss_ejemplo_var.get().strip()
            patron = patron_desde_ejemplo_nss(ejemplo)
            prueba = re.sub(r"\D", "", nss_prueba_sin_guiones_var.get().strip())
            if not patron:
                nss_patron_detectado_var.set("Resultado: escriba el ejemplo con guiones. Ej.: 00896-00258-03")
                return
            resultado = aplicar_patron_nss(prueba, patron) if prueba else ejemplo
            nss_patron_detectado_var.set(f"Patrón detectado: {patron}   |   Vista PDF: {resultado}")

        def guardar_formato_nss():
            ars_key = _limpiar_texto_seguro(nss_ars_var.get())
            ejemplo = nss_ejemplo_var.get().strip()
            patron = patron_desde_ejemplo_nss(ejemplo)

            if not ars_key:
                messagebox.showwarning("Formato NSS", "Seleccione o escriba la ARS.")
                combo_nss_ars.focus_set()
                return

            if not patron:
                messagebox.showwarning(
                    "Formato NSS",
                    "Escriba un ejemplo con guiones para que el sistema detecte el formato.\n\n"
                    "Ejemplo: 00896-00258-03"
                )
                ent_nss_ejemplo.focus_set()
                return

            nss_fmt_data[ars_key] = patron
            guardar_formatos_nss_ars(nss_fmt_data)
            if ars_key not in ars_disponibles:
                ars_disponibles.append(ars_key)
            refrescar_formatos_nss()
            self.set_status("Formato NSS guardado", "ok")
            messagebox.showinfo(
                "Formato NSS",
                f"Formato guardado para {ars_key}.\n\n"
                f"Ejemplo: {ejemplo}\n"
                f"Patrón detectado: {patron}\n\n"
                "Solo afectará el NSS mostrado en PDF."
            )

        def eliminar_formato_nss():
            ars_key = _limpiar_texto_seguro(nss_ars_var.get())
            if not ars_key or ars_key not in nss_fmt_data:
                messagebox.showwarning("Formato NSS", "Seleccione o escriba una ARS configurada.")
                return
            if not messagebox.askyesno("Formato NSS", f"¿Eliminar el formato de NSS para {ars_key}?"):
                return
            nss_fmt_data.pop(ars_key, None)
            guardar_formatos_nss_ars(nss_fmt_data)
            nss_ejemplo_var.set("")
            refrescar_formatos_nss()

        tree_nss_fmt.bind("<<TreeviewSelect>>", seleccionar_formato_nss)
        ent_nss_ejemplo.bind("<KeyRelease>", probar_ejemplo_nss)
        ent_nss_prueba.bind("<KeyRelease>", lambda e: refrescar_formatos_nss())

        barra_nss = tb.Frame(tab_nss, style="Card.TFrame")
        barra_nss.grid(row=6, column=0, columnspan=2, sticky="ew", pady=8)
        tb.Button(barra_nss, text="💾  Guardar formato", bootstyle=SUCCESS, command=guardar_formato_nss, width=18).pack(side="left", padx=5, ipady=5)
        tb.Button(barra_nss, text="🗑️  Eliminar formato", bootstyle=DANGER, command=eliminar_formato_nss, width=18).pack(side="left", padx=5, ipady=5)
        tb.Button(barra_nss, text="↻  Recargar", bootstyle=SECONDARY, command=refrescar_formatos_nss, width=14).pack(side="left", padx=5, ipady=5)

        refrescar_formatos_nss()

        # ---------------- TAB REVISION NSS ----------------
        tab_revision_nss = tb.Frame(notebook, padding=12, style="Card.TFrame")
        notebook.add(tab_revision_nss, text="Revisión NSS")
        tab_revision_nss.columnconfigure(0, weight=1)
        tab_revision_nss.rowconfigure(1, weight=1)
        revision_nss_status = tk.StringVar(
            value="Casos sin cédula cuyo NSS aparece en fichas con datos diferentes."
        )
        tb.Label(
            tab_revision_nss,
            textvariable=revision_nss_status,
            style="Muted.TLabel",
            wraplength=1050,
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        columnas_revision = (
            "caso","nss","ficha_nueva","paciente_nuevo","ficha_existente",
            "paciente_existente","atencion","fecha",
        )
        revision_nss_tree = ttk.Treeview(
            tab_revision_nss,
            columns=columnas_revision,
            show="headings",
            style="Modern.Treeview",
            height=14,
        )
        for columna, titulo, ancho in (
            ("caso","Caso",65),("nss","NSS",130),("ficha_nueva","Ficha nueva",90),
            ("paciente_nuevo","Paciente nuevo",210),("ficha_existente","Ficha existente",105),
            ("paciente_existente","Paciente existente",210),("atencion","Atención",80),
            ("fecha","Fecha",100),
        ):
            revision_nss_tree.heading(columna, text=titulo)
            revision_nss_tree.column(columna, width=ancho, anchor="w")
        revision_nss_tree.grid(row=1, column=0, sticky="nsew")
        revisiones_nss = {}

        def cargar_revisiones_nss():
            revisiones_nss.clear()
            revision_nss_tree.delete(*revision_nss_tree.get_children())
            for row in self.db.listar_revisiones_nss(True, 1000):
                revision_id = int(row["id"])
                revisiones_nss[revision_id] = row
                revision_nss_tree.insert(
                    "","end",iid=str(revision_id),
                    values=(
                        revision_id,row.get("nss_normalizado") or "",
                        row.get("paciente_nuevo_id") or "",row.get("nombre_nuevo") or "",
                        row.get("paciente_referencia_id") or "",row.get("nombre_referencia") or "",
                        row.get("atencion_id") or "",row.get("fecha") or "",
                    ),
                )
            revision_nss_status.set(
                f"{len(revisiones_nss)} caso(s) pendiente(s). Esta revisión nunca detiene la admisión."
            )

        def revision_nss_seleccionada():
            seleccion = revision_nss_tree.selection()
            if not seleccion:
                messagebox.showwarning(
                    "Revisión NSS","Seleccione un caso.",parent=win
                )
                return None
            return revisiones_nss.get(int(seleccion[0]))

        def resolver_nss(tipo, descripcion):
            row = revision_nss_seleccionada()
            if not row:
                return
            motivo = simpledialog.askstring(
                "Motivo administrativo",
                f"Acción: {descripcion}\n\nExplique el criterio utilizado:",
                parent=win,
            )
            motivo = (motivo or "").strip()
            if len(motivo) < 8:
                messagebox.showwarning(
                    "Revisión NSS","El motivo debe tener al menos 8 caracteres.",parent=win
                )
                return
            if not messagebox.askyesno(
                "Confirmar revisión NSS",
                f"Caso #{row['id']} · NSS {row['nss_normalizado']}\n\n{descripcion}\n\n¿Continuar?",
                parent=win,
            ):
                return
            try:
                actor = self._admin_authorized_actor or self._actor_actual()
                self.db.resolver_revision_nss(int(row["id"]),tipo,actor,motivo)
                self.security.audit(
                    "NSS_REVIEW_RESOLVED",actor=actor,success=True,
                    detail=f"revision={row['id']}; resolucion={tipo}",
                )
                cargar_revisiones_nss()
                self._invalidar_caches_datos()
                self.set_status(f"Revisión NSS #{row['id']} resuelta", "ok")
            except Exception as exc:
                APP_LOG.exception("No se pudo resolver la revisión NSS #%s", row["id"])
                messagebox.showerror("Revisión NSS",str(exc),parent=win)

        def abrir_atencion_revision_nss():
            row = revision_nss_seleccionada()
            if not row:
                return
            if not row.get("atencion_id"):
                messagebox.showinfo(
                    "Revisión NSS","Este caso no tiene una atención asociada.",parent=win
                )
                return
            self._abrir_editor_atencion(int(row["atencion_id"]))

        acciones_revision_nss = tb.Frame(tab_revision_nss, style="Card.TFrame")
        acciones_revision_nss.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        tb.Button(
            acciones_revision_nss,text="Conservar ambas fichas",bootstyle=SECONDARY,
            command=lambda: resolver_nss("MANTENER_AMBOS","Conservar las dos fichas con el mismo NSS"),
        ).pack(side="left", padx=4)
        tb.Button(
            acciones_revision_nss,text="Quitar NSS de ficha nueva",bootstyle=WARNING,
            command=lambda: resolver_nss("DESVINCULAR_NSS","Retirar el NSS de la ficha y atención nuevas"),
        ).pack(side="left", padx=4)
        tb.Button(
            acciones_revision_nss,text="Fusionar y eliminar duplicada",bootstyle=DANGER,
            command=lambda: resolver_nss("FUSIONAR_CON_EXISTENTE","Fusionar con la ficha existente y eliminar la ficha duplicada"),
        ).pack(side="left", padx=4)
        tb.Button(
            acciones_revision_nss,text="Abrir atención",
            command=abrir_atencion_revision_nss,
        ).pack(side="left", padx=4)
        tb.Button(
            acciones_revision_nss,text="Actualizar",command=cargar_revisiones_nss
        ).pack(side="right", padx=4)
        cargar_revisiones_nss()

        # ---------------- TAB PREFERENCIAS ----------------
        # ---------------- TAB RESPALDOS ----------------
        tab_backups = tb.Frame(notebook, padding=12, style="Card.TFrame")
        notebook.add(tab_backups, text="Respaldos")
        tab_backups.columnconfigure(0, weight=1)
        tab_backups.rowconfigure(1, weight=1)
        backup_status = tk.StringVar(
            value="Respaldos verificados; se eliminan automáticamente después de 4 días"
        )
        tb.Label(tab_backups, textvariable=backup_status, style="Muted.TLabel").grid(
            row=0, column=0, sticky="w", pady=(0, 8)
        )
        backup_tree = ttk.Treeview(
            tab_backups,
            columns=("fecha", "motivo", "estado"),
            show="headings",
            height=15,
            style="Modern.Treeview",
        )
        backup_tree.heading("fecha", text="Fecha")
        backup_tree.heading("motivo", text="Motivo")
        backup_tree.heading("estado", text="Verificación")
        backup_tree.column("fecha", width=175, anchor="w")
        backup_tree.column("motivo", width=360, anchor="w")
        backup_tree.column("estado", width=130, anchor="center")
        backup_tree.grid(row=1, column=0, sticky="nsew")

        def refrescar_respaldos():
            for item in backup_tree.get_children():
                backup_tree.delete(item)
            for folder in self.db.backup_manager.list_backups():
                try:
                    manifest = self.db.backup_manager.verify(folder)
                    created = str(manifest.get("created_at", "")).replace("T", " ")
                    reason = str(manifest.get("reason", ""))
                    status = "Válido"
                except Exception as exc:
                    created = folder.name[:15]
                    reason = str(exc)
                    status = "Inválido"
                backup_tree.insert("", "end", iid=str(folder), values=(created, reason, status))
            backup_status.set(f"{len(backup_tree.get_children())} respaldo(s) disponible(s)")

        def respaldo_seleccionado():
            selection = backup_tree.selection()
            return selection[0] if selection else ""

        def crear_respaldo_manual():
            try:
                folder = self.db.backup_manager.create(
                    "respaldo_manual", label=f"actor={self._actor_actual()}"
                )
                backup_status.set(f"Respaldo creado y verificado: {os.path.basename(folder)}")
                refrescar_respaldos()
            except Exception as exc:
                APP_LOG.exception("Falló el respaldo manual")
                messagebox.showerror("Respaldos", str(exc), parent=win)

        def verificar_respaldo_ui():
            folder = respaldo_seleccionado()
            if not folder:
                backup_status.set("Seleccione un respaldo para verificar")
                return
            try:
                self.db.backup_manager.verify(folder)
                backup_status.set("El respaldo seleccionado es íntegro y restaurable")
            except Exception as exc:
                backup_status.set(f"Respaldo inválido: {exc}")

        def restaurar_respaldo_ui():
            folder = respaldo_seleccionado()
            if not folder:
                backup_status.set("Seleccione un respaldo para restaurar")
                return
            actor = self._solicitar_autorizacion_admin(
                "RESTAURAR_BASE_DE_DATOS", parent=win, force=True
            )
            if not actor:
                return
            if not messagebox.askyesno(
                "Restaurar base de datos",
                "La base actual se respaldará antes de restaurar. La aplicación se reiniciará.\n\n"
                "ADVERTENCIA: un respaldo anterior puede reintroducir pacientes de prueba "
                "que fueron purgados después de su creación.\n\n"
                "¿Desea continuar?",
                icon="warning",
                parent=win,
            ):
                return
            confirmacion = simpledialog.askstring(
                "Confirmación de restauración",
                "Escriba RESTAURAR para confirmar que comprende el riesgo:",
                parent=win,
            )
            if (confirmacion or "").strip().upper() != "RESTAURAR":
                messagebox.showinfo("Restauración cancelada", "No se modificó la base.", parent=win)
                return
            try:
                self.db.backup_manager.restore_database(folder)
                self.security.audit(
                    "DATABASE_RESTORED", actor=actor, success=True, detail=os.path.basename(folder)
                )
                messagebox.showinfo(
                    "Restauración completada",
                    "La base fue restaurada y verificada. La aplicación se reiniciará.",
                    parent=win,
                )
                command = [sys.executable] + ([] if getattr(sys, "frozen", False) else [os.path.abspath(__file__)])
                subprocess.Popen(command, cwd=os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else BASE_DIR)
                self.root.destroy()
            except Exception as exc:
                APP_LOG.exception("Falló la restauración de la base")
                self.security.audit("DATABASE_RESTORED", actor=actor, success=False, detail=str(exc))
                messagebox.showerror("Restauración", str(exc), parent=win)

        backup_actions = tb.Frame(tab_backups, style="Card.TFrame")
        backup_actions.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        tb.Button(backup_actions, text="Crear respaldo", command=crear_respaldo_manual, bootstyle=SUCCESS).pack(side="left", padx=4)
        tb.Button(backup_actions, text="Verificar", command=verificar_respaldo_ui, bootstyle=INFO).pack(side="left", padx=4)
        tb.Button(backup_actions, text="Restaurar", command=restaurar_respaldo_ui, bootstyle=DANGER).pack(side="left", padx=4)
        refrescar_respaldos()

        # ---------------- TAB USUARIOS ----------------
        tab_usuarios = tb.Frame(notebook, padding=12, style="Card.TFrame")
        notebook.add(tab_usuarios, text="Usuarios")
        tab_usuarios.columnconfigure(0, weight=1)
        tab_usuarios.rowconfigure(2, weight=1)

        usuarios_estado = tk.StringVar(
            value="Administra los nombres disponibles para representar cada turno."
        )
        tb.Label(
            tab_usuarios,
            textvariable=usuarios_estado,
            style="Muted.TLabel",
            background="#0E1B2B",
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        usuario_var = tk.StringVar()
        usuario_form = tb.Frame(tab_usuarios, style="Card.TFrame")
        usuario_form.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        usuario_form.columnconfigure(1, weight=1)
        tb.Label(
            usuario_form,
            text="Nombre del representante:",
            background="#0E1B2B",
        ).grid(row=0, column=0, sticky="w", padx=(0, 8))
        usuario_entry = tb.Entry(usuario_form, textvariable=usuario_var)
        usuario_entry.grid(row=0, column=1, sticky="ew", ipady=4)

        usuarios_tree = ttk.Treeview(
            tab_usuarios,
            columns=("nombre", "estado"),
            show="headings",
            height=13,
            style="Modern.Treeview",
        )
        usuarios_tree.heading("nombre", text="Representante")
        usuarios_tree.heading("estado", text="Uso")
        usuarios_tree.column("nombre", width=520, anchor="w")
        usuarios_tree.column("estado", width=170, anchor="center")
        usuarios_tree.grid(row=2, column=0, sticky="nsew")

        def usuario_actual():
            cfg = cargar_turno_config(permitir_vencido=True) or {}
            nombre = limpiar_nombre_representante(cfg.get("representante", ""))
            return nombre if es_representante_valido(nombre) else ""

        def cargar_usuarios():
            usuarios_tree.delete(*usuarios_tree.get_children())
            actual = usuario_actual()
            nombres = cargar_representantes(self.db)
            for indice, nombre in enumerate(nombres):
                estado = "Turno actual" if nombre.casefold() == actual.casefold() else "Disponible"
                usuarios_tree.insert(
                    "", "end", iid=f"usuario-{indice}", values=(nombre, estado)
                )
            usuarios_estado.set(
                f"{len(nombres)} usuario(s). "
                + (f"Turno actual: {actual}" if actual else "El turno no tiene un representante válido.")
            )

        def usuario_seleccionado():
            seleccion = usuarios_tree.selection()
            if not seleccion:
                return ""
            return str(usuarios_tree.item(seleccion[0], "values")[0]).strip()

        def seleccionar_usuario(_evento=None):
            nombre = usuario_seleccionado()
            if nombre:
                usuario_var.set(nombre)

        def validar_usuario_entrada():
            nombre = limpiar_nombre_representante(usuario_var.get())
            if not es_representante_valido(nombre):
                messagebox.showwarning(
                    "Usuarios",
                    "Escriba un nombre válido. 'No disponible' no se admite.",
                    parent=win,
                )
                usuario_entry.focus_set()
                return ""
            return nombre

        def agregar_usuario():
            nombre = validar_usuario_entrada()
            if not nombre:
                return
            nombres = cargar_representantes(self.db)
            if nombre.casefold() in {item.casefold() for item in nombres}:
                usuarios_estado.set("Ese usuario ya existe; puede seleccionarlo en la lista.")
                return
            nombres.append(nombre)
            if not guardar_catalogo_representantes(nombres):
                messagebox.showerror("Usuarios", "No se pudo guardar el catálogo.", parent=win)
                return
            self.security.audit(
                "SHIFT_USER_ADDED",
                actor=self._admin_authorized_actor or self._actor_actual(),
                success=True,
                detail=nombre,
            )
            usuario_var.set("")
            cargar_usuarios()

        def aplicar_usuario_turno(nombre):
            actualizado = actualizar_representante_turno_actual(self.db, nombre)
            self.security.audit(
                "SHIFT_USER_CHANGED",
                actor=self._admin_authorized_actor or self._actor_actual(),
                success=True,
                detail=f"turno={actualizado['turno_id']}; representante={nombre}",
            )
            self._actualizar_turno_visual_en_vivo()
            self._invalidar_caches_datos()
            cargar_usuarios()
            return actualizado

        def editar_usuario():
            anterior = usuario_seleccionado()
            nuevo = validar_usuario_entrada()
            if not anterior or not nuevo:
                if not anterior:
                    messagebox.showwarning(
                        "Usuarios", "Seleccione el usuario que desea editar.", parent=win
                    )
                return
            nombres = cargar_representantes(self.db)
            if (
                nuevo.casefold() != anterior.casefold()
                and nuevo.casefold() in {item.casefold() for item in nombres}
            ):
                messagebox.showwarning(
                    "Usuarios", "Ya existe otro usuario con ese nombre.", parent=win
                )
                return
            if anterior.casefold() == usuario_actual().casefold():
                try:
                    aplicar_usuario_turno(nuevo)
                except PermissionError:
                    messagebox.showwarning(
                        "Excel abierto",
                        "Cierre el listado de Excel y vuelva a intentarlo. "
                        "No se modificó el usuario activo.",
                        parent=win,
                    )
                    return
                except Exception as exc:
                    APP_LOG.exception("No se pudo editar el usuario del turno")
                    messagebox.showerror("Usuarios", str(exc), parent=win)
                    return
            actualizados = [
                nuevo if item.casefold() == anterior.casefold() else item
                for item in nombres
            ]
            if not guardar_catalogo_representantes(actualizados):
                messagebox.showerror("Usuarios", "No se pudo editar el catálogo.", parent=win)
                return
            self.security.audit(
                "SHIFT_USER_RENAMED",
                actor=self._admin_authorized_actor or self._actor_actual(),
                success=True,
                detail=f"{anterior}->{nuevo}",
            )
            usuario_var.set(nuevo)
            cargar_usuarios()

        def eliminar_usuario():
            nombre = usuario_seleccionado()
            if not nombre:
                messagebox.showwarning(
                    "Usuarios", "Seleccione el usuario que desea eliminar.", parent=win
                )
                return
            if nombre.casefold() == usuario_actual().casefold():
                messagebox.showwarning(
                    "Usuarios",
                    "No puede eliminar el representante del turno actual. "
                    "Seleccione otro usuario para el turno y vuelva a intentarlo.",
                    parent=win,
                )
                return
            if not messagebox.askyesno(
                "Eliminar usuario",
                f"¿Eliminar '{nombre}' de las sugerencias disponibles?\n\n"
                "Los turnos históricos no se modificarán.",
                parent=win,
            ):
                return
            nombres = [
                item for item in cargar_representantes(self.db)
                if item.casefold() != nombre.casefold()
            ]
            if not guardar_catalogo_representantes(nombres):
                messagebox.showerror("Usuarios", "No se pudo guardar el catálogo.", parent=win)
                return
            self.security.audit(
                "SHIFT_USER_REMOVED",
                actor=self._admin_authorized_actor or self._actor_actual(),
                success=True,
                detail=nombre,
            )
            usuario_var.set("")
            cargar_usuarios()

        def usar_usuario_actual():
            nombre = usuario_seleccionado() or validar_usuario_entrada()
            if not nombre:
                return
            if not messagebox.askyesno(
                "Cambiar representante",
                f"¿Usar '{nombre}' en el turno actual?\n\n"
                "Solo cambiarán el encabezado del Excel, los reportes futuros y "
                "el nombre mostrado en la GUI. Los pacientes no se modificarán.",
                parent=win,
            ):
                return
            try:
                aplicar_usuario_turno(nombre)
                messagebox.showinfo(
                    "Usuarios",
                    "Representante actualizado sin modificar pacientes ni atenciones.",
                    parent=win,
                )
            except PermissionError:
                messagebox.showwarning(
                    "Excel abierto",
                    "Cierre el listado de Excel y vuelva a intentarlo. No se cambió el usuario.",
                    parent=win,
                )
            except Exception as exc:
                APP_LOG.exception("No se pudo cambiar el representante del turno")
                messagebox.showerror("Usuarios", str(exc), parent=win)

        usuarios_tree.bind("<<TreeviewSelect>>", seleccionar_usuario)
        usuarios_acciones = tb.Frame(tab_usuarios, style="Card.TFrame")
        usuarios_acciones.grid(row=3, column=0, sticky="ew", pady=(10, 0))
        tb.Button(
            usuarios_acciones, text="Añadir", command=agregar_usuario, bootstyle=SUCCESS
        ).pack(side="left", padx=4)
        tb.Button(
            usuarios_acciones, text="Editar", command=editar_usuario, bootstyle=INFO
        ).pack(side="left", padx=4)
        tb.Button(
            usuarios_acciones, text="Eliminar", command=eliminar_usuario, bootstyle=DANGER
        ).pack(side="left", padx=4)
        tb.Button(
            usuarios_acciones,
            text="Usar en turno actual",
            command=usar_usuario_actual,
            bootstyle=PRIMARY,
        ).pack(side="right", padx=4)
        cargar_usuarios()

        # ---------------- TAB PREFERENCIAS ----------------
        tab_pref = tb.Frame(notebook, padding=12, style="Card.TFrame")
        notebook.add(tab_pref, text="Preferencias")
        pref = dict(self.app_settings)

        pref_nb = ttk.Notebook(tab_pref)
        pref_nb.pack(fill="both", expand=True)

        def add_labeled_combo(parent, row, label, var, values, width=24):
            tb.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=6)
            cb = tb.Combobox(parent, textvariable=var, state="readonly", values=values, width=width)
            cb.grid(row=row, column=1, sticky="w", padx=6, pady=6)
            return cb

        def add_labeled_entry(parent, row, label, var, width=10):
            tb.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=6, pady=6)
            ent = tb.Entry(parent, textvariable=var, width=width)
            ent.grid(row=row, column=1, sticky="w", padx=6, pady=6)
            return ent

        font_var = tk.StringVar(value=str(pref.get("font_size", 11)))
        theme_var = tk.StringVar(value=pref.get("theme", "oscuro"))
        contrast_var = tk.BooleanVar(value=bool(pref.get("high_contrast", False)))
        autosize_var = tk.StringVar(value=pref.get("window_size", "1280x740"))
        auto_print_var = tk.BooleanVar(value=bool(pref.get("auto_print", True)))

        print_auto_hoja_var = tk.BooleanVar(value=bool(pref.get("print_auto_hoja", True)))
        print_auto_reporte_var = tk.BooleanVar(value=bool(pref.get("print_auto_reporte_turno", True)))
        print_auto_excel_var = tk.BooleanVar(value=bool(pref.get("print_auto_excel_turno", True)))
        copies_hoja_var = tk.StringVar(value=str(pref.get("print_copies_hoja", 1)))
        copies_reporte_var = tk.StringVar(value=str(pref.get("print_copies_reporte", 2)))
        copies_excel_var = tk.StringVar(value=str(pref.get("print_copies_excel", 2)))
        pdf_orientation_var = tk.StringVar(value=pref.get("print_pdf_orientation", "Horizontal"))
        excel_orientation_var = tk.StringVar(value=pref.get("print_excel_orientation", "Horizontal"))
        print_behavior_var = tk.StringVar(value=pref.get("print_behavior_hoja", "Imprimir y abrir PDF"))

        confirm_var = tk.BooleanVar(value=bool(pref.get("validation_confirm_before_generate", True)))
        warn_nss_var = tk.BooleanVar(value=bool(pref.get("validation_warn_nss_incomplete", True)))
        warn_ars_var = tk.BooleanVar(value=bool(pref.get("validation_warn_ars_sin_seguro", True)))
        block_short_ars_var = tk.BooleanVar(value=bool(pref.get("validation_block_short_ars", True)))
        allow_cedula_var = tk.BooleanVar(value=bool(pref.get("validation_allow_missing_cedula", True)))
        allow_phone_var = tk.BooleanVar(value=bool(pref.get("validation_allow_missing_phone", False)))
        warn_dup_var = tk.BooleanVar(value=bool(pref.get("validation_warn_duplicate_turno", True)))

        rn_strip_var = tk.BooleanVar(value=bool(pref.get("rn_strip_db", True)))
        rn_show_pdf_var = tk.BooleanVar(value=bool(pref.get("rn_show_pdf", True)))
        rn_warn_var = tk.BooleanVar(value=bool(pref.get("rn_warn", True)))
        rn_format_var = tk.StringVar(value=pref.get("rn_format_display", "RN- NOMBRE DE LA MADRE"))

        button_size_var = tk.StringVar(value=pref.get("button_size", "Normal"))
        table_row_height_var = tk.StringVar(value=str(pref.get("table_row_height", 29)))
        compact_mode_var = tk.BooleanVar(value=bool(pref.get("compact_mode", False)))
        small_screen_var = tk.BooleanVar(value=bool(pref.get("small_screen_mode", False)))
        show_side_panel_var = tk.BooleanVar(value=bool(pref.get("show_side_panel", True)))
        show_summary_var = tk.BooleanVar(value=bool(pref.get("show_turno_summary", True)))
        accent_color_var = tk.StringVar(value=nombre_color_principal(pref.get("accent_color", "Azul hospitalario")))

        hist_initial_var = tk.StringVar(value=str(pref.get("hist_initial_limit", 100)))
        hist_next_var = tk.StringVar(value=str(pref.get("hist_next_limit", 150)))
        hist_filter_var = tk.StringVar(value=pref.get("hist_default_filter", "Todos"))
        hist_order_var = tk.StringVar(value=pref.get("hist_order", "Más reciente primero"))

        turno_default_var = tk.StringVar(value=pref.get("turno_default", "8AM_8AM"))
        ask_rep_var = tk.BooleanVar(value=bool(pref.get("turnos_ask_representante_start", False)))
        gen_report_var = tk.BooleanVar(value=bool(pref.get("turnos_generate_report", True)))
        save_excel_copy_var = tk.BooleanVar(value=bool(pref.get("turnos_save_excel_copy", True)))
        print_empty_report_var = tk.BooleanVar(value=bool(pref.get("turnos_print_empty_report", False)))
        open_archive_var = tk.BooleanVar(value=bool(pref.get("turnos_open_archive_folder", False)))

        pdf_nss_guiones_var = tk.BooleanVar(value=bool(pref.get("pdf_nss_guiones", True)))
        pdf_ars_mode_var = tk.StringVar(value=pref.get("pdf_ars_display_mode", "Abreviada"))
        pdf_nombre_font_var = tk.StringVar(value=str(pref.get("pdf_nombre_font_size", 12)))
        pdf_dir_font_var = tk.StringVar(value=str(pref.get("pdf_direccion_font_size", 12)))
        pdf_open_var = tk.BooleanVar(value=bool(pref.get("pdf_open_after_generate", True)))
        pdf_keep_temp_var = tk.BooleanVar(value=bool(pref.get("pdf_keep_temp", False)))

        p_print = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
        pref_nb.add(p_print, text="Impresión")
        p_print.columnconfigure(1, weight=1)

        tb.Checkbutton(p_print, text="Activar impresión automática (reportes y Excel)", variable=auto_print_var).grid(row=0, column=0, columnspan=2, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_print, text="Permitir impresión operativa de hoja", variable=print_auto_hoja_var).grid(row=1, column=0, columnspan=2, sticky="w", padx=6, pady=6)
        add_labeled_entry(p_print, 2, "Copias (hoja):", copies_hoja_var)
        add_labeled_combo(p_print, 3, "Al generar hoja:", print_behavior_var, ["Solo imprimir", "Imprimir y abrir PDF"], 26)
        
        tb.Checkbutton(p_print, text="Imprimir reporte al cambiar turno", variable=print_auto_reporte_var).grid(row=4, column=0, columnspan=2, sticky="w", padx=6, pady=6)
        add_labeled_entry(p_print, 5, "Copias (reporte):", copies_reporte_var)
        add_labeled_combo(p_print, 6, "Orientación reporte PDF:", pdf_orientation_var, ["Horizontal", "Vertical"], 18)
        
        tb.Checkbutton(p_print, text="Imprimir Excel al cambiar turno", variable=print_auto_excel_var).grid(row=7, column=0, columnspan=2, sticky="w", padx=6, pady=6)
        add_labeled_entry(p_print, 8, "Copias (Excel):", copies_excel_var)
        add_labeled_combo(p_print, 9, "Orientación Excel:", excel_orientation_var, ["Horizontal", "Vertical"], 18)

        p_val = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
        pref_nb.add(p_val, text="Validación")
        checks_val = [
            ("Confirmar antes de generar PDF", confirm_var),
            ("Alertar si NSS parece incompleto", warn_nss_var),
            ("Alertar si ARS se detecta como SIN SEGURO", warn_ars_var),
            ("Bloquear ARS de una letra o solo números", block_short_ars_var),
            ("Permitir paciente sin cédula", allow_cedula_var),
            ("Permitir paciente sin teléfono", allow_phone_var),
            ("Alertar posible duplicado en turno actual", warn_dup_var),
        ]
        for i, (txt, var) in enumerate(checks_val):
            tb.Checkbutton(p_val, text=txt, variable=var).grid(row=i, column=0, sticky="w", padx=6, pady=5)

        p_rn = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
        pref_nb.add(p_rn, text="Recién nacido")
        tb.Checkbutton(p_rn, text="No guardar RN- en la base de datos", variable=rn_strip_var).grid(row=0, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_rn, text="Mostrar RN- solo en el PDF", variable=rn_show_pdf_var).grid(row=1, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_rn, text="Mostrar aviso cuando el nombre tenga RN-", variable=rn_warn_var).grid(row=2, column=0, sticky="w", padx=6, pady=6)
        add_labeled_combo(p_rn, 3, "Formato visual:", rn_format_var, ["RN- NOMBRE DE LA MADRE", "RN DE NOMBRE DE LA MADRE", "RECIÉN NACIDO DE NOMBRE"], 32)

        p_visual = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
        pref_nb.add(p_visual, text="Visual")
        add_labeled_combo(p_visual, 0, "Tamaño de letra:", font_var, ["10", "11", "12", "13", "14", "15", "16", "18"], 18)
        add_labeled_combo(p_visual, 1, "Tema:", theme_var, ["oscuro", "claro"], 18)
        add_labeled_combo(p_visual, 2, "Tamaño de ventana:", autosize_var, ["1220x700", "1280x740", "1366x768", "1440x820", "1600x900"], 18)
        add_labeled_combo(p_visual, 3, "Tamaño de botones:", button_size_var, ["Compacto", "Normal", "Grande"], 18)
        add_labeled_entry(p_visual, 4, "Alto de filas:", table_row_height_var, 10)
        add_labeled_combo(p_visual, 5, "Color principal:", accent_color_var, list(ACCENT_COLOR_PRESETS.keys()), 22)
        tb.Checkbutton(p_visual, text="Modo compacto", variable=compact_mode_var).grid(row=6, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_visual, text="Modo pantalla pequeña", variable=small_screen_var).grid(row=6, column=1, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_visual, text="Modo alto contraste", variable=contrast_var).grid(row=7, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_visual, text="Mostrar panel derecho", variable=show_side_panel_var).grid(row=8, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_visual, text="Mostrar resumen del turno", variable=show_summary_var).grid(row=8, column=1, sticky="w", padx=6, pady=6)

        p_hist = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
        pref_nb.add(p_hist, text="Historial")
        add_labeled_entry(p_hist, 0, "Registros al abrir:", hist_initial_var, 10)
        add_labeled_entry(p_hist, 1, "Cargar más:", hist_next_var, 10)
        add_labeled_combo(p_hist, 2, "Filtro inicial:", hist_filter_var, ["Todos", "Hoy", "Turno actual", "Sin seguro", "Por ARS", "Por especialidad", "Por fecha"], 20)
        add_labeled_combo(p_hist, 3, "Orden:", hist_order_var, ["Más reciente primero", "Más antiguo primero"], 22)

        p_turno = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
        pref_nb.add(p_turno, text="Turnos")
        add_labeled_combo(p_turno, 0, "Turno predeterminado:", turno_default_var, ["8AM_8AM", "8AM_8PM", "8PM_8AM"], 18)
        tb.Checkbutton(p_turno, text="Preguntar representante al iniciar", variable=ask_rep_var).grid(row=1, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_turno, text="Generar reporte al cambiar turno", variable=gen_report_var).grid(row=2, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_turno, text="Guardar copia del Excel al cambiar turno", variable=save_excel_copy_var).grid(row=3, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_turno, text="Imprimir reporte aunque no haya pacientes", variable=print_empty_report_var).grid(row=4, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_turno, text="Abrir carpeta del archivo diario después del cambio", variable=open_archive_var).grid(row=5, column=0, sticky="w", padx=6, pady=6)

        p_pdf = tb.Frame(pref_nb, padding=14, style="Card.TFrame")
        pref_nb.add(p_pdf, text="PDF")
        tb.Checkbutton(p_pdf, text="Mostrar NSS con guiones según ARS", variable=pdf_nss_guiones_var).grid(row=0, column=0, sticky="w", padx=6, pady=6)
        add_labeled_combo(p_pdf, 1, "Mostrar ARS:", pdf_ars_mode_var, ["Abreviada", "Completa"], 18)
        add_labeled_entry(p_pdf, 2, "Letra nombre:", pdf_nombre_font_var, 10)
        add_labeled_entry(p_pdf, 3, "Letra dirección:", pdf_dir_font_var, 10)
        tb.Checkbutton(p_pdf, text="Abrir PDF después de generarlo", variable=pdf_open_var).grid(row=4, column=0, sticky="w", padx=6, pady=6)
        tb.Checkbutton(p_pdf, text="Conservar PDF temporal", variable=pdf_keep_temp_var).grid(row=5, column=0, sticky="w", padx=6, pady=6)

        pref_estado_var = tk.StringVar(value="Organiza aquí impresión, validación, RN, visual, historial, turnos y PDF.")

        def _int_pref(var, default, min_value=1):
            try:
                value = int(str(var.get()).strip())
                return max(min_value, value)
            except Exception:
                return default

        def guardar_pref():
            self.app_settings.update({
                "font_size": _int_pref(font_var, 11, 10),
                "theme": theme_var.get(),
                "high_contrast": bool(contrast_var.get()),
                "window_size": autosize_var.get(),
                "auto_print": bool(auto_print_var.get()),

                "print_auto_hoja": bool(print_auto_hoja_var.get()),
                "print_auto_reporte_turno": bool(print_auto_reporte_var.get()),
                "print_auto_excel_turno": bool(print_auto_excel_var.get()),
                "print_copies_hoja": _int_pref(copies_hoja_var, 1, 1),
                "print_copies_reporte": _int_pref(copies_reporte_var, 2, 1),
                "print_copies_excel": _int_pref(copies_excel_var, 2, 1),
                "print_pdf_orientation": pdf_orientation_var.get(),
                "print_excel_orientation": excel_orientation_var.get(),
                "print_behavior_hoja": print_behavior_var.get(),

                "validation_confirm_before_generate": bool(confirm_var.get()),
                "validation_warn_nss_incomplete": bool(warn_nss_var.get()),
                "validation_warn_ars_sin_seguro": bool(warn_ars_var.get()),
                "validation_block_short_ars": bool(block_short_ars_var.get()),
                "validation_allow_missing_cedula": bool(allow_cedula_var.get()),
                "validation_allow_missing_phone": bool(allow_phone_var.get()),
                "validation_warn_duplicate_turno": bool(warn_dup_var.get()),

                "rn_strip_db": bool(rn_strip_var.get()),
                "rn_show_pdf": bool(rn_show_pdf_var.get()),
                "rn_warn": bool(rn_warn_var.get()),
                "rn_format_display": rn_format_var.get(),

                "button_size": button_size_var.get(),
                "table_row_height": _int_pref(table_row_height_var, 29, 20),
                "compact_mode": bool(compact_mode_var.get()),
                "small_screen_mode": bool(small_screen_var.get()),
                "show_side_panel": bool(show_side_panel_var.get()),
                "show_turno_summary": bool(show_summary_var.get()),
                "accent_color": nombre_color_principal(accent_color_var.get()),

                "hist_initial_limit": _int_pref(hist_initial_var, 100, 50),
                "hist_next_limit": _int_pref(hist_next_var, 150, 50),
                "hist_default_filter": hist_filter_var.get(),
                "hist_order": hist_order_var.get(),

                "turno_default": normalizar_turno_codigo(turno_default_var.get()),
                "turnos_ask_representante_start": bool(ask_rep_var.get()),
                "turnos_generate_report": bool(gen_report_var.get()),
                "turnos_save_excel_copy": bool(save_excel_copy_var.get()),
                "turnos_print_empty_report": bool(print_empty_report_var.get()),
                "turnos_open_archive_folder": bool(open_archive_var.get()),

                "pdf_nss_guiones": bool(pdf_nss_guiones_var.get()),
                "pdf_ars_display_mode": pdf_ars_mode_var.get(),
                "pdf_nombre_font_size": _int_pref(pdf_nombre_font_var, 12, 6),
                "pdf_direccion_font_size": _int_pref(pdf_dir_font_var, 12, 6),
                "pdf_open_after_generate": bool(pdf_open_var.get()),
                "pdf_keep_temp": bool(pdf_keep_temp_var.get()),
            })

            if not guardar_app_settings(self.app_settings):
                messagebox.showwarning("Preferencias", "No se pudieron guardar las preferencias en disco.")
                return

            self._aplicar_preferencias_en_vivo(win)
            self._actualizar_resumen_turno_panel()
            pref_estado_var.set("Preferencias guardadas y aplicadas.")
            self.set_status("Preferencias aplicadas", "ok")
            messagebox.showinfo(
                "Preferencias",
                "Preferencias guardadas correctamente. Algunos cambios visuales estructurales pueden requerir cerrar y abrir la app."
            )

        bottom_pref = tb.Frame(tab_pref, padding=(8, 10), style="Card.TFrame")
        bottom_pref.pack(fill="x", side="bottom", pady=(10, 0))
        tb.Label(bottom_pref, textvariable=pref_estado_var, style="Muted.TLabel").pack(side="left", padx=5)
        tb.Button(bottom_pref, text="💾  Guardar preferencias", bootstyle=SUCCESS, command=guardar_pref, width=22).pack(side="right", padx=5, ipady=5)



    def mostrar_menu_contextual(self, event):
        try:
            self._widget_actual = event.widget
            self.menu_contextual.post(event.x_root, event.y_root)
        except Exception:
            pass

    def _copiar(self):
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(self._widget_actual.selection_get())
        except Exception:
            pass

    def _pegar(self):
        try:
            self._widget_actual.insert(tk.INSERT, self.root.clipboard_get())
        except Exception:
            pass

    def _cortar(self):
        try:
            sel = self._widget_actual.selection_get()
            self.root.clipboard_clear()
            self.root.clipboard_append(sel)
            self._widget_actual.delete("sel.first", "sel.last")
        except Exception:
            pass

    def reiniciar_datos_excel(self, *_):
        if not messagebox.askyesno("Confirmación", "¿Está seguro de que desea cambiar de turno y reiniciar los datos del Excel?"):
            return
        self._dialogo_turno()

    def _dialogo_turno(self):
        win = self._crear_toplevel_estable("Configurar Turno y Encabezado", "680x540", "turno_win")
        if win is None:
            return

        self._bind_esc_cerrar(win)

        cont = tb.Frame(win, padding=14, style="Root.TFrame")
        cont.pack(fill="both", expand=True)

        self._crear_header_ventana(
            cont,
            "Configurar Encabezado y Turno",
            "Cambia el turno, guarda e imprime el reporte saliente y reconstruye el listado del nuevo turno.",
            "⚙"
        )

        form_card = tb.Frame(cont, padding=12, style="Card.TFrame")
        form_card.pack(fill="both", expand=True)

        tb.Label(form_card, text="Representante:", background="#0E1B2B", foreground="#EAF2FF").grid(row=0, column=0, sticky="w", pady=6)
        turno_actual = cargar_turno_config() or {}
        representantes = cargar_representantes(self.db)
        rep_var = tk.StringVar(value=limpiar_nombre_representante(turno_actual.get("representante", "")))
        rep_box = tb.Frame(form_card, style="Card.TFrame")
        rep_box.grid(row=0, column=1, sticky="ew", pady=6)
        rep_entry = tb.Entry(rep_box, textvariable=rep_var, width=36)
        rep_entry.pack(fill="x")
        similitudes = tk.Listbox(
            rep_box,
            height=3,
            bg="#0B1624",
            fg="#EAF2FF",
            selectbackground="#1D6EFF",
            selectforeground="#FFFFFF",
            highlightthickness=1,
            highlightbackground="#254260",
            relief="flat",
        )

        def filtrar_representantes(_evento=None):
            texto = rep_var.get().strip().casefold()
            coincidencias = [
                nombre
                for nombre in representantes
                if texto and texto in nombre.casefold()
                and nombre.casefold() != texto
            ][:5]
            similitudes.delete(0, tk.END)
            for nombre in coincidencias:
                similitudes.insert(tk.END, nombre)
            if coincidencias:
                similitudes.pack(fill="x", pady=(3, 0))
            else:
                similitudes.pack_forget()

        def elegir_similitud(_evento=None):
            seleccion = similitudes.curselection()
            if not seleccion:
                return
            rep_var.set(similitudes.get(seleccion[0]))
            similitudes.pack_forget()
            rep_entry.icursor(tk.END)
            rep_entry.focus_set()

        rep_entry.bind("<KeyRelease>", filtrar_representantes)
        similitudes.bind("<ButtonRelease-1>", elegir_similitud)
        similitudes.bind("<Return>", elegir_similitud)

        tb.Label(form_card, text="Turno:", background="#0E1B2B", foreground="#EAF2FF").grid(row=1, column=0, sticky="w", pady=6)

        turno_var = tk.StringVar(value=self.app_settings.get("turno_default", "8AM_8AM"))
        combo_turno = tb.Combobox(
            form_card,
            textvariable=turno_var,
            state="readonly",
            values=[
                "8:00 AM → 8:00 AM",
                "8:00 AM → 8:00 PM",
                "8:00 PM → 8:00 AM",
            ],
            width=30
        )
        combo_turno.grid(row=1, column=1, sticky="w", pady=6)
        _td = normalizar_turno_codigo(self.app_settings.get("turno_default", "8AM_8AM"))
        combo_turno.set("8:00 AM → 8:00 PM" if _td == "8AM_8PM" else ("8:00 PM → 8:00 AM" if _td == "8PM_8AM" else "8:00 AM → 8:00 AM"))

        aviso_var = tk.StringVar(value="")
        aviso_lbl = tb.Label(form_card, textvariable=aviso_var, bootstyle=INFO, wraplength=470, justify="left")
        aviso_lbl.grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 10))

        vista_turno_var = tk.StringVar(value="")
        vista_fecha_var = tk.StringVar(value="")
        vista_inicio_real_var = tk.StringVar(value="")

        tb.Label(form_card, text="Vista previa del turno:", bootstyle=SECONDARY, background="#0E1B2B").grid(row=3, column=0, sticky="w", pady=(5, 2))
        tb.Label(form_card, textvariable=vista_turno_var, font=("Arial", 11, "bold"), background="#0E1B2B", foreground="#FFFFFF").grid(row=3, column=1, sticky="w", pady=(5, 2))

        tb.Label(form_card, text="Vista previa de fecha:", bootstyle=SECONDARY, background="#0E1B2B").grid(row=4, column=0, sticky="w", pady=(2, 2))
        tb.Label(form_card, textvariable=vista_fecha_var, font=("Arial", 11), background="#0E1B2B", foreground="#EAF2FF").grid(row=4, column=1, sticky="w", pady=(2, 2))

        tb.Label(form_card, text="Conteo real desde:", bootstyle=SECONDARY, background="#0E1B2B").grid(row=5, column=0, sticky="w", pady=(2, 8))
        tb.Label(form_card, textvariable=vista_inicio_real_var, font=("Arial", 11), background="#0E1B2B", foreground="#EAF2FF").grid(row=5, column=1, sticky="w", pady=(2, 8))

        fecha_base = datetime.now().date()

        def es_domingo(f: date) -> bool:
            return f.weekday() == 6

        def es_sabado(f: date) -> bool:
            return f.weekday() == 5

        def normalizar_turno_desde_combo(texto_combo: str) -> str:
            mapa = {
                "8:00 AM → 8:00 AM": "8AM_8AM",
                "8:00 AM → 8:00 PM": "8AM_8PM",
                "8:00 PM → 8:00 AM": "8PM_8AM",
            }
            return mapa.get(texto_combo, "8AM_8AM")

        def actualizar_vista_previa():
            turno_codigo = normalizar_turno_desde_combo(combo_turno.get())
            if not es_domingo(fecha_base) and not es_sabado(fecha_base):
                turno_codigo = "8AM_8AM"
            datos_turno = obtener_datos_turno_visual(fecha_base, turno_codigo)
            vista_turno_var.set(datos_turno["turno_label"])
            vista_fecha_var.set(datos_turno["fecha_label"])
            vista_inicio_real_var.set(datetime.now().strftime("%d/%m/%Y %I:%M %p"))

        def refrescar_turnos():
            if es_domingo(fecha_base) or es_sabado(fecha_base):
                combo_turno.configure(values=[
                    "8:00 AM → 8:00 AM",
                    "8:00 AM → 8:00 PM",
                    "8:00 PM → 8:00 AM",
                ])
                aviso_var.set(
                    "Hoy es domingo o sábado. Están habilitados el turno completo y los dos turnos divididos."
                )
            else:
                combo_turno.configure(values=["8:00 AM → 8:00 AM"])
                combo_turno.set("8:00 AM → 8:00 AM")
                aviso_var.set(
                    "Hoy no es domingo ni sábado. Solo está permitido el turno completo 8:00 AM → 8:00 AM del día siguiente."
                )
            actualizar_vista_previa()

        combo_turno.bind("<<ComboboxSelected>>", lambda e: actualizar_vista_previa())
        refrescar_turnos()

        def aplicar():
            representante = limpiar_nombre_representante(rep_var.get())
            if not es_representante_valido(representante):
                messagebox.showwarning(
                    "Representante",
                    "Escriba un nombre válido. 'No disponible' no puede guardarse.",
                    parent=win,
                )
                rep_entry.focus_set()
                return

            turno_codigo = normalizar_turno_desde_combo(combo_turno.get())
            if not es_domingo(fecha_base) and not es_sabado(fecha_base):
                turno_codigo = "8AM_8AM"

            momento_cambio = datetime.now()
            turno_saliente = cargar_turno_config(permitir_vencido=True)

            mismo_turno = bool(
                turno_saliente
                and turno_saliente.get("fecha_base") == fecha_base
                and normalizar_turno_codigo(turno_saliente.get("turno_codigo"))
                == turno_codigo
            )
            representante_anterior = limpiar_nombre_representante(
                (turno_saliente or {}).get("representante", "")
            )
            if (
                mismo_turno
                and representante.casefold() != representante_anterior.casefold()
            ):
                try:
                    actualizado = actualizar_representante_turno_actual(
                        self.db, representante
                    )
                    self.security.audit(
                        "SHIFT_USER_CORRECTED",
                        actor=self._admin_authorized_actor or self._actor_actual(),
                        success=True,
                        detail=(
                            f"turno={actualizado['turno_id']}; "
                            f"{representante_anterior}->{representante}"
                        ),
                    )
                    self._actualizar_turno_visual_en_vivo()
                    messagebox.showinfo(
                        "Representante actualizado",
                        "Se corrigió el representante en la GUI, el encabezado "
                        "del Excel y los reportes futuros.\n\n"
                        "No se reinició el turno ni se modificaron pacientes.",
                        parent=win,
                    )
                    self.turno_win = None
                    win.destroy()
                    return
                except PermissionError:
                    messagebox.showwarning(
                        "Excel abierto",
                        "Cierre el listado de Excel y vuelva a intentarlo. "
                        "El turno y sus datos permanecen sin cambios.",
                        parent=win,
                    )
                    return
                except Exception as exc:
                    APP_LOG.exception("No se pudo corregir el representante del turno")
                    messagebox.showerror(
                        "Representante", str(exc), parent=win
                    )
                    return

            try:
                self.db.backup_manager.create(
                    "cierre_turno",
                    label=f"representante={representante}; momento={momento_cambio.isoformat(timespec='seconds')}",
                )
            except Exception as exc:
                APP_LOG.exception("No se pudo crear el respaldo previo al cambio de turno")
                messagebox.showerror(
                    "Respaldo requerido",
                    f"No se cambiará el turno porque falló el respaldo previo:\n{exc}",
                    parent=win,
                )
                return

            if turno_saliente and bool(self.app_settings.get("turnos_generate_report", True)):
                try:
                    self._generar_y_abrir_reporte_turno(turno_saliente, fin_corte=momento_cambio)
                except Exception as e:
                    messagebox.showwarning(
                        "Aviso",
                        f"No se pudo generar/imprimir automáticamente el reporte del turno saliente:\n{str(e)}"
                    )

            if turno_saliente and bool(self.app_settings.get("turnos_save_excel_copy", True)):
                try:
                    guardar_copia_excel_turno(turno_saliente, EXCEL_PATH)
                except Exception as e:
                    messagebox.showwarning(
                        "Archivo diario",
                        f"No se pudo guardar la copia clasificada del Excel antes del cambio de turno:\n{str(e)}"
                    )

            if self.app_settings.get("auto_print", True) and bool(self.app_settings.get("print_auto_excel_turno", True)):
                try:
                    imprimir_excel(EXCEL_PATH, copias=max(1, int(self.app_settings.get("print_copies_excel", 2) or 2)))
                except Exception as e:
                    messagebox.showwarning(
                        "Aviso",
                        f"No se pudo imprimir automáticamente el listado de Excel antes del cambio de turno:\n{str(e)}"
                    )

            candidato = {
                "representante": representante,
                "turno_codigo": turno_codigo,
                "fecha_base": fecha_base,
                "inicio_real": format_datetime_local(momento_cambio),
                "inicio_real_dt": momento_cambio,
            }
            if not turno_config_es_vigente(candidato, momento=momento_cambio):
                messagebox.showwarning(
                    "Turno no vigente",
                    "El horario seleccionado ya finalizó o todavía no comienza para el día operativo actual.",
                    parent=win,
                )
                return
            if not guardar_turno_config(
                representante, turno_codigo, fecha_base, inicio_real=momento_cambio
            ):
                messagebox.showerror(
                    "Turno",
                    "No se pudo guardar la configuración. El turno anterior permanece sin cambios.",
                    parent=win,
                )
                return
            if turno_saliente:
                self.db.cerrar_turno_existente(turno_saliente, momento_cambio)
            guardar_representante_catalogo(representante, self.db)

            turno_cfg_nuevo = cargar_turno_config()
            if not turno_cfg_nuevo:
                raise TurnoNoVigenteError("El turno guardado no quedó vigente.")
            self.db.obtener_o_crear_turno(turno_cfg_nuevo)
            reconstruir_excel_turno(self.db, turno_cfg_nuevo)

            self._actualizar_turno_visual_en_vivo()

            messagebox.showinfo(
                "Listo",
                "Turno actualizado, archivos clasificados guardados, Excel reiniciado y reporte del turno saliente generado.\n\n"
                "El conteo del nuevo turno comenzará desde la hora real del cambio."
            )

            if bool(self.app_settings.get("turnos_open_archive_folder", False)):
                try:
                    carpeta = carpeta_archivo_turno(turno_saliente) if turno_saliente else ARCHIVO_DIARIO_DIR
                    if platform.system() == "Windows":
                        os.startfile(carpeta)
                    elif platform.system() == "Darwin":
                        subprocess.run(["open", carpeta], check=False)
                    else:
                        subprocess.run(["xdg-open", carpeta], check=False)
                except Exception:
                    pass

            try:
                self.turno_win = None
                win.destroy()
            except Exception:
                self.turno_win = None

        tb.Button(form_card, text="Aplicar", bootstyle=SUCCESS, command=aplicar).grid(row=6, column=0, pady=12)
        tb.Button(form_card, text="Cancelar", bootstyle=SECONDARY, command=win.destroy).grid(row=6, column=1, sticky="w", pady=12)

    def run(self):
        self.root.mainloop()


# -------------------------------
# MAIN
# -------------------------------
if __name__ == "__main__":
    try:
        if SELF_TEST_MODE:
            manager = DatabaseManager()
            with closing(manager._connect()) as conn:
                integrity = conn.execute("PRAGMA integrity_check").fetchone()
                version = conn.execute("SELECT version FROM schema_version WHERE id=1").fetchone()
            if not integrity or integrity[0] != "ok" or not version or int(version[0]) != LATEST_SCHEMA_VERSION:
                raise RuntimeError("La autoprueba de SQLite no fue satisfactoria.")
            missing_templates = [path for path in RUTA_HOJAS.values() if not os.path.isfile(path)]
            if missing_templates:
                raise FileNotFoundError("Faltan plantillas: " + ", ".join(missing_templates))
            packaged_logo = resource_path("istipo_hospitales.png")
            if not os.path.isfile(packaged_logo):
                raise FileNotFoundError("El logo principal no quedó incluido en el ejecutable.")
            raise SystemExit(0)
        app = App()
        app.run()
    except Exception as exc:
        APP_LOG.exception("Error fatal durante el inicio de la aplicación")
        try:
            messagebox.showerror("Error al iniciar", f"No se pudo iniciar la aplicación:\n{str(exc)}")
        except Exception:
            pass
        raise
    finally:
        if SELF_TEST_DATA_DIR:
            shutil.rmtree(SELF_TEST_DATA_DIR, ignore_errors=True)
